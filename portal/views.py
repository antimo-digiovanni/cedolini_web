import os
import csv
import json
import io
import math
import calendar
from copy import deepcopy
import re
import tempfile
import zipfile
from urllib.parse import parse_qsl, quote, urlencode, urlsplit, urlunsplit
import uuid
import unicodedata
from collections import OrderedDict
from datetime import date
from datetime import datetime
from datetime import timedelta
from decimal import Decimal
from pathlib import Path
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse, FileResponse, HttpResponseRedirect, HttpResponsePermanentRedirect
from django.db import transaction, IntegrityError
from django.db.models import Count, Q, Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from django.conf import settings
from django.core.files.base import File
from django.core.files.storage import default_storage
from django.core.files.storage import FileSystemStorage
from django.core.mail import EmailMultiAlternatives
from django.utils import timezone
from django.templatetags.static import static
from django.urls import reverse

from .models import (
    Employee,
    Payslip,
    PayslipView,
    ImportJob,
    InviteToken,
    Cud,
    CudView,
    WorkZone,
    EmployeeWorkZone,
    WorkSession,
    WorkMarkRequest,
    VacationRequest,
    TurniPlannerWeekState,
    PersonalAssetEntry,
)
from .models import AuditEvent
from django.core.paginator import Paginator
from .forms import PersonalAssetEntryForm

import logging
import secrets
from PIL import Image

from turni_app.pdf_export import (
    JOLLY_WEEKEND_IMAGE_NAME,
    JOLLY_WEEKEND_PDF_NAME,
    PORTINERIA_WEEKEND_IMAGE_NAME,
    PORTINERIA_WEEKEND_PDF_NAME,
    PORTINERIA_WEEKLY_IMAGE_NAME,
    PORTINERIA_WEEKLY_PDF_NAME,
    SCORRIMENTO_IMAGE_NAME,
    SCORRIMENTO_PDF_NAME,
    SATURDAY_IMAGE_NAME,
    SATURDAY_PDF_NAME,
    SUNDAY_IMAGE_NAME,
    SUNDAY_PDF_NAME,
    ScorrimentoExportData,
    WEEKLY_IMAGE_NAME,
    WEEKLY_PDF_NAME,
    WeekendExportData,
    export_scorrimento_images,
    export_scorrimento_pdf,
    export_weekend_images,
    export_weekend_pdf,
    export_weekly_images,
    export_weekly_pdf,
)
from turni_app.workbook import WeeklySectionData

from .utils_import import parse_payslip_filename
from .access import (
    PATRIMONIO_GROUP_NAME,
    RICONFEZIONAMENTO_GROUP_NAME,
    user_has_full_admin_access,
    user_has_patrimonio_access,
    user_has_riconfezionamento_access,
    user_has_turni_planner_access,
    user_has_today_markings_access,
    user_has_today_markings_only_access,
    user_home_url_name,
)

logger = logging.getLogger(__name__)


MONTH_LABELS_IT = {
    1: 'Gennaio',
    2: 'Febbraio',
    3: 'Marzo',
    4: 'Aprile',
    5: 'Maggio',
    6: 'Giugno',
    7: 'Luglio',
    8: 'Agosto',
    9: 'Settembre',
    10: 'Ottobre',
    11: 'Novembre',
    12: 'Dicembre',
}

MAX_SHIFT_DURATION_HOURS = 18


def _turni_planner_allowed_or_403(request):
    if not user_has_turni_planner_access(request.user):
        return HttpResponse('Turni Planner non disponibile per questo account.', status=403)
    return None


def _riconfezionamento_allowed_or_403(request):
    if not settings.RICONFEZIONAMENTO_ONLINE_ENABLED:
        return HttpResponse('Riconfezionamento online non disponibile.', status=404)
    if not user_has_riconfezionamento_access(request.user):
        return HttpResponse('Riconfezionamento non disponibile per questo account.', status=403)
    return None


def _patrimonio_allowed_or_403(request):
    if not user_has_patrimonio_access(request.user):
        return HttpResponse('Gestione patrimonio non disponibile per questo account.', status=403)
    return None


def _personal_asset_history_queryset(user):
    return PersonalAssetEntry.objects.filter(user=user).order_by('-occurred_on', '-created_at', '-id')


def _personal_asset_reimbursement_entries_queryset(user):
    return (
        PersonalAssetEntry.objects
        .filter(user=user, operation_type=PersonalAssetEntry.TYPE_REIMBURSABLE_EXPENSE)
        .order_by('-occurred_on', '-created_at', '-id')
    )


def _personal_asset_report_display_name(user):
    employee = getattr(user, 'employee', None)
    if employee and employee.full_name:
        return employee.full_name
    full_name = user.get_full_name().strip()
    if full_name:
        return full_name
    return user.get_username()


def _personal_asset_reimbursement_report_email_recipients():
    recipients = getattr(settings, 'EXPENSE_REIMBURSEMENT_EMAILS', None)
    if recipients:
        return [email for email in recipients if email]
    return list(getattr(settings, 'ADMIN_NOTIFICATION_EMAILS', []))


def _build_personal_asset_reimbursement_report_image(user):
    from PIL import ImageDraw, ImageFont

    entries = list(_personal_asset_reimbursement_entries_queryset(user))
    display_name = _personal_asset_report_display_name(user)
    total_amount = sum((entry.reimbursement_amount or entry.amount or Decimal('0.00')) for entry in entries)
    logo_candidates = [
        Path(settings.BASE_DIR) / 'riconfezionamento_app' / 'static' / 'assets' / 'logo-san-vincenzo.png',
        Path(settings.BASE_DIR) / 'portal' / 'static' / 'portal' / 'logo.png',
    ]

    def load_font(size, bold=False):
        candidates = [
            Path('C:/Windows/Fonts/arialbd.ttf'),
            Path('C:/Windows/Fonts/Arialbd.ttf'),
            Path('C:/Windows/Fonts/segoeuib.ttf'),
            Path('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
            Path('/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf'),
            'arialbd.ttf',
            'Arial Bold.ttf',
        ] if bold else [
            Path('C:/Windows/Fonts/arial.ttf'),
            Path('C:/Windows/Fonts/Arial.ttf'),
            Path('C:/Windows/Fonts/segoeui.ttf'),
            Path('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'),
            Path('/usr/share/fonts/dejavu/DejaVuSans.ttf'),
            'arial.ttf',
            'Arial.ttf',
        ]
        for candidate in candidates:
            try:
                return ImageFont.truetype(str(candidate), size)
            except OSError:
                continue
        return ImageFont.load_default()

    def format_amount(value):
        return f'{value:.2f}'.replace('.', ',') + ' €'

    def format_date(value):
        return f'{value.day}/{value.month}/{value.year}'

    def text_height(font):
        bbox = font.getbbox('Ag')
        return bbox[3] - bbox[1]

    def draw_right_text(x_right, y, text, font, fill):
        box = draw.textbbox((0, 0), text, font=font)
        text_width = box[2] - box[0]
        draw.text((x_right - text_width, y), text, fill=fill, font=font)

    logo_image = None
    for logo_path in logo_candidates:
        if not logo_path.exists():
            continue
        try:
            logo_image = Image.open(logo_path).convert('RGBA')
            logo_image.thumbnail((88, 32))
            break
        except OSError:
            continue

    width = 1400
    margin = 24
    title_row_height = 60
    header_row_height = 46
    row_height = 50
    total_row_height = 50
    calculated_height = margin + title_row_height + header_row_height + (len(entries) * row_height) + total_row_height + margin
    height = max(calculated_height, margin + title_row_height + header_row_height + total_row_height + margin)

    image = Image.new('RGB', (width, height), '#ffffff')
    draw = ImageDraw.Draw(image)

    grid_color = '#111111'
    total_fill = '#12c25b'
    text_dark = '#111111'

    title_font = load_font(26, bold=False)
    header_font = load_font(20, bold=True)
    body_font = load_font(19)
    total_font = load_font(22, bold=True)

    table_top = margin
    date_col_width = 175
    amount_col_width = 220
    desc_col_width = width - (margin * 2) - date_col_width - amount_col_width
    x_date = margin
    x_desc = x_date + date_col_width
    x_amount = x_desc + desc_col_width

    title_text = f'Rimborso spese {display_name.split()[0] if display_name else ""}'.strip()
    draw.rectangle((margin, table_top, width - margin, table_top + title_row_height), outline=grid_color, width=2)
    if logo_image is not None:
        logo_y = table_top + max(0, (title_row_height - logo_image.height) // 2)
        image.paste(logo_image, (margin + 10, logo_y), logo_image)
    title_box = draw.textbbox((0, 0), title_text, font=title_font)
    title_width = title_box[2] - title_box[0]
    draw.text(((width - title_width) / 2, table_top + 11), title_text, fill=text_dark, font=title_font)

    header_top = table_top + title_row_height
    draw.rectangle((margin, header_top, width - margin, header_top + header_row_height), outline=grid_color, width=2)
    draw.line((x_desc, header_top, x_desc, header_top + header_row_height), fill=grid_color, width=2)
    draw.line((x_amount, header_top, x_amount, header_top + header_row_height), fill=grid_color, width=2)
    draw.text((x_date + 16, header_top + 10), 'DATA', fill=text_dark, font=header_font)
    draw.text((x_desc + 16, header_top + 10), 'SPESA', fill=text_dark, font=header_font)
    draw_right_text(width - margin - 16, header_top + 10, 'IMPORTO', header_font, text_dark)

    current_y = header_top + header_row_height
    for index, entry in enumerate(entries):
        description = (entry.description or '').strip() or entry.category
        report_amount = entry.reimbursement_amount or entry.amount or Decimal('0.00')
        draw.rectangle((margin, current_y, width - margin, current_y + row_height), outline=grid_color, width=1)
        draw.line((x_desc, current_y, x_desc, current_y + row_height), fill=grid_color, width=1)
        draw.line((x_amount, current_y, x_amount, current_y + row_height), fill=grid_color, width=1)
        draw.text((x_date + 10, current_y + 11), format_date(entry.occurred_on), fill=text_dark, font=body_font)
        draw.text((x_desc + 10, current_y + 11), description, fill=text_dark, font=body_font)
        draw_right_text(width - margin - 16, current_y + 11, format_amount(report_amount), body_font, text_dark)
        current_y += row_height

    total_top = current_y
    draw.rectangle((margin, total_top, width - margin, total_top + total_row_height), fill=total_fill, outline=grid_color, width=2)
    draw.line((x_amount, total_top, x_amount, total_top + total_row_height), fill=grid_color, width=2)
    draw.text((x_desc + 14, total_top + 10), 'TOTALE RIMBORSO SPESE', fill=text_dark, font=total_font)
    total_text = format_amount(total_amount)
    draw_right_text(width - margin - 16, total_top + 10, total_text, total_font, text_dark)

    output = io.BytesIO()
    image.save(output, format='JPEG', quality=94)
    output.seek(0)
    return output.getvalue(), entries, total_amount


def _build_personal_asset_reimbursement_report_response(user):
    image_bytes, _, _ = _build_personal_asset_reimbursement_report_image(user)
    safe_username = re.sub(r'[^a-zA-Z0-9_-]+', '_', user.get_username())
    filename = f'rimborso_spese_{safe_username}.jpg'
    response = HttpResponse(content_type='image/jpeg')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(image_bytes)
    return response, filename


def _send_personal_asset_reimbursement_report_email(user):
    image_bytes, entries, total_amount = _build_personal_asset_reimbursement_report_image(user)
    recipients = _personal_asset_reimbursement_report_email_recipients()
    if not recipients:
        return False, 'Nessun destinatario configurato per l\'invio del rimborso spese.'

    safe_username = re.sub(r'[^a-zA-Z0-9_-]+', '_', user.get_username())
    filename = f'rimborso_spese_{safe_username}.jpg'

    display_name = _personal_asset_report_display_name(user)
    subject = f'Rimborso spese {display_name}'
    text_content = (
        f'In allegato trovi il report rimborso spese di {display_name}.\n\n'
        f'Voci incluse: {len(entries)}\n'
        f'Totale da rimborsare: {total_amount:.2f} EUR\n'
    )

    email = EmailMultiAlternatives(
        subject,
        text_content,
        settings.DEFAULT_FROM_EMAIL,
        recipients,
    )
    email.attach(filename, image_bytes, 'image/jpeg')
    email.send(fail_silently=False)
    return True, f'Report rimborso spese inviato a: {", ".join(recipients)}.'


def _personal_asset_summary(user, *, reference_date=None):
    reference_date = reference_date or timezone.localdate()
    zero = Decimal('0.00')
    decimal_zero = Value(zero, output_field=DecimalField(max_digits=12, decimal_places=2))
    history_qs = _personal_asset_history_queryset(user)

    totals = history_qs.aggregate(
        account_balance=Coalesce(Sum('account_delta'), decimal_zero),
        piggy_bank_balance=Coalesce(Sum('piggy_bank_delta'), decimal_zero),
        reimbursement_balance=Coalesce(Sum('reimbursement_delta'), decimal_zero),
        advance_balance=Coalesce(Sum('advance_delta'), decimal_zero),
    )

    monthly_qs = history_qs.filter(
        occurred_on__year=reference_date.year,
        occurred_on__month=reference_date.month,
    )
    monthly_totals = monthly_qs.aggregate(
        income=Coalesce(
            Sum(
                'amount',
                filter=Q(operation_type__in=[
                    PersonalAssetEntry.TYPE_INCOME,
                    PersonalAssetEntry.TYPE_REIMBURSEMENT_RECEIVED,
                ]),
            ),
            decimal_zero,
        ),
        expense=Coalesce(
            Sum(
                'amount',
                filter=Q(operation_type__in=[
                    PersonalAssetEntry.TYPE_EXPENSE,
                    PersonalAssetEntry.TYPE_REIMBURSABLE_EXPENSE,
                ]),
            ),
            decimal_zero,
        ),
        account_delta=Coalesce(Sum('account_delta'), decimal_zero),
        piggy_bank_delta=Coalesce(Sum('piggy_bank_delta'), decimal_zero),
        reimbursement_delta=Coalesce(Sum('reimbursement_delta'), decimal_zero),
    )

    total_assets = (
        totals['account_balance']
        + totals['piggy_bank_balance']
        + totals['reimbursement_balance']
    )
    monthly_saving = (
        monthly_totals['account_delta']
        + monthly_totals['piggy_bank_delta']
        + monthly_totals['reimbursement_delta']
    )

    return {
        'account_balance': totals['account_balance'],
        'piggy_bank_balance': totals['piggy_bank_balance'],
        'reimbursement_balance': totals['reimbursement_balance'],
        'advance_balance': totals['advance_balance'],
        'total_assets': total_assets,
        'monthly_income': monthly_totals['income'],
        'monthly_expense': monthly_totals['expense'],
        'monthly_saving': monthly_saving,
        'month_label': f"{MONTH_LABELS_IT[reference_date.month]} {reference_date.year}",
    }


def _turni_planner_data_has_content(value):
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, dict):
        return any(_turni_planner_data_has_content(item) for item in value.values())
    if isinstance(value, (list, tuple)):
        return any(_turni_planner_data_has_content(item) for item in value)
    return False


def _turni_planner_initial_data_for_new_week(*, week_label):
    previous_states = TurniPlannerWeekState.objects.exclude(week_label=week_label).order_by('-updated_at', '-id')
    for previous_state in previous_states:
        planner_data = previous_state.planner_data or {}
        if _turni_planner_data_has_content(planner_data):
            cloned_data = deepcopy(planner_data)
            cloned_data['weekly_export_week_label'] = week_label
            cloned_data['portineria_weekly_export_week_label'] = week_label
            return cloned_data
    return {}


def _resolve_turni_export_week_label(planner_data, week_label, *, key):
    if isinstance(planner_data, dict):
        custom_label = str(planner_data.get(key) or '').strip()
        if custom_label:
            return custom_label
        custom_label = str(planner_data.get('export_week_label') or '').strip()
        if custom_label:
            return custom_label
    return week_label


TURNI_WEEKLY_HEADER_COUNT = 10
TURNI_WEEKLY_SECTION_LABELS = (
    '1 turno',
    '2 turno',
    '3 turno',
    'turno centrale',
)
TURNI_WEEKLY_ROWS_PER_SECTION = 3
TURNI_WEEKEND_COLUMN_LABELS = (
    'Data',
    'Turno',
    'Nominativo',
    'Preposto',
    'Attivita',
    'Reparto',
)
TURNI_WEEKEND_MIN_ROW_COUNT = 1
TURNI_WEEKEND_DEFAULT_ROW_COUNT = 20
TURNI_WEEKEND_MAX_SINGLE_PAGE_ROW_COUNT = 43
TURNI_PORTINERIA_HEADERS = (
    'PORTINERIA CENTRALE',
    'CENTRALINISTA',
    'PORTINERIA CELLA',
)
TURNI_PORTINERIA_DEFAULT_TIMES = (
    ('06:14', '08:17', '06:14'),
    ('14:22', '', '14:22'),
    ('22:06', '', '22:06'),
)
TURNI_PORTINERIA_SECTION_LABELS = (
    '1 turno',
    '2 turno',
    '3 turno',
)
TURNI_PORTINERIA_ROWS_PER_SECTION = 3
TURNI_PORTINERIA_WEEKEND_DEFAULT_ROW_COUNT = 34
TURNI_SCORRIMENTO_BLOCK_COUNT = 4
TURNI_SCORRIMENTO_DAYS_PER_BLOCK = 7
TURNI_SCORRIMENTO_TOTAL_COLUMNS = TURNI_SCORRIMENTO_BLOCK_COUNT * TURNI_SCORRIMENTO_DAYS_PER_BLOCK
TURNI_SCORRIMENTO_SQUAD_COUNT = 4
TURNI_SCORRIMENTO_DEFAULT_TITLE = 'SANVINCENZO S.R.L :ORGANIZZAZIONE TURNI A SCORRIMENTO'
TURNI_SCORRIMENTO_DEFAULT_BLOCK_LABELS = (
    '3:7:11:15:19:23:27:31:35:39:43:47:51',
    '4:8:12:16:20:24:28:32:36:40:44:48:52',
    '1:5:9:13:17:21:25:29:33:37:41:45:49',
    '2:6:10:14:18:22:26:30:34:38:42:46:50',
)
TURNI_SCORRIMENTO_DEFAULT_DAY_LABELS = ('L', 'M', 'M', 'G', 'V', 'S', 'D')
TURNI_SCORRIMENTO_DEFAULT_SQUAD_LABELS = ('SQUADRA 1', 'SQUADRA 2', 'SQUADRA 3', 'SQUADRA 4')
TURNI_SCORRIMENTO_DEFAULT_DEPARTMENT_TITLES = ('PRODUZIONE', 'NAVETTA', 'IMPIANTO TIGER', 'STANZETTE')
TURNI_SCORRIMENTO_DEFAULT_DEPARTMENT_NAMES = (
    ('CAIA', 'FERRARA', 'LANDOLFO', 'PROSITTO'),
    ('SALZILLO', 'CERRONE F.', 'ARTUSO', 'CHIACCHIO'),
    ('ACERRA R.', 'GATTA', 'CIPOLLETTA', 'MENNILLO'),
    ('', '', '', ''),
)
TURNI_DEFAULT_WEEKLY_PDF_TITLE = 'SAN VINCENZO S.R.L.:ORGANIZZAZIONE TURNI'
TURNI_EXPORT_APP_LOGO_PATH = Path(settings.BASE_DIR) / 'portal' / 'static' / 'portal' / 'logo.png'
TURNI_EXPORT_WEEKEND_ANCIS_LOGO_PATH = Path(settings.BASE_DIR) / 'portal' / 'static' / 'portal' / 'ancis-sgq-sga-2026.png'
TURNI_EXPORT_WEEKEND_ANID_LOGO_PATH = Path(settings.BASE_DIR) / 'portal' / 'static' / 'portal' / 'logo-anid.jpg'
TURNI_EMPLOYEE_SECTION_META = OrderedDict([
    ('weekly', {'label': 'Settimana'}),
    ('saturday', {'label': 'Sabato'}),
    ('sunday', {'label': 'Domenica'}),
    ('jolly_weekend', {'label': 'Jolly'}),
    ('scorrimento', {'label': 'Scorrimento'}),
])
TURNI_MARKINGS_SECTION_META = OrderedDict([
    ('weekly', {'label': 'Settimana'}),
    ('portineria_weekly', {'label': 'Portineria settimana'}),
    ('saturday', {'label': 'Sabato'}),
    ('sunday', {'label': 'Domenica'}),
    ('jolly_weekend', {'label': 'Jolly'}),
    ('scorrimento', {'label': 'Scorrimento'}),
    ('portineria_weekend', {'label': 'Portineria weekend'}),
])
TURNI_PUBLISHED_SECTIONS_KEY = 'published_sections'


def _default_turni_weekly_data():
    return {
        'headers': ['' for _ in range(TURNI_WEEKLY_HEADER_COUNT)],
        'central_departments': ['' for _ in range(TURNI_WEEKLY_HEADER_COUNT)],
        'sections': [
            {
                'label': label,
                'time_values': ['' for _ in range(TURNI_WEEKLY_HEADER_COUNT)],
                'rows': [
                    ['' for _ in range(TURNI_WEEKLY_HEADER_COUNT)]
                    for _ in range(TURNI_WEEKLY_ROWS_PER_SECTION)
                ],
            }
            for label in TURNI_WEEKLY_SECTION_LABELS
        ],
    }


def _merge_turni_weekly_data(raw_weekly_data):
    weekly_data = _default_turni_weekly_data()
    if not isinstance(raw_weekly_data, dict):
        return weekly_data

    raw_headers = raw_weekly_data.get('headers')
    if isinstance(raw_headers, list):
        for index in range(min(len(raw_headers), TURNI_WEEKLY_HEADER_COUNT)):
            weekly_data['headers'][index] = str(raw_headers[index] or '').strip()

    raw_central_departments = raw_weekly_data.get('central_departments')
    if isinstance(raw_central_departments, list):
        for index in range(min(len(raw_central_departments), TURNI_WEEKLY_HEADER_COUNT)):
            weekly_data['central_departments'][index] = str(raw_central_departments[index] or '').strip()

    raw_sections = raw_weekly_data.get('sections')
    if isinstance(raw_sections, list):
        for section_index in range(min(len(raw_sections), len(TURNI_WEEKLY_SECTION_LABELS))):
            raw_section = raw_sections[section_index]
            if not isinstance(raw_section, dict):
                continue
            weekly_section = weekly_data['sections'][section_index]
            raw_time_values = raw_section.get('time_values')
            if isinstance(raw_time_values, list):
                for value_index in range(min(len(raw_time_values), TURNI_WEEKLY_HEADER_COUNT)):
                    weekly_section['time_values'][value_index] = str(raw_time_values[value_index] or '').strip()
            raw_rows = raw_section.get('rows')
            if isinstance(raw_rows, list):
                for row_index in range(min(len(raw_rows), TURNI_WEEKLY_ROWS_PER_SECTION)):
                    raw_row = raw_rows[row_index]
                    if not isinstance(raw_row, list):
                        continue
                    for value_index in range(min(len(raw_row), TURNI_WEEKLY_HEADER_COUNT)):
                        weekly_section['rows'][row_index][value_index] = str(raw_row[value_index] or '').strip()

    if not any(weekly_data['central_departments']) and len(weekly_data['sections']) >= 4:
        fallback_values = list(weekly_data['sections'][2]['rows'][2])
        if any(fallback_values):
            weekly_data['central_departments'] = fallback_values
            weekly_data['sections'][2]['rows'][2] = ['' for _ in range(TURNI_WEEKLY_HEADER_COUNT)]
    return weekly_data


def _extract_turni_weekly_data_from_post(post_data):
    weekly_data = _default_turni_weekly_data()
    raw_headers = post_data.getlist('weekly_headers')
    for index in range(min(len(raw_headers), TURNI_WEEKLY_HEADER_COUNT)):
        weekly_data['headers'][index] = raw_headers[index].strip()

    raw_central_departments = post_data.getlist('weekly_central_departments')
    for index in range(min(len(raw_central_departments), TURNI_WEEKLY_HEADER_COUNT)):
        weekly_data['central_departments'][index] = raw_central_departments[index].strip()

    for section_index, section in enumerate(weekly_data['sections']):
        raw_time_values = post_data.getlist(f'weekly_time_{section_index}')
        for value_index in range(min(len(raw_time_values), TURNI_WEEKLY_HEADER_COUNT)):
            section['time_values'][value_index] = raw_time_values[value_index].strip()

        for row_index in range(TURNI_WEEKLY_ROWS_PER_SECTION):
            raw_row_values = post_data.getlist(f'weekly_row_{section_index}_{row_index}')
            for value_index in range(min(len(raw_row_values), TURNI_WEEKLY_HEADER_COUNT)):
                section['rows'][row_index][value_index] = raw_row_values[value_index].strip()

    return weekly_data


def _default_turni_weekend_data(row_count=TURNI_WEEKEND_DEFAULT_ROW_COUNT):
    row_count = _clamp_turni_weekend_row_count(
        row_count,
        default=TURNI_WEEKEND_DEFAULT_ROW_COUNT,
        minimum=TURNI_WEEKEND_MIN_ROW_COUNT,
        maximum=TURNI_WEEKEND_MAX_SINGLE_PAGE_ROW_COUNT,
    )
    return {
        'title': '',
        'base_date': '',
        'rows': [
            ['' for _ in range(len(TURNI_WEEKEND_COLUMN_LABELS))]
            for _ in range(row_count)
        ],
    }


def _clamp_turni_weekend_row_count(raw_value, *, default, minimum, maximum):
    try:
        resolved = int(raw_value)
    except (TypeError, ValueError):
        resolved = default

    resolved = max(minimum, resolved)
    if maximum is not None:
        resolved = min(maximum, resolved)
    return resolved


def _merge_turni_weekend_data(
    raw_weekend_data,
    row_count=None,
    *,
    default_row_count=TURNI_WEEKEND_DEFAULT_ROW_COUNT,
    minimum=TURNI_WEEKEND_MIN_ROW_COUNT,
    maximum=TURNI_WEEKEND_MAX_SINGLE_PAGE_ROW_COUNT,
):
    if row_count is None and isinstance(raw_weekend_data, dict):
        raw_rows = raw_weekend_data.get('rows')
        if isinstance(raw_rows, list) and raw_rows:
            row_count = len(raw_rows)

    row_count = _clamp_turni_weekend_row_count(
        row_count,
        default=default_row_count,
        minimum=minimum,
        maximum=maximum,
    )
    weekend_data = _default_turni_weekend_data(row_count=row_count)
    if not isinstance(raw_weekend_data, dict):
        return weekend_data

    weekend_data['title'] = str(raw_weekend_data.get('title') or '').strip()
    weekend_data['base_date'] = str(raw_weekend_data.get('base_date') or '').strip()
    raw_rows = raw_weekend_data.get('rows')
    if isinstance(raw_rows, list):
        for row_index in range(min(len(raw_rows), len(weekend_data['rows']))):
            raw_row = raw_rows[row_index]
            if not isinstance(raw_row, list):
                continue
            for value_index in range(min(len(raw_row), len(TURNI_WEEKEND_COLUMN_LABELS))):
                weekend_data['rows'][row_index][value_index] = str(raw_row[value_index] or '').strip()
    return weekend_data


def _extract_turni_weekend_data_from_post(
    post_data,
    prefix,
    row_count=None,
    *,
    default_row_count=TURNI_WEEKEND_DEFAULT_ROW_COUNT,
    minimum=TURNI_WEEKEND_MIN_ROW_COUNT,
    maximum=TURNI_WEEKEND_MAX_SINGLE_PAGE_ROW_COUNT,
):
    if row_count is None:
        row_count = post_data.get(f'{prefix}_row_count')

    row_count = _clamp_turni_weekend_row_count(
        row_count,
        default=default_row_count,
        minimum=minimum,
        maximum=maximum,
    )
    weekend_data = _default_turni_weekend_data(row_count=row_count)
    weekend_data['title'] = (post_data.get(f'{prefix}_title') or '').strip()
    weekend_data['base_date'] = (post_data.get(f'{prefix}_base_date') or '').strip()
    for row_index in range(row_count):
        raw_row_values = post_data.getlist(f'{prefix}_row_{row_index}')
        for value_index in range(min(len(raw_row_values), len(TURNI_WEEKEND_COLUMN_LABELS))):
            weekend_data['rows'][row_index][value_index] = raw_row_values[value_index].strip()
    return weekend_data


def _default_turni_portineria_weekly_data():
    return {
        'headers': list(TURNI_PORTINERIA_HEADERS),
        'sections': [
            {
                'label': label,
                'time_values': list(TURNI_PORTINERIA_DEFAULT_TIMES[index]),
                'rows': [
                    ['' for _ in range(len(TURNI_PORTINERIA_HEADERS))]
                    for _ in range(TURNI_PORTINERIA_ROWS_PER_SECTION)
                ],
            }
            for index, label in enumerate(TURNI_PORTINERIA_SECTION_LABELS)
        ],
    }


def _normalize_turni_portineria_weekly_values(raw_values):
    if not isinstance(raw_values, list):
        return []
    if len(raw_values) >= 3:
        return list(raw_values[:3])
    if len(raw_values) == 2:
        return [raw_values[0], '', raw_values[1]]
    if len(raw_values) == 1:
        return [raw_values[0], '', '']
    return []


def _merge_turni_portineria_weekly_data(raw_portineria_weekly_data):
    weekly_data = _default_turni_portineria_weekly_data()
    if not isinstance(raw_portineria_weekly_data, dict):
        return weekly_data

    raw_headers = _normalize_turni_portineria_weekly_values(raw_portineria_weekly_data.get('headers'))
    if isinstance(raw_headers, list):
        for index in range(min(len(raw_headers), len(TURNI_PORTINERIA_HEADERS))):
            weekly_data['headers'][index] = str(raw_headers[index] or '').strip()

    raw_sections = raw_portineria_weekly_data.get('sections')
    if isinstance(raw_sections, list):
        for section_index in range(min(len(raw_sections), len(TURNI_PORTINERIA_SECTION_LABELS))):
            raw_section = raw_sections[section_index]
            if not isinstance(raw_section, dict):
                continue
            weekly_section = weekly_data['sections'][section_index]
            raw_time_values = _normalize_turni_portineria_weekly_values(raw_section.get('time_values'))
            if isinstance(raw_time_values, list):
                for value_index in range(min(len(raw_time_values), len(TURNI_PORTINERIA_HEADERS))):
                    weekly_section['time_values'][value_index] = str(raw_time_values[value_index] or '').strip()
            raw_rows = raw_section.get('rows')
            if isinstance(raw_rows, list):
                for row_index in range(min(len(raw_rows), TURNI_PORTINERIA_ROWS_PER_SECTION)):
                    raw_row = _normalize_turni_portineria_weekly_values(raw_rows[row_index])
                    if not isinstance(raw_row, list):
                        continue
                    for value_index in range(min(len(raw_row), len(TURNI_PORTINERIA_HEADERS))):
                        weekly_section['rows'][row_index][value_index] = str(raw_row[value_index] or '').strip()
    return weekly_data


def _default_turni_scorrimento_data():
    return {
        'title': TURNI_SCORRIMENTO_DEFAULT_TITLE,
        'base_date': '',
        'block_labels': list(TURNI_SCORRIMENTO_DEFAULT_BLOCK_LABELS),
        'day_labels': list(TURNI_SCORRIMENTO_DEFAULT_DAY_LABELS),
        'squad_labels': list(TURNI_SCORRIMENTO_DEFAULT_SQUAD_LABELS),
        'matrix': [
            ['N', 'N', 'N', 'R', 'P', 'P', 'R', 'M', 'M', 'R', 'N', 'N', 'N', 'N', 'R', 'P', 'P', 'P', 'R', 'M', 'M', 'P', 'R', 'M', 'M', 'M', 'R', 'P'],
            ['M', 'M', 'R', 'N', 'N', 'N', 'N', 'R', 'P', 'P', 'P', 'R', 'M', 'M', 'P', 'R', 'M', 'M', 'M', 'R', 'P', 'N', 'N', 'N', 'R', 'P', 'P', 'R'],
            ['R', 'P', 'P', 'P', 'R', 'M', 'M', 'P', 'R', 'M', 'M', 'M', 'R', 'P', 'N', 'N', 'N', 'R', 'P', 'P', 'R', 'M', 'M', 'R', 'N', 'N', 'N', 'N'],
            ['P', 'R', 'M', 'M', 'M', 'R', 'P', 'N', 'N', 'N', 'R', 'P', 'P', 'R', 'M', 'M', 'R', 'N', 'N', 'N', 'N', 'R', 'P', 'P', 'P', 'R', 'M', 'M'],
        ],
        'department_titles': list(TURNI_SCORRIMENTO_DEFAULT_DEPARTMENT_TITLES),
        'department_names': [list(names) for names in TURNI_SCORRIMENTO_DEFAULT_DEPARTMENT_NAMES],
        'rows': [],
    }


def _merge_turni_scorrimento_data(raw_scorrimento_data):
    scorrimento_data = _default_turni_scorrimento_data()
    if not isinstance(raw_scorrimento_data, dict):
        return scorrimento_data

    title = str(raw_scorrimento_data.get('title') or '').strip()
    if title:
        scorrimento_data['title'] = title
    scorrimento_data['base_date'] = str(raw_scorrimento_data.get('base_date') or '').strip()

    raw_block_labels = raw_scorrimento_data.get('block_labels')
    if isinstance(raw_block_labels, list):
        for index in range(min(len(raw_block_labels), TURNI_SCORRIMENTO_BLOCK_COUNT)):
            scorrimento_data['block_labels'][index] = str(raw_block_labels[index] or '').strip()

    raw_day_labels = raw_scorrimento_data.get('day_labels')
    if isinstance(raw_day_labels, list):
        for index in range(min(len(raw_day_labels), TURNI_SCORRIMENTO_DAYS_PER_BLOCK)):
            scorrimento_data['day_labels'][index] = str(raw_day_labels[index] or '').strip()

    raw_squad_labels = raw_scorrimento_data.get('squad_labels')
    if isinstance(raw_squad_labels, list):
        for index in range(min(len(raw_squad_labels), TURNI_SCORRIMENTO_SQUAD_COUNT)):
            scorrimento_data['squad_labels'][index] = str(raw_squad_labels[index] or '').strip()

    raw_matrix = raw_scorrimento_data.get('matrix')
    if isinstance(raw_matrix, list):
        for row_index in range(min(len(raw_matrix), TURNI_SCORRIMENTO_SQUAD_COUNT)):
            raw_row = raw_matrix[row_index]
            if not isinstance(raw_row, list):
                continue
            for col_index in range(min(len(raw_row), TURNI_SCORRIMENTO_TOTAL_COLUMNS)):
                scorrimento_data['matrix'][row_index][col_index] = str(raw_row[col_index] or '').strip().upper()

    raw_department_titles = raw_scorrimento_data.get('department_titles')
    if isinstance(raw_department_titles, list):
        for index in range(min(len(raw_department_titles), TURNI_SCORRIMENTO_BLOCK_COUNT)):
            scorrimento_data['department_titles'][index] = str(raw_department_titles[index] or '').strip()

    raw_department_names = raw_scorrimento_data.get('department_names')
    if isinstance(raw_department_names, list):
        for block_index in range(min(len(raw_department_names), TURNI_SCORRIMENTO_BLOCK_COUNT)):
            raw_block_names = raw_department_names[block_index]
            if not isinstance(raw_block_names, list):
                continue
            for row_index in range(min(len(raw_block_names), TURNI_SCORRIMENTO_SQUAD_COUNT)):
                scorrimento_data['department_names'][block_index][row_index] = str(raw_block_names[row_index] or '').strip()

    raw_rows = raw_scorrimento_data.get('rows')
    if isinstance(raw_rows, list):
        compat_rows = []
        for raw_row in raw_rows:
            if not isinstance(raw_row, list):
                continue
            compat_rows.append([str(value or '').strip() for value in raw_row[:len(TURNI_WEEKEND_COLUMN_LABELS)]])
        scorrimento_data['rows'] = compat_rows
    return scorrimento_data


def _extract_turni_scorrimento_data_from_post(post_data):
    scorrimento_data = _default_turni_scorrimento_data()
    scorrimento_data['title'] = (post_data.get('scorrimento_title') or '').strip() or TURNI_SCORRIMENTO_DEFAULT_TITLE
    scorrimento_data['base_date'] = (post_data.get('scorrimento_base_date') or '').strip()

    raw_block_labels = post_data.getlist('scorrimento_block_labels')
    for index in range(min(len(raw_block_labels), TURNI_SCORRIMENTO_BLOCK_COUNT)):
        scorrimento_data['block_labels'][index] = raw_block_labels[index].strip()

    raw_day_labels = post_data.getlist('scorrimento_day_labels')
    for index in range(min(len(raw_day_labels), TURNI_SCORRIMENTO_DAYS_PER_BLOCK)):
        scorrimento_data['day_labels'][index] = raw_day_labels[index].strip()

    raw_squad_labels = post_data.getlist('scorrimento_squad_labels')
    for index in range(min(len(raw_squad_labels), TURNI_SCORRIMENTO_SQUAD_COUNT)):
        scorrimento_data['squad_labels'][index] = raw_squad_labels[index].strip()

    raw_department_titles = post_data.getlist('scorrimento_department_titles')
    for index in range(min(len(raw_department_titles), TURNI_SCORRIMENTO_BLOCK_COUNT)):
        scorrimento_data['department_titles'][index] = raw_department_titles[index].strip()

    for row_index in range(TURNI_SCORRIMENTO_SQUAD_COUNT):
        raw_row_values = post_data.getlist(f'scorrimento_matrix_{row_index}')
        for col_index in range(min(len(raw_row_values), TURNI_SCORRIMENTO_TOTAL_COLUMNS)):
            scorrimento_data['matrix'][row_index][col_index] = raw_row_values[col_index].strip().upper()

    for block_index in range(TURNI_SCORRIMENTO_BLOCK_COUNT):
        raw_names = post_data.getlist(f'scorrimento_department_names_{block_index}')
        for row_index in range(min(len(raw_names), TURNI_SCORRIMENTO_SQUAD_COUNT)):
            scorrimento_data['department_names'][block_index][row_index] = raw_names[row_index].strip()

    legacy_row_count = post_data.get('scorrimento_row_count')
    if legacy_row_count is not None or post_data.getlist('scorrimento_row_0'):
        compat_rows = _extract_turni_weekend_data_from_post(post_data, 'scorrimento')['rows']
        scorrimento_data['rows'] = compat_rows

    return scorrimento_data


def _turni_scorrimento_export_data(raw_scorrimento_data):
    scorrimento_data = _merge_turni_scorrimento_data(raw_scorrimento_data)
    return ScorrimentoExportData(
        title=scorrimento_data['title'],
        block_labels=list(scorrimento_data['block_labels']),
        day_labels=list(scorrimento_data['day_labels']),
        squad_labels=list(scorrimento_data['squad_labels']),
        matrix=[list(row) for row in scorrimento_data['matrix']],
        department_titles=list(scorrimento_data['department_titles']),
        department_names=[list(row) for row in scorrimento_data['department_names']],
    )


def _turni_scorrimento_template_blocks(scorrimento_data):
    blocks = []
    for block_index, block_label in enumerate(scorrimento_data['block_labels']):
        start = block_index * TURNI_SCORRIMENTO_DAYS_PER_BLOCK
        columns = []
        for offset, day_label in enumerate(scorrimento_data['day_labels']):
            columns.append({
                'index': start + offset,
                'day_label': day_label,
            })
        blocks.append({
            'index': block_index,
            'label': block_label,
            'columns': columns,
        })
    return blocks


def _turni_scorrimento_template_rows(scorrimento_data):
    rows = []
    for squad_index, squad_label in enumerate(scorrimento_data['squad_labels']):
        current_matrix_row = scorrimento_data['matrix'][squad_index] if squad_index < len(scorrimento_data['matrix']) else []
        cells = []
        for block_index in range(TURNI_SCORRIMENTO_BLOCK_COUNT):
            start = block_index * TURNI_SCORRIMENTO_DAYS_PER_BLOCK
            for offset in range(TURNI_SCORRIMENTO_DAYS_PER_BLOCK):
                column_index = start + offset
                value = str(current_matrix_row[column_index] or '').strip().upper() if column_index < len(current_matrix_row) else ''
                cells.append([
                    value,
                    block_index > 0 and offset == 0,
                    value == 'R',
                ])
        rows.append({
            'index': squad_index,
            'label': squad_label,
            'cells': cells,
        })
    return rows


def _turni_scorrimento_template_department_blocks(scorrimento_data):
    blocks = []
    department_titles = scorrimento_data.get('department_titles') or []
    for block_index, block_title in enumerate(department_titles):
        block_names = scorrimento_data['department_names'][block_index] if block_index < len(scorrimento_data['department_names']) else []
        rows = []
        for squad_index, squad_label in enumerate(scorrimento_data['squad_labels']):
            rows.append([
                squad_label,
                block_names[squad_index] if squad_index < len(block_names) else '',
            ])
        blocks.append({
            'index': block_index,
            'title': block_title,
            'rows': rows,
        })
    return blocks


def _extract_turni_portineria_weekly_data_from_post(post_data):
    weekly_data = _default_turni_portineria_weekly_data()
    raw_headers = post_data.getlist('portineria_weekly_headers')
    for index in range(min(len(raw_headers), len(TURNI_PORTINERIA_HEADERS))):
        weekly_data['headers'][index] = raw_headers[index].strip()

    for section_index, section in enumerate(weekly_data['sections']):
        raw_time_values = post_data.getlist(f'portineria_weekly_time_{section_index}')
        for value_index in range(min(len(raw_time_values), len(TURNI_PORTINERIA_HEADERS))):
            section['time_values'][value_index] = raw_time_values[value_index].strip()

        for row_index in range(TURNI_PORTINERIA_ROWS_PER_SECTION):
            raw_row_values = post_data.getlist(f'portineria_weekly_row_{section_index}_{row_index}')
            for value_index in range(min(len(raw_row_values), len(TURNI_PORTINERIA_HEADERS))):
                section['rows'][row_index][value_index] = raw_row_values[value_index].strip()

    return weekly_data


def _existing_turni_export_path(path):
    if path is not None and path.exists():
        return path
    return None


def _turni_weekly_sections_for_export(raw_weekly_data):
    weekly_data = _merge_turni_weekly_data(raw_weekly_data)
    if len(weekly_data['sections']) >= 4 and any(weekly_data.get('central_departments', [])):
        weekly_data['sections'][2]['rows'][2] = list(weekly_data['central_departments'])
    sections = []
    for section in weekly_data['sections']:
        time_values = list(section['time_values'])
        sections.append(
            WeeklySectionData(
                label=section['label'],
                time_label=time_values[0] if time_values else '',
                time_values=time_values,
                rows=[list(row) for row in section['rows']],
            )
        )
    return weekly_data['headers'], sections


def _turni_portineria_weekly_sections_for_export(raw_weekly_data):
    weekly_data = _merge_turni_portineria_weekly_data(raw_weekly_data)
    sections = []
    for section in weekly_data['sections']:
        time_values = list(section['time_values'])
        sections.append(
            WeeklySectionData(
                label=section['label'],
                time_label=time_values[0] if time_values else '',
                time_values=time_values,
                rows=[list(row) for row in section['rows']],
            )
        )
    return weekly_data['headers'], sections


def _turni_weekend_export_data(
    raw_weekend_data,
    *,
    title,
    row_count=None,
    default_row_count=TURNI_WEEKEND_DEFAULT_ROW_COUNT,
    minimum=TURNI_WEEKEND_MIN_ROW_COUNT,
    maximum=TURNI_WEEKEND_MAX_SINGLE_PAGE_ROW_COUNT,
):
    weekend_data = _merge_turni_weekend_data(
        raw_weekend_data,
        row_count=row_count,
        default_row_count=default_row_count,
        minimum=minimum,
        maximum=maximum,
    )
    return WeekendExportData(
        title=str(weekend_data.get('title') or '').strip() or title,
        authorization_date=weekend_data['base_date'],
        rows=[list(row) for row in weekend_data['rows']],
    )


def _turni_download_response(content, *, content_type, filename):
    response = HttpResponse(content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _turni_combined_jpg_bytes(image_paths):
    if not image_paths:
        raise ValueError('Nessuna immagine generata per l\'export.')

    images = []
    combined_image = None
    try:
        for image_path in image_paths:
            with Image.open(image_path) as image:
                images.append(image.convert('RGB'))

        if len(images) == 1:
            output = io.BytesIO()
            images[0].save(output, format='JPEG', quality=95)
            return output.getvalue()

        max_width = max(image.width for image in images)
        total_height = sum(image.height for image in images)
        combined_image = Image.new('RGB', (max_width, total_height), 'white')
        offset_y = 0
        for image in images:
            offset_x = max((max_width - image.width) // 2, 0)
            combined_image.paste(image, (offset_x, offset_y))
            offset_y += image.height

        output = io.BytesIO()
        combined_image.save(output, format='JPEG', quality=95)
        return output.getvalue()
    finally:
        for image in images:
            image.close()
        if combined_image is not None:
            combined_image.close()


def _turni_planner_export_response(state, *, export_format, export_target):
    planner_data = dict(state.planner_data or {})
    logo_path = _existing_turni_export_path(TURNI_EXPORT_APP_LOGO_PATH)
    ancis_logo_path = _existing_turni_export_path(TURNI_EXPORT_WEEKEND_ANCIS_LOGO_PATH)
    anid_logo_path = _existing_turni_export_path(TURNI_EXPORT_WEEKEND_ANID_LOGO_PATH)

    weekly_configs = {
        'weekly': {
            'pdf_name': WEEKLY_PDF_NAME,
            'image_name': WEEKLY_IMAGE_NAME,
            'layout': 'default',
            'builder': lambda: _turni_weekly_sections_for_export(planner_data.get('weekly')),
        },
        'portineria_weekly': {
            'pdf_name': PORTINERIA_WEEKLY_PDF_NAME,
            'image_name': PORTINERIA_WEEKLY_IMAGE_NAME,
            'layout': 'portineria',
            'builder': lambda: _turni_portineria_weekly_sections_for_export(planner_data.get('portineria_weekly')),
        },
    }
    weekend_configs = {
        'saturday': {
            'pdf_name': SATURDAY_PDF_NAME,
            'image_name': SATURDAY_IMAGE_NAME,
            'title': 'Comandata sabato',
            'row_count': None,
        },
        'sunday': {
            'pdf_name': SUNDAY_PDF_NAME,
            'image_name': SUNDAY_IMAGE_NAME,
            'title': 'Comandata domenica',
            'row_count': None,
        },
        'jolly_weekend': {
            'pdf_name': JOLLY_WEEKEND_PDF_NAME,
            'image_name': JOLLY_WEEKEND_IMAGE_NAME,
            'title': 'Comandata jolly',
            'row_count': None,
        },
        'scorrimento': {
            'pdf_name': SCORRIMENTO_PDF_NAME,
            'image_name': SCORRIMENTO_IMAGE_NAME,
            'title': 'Scorrimento',
            'row_count': None,
        },
        'portineria_weekend': {
            'pdf_name': PORTINERIA_WEEKEND_PDF_NAME,
            'image_name': PORTINERIA_WEEKEND_IMAGE_NAME,
            'title': 'Sabato - Domenica e festivi Portineria',
            'row_count': None,
        },
    }

    with tempfile.TemporaryDirectory(prefix='turni_planner_export_') as temp_dir:
        export_dir = Path(temp_dir)

        if export_target == 'scorrimento':
            export_data = _turni_scorrimento_export_data(planner_data.get('scorrimento'))
            if export_format == 'pdf':
                exported_path = export_scorrimento_pdf(
                    export_dir / SCORRIMENTO_PDF_NAME,
                    data=export_data,
                    logo_path=logo_path,
                    cert_logo_path=ancis_logo_path,
                    anid_logo_path=anid_logo_path,
                )
                return _turni_download_response(
                    exported_path.read_bytes(),
                    content_type='application/pdf',
                    filename=SCORRIMENTO_PDF_NAME,
                )

            exported_paths = export_scorrimento_images(
                export_dir / SCORRIMENTO_IMAGE_NAME,
                data=export_data,
                logo_path=logo_path,
                cert_logo_path=ancis_logo_path,
                anid_logo_path=anid_logo_path,
            )
            return _turni_download_response(
                _turni_combined_jpg_bytes(exported_paths),
                content_type='image/jpeg',
                filename=SCORRIMENTO_IMAGE_NAME,
            )

        if export_target in weekly_configs:
            config = weekly_configs[export_target]
            export_week_label = _resolve_turni_export_week_label(
                planner_data,
                state.week_label,
                key='weekly_export_week_label' if export_target == 'weekly' else 'portineria_weekly_export_week_label',
            )
            headers, sections = config['builder']()
            if export_format == 'pdf':
                exported_path = export_weekly_pdf(
                    export_dir / config['pdf_name'],
                    title_text=TURNI_DEFAULT_WEEKLY_PDF_TITLE,
                    week_label=export_week_label,
                    signature='',
                    headers=headers,
                    sections=sections,
                    logo_path=logo_path,
                    layout=config['layout'],
                )
                return _turni_download_response(
                    exported_path.read_bytes(),
                    content_type='application/pdf',
                    filename=config['pdf_name'],
                )

            exported_paths = export_weekly_images(
                export_dir / config['image_name'],
                title_text=TURNI_DEFAULT_WEEKLY_PDF_TITLE,
                week_label=export_week_label,
                signature='',
                headers=headers,
                sections=sections,
                logo_path=logo_path,
                temp_pdf_name=config['pdf_name'],
                layout=config['layout'],
            )
            return _turni_download_response(
                _turni_combined_jpg_bytes(exported_paths),
                content_type='image/jpeg',
                filename=config['image_name'],
            )

        if export_target in weekend_configs:
            config = weekend_configs[export_target]
            export_data = _turni_weekend_export_data(
                planner_data.get(export_target),
                title=config['title'],
                row_count=config['row_count'],
            )
            if export_format == 'pdf':
                exported_path = export_weekend_pdf(
                    export_dir / config['pdf_name'],
                    data=export_data,
                    logo_path=logo_path,
                    cert_logo_path=ancis_logo_path,
                    anid_logo_path=anid_logo_path,
                )
                return _turni_download_response(
                    exported_path.read_bytes(),
                    content_type='application/pdf',
                    filename=config['pdf_name'],
                )

            exported_paths = export_weekend_images(
                export_dir / config['image_name'],
                data=export_data,
                logo_path=logo_path,
                cert_logo_path=ancis_logo_path,
                anid_logo_path=anid_logo_path,
            )
            return _turni_download_response(
                _turni_combined_jpg_bytes(exported_paths),
                content_type='image/jpeg',
                filename=config['image_name'],
            )

    return HttpResponse('Export non disponibile per questa sezione.', status=400)


def _turni_planner_published_state():
    return TurniPlannerWeekState.objects.filter(visible_to_employees=True).order_by('-updated_at', '-id').first()


def _normalize_turni_published_sections(raw_sections, *, include_portineria=False):
    section_meta_map = TURNI_MARKINGS_SECTION_META if include_portineria else TURNI_EMPLOYEE_SECTION_META
    if not isinstance(raw_sections, list):
        return []

    normalized_sections = []
    for section_key in section_meta_map.keys():
        if section_key in raw_sections:
            normalized_sections.append(section_key)
    return normalized_sections


def _turni_planner_selected_section_keys(state, *, include_portineria=False):
    if not state:
        return []

    planner_data = dict(state.planner_data or {})
    allowed_targets = _turni_planner_allowed_targets(state, include_portineria=include_portineria)
    selected_sections = _normalize_turni_published_sections(
        planner_data.get(TURNI_PUBLISHED_SECTIONS_KEY),
        include_portineria=include_portineria,
    )
    if not selected_sections:
        return allowed_targets if state.visible_to_employees else []
    return [section_key for section_key in selected_sections if section_key in allowed_targets]


def _turni_planner_published_section_keys(state, *, include_portineria=False):
    if not state or not state.visible_to_employees:
        return []
    return _turni_planner_selected_section_keys(state, include_portineria=include_portineria)


def _turni_planner_employee_sections(state, *, include_portineria=False):
    sections = []
    section_meta_map = TURNI_MARKINGS_SECTION_META if include_portineria else TURNI_EMPLOYEE_SECTION_META
    for section_key in _turni_planner_published_section_keys(state, include_portineria=include_portineria):
        section_meta = section_meta_map.get(section_key)
        if not section_meta:
            continue
        sections.append({
            'key': section_key,
            'label': section_meta['label'],
            'image_url': reverse('employee_turni_published_image', args=[section_key]),
        })
    return sections


def _turni_planner_allowed_targets(state, *, include_portineria=False):
    section_meta_map = TURNI_MARKINGS_SECTION_META if include_portineria else TURNI_EMPLOYEE_SECTION_META
    allowed_targets = [key for key in section_meta_map.keys() if key != 'jolly_weekend']
    if state and _turni_planner_data_has_content(dict(state.planner_data or {}).get('jolly_weekend')):
        allowed_targets.append('jolly_weekend')
    return allowed_targets


def _user_can_view_published_turni(user, employee=None):
    if user_has_full_admin_access(user):
        return True
    if employee and not employee.show_published_turni:
        return False
    setting = getattr(user, 'portal_setting', None)
    if setting is not None:
        return bool(setting.show_published_turni)
    return True


def _turni_planner_employee_jpg_payload(state, *, export_target, include_portineria=False):
    if export_target not in _turni_planner_published_section_keys(state, include_portineria=include_portineria):
        raise ValueError('Sezione turni non disponibile per i dipendenti.')

    planner_data = dict(state.planner_data or {})
    logo_path = _existing_turni_export_path(TURNI_EXPORT_APP_LOGO_PATH)
    ancis_logo_path = _existing_turni_export_path(TURNI_EXPORT_WEEKEND_ANCIS_LOGO_PATH)
    anid_logo_path = _existing_turni_export_path(TURNI_EXPORT_WEEKEND_ANID_LOGO_PATH)

    weekly_configs = {
        'weekly': {
            'image_name': WEEKLY_IMAGE_NAME,
            'pdf_name': WEEKLY_PDF_NAME,
            'layout': 'default',
            'builder': lambda: _turni_weekly_sections_for_export(planner_data.get('weekly')),
            'label_key': 'weekly_export_week_label',
        },
        'portineria_weekly': {
            'image_name': PORTINERIA_WEEKLY_IMAGE_NAME,
            'pdf_name': PORTINERIA_WEEKLY_PDF_NAME,
            'layout': 'portineria',
            'builder': lambda: _turni_portineria_weekly_sections_for_export(planner_data.get('portineria_weekly')),
            'label_key': 'portineria_weekly_export_week_label',
        },
    }
    weekend_configs = {
        'saturday': {
            'image_name': SATURDAY_IMAGE_NAME,
            'title': 'Comandata sabato',
        },
        'sunday': {
            'image_name': SUNDAY_IMAGE_NAME,
            'title': 'Comandata domenica',
        },
        'jolly_weekend': {
            'image_name': JOLLY_WEEKEND_IMAGE_NAME,
            'title': 'Comandata jolly',
        },
        'scorrimento': {
            'image_name': SCORRIMENTO_IMAGE_NAME,
            'title': 'Scorrimento',
        },
        'portineria_weekend': {
            'image_name': PORTINERIA_WEEKEND_IMAGE_NAME,
            'title': 'Sabato - Domenica e festivi Portineria',
        },
    }

    with tempfile.TemporaryDirectory(prefix='turni_planner_employee_') as temp_dir:
        export_dir = Path(temp_dir)

        if export_target == 'scorrimento':
            exported_paths = export_scorrimento_images(
                export_dir / SCORRIMENTO_IMAGE_NAME,
                data=_turni_scorrimento_export_data(planner_data.get('scorrimento')),
            )
            return _turni_combined_jpg_bytes(exported_paths), SCORRIMENTO_IMAGE_NAME

        if export_target in weekly_configs:
            config = weekly_configs[export_target]
            export_week_label = _resolve_turni_export_week_label(
                planner_data,
                state.week_label,
                key=config['label_key'],
            )
            headers, sections = config['builder']()
            exported_paths = export_weekly_images(
                export_dir / config['image_name'],
                title_text=TURNI_DEFAULT_WEEKLY_PDF_TITLE,
                week_label=export_week_label,
                signature='',
                headers=headers,
                sections=sections,
                logo_path=logo_path,
                temp_pdf_name=config['pdf_name'],
                layout=config['layout'],
            )
            return _turni_combined_jpg_bytes(exported_paths), config['image_name']

        if export_target in weekend_configs:
            config = weekend_configs[export_target]
            exported_paths = export_weekend_images(
                export_dir / config['image_name'],
                data=_turni_weekend_export_data(planner_data.get(export_target), title=config['title']),
                logo_path=logo_path,
                cert_logo_path=ancis_logo_path,
                anid_logo_path=anid_logo_path,
            )
            return _turni_combined_jpg_bytes(exported_paths), config['image_name']

    raise ValueError('Sezione turni non disponibile per i dipendenti.')


@login_required
def employee_turni_published_image(request, section_key):
    employee = Employee.objects.filter(user=request.user).first()
    include_portineria = bool(employee) or user_has_today_markings_access(request.user) or user_has_full_admin_access(request.user)
    if not request.user.is_staff and not employee and not user_has_today_markings_access(request.user):
        return HttpResponse('Non autorizzato', status=403)
    if not request.user.is_staff and not _user_can_view_published_turni(request.user, employee=employee):
        return HttpResponse('Sezione turni non disponibile.', status=404)

    state = _turni_planner_published_state()
    if not state:
        return HttpResponse('Nessun turno pubblicato.', status=404)

    try:
        image_bytes, image_name = _turni_planner_employee_jpg_payload(
            state,
            export_target=section_key,
            include_portineria=include_portineria,
        )
    except ValueError:
        return HttpResponse('Sezione turni non disponibile.', status=404)

    response = HttpResponse(image_bytes, content_type='image/jpeg')
    response['Content-Disposition'] = f'inline; filename="{image_name}"'
    return _disable_response_cache(response)


def _turni_planner_bulk_export_response(state, *, export_format):
    planner_data = dict(state.planner_data or {})
    logo_path = _existing_turni_export_path(TURNI_EXPORT_APP_LOGO_PATH)
    ancis_logo_path = _existing_turni_export_path(TURNI_EXPORT_WEEKEND_ANCIS_LOGO_PATH)
    anid_logo_path = _existing_turni_export_path(TURNI_EXPORT_WEEKEND_ANID_LOGO_PATH)

    weekly_configs = {
        'weekly': {
            'pdf_name': WEEKLY_PDF_NAME,
            'image_name': WEEKLY_IMAGE_NAME,
            'layout': 'default',
            'builder': lambda: _turni_weekly_sections_for_export(planner_data.get('weekly')),
        },
        'portineria_weekly': {
            'pdf_name': PORTINERIA_WEEKLY_PDF_NAME,
            'image_name': PORTINERIA_WEEKLY_IMAGE_NAME,
            'layout': 'portineria',
            'builder': lambda: _turni_portineria_weekly_sections_for_export(planner_data.get('portineria_weekly')),
        },
    }
    weekend_configs = {
        'saturday': {
            'pdf_name': SATURDAY_PDF_NAME,
            'image_name': SATURDAY_IMAGE_NAME,
            'title': 'Comandata sabato',
            'row_count': None,
        },
        'sunday': {
            'pdf_name': SUNDAY_PDF_NAME,
            'image_name': SUNDAY_IMAGE_NAME,
            'title': 'Comandata domenica',
            'row_count': None,
        },
        'jolly_weekend': {
            'pdf_name': JOLLY_WEEKEND_PDF_NAME,
            'image_name': JOLLY_WEEKEND_IMAGE_NAME,
            'title': 'Comandata jolly',
            'row_count': None,
        },
        'scorrimento': {
            'pdf_name': SCORRIMENTO_PDF_NAME,
            'image_name': SCORRIMENTO_IMAGE_NAME,
            'title': 'Scorrimento',
            'row_count': None,
        },
        'portineria_weekend': {
            'pdf_name': PORTINERIA_WEEKEND_PDF_NAME,
            'image_name': PORTINERIA_WEEKEND_IMAGE_NAME,
            'title': 'Sabato - Domenica e festivi Portineria',
            'row_count': None,
        },
    }

    formats_to_include = ['pdf', 'jpg'] if export_format == 'all' else [export_format]
    filename_map = {
        'pdf': f'turni-planner-{state.week_label}-pdf.zip',
        'jpg': f'turni-planner-{state.week_label}-jpg.zip',
        'all': f'turni-planner-{state.week_label}-pdf-jpg.zip',
    }

    with tempfile.TemporaryDirectory(prefix='turni_planner_bulk_export_') as temp_dir:
        export_dir = Path(temp_dir)
        archive_buffer = io.BytesIO()

        with zipfile.ZipFile(archive_buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as archive:
            scorrimento_export_data = _turni_scorrimento_export_data(planner_data.get('scorrimento'))
            if 'pdf' in formats_to_include:
                pdf_path = export_scorrimento_pdf(
                    export_dir / SCORRIMENTO_PDF_NAME,
                    data=scorrimento_export_data,
                    logo_path=logo_path,
                    cert_logo_path=ancis_logo_path,
                    anid_logo_path=anid_logo_path,
                )
                archive.writestr(SCORRIMENTO_PDF_NAME, pdf_path.read_bytes())
            if 'jpg' in formats_to_include:
                image_paths = export_scorrimento_images(
                    export_dir / SCORRIMENTO_IMAGE_NAME,
                    data=scorrimento_export_data,
                    logo_path=logo_path,
                    cert_logo_path=ancis_logo_path,
                    anid_logo_path=anid_logo_path,
                )
                for image_path in image_paths:
                    archive.writestr(image_path.name, image_path.read_bytes())

            for config_key, config in weekly_configs.items():
                export_week_label = _resolve_turni_export_week_label(
                    planner_data,
                    state.week_label,
                    key='weekly_export_week_label' if config_key == 'weekly' else 'portineria_weekly_export_week_label',
                )
                headers, sections = config['builder']()
                if 'pdf' in formats_to_include:
                    pdf_path = export_weekly_pdf(
                        export_dir / config['pdf_name'],
                        title_text=TURNI_DEFAULT_WEEKLY_PDF_TITLE,
                        week_label=export_week_label,
                        signature='',
                        headers=headers,
                        sections=sections,
                        logo_path=logo_path,
                        layout=config['layout'],
                    )
                    archive.writestr(config['pdf_name'], pdf_path.read_bytes())
                if 'jpg' in formats_to_include:
                    image_paths = export_weekly_images(
                        export_dir / config['image_name'],
                        title_text=TURNI_DEFAULT_WEEKLY_PDF_TITLE,
                        week_label=export_week_label,
                        signature='',
                        headers=headers,
                        sections=sections,
                        logo_path=logo_path,
                        temp_pdf_name=config['pdf_name'],
                        layout=config['layout'],
                    )
                    for image_path in image_paths:
                        archive.writestr(image_path.name, image_path.read_bytes())

            for config_key, config in weekend_configs.items():
                export_data = _turni_weekend_export_data(
                    planner_data.get(config_key),
                    title=config['title'],
                    row_count=config['row_count'],
                )
                if 'pdf' in formats_to_include:
                    pdf_path = export_weekend_pdf(
                        export_dir / config['pdf_name'],
                        data=export_data,
                        logo_path=logo_path,
                        cert_logo_path=ancis_logo_path,
                        anid_logo_path=anid_logo_path,
                    )
                    archive.writestr(config['pdf_name'], pdf_path.read_bytes())
                if 'jpg' in formats_to_include:
                    image_paths = export_weekend_images(
                        export_dir / config['image_name'],
                        data=export_data,
                        logo_path=logo_path,
                        cert_logo_path=ancis_logo_path,
                        anid_logo_path=anid_logo_path,
                    )
                    for image_path in image_paths:
                        archive.writestr(image_path.name, image_path.read_bytes())

        return _turni_download_response(
            archive_buffer.getvalue(),
            content_type='application/zip',
            filename=filename_map[export_format],
        )



def _turni_weekend_outlook_package_filename(week_label):
    normalized = unicodedata.normalize('NFKD', week_label or '')
    normalized = normalized.encode('ascii', 'ignore').decode('ascii')
    normalized = re.sub(r'[^a-zA-Z0-9]+', '-', normalized).strip('-').lower()
    return f'turni-weekend-outlook-{normalized or "settimana"}.zip'


def _turni_email_recipients_from_text(raw_value):
    return [item.strip() for item in re.split(r'[;,]+', raw_value or '') if item.strip()]


TURNI_PLANNER_MAIL_ATTACHMENT_OPTIONS = [
    {'key': 'weekly', 'label': 'Turno settimanale'},
    {'key': 'portineria_weekly', 'label': 'Turno settimanale portineria'},
    {'key': 'saturday', 'label': 'Comandata sabato'},
    {'key': 'sunday', 'label': 'Comandata domenica'},
    {'key': 'jolly_weekend', 'label': 'Comandata jolly'},
    {'key': 'scorrimento', 'label': 'Scorrimento'},
    {'key': 'portineria_weekend', 'label': 'Sabato - Domenica e festivi Portineria'},
]


def _turni_planner_weekend_mail_response(state, *, recipient_text='', subject_text='', body_text='', selected_attachment_keys=None, selected_file_types=None):
    planner_data = dict(state.planner_data or {})
    logo_path = _existing_turni_export_path(TURNI_EXPORT_APP_LOGO_PATH)
    ancis_logo_path = _existing_turni_export_path(TURNI_EXPORT_WEEKEND_ANCIS_LOGO_PATH)
    anid_logo_path = _existing_turni_export_path(TURNI_EXPORT_WEEKEND_ANID_LOGO_PATH)

    weekly_configs = [
        {
            'key': 'weekly',
            'pdf_name': WEEKLY_PDF_NAME,
            'image_name': WEEKLY_IMAGE_NAME,
            'title': 'Turno settimanale',
            'layout': 'default',
            'builder': lambda: _turni_weekly_sections_for_export(planner_data.get('weekly')),
            'label_key': 'weekly_export_week_label',
        },
        {
            'key': 'portineria_weekly',
            'pdf_name': PORTINERIA_WEEKLY_PDF_NAME,
            'image_name': PORTINERIA_WEEKLY_IMAGE_NAME,
            'title': 'Turno settimanale portineria',
            'layout': 'portineria',
            'builder': lambda: _turni_portineria_weekly_sections_for_export(planner_data.get('portineria_weekly')),
            'label_key': 'portineria_weekly_export_week_label',
        },
    ]

    weekend_configs = [
        {
            'key': 'saturday',
            'pdf_name': SATURDAY_PDF_NAME,
            'image_name': SATURDAY_IMAGE_NAME,
            'title': 'Comandata sabato',
        },
        {
            'key': 'sunday',
            'pdf_name': SUNDAY_PDF_NAME,
            'image_name': SUNDAY_IMAGE_NAME,
            'title': 'Comandata domenica',
        },
        {
            'key': 'jolly_weekend',
            'pdf_name': JOLLY_WEEKEND_PDF_NAME,
            'image_name': JOLLY_WEEKEND_IMAGE_NAME,
            'title': 'Comandata jolly',
        },
        {
            'key': 'scorrimento',
            'pdf_name': SCORRIMENTO_PDF_NAME,
            'image_name': SCORRIMENTO_IMAGE_NAME,
            'title': 'Scorrimento',
        },
        {
            'key': 'portineria_weekend',
            'pdf_name': PORTINERIA_WEEKEND_PDF_NAME,
            'image_name': PORTINERIA_WEEKEND_IMAGE_NAME,
            'title': 'Sabato - Domenica e festivi Portineria',
        },
    ]

    attachments = []
    attachment_labels = []
    normalized_attachment_keys = None
    if selected_attachment_keys is not None:
        normalized_attachment_keys = {
            value.strip()
            for value in selected_attachment_keys
            if value and value.strip()
        }
        if not normalized_attachment_keys:
            raise ValueError('Seleziona almeno un allegato da inviare.')
    normalized_file_types = None
    if selected_file_types is not None:
        normalized_file_types = {
            value.strip().lower()
            for value in selected_file_types
            if value and value.strip()
        }
        normalized_file_types &= {'pdf', 'jpg'}
        if not normalized_file_types:
            raise ValueError('Seleziona almeno un formato da inviare.')

    include_pdf = normalized_file_types is None or 'pdf' in normalized_file_types
    include_jpg = normalized_file_types is None or 'jpg' in normalized_file_types

    with tempfile.TemporaryDirectory(prefix='turni_planner_weekend_mail_') as temp_dir:
        export_dir = Path(temp_dir)
        for config in weekly_configs:
            if normalized_attachment_keys is not None and config['key'] not in normalized_attachment_keys:
                continue
            export_week_label = _resolve_turni_export_week_label(
                planner_data,
                state.week_label,
                key=config['label_key'],
            )
            headers, sections = config['builder']()
            pdf_path = export_weekly_pdf(
                export_dir / config['pdf_name'],
                title_text=TURNI_DEFAULT_WEEKLY_PDF_TITLE,
                week_label=export_week_label,
                signature='',
                headers=headers,
                sections=sections,
                logo_path=logo_path,
                layout=config['layout'],
            )
            if include_pdf:
                attachments.append((config['pdf_name'], pdf_path.read_bytes(), 'application/pdf'))

            if include_jpg:
                image_paths = export_weekly_images(
                    export_dir / config['image_name'],
                    title_text=TURNI_DEFAULT_WEEKLY_PDF_TITLE,
                    week_label=export_week_label,
                    signature='',
                    headers=headers,
                    sections=sections,
                    logo_path=logo_path,
                    temp_pdf_name=config['pdf_name'],
                    layout=config['layout'],
                )
                attachments.append((config['image_name'], _turni_combined_jpg_bytes(image_paths), 'image/jpeg'))
            if include_pdf or include_jpg:
                attachment_labels.append(config['title'])

        for config in weekend_configs:
            if normalized_attachment_keys is not None and config['key'] not in normalized_attachment_keys:
                continue
            raw_weekend_data = planner_data.get(config['key'])
            export_data = _turni_weekend_export_data(
                raw_weekend_data,
                title=config['title'],
                row_count=None,
            )
            exported_path = export_weekend_pdf(
                export_dir / config['pdf_name'],
                data=export_data,
                logo_path=logo_path,
                cert_logo_path=ancis_logo_path,
                anid_logo_path=anid_logo_path,
            )
            image_paths = []
            if include_jpg:
                image_paths = export_weekend_images(
                    export_dir / config['image_name'],
                    data=export_data,
                    logo_path=logo_path,
                    cert_logo_path=ancis_logo_path,
                    anid_logo_path=anid_logo_path,
                )
            if config['key'] == 'scorrimento':
                exported_path = export_scorrimento_pdf(
                    export_dir / config['pdf_name'],
                    data=_turni_scorrimento_export_data(raw_weekend_data),
                    logo_path=logo_path,
                    cert_logo_path=ancis_logo_path,
                    anid_logo_path=anid_logo_path,
                )
                if include_jpg:
                    image_paths = export_scorrimento_images(
                        export_dir / config['image_name'],
                        data=_turni_scorrimento_export_data(raw_weekend_data),
                        logo_path=logo_path,
                        cert_logo_path=ancis_logo_path,
                        anid_logo_path=anid_logo_path,
                    )
            if include_pdf:
                attachments.append((config['pdf_name'], exported_path.read_bytes(), 'application/pdf'))
            if include_jpg:
                attachments.append((config['image_name'], _turni_combined_jpg_bytes(image_paths), 'image/jpeg'))
            if include_pdf or include_jpg:
                attachment_labels.append(export_data.title or config['title'])

    recipients = _turni_email_recipients_from_text(recipient_text)
    if not recipients:
        raise ValueError('Inserisci almeno un destinatario email.')
    subject = subject_text.strip() or f'Turni planner {state.week_label}'
    normalized_body = (body_text or '').replace('\r\n', '\n').replace('\r', '\n').strip('\n')
    format_label = 'PDF e JPG'
    if include_pdf and not include_jpg:
        format_label = 'PDF'
    elif include_jpg and not include_pdf:
        format_label = 'JPG'
    if normalized_body:
        body_lines = normalized_body.split('\n')
        body_lines.extend([
            '',
            'Allegati inclusi:',
        ])
    else:
        body_lines = [
            'Buongiorno,',
            '',
            f'in allegato trovate i {format_label} del planner della settimana {state.week_label}.',
            '',
            'Allegati inclusi:',
        ]
    body_lines.extend(f'- {label}' for label in attachment_labels)
    body_lines.extend([
        '',
        'Cordiali saluti',
    ])

    email = EmailMultiAlternatives(
        subject=subject,
        body='\n'.join(body_lines),
        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', '') or 'cedolini@sanvincenzosrl.com',
        to=recipients,
    )
    for filename, content, content_type in attachments:
        email.attach(filename, content, content_type)
    email.send(fail_silently=False)
    return recipients


def _employee_admin_display_name(employee):
    parts = [part.strip() for part in (employee.first_name, employee.last_name) if part and part.strip()]
    if parts:
        return ' '.join(parts)
    return getattr(getattr(employee, 'user', None), 'username', '') or ''


def _employee_name_sort_key(employee):
    first_name = (employee.first_name or '').strip().casefold()
    last_name = (employee.last_name or '').strip().casefold()
    username = (getattr(getattr(employee, 'user', None), 'username', '') or '').strip().casefold()
    return (first_name, last_name, username, employee.id)


def _decorate_employee_display_names(employees):
    for employee in employees:
        employee.display_name = _employee_admin_display_name(employee)
    return employees


def _normalize_import_name(value):
    value = unicodedata.normalize('NFKD', value or '').encode('ascii', 'ignore').decode('ascii')
    value = value.lower().replace("'", ' ')
    value = re.sub(r'[^a-z0-9]+', ' ', value)
    return ' '.join(value.split())


def _import_lookup_keys(left, right):
    left_norm = _normalize_import_name(left)
    right_norm = _normalize_import_name(right)
    keys = []
    if left_norm and right_norm:
        keys.append(f'{left_norm}|{right_norm}')
        left_compact = left_norm.replace(' ', '')
        right_compact = right_norm.replace(' ', '')
        keys.append(f'{left_compact}|{right_compact}')
    return list(dict.fromkeys(keys))


def _employee_candidate_priority(employee):
    user = employee.user
    return (
        1 if user.is_active else 0,
        1 if employee.privacy_accepted else 0,
        1 if employee.invito_inviato else 0,
        employee.payslips.count(),
        -employee.id,
    )


def _build_employee_import_lookup():
    lookup = {}
    employees = list(Employee.objects.select_related('user').prefetch_related('payslips'))

    def register(left, right, employee):
        for key in _import_lookup_keys(left, right):
            bucket = lookup.setdefault(key, [])
            if employee not in bucket:
                bucket.append(employee)

    for employee in employees:
        pairs = [
            (employee.first_name, employee.last_name),
            (employee.user.first_name, employee.user.last_name),
        ]
        for first_name, last_name in pairs:
            if not first_name or not last_name:
                continue
            register(last_name, first_name, employee)
            register(first_name, last_name, employee)

        username_parts = [part for part in re.split(r'[-_\s]+', employee.user.username or '') if part]
        if len(username_parts) >= 2:
            for i in range(1, len(username_parts)):
                left = ' '.join(username_parts[:i]).strip()
                right = ' '.join(username_parts[i:]).strip()
                register(left, right, employee)
                register(right, left, employee)

    return lookup


def _find_employee_for_import_tokens(name_tokens, employee_lookup):
    candidates = []

    for i in range(1, len(name_tokens)):
        left = ' '.join(name_tokens[:i]).strip()
        right = ' '.join(name_tokens[i:]).strip()
        if not left or not right:
            continue
        for key in _import_lookup_keys(left, right):
            candidates.extend(employee_lookup.get(key, []))

    unique_candidates = []
    seen_ids = set()
    for employee in candidates:
        if employee.id in seen_ids:
            continue
        seen_ids.add(employee.id)
        unique_candidates.append(employee)

    if not unique_candidates:
        return None

    unique_candidates.sort(key=_employee_candidate_priority, reverse=True)
    return unique_candidates[0]


def _parse_cud_import_filename(filename):
    stem = os.path.splitext(filename or '')[0].strip()
    normalized = stem.replace('_', ' ').replace('-', ' ')

    year_match = re.search(r'(19|20)\d{2}', normalized)
    if not year_match:
        return None, None, 'anno non trovato nel filename'

    year = int(year_match.group(0))
    normalized = normalized[:year_match.start()] + ' ' + normalized[year_match.end():]
    normalized = re.sub(r'\bcertificazione\s+unica\b', ' ', normalized, flags=re.IGNORECASE)
    normalized = re.sub(r'\bcu\b|\bcud\b', ' ', normalized, flags=re.IGNORECASE)
    normalized = re.sub(r'\s+', ' ', normalized).strip()

    name_tokens = [token for token in normalized.split(' ') if token]
    if len(name_tokens) < 2:
        return None, None, 'nome dipendente non riconosciuto nel filename'

    return year, name_tokens, None


def _pending_import_session_key(document_type):
    return f'pending_{document_type}_import'


def _pending_import_storage_name(document_type, filename):
    safe_name = os.path.basename(filename or 'documento.pdf')
    safe_name = re.sub(r'[^A-Za-z0-9._-]+', '_', safe_name)
    return f'pending/{document_type}/{uuid.uuid4().hex}_{safe_name}'


def _get_pending_import_storage():
    pending_root = os.path.join(settings.MEDIA_ROOT, 'pending_imports')
    os.makedirs(pending_root, exist_ok=True)
    return FileSystemStorage(location=pending_root, base_url=None)


def _store_pending_uploaded_file(uploaded_file, document_type):
    pending_storage = _get_pending_import_storage()
    return pending_storage.save(_pending_import_storage_name(document_type, uploaded_file.name), uploaded_file)


def _delete_pending_files(records):
    pending_storage = _get_pending_import_storage()
    for record in records or []:
        temp_path = record.get('temp_path')
        if not temp_path:
            continue
        try:
            if pending_storage.exists(temp_path):
                pending_storage.delete(temp_path)
        except Exception:
            logger.exception('Error deleting pending upload file %s', temp_path)


def _clear_pending_import(request, document_type):
    session_key = _pending_import_session_key(document_type)
    pending_data = request.session.pop(session_key, None)
    if pending_data:
        _delete_pending_files(pending_data.get('records', []))


def _display_name_from_tokens(name_tokens):
    return ' '.join(token for token in (name_tokens or []) if token).strip()


def _candidate_key_from_tokens(name_tokens):
    normalized = _normalize_import_name(_display_name_from_tokens(name_tokens))
    return normalized.replace(' ', '-') or uuid.uuid4().hex


def _guess_identity_from_tokens(name_tokens):
    parts = [part.strip() for part in (name_tokens or []) if part.strip()]
    if not parts:
        return '', ''
    if len(parts) == 1:
        return parts[0].title(), ''
    last_name = parts[0].title()
    first_name = ' '.join(parts[1:]).title()
    return first_name, last_name


def _username_slug_part(value):
    value = unicodedata.normalize('NFKD', value or '').encode('ascii', 'ignore').decode('ascii')
    value = value.lower().replace("'", '')
    value = re.sub(r'[^a-z0-9]+', '-', value)
    return value.strip('-')


def _generate_unique_import_username(first_name, last_name):
    base = '-'.join(part for part in [_username_slug_part(last_name), _username_slug_part(first_name)] if part)
    if not base:
        base = f'dipendente-{uuid.uuid4().hex[:8]}'

    candidate = base
    counter = 2
    while User.objects.filter(username=candidate).exists():
        candidate = f'{base}-{counter}'
        counter += 1
    return candidate


def _create_employee_for_import(first_name, last_name):
    username = _generate_unique_import_username(first_name, last_name)
    user = User.objects.create_user(
        username=username,
        password=secrets.token_urlsafe(16),
        is_active=False,
        first_name=first_name,
        last_name=last_name,
    )
    employee = Employee.objects.create(
        user=user,
        first_name=first_name,
        last_name=last_name,
    )
    return employee, username


def _build_pending_import_data(records, document_type):
    missing_candidates = OrderedDict()
    for record in records:
        if record.get('status') != 'missing':
            continue
        key = record['candidate_key']
        candidate = missing_candidates.setdefault(key, {
            'key': key,
            'display_name': record['display_name'],
            'suggested_first_name': record['suggested_first_name'],
            'suggested_last_name': record['suggested_last_name'],
            'file_count': 0,
        })
        candidate['file_count'] += 1

    return {
        'document_type': document_type,
        'records': records,
        'missing_candidates': list(missing_candidates.values()),
    }


def _build_import_record(document_type, uploaded_file, *, name_tokens=None, employee=None, reason=None, year=None, month=None):
    display_name = _display_name_from_tokens(name_tokens)
    suggested_first_name, suggested_last_name = _guess_identity_from_tokens(name_tokens)

    if reason:
        status = 'skipped'
    elif employee and employee.user_id and employee.user.is_active:
        status = 'existing_active'
    elif employee:
        status = 'existing_inactive'
        reason = 'account non attivo'
    else:
        status = 'missing'

    temp_path = None
    if status in {'existing_active', 'missing'}:
        temp_path = _store_pending_uploaded_file(uploaded_file, document_type)

    return {
        'document_type': document_type,
        'filename': uploaded_file.name,
        'temp_path': temp_path,
        'name_tokens': list(name_tokens or []),
        'display_name': display_name,
        'suggested_first_name': suggested_first_name,
        'suggested_last_name': suggested_last_name,
        'candidate_key': _candidate_key_from_tokens(name_tokens or []),
        'employee_id': employee.id if employee else None,
        'year': year,
        'month': month,
        'status': status,
        'reason': reason,
    }


def _save_payslip_from_record(record, employee):
    pending_storage = _get_pending_import_storage()
    with transaction.atomic():
        existing = Payslip.objects.filter(employee=employee, year=record['year'], month=record['month']).first()
        if existing:
            try:
                existing.pdf.delete(save=False)
            except Exception:
                logger.exception('Error deleting existing payslip file for %s', existing.id)
            existing.delete()

        with pending_storage.open(record['temp_path'], 'rb') as handle:
            payslip = Payslip(employee=employee, year=record['year'], month=record['month'])
            payslip.pdf.save(record['filename'], File(handle), save=True)
        return payslip.id


def _save_cud_from_record(record, employee):
    replaced = False
    pending_storage = _get_pending_import_storage()
    with transaction.atomic():
        existing = Cud.objects.filter(employee=employee, year=record['year']).first()
        if existing:
            try:
                existing.pdf.delete(save=False)
            except Exception:
                logger.exception('Error deleting existing CUD file for %s', existing.id)
            existing.delete()
            replaced = True

        with pending_storage.open(record['temp_path'], 'rb') as handle:
            cud_obj = Cud(employee=employee, year=record['year'])
            cud_obj.pdf.save(record['filename'], File(handle), save=True)
        return cud_obj.id, replaced


def _render_missing_account_resolution(request, document_type, pending_data):
    document_label = 'Cedolini' if document_type == 'payslip' else 'CUD'
    return render(request, 'portal/admin_resolve_missing_accounts.html', {
        'document_type': document_type,
        'document_label': document_label,
        'missing_candidates': pending_data.get('missing_candidates', []),
        'records': pending_data.get('records', []),
        'action_url_name': 'admin_upload_period_folder' if document_type == 'payslip' else 'admin_upload_cud',
    })


def _finalize_pending_import(request, document_type, pending_data, create_selected_keys, posted_data=None):
    posted_data = posted_data or {}
    created_usernames = []
    created_payslip_ids = []
    created_cuds = 0
    replaced_cuds = 0
    skipped = []
    created_employees = {}
    import_job = None
    total_files = len(pending_data.get('records', []))
    processed_files = 0

    if document_type == 'payslip':
        import_job = ImportJob.objects.create(total_files=total_files, status='processing')

    for candidate in pending_data.get('missing_candidates', []):
        key = candidate['key']
        if key not in create_selected_keys:
            continue

        first_name = (posted_data.get(f'first_name_{key}') or candidate.get('suggested_first_name') or '').strip()
        last_name = (posted_data.get(f'last_name_{key}') or candidate.get('suggested_last_name') or '').strip()

        if not first_name or not last_name:
            created_employees[key] = {'error': 'nome o cognome mancanti per la creazione account'}
            continue

        employee, username = _create_employee_for_import(first_name, last_name)
        created_employees[key] = {'employee': employee}
        created_usernames.append(username)

    try:
        for record in pending_data.get('records', []):
            processed_files += 1
            status = record.get('status')

            if status == 'skipped':
                skipped.append((record['filename'], record.get('reason') or 'file scartato'))
                continue

            if status == 'existing_inactive':
                skipped.append((record['filename'], record.get('reason') or 'account non attivo'))
                continue

            employee = None
            if status == 'existing_active' and record.get('employee_id'):
                employee = Employee.objects.select_related('user').filter(id=record['employee_id']).first()
            elif status == 'missing':
                resolution = created_employees.get(record['candidate_key'])
                if resolution and resolution.get('employee'):
                    employee = resolution['employee']
                elif resolution and resolution.get('error'):
                    skipped.append((record['filename'], resolution['error']))
                    continue
                else:
                    skipped.append((record['filename'], 'account non creato'))
                    continue

            if not employee:
                skipped.append((record['filename'], 'dipendente non disponibile'))
                continue

            try:
                if document_type == 'payslip':
                    created_payslip_ids.append(_save_payslip_from_record(record, employee))
                else:
                    _, replaced = _save_cud_from_record(record, employee)
                    created_cuds += 1
                    if replaced:
                        replaced_cuds += 1
            except IntegrityError:
                logger.exception('Integrity error during %s import for file %s', document_type, record['filename'])
                skipped.append((record['filename'], 'integrity error'))
            except Exception as exc:
                logger.exception('Error during %s import for file %s', document_type, record['filename'])
                skipped.append((record['filename'], str(exc)))
    finally:
        _delete_pending_files(pending_data.get('records', []))

    if document_type == 'payslip':
        import_job.processed_files = processed_files
        import_job.created_users = len(created_usernames)
        import_job.created_payslips = len(created_payslip_ids)
        import_job.skipped = len(skipped)
        import_job.status = 'completed'
        import_job.save()

        _create_audit_event(
            request,
            'payslip_import_completed',
            metadata={
                'import_job_id': import_job.id,
                'total_files': total_files,
                'processed_files': processed_files,
                'created_users': len(created_usernames),
                'created_payslips': len(created_payslip_ids),
                'skipped': len(skipped),
                'status': import_job.status,
            },
        )

        request.session['last_import_created_users'] = created_usernames
        request.session['last_import_created_payslips'] = created_payslip_ids

        return render(request, 'portal/admin_confirm_import.html', {
            'created_users': created_usernames,
            'created_payslips': created_payslip_ids,
            'skipped': skipped,
            'import_job': import_job,
        })

    _create_audit_event(
        request,
        'cud_import_completed',
        metadata={
            'total_files': total_files,
            'processed_files': processed_files,
            'created_users': len(created_usernames),
            'created_cuds': created_cuds,
            'replaced': replaced_cuds,
            'skipped': len(skipped),
        },
    )

    return render(request, 'portal/admin_confirm_cud_import.html', {
        'total_files': total_files,
        'processed_files': processed_files,
        'created_cuds': created_cuds,
        'created_users': created_usernames,
        'replaced_count': replaced_cuds,
        'skipped': skipped,
    })


def _disable_response_cache(response):
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


# =========================================================
# Utility: Audit logging
# =========================================================

def _create_audit_event(request, action, *, employee=None, payslip=None, metadata=None):
    """Crea un AuditEvent di base con IP e user-agent.

    Usato per popolare la pagina "Storico azioni" con eventi
    significativi (apertura cedolino, import massivi, inviti, ecc.).
    """
    try:
        ip = request.META.get('REMOTE_ADDR')
        ua = request.META.get('HTTP_USER_AGENT', '')
        AuditEvent.objects.create(
            action=action,
            actor_user=request.user if getattr(request, 'user', None) and request.user.is_authenticated else None,
            employee=employee,
            payslip=payslip,
            ip_address=ip,
            user_agent=ua,
            metadata=metadata or {},
        )
    except Exception:
        # Non bloccare il flusso se il log fallisce
        logger.exception("Errore nella creazione di AuditEvent (%s)", action)


def _send_out_of_zone_request_admin_notification(request_obj):
    recipients = [email for email in getattr(settings, 'ADMIN_NOTIFICATION_EMAILS', []) if email]
    if not recipients:
        return

    employee = request_obj.employee
    mark_type_label = dict(WorkMarkRequest.MARK_TYPE_CHOICES).get(request_obj.mark_type, request_obj.mark_type)
    app_base_url = getattr(settings, 'APP_BASE_URL', '').rstrip('/')
    admin_url = f"{app_base_url}/portal/admin-richieste-fuori-zona/" if app_base_url else '/portal/admin-richieste-fuori-zona/'

    subject = f"Nuova richiesta fuori zona - {employee.full_name}"
    body = (
        "E' stata inviata una nuova richiesta di marcatura fuori zona.\n\n"
        f"Dipendente: {employee.full_name}\n"
        f"Data lavoro: {request_obj.work_date:%d/%m/%Y}\n"
        f"Tipo: {mark_type_label}\n"
        f"Motivazione: {request_obj.reason}\n"
        f"Richiesta ID: {request_obj.id}\n\n"
        f"Apri lo storico richieste: {admin_url}\n"
    )

    try:
        email = EmailMultiAlternatives(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipients,
        )
        email.send(fail_silently=False)
    except Exception:
        logger.exception(
            'Errore invio notifica admin per richiesta fuori zona id=%s',
            request_obj.id,
        )


def _send_vacation_request_admin_notification(request_obj):
    recipients = [email for email in getattr(settings, 'ADMIN_NOTIFICATION_EMAILS', []) if email]
    if not recipients:
        return

    employee = request_obj.employee
    app_base_url = getattr(settings, 'APP_BASE_URL', '').rstrip('/')
    admin_url = f"{app_base_url}/portal/admin-richieste-ferie/" if app_base_url else '/portal/admin-richieste-ferie/'

    subject = f"Nuova richiesta ferie - {employee.full_name}"
    body = (
        "E' stata inviata una nuova richiesta ferie.\n\n"
        f"Dipendente: {employee.full_name}\n"
        f"Periodo: {request_obj.start_date:%d/%m/%Y} - {request_obj.end_date:%d/%m/%Y}\n"
        f"Giorni richiesti: {request_obj.day_count()}\n"
        f"Motivazione: {request_obj.reason}\n"
        f"Richiesta ID: {request_obj.id}\n\n"
        f"Apri lo storico richieste ferie: {admin_url}\n"
    )

    try:
        email = EmailMultiAlternatives(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipients,
        )
        email.send(fail_silently=False)
    except Exception:
        logger.exception(
            'Errore invio notifica admin per richiesta ferie id=%s',
            request_obj.id,
        )


def _parse_date_or_none(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _attach_payslip_display_period(payslip):
    display_month = payslip.month
    display_year = payslip.year

    try:
        _, _, parsed_month, parsed_year = parse_payslip_filename(os.path.basename(payslip.pdf.name))
        display_month = parsed_month
        display_year = parsed_year
    except Exception:
        pass

    payslip.display_month = display_month
    payslip.display_year = display_year
    payslip.display_month_name = MONTH_LABELS_IT.get(display_month, str(display_month))
    return payslip


# =========================================================
# HOME
# =========================================================

def home(request):
    if request.user.is_authenticated:
        return redirect(user_home_url_name(request.user))
    return render(request, 'site/home.html')


def public_home(request):
    return render(request, 'site/home.html')


def public_services(request):
    return render(request, 'site/services.html')


def public_digital_services(request):
    return render(request, 'site/digital_services.html')


def public_about(request):
    return render(request, 'site/about.html')


def public_machinery(request):
    machinery_items = [
        {
            'title': 'Bobcat',
            'category': 'Compatto operativo',
            'description': 'Mezzo compatto pensato per movimentazione rapida, riassetto aree di lavoro e interventi agili in spazi contenuti.',
            'image': 'portal/machinery/bobcat.jpg',
            'gallery_images': ['portal/machinery/bobcat.jpg'],
            'delay': '.04s',
        },
        {
            'title': 'Carrello elevatore retrattile',
            'category': 'Logistica verticale',
            'description': 'Ideale per stoccaggio in altezza e movimentazione precisa in corsie di magazzino ad alta densita operativa.',
            'image': 'portal/machinery/carrello-elevatore-retrattile.jpg',
            'gallery_images': ['portal/machinery/carrello-elevatore-retrattile.jpg'],
            'delay': '.08s',
        },
        {
            'title': 'Carrello elevatore',
            'category': 'Movimentazione carichi',
            'description': 'Supporto affidabile per carico, scarico e trasferimento di pallet, materiali e forniture nei flussi quotidiani.',
            'image': 'portal/machinery/carrello-elevatore.jpg',
            'gallery_images': ['portal/machinery/carrello-elevatore.jpg'],
            'delay': '.12s',
        },
        {
            'title': 'Escavatore compatto',
            'category': 'Interventi di precisione',
            'description': 'Soluzione efficace per scavi localizzati, rifiniture tecniche e cantieri con accessi piu contenuti.',
            'image': 'portal/machinery/escavatore-piccolo.jpg',
            'gallery_images': ['portal/machinery/escavatore-piccolo.jpg'],
            'delay': '.16s',
        },
        {
            'title': 'Escavatore',
            'category': 'Scavo e movimentazione',
            'description': 'Mezzo operativo per lavori di scavo, movimentazione terra e attivita strutturate con maggiore capacita di azione.',
            'image': 'portal/machinery/escavatore.jpg',
            'gallery_images': ['portal/machinery/escavatore.jpg'],
            'delay': '.20s',
        },
        {
            'title': 'Magazzino operativo con carrelli',
            'category': 'Logistica interna',
            'description': 'Area organizzata per deposito, preparazione materiali e movimentazione con carrelli, pensata per garantire ordine operativo e continuita alle attivita di supporto.',
            'image': 'portal/machinery/magazzino.jpg',
            'gallery_images': [
                'portal/machinery/magazzino.jpg',
                'portal/machinery/magazzino-con-carrello.jpg',
            ],
            'delay': '.24s',
        },
        {
            'title': 'Officina mobile',
            'category': 'Assistenza sul campo',
            'description': 'Unita mobile attrezzata per supporto tecnico, manutenzioni rapide e interventi direttamente in sede operativa.',
            'image': 'portal/machinery/officina-mobile.jpg',
            'gallery_images': [
                'portal/machinery/officina-mobile.jpg',
                'portal/machinery/officina-mobile-2.jpg',
                'portal/machinery/officina-mobile-3.jpg',
            ],
            'delay': '.28s',
        },
        {
            'title': 'Piattaforma aerea a pantografo',
            'category': 'Lavori in quota',
            'description': 'Soluzione stabile e funzionale per lavorazioni in elevazione, manutenzioni e accessi sicuri su superfici verticali.',
            'image': 'portal/machinery/piattaforma-aerea-a-pantografo.jpg',
            'gallery_images': ['portal/machinery/piattaforma-aerea-a-pantografo.jpg'],
            'delay': '.32s',
        },
        {
            'title': 'Piattaforma aerea autocarrata',
            'category': 'Quota e mobilita',
            'description': 'Mezzo versatile per raggiungere rapidamente aree di lavoro in altezza con flessibilita di posizionamento.',
            'image': 'portal/machinery/piattaforma-aerea-autocarrata.jpg',
            'gallery_images': [
                'portal/machinery/piattaforma-aerea-autocarrata.jpg',
                'portal/machinery/piattaforma-aerea-autocarrata-1.jpg',
                'portal/machinery/piattaforma-aerea-autocarrata-2.jpg',
            ],
            'delay': '.36s',
        },
        {
            'title': 'Piattaforma aerea semovente a braccio articolato JLG E300',
            'category': 'Quota e precisione',
            'description': 'Piattaforma semovente a braccio articolato pensata per raggiungere punti complessi in quota con manovrabilita, precisione di posizionamento e sicurezza operativa.',
            'image': 'portal/machinery/piattaforma-aerea-semovente-jlg-e300-1.jpg',
            'gallery_images': [
                'portal/machinery/piattaforma-aerea-semovente-jlg-e300-1.jpg',
                'portal/machinery/piattaforma-aerea-semovente-jlg-e300-2.jpg',
            ],
            'delay': '.40s',
        },
        {
            'title': 'Terna',
            'category': 'Multiuso cantiere',
            'description': 'Macchina polivalente per scavo, carico e movimentazione, utile quando servono operativita e rapidita nello stesso mezzo.',
            'image': 'portal/machinery/terna.jpg',
            'gallery_images': ['portal/machinery/terna.jpg'],
            'delay': '.44s',
        },
        {
            'title': 'Trattore stradale con semirimorchio',
            'category': 'Trasporto dedicato',
            'description': 'Configurazione per movimentazione su strada e supporto logistico nelle attivita che richiedono continuita e portata.',
            'image': 'portal/machinery/trattore-stradale-semirimorchio.jpg',
            'gallery_images': [
                'portal/machinery/trattore-stradale-semirimorchio.jpg',
                'portal/machinery/trattore-stradale-semirimorchio-2.jpg',
            ],
            'delay': '.48s',
        },
        {
            'title': 'Gruppo elettrogeno industriale',
            'category': 'Energia di supporto',
            'description': 'Sistema di alimentazione industriale pensato per garantire continuita operativa, autonomia energetica e supporto alle attivita in campo o in aree non servite.',
            'image': 'portal/machinery/gruppo-elettrogeno-industriale-2.jpg',
            'gallery_images': [
                'portal/machinery/gruppo-elettrogeno-industriale-2.jpg',
                'portal/machinery/gruppo-elettrogeno-industriale-1.jpg',
                'portal/machinery/gruppo-elettrogeno-industriale-3.jpg',
                'portal/machinery/gruppo-elettrogeno-industriale-4.jpg',
            ],
            'delay': '.52s',
        },
        {
            'title': 'Autospurgo canal-jet su Iveco Stralis a 4 assi',
            'category': 'Spurgo e lavaggi tecnici',
            'description': 'Mezzo specializzato per interventi di spurgo, pulizia condotte e lavaggi ad alta efficienza, configurato su telaio a quattro assi per operativita strutturate.',
            'image': 'portal/machinery/autospurgo-canal-jet-iveco-stralis-1.jpg',
            'gallery_images': [
                'portal/machinery/autospurgo-canal-jet-iveco-stralis-1.jpg',
                'portal/machinery/autospurgo-canal-jet-iveco-stralis-2.jpg',
            ],
            'delay': '.56s',
        },
        {
            'title': 'Atomizzatore per pest control',
            'category': 'Nebulizzazione tecnica',
            'description': 'Attrezzatura ad alta resa per nebulizzazione e controllo dei parassiti, indicata per interventi mirati di pest control in contesti operativi e ambientali.',
            'image': 'portal/machinery/atomizzatore-pest-control-1.jpg',
            'gallery_images': [
                'portal/machinery/atomizzatore-pest-control-1.jpg',
                'portal/machinery/atomizzatore-pest-control-2.jpg',
            ],
            'delay': '.60s',
        },
        {
            'title': 'Spazzatrice stradale aspirante Dulevo D6',
            'category': 'Igiene urbana',
            'description': 'Spazzatrice aspirante per pulizia stradale e decoro urbano, adatta a interventi continui su viabilita, piazzali e superfici estese.',
            'image': 'portal/machinery/spazzatrice-stradale-dulevo-d6-1.jpg',
            'gallery_images': [
                'portal/machinery/spazzatrice-stradale-dulevo-d6-1.jpg',
                'portal/machinery/spazzatrice-stradale-dulevo-d6-2.jpg',
            ],
            'delay': '.64s',
        },
        {
            'title': 'Autocarro con gru retrocabina',
            'category': 'Sollevamento e trasporto',
            'description': 'Mezzo operativo pensato per carico, scarico e movimentazione di materiali con autonomia di sollevamento direttamente sul punto di intervento, con piu viste disponibili nella galleria dedicata.',
            'image': 'portal/machinery/autocarro-con-gru-retrocabina.jpg',
            'gallery_images': [
                'portal/machinery/autocarro-con-gru-retrocabina.jpg',
                'portal/machinery/autocarro-con-gru-retrocabina-2.jpg',
            ],
            'delay': '.68s',
        },
    ]
    return render(request, 'site/machinery.html', {
        'machinery_items': machinery_items,
    })


def public_contacts(request):
    contact_email = 'antimo.digiovanni@sanvincenzosrl.com'
    form_data = {
        'name': '',
        'email': '',
        'phone': '',
        'subject': '',
        'message': '',
    }
    contact_error = None
    contact_success = False

    if request.method == 'POST':
        form_data = {
            'name': request.POST.get('name', '').strip(),
            'email': request.POST.get('email', '').strip(),
            'phone': request.POST.get('phone', '').strip(),
            'subject': request.POST.get('subject', '').strip(),
            'message': request.POST.get('message', '').strip(),
        }

        if not form_data['name'] or not form_data['email'] or not form_data['message']:
            contact_error = 'Compila almeno nome, email e messaggio.'
        else:
            email_subject = form_data['subject'] or 'Nuova richiesta dal sito San Vincenzo SRL'
            email_body = (
                'Nuova richiesta ricevuta dal sito web.\n\n'
                f"Nome: {form_data['name']}\n"
                f"Email: {form_data['email']}\n"
                f"Telefono: {form_data['phone'] or 'Non indicato'}\n"
                f"Oggetto: {form_data['subject'] or 'Non indicato'}\n\n"
                'Messaggio:\n'
                f"{form_data['message']}\n"
            )

            from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', '') or 'noreply@sanvincenzosrl.com'

            try:
                mail = EmailMultiAlternatives(
                    subject=email_subject,
                    body=email_body,
                    from_email=from_email,
                    to=[contact_email],
                    reply_to=[form_data['email']],
                )
                mail.send(fail_silently=False)
                contact_success = True
                form_data = {
                    'name': '',
                    'email': '',
                    'phone': '',
                    'subject': '',
                    'message': '',
                }
            except Exception:
                logger.exception('Errore invio richiesta contatti sito pubblico')
                contact_error = 'Invio non riuscito al momento. Puoi scriverci direttamente via email o telefono.'

    return render(request, 'site/contacts.html', {
        'contact_email': contact_email,
        'form_data': form_data,
        'contact_error': contact_error,
        'contact_success': contact_success,
    })


def sitemap_xml(request):
    pages = [
        '',
        'chi-siamo/',
        'servizi/',
        'servizi-digitali/',
        'macchinari/',
        'contatti/',
        'login/',
    ]
    base_url = request.build_absolute_uri('/').rstrip('/')
    xml_items = []
    for page in pages:
        xml_items.append(
            f"<url><loc>{base_url}/{page}</loc><changefreq>weekly</changefreq><priority>0.8</priority></url>"
        )

    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">'
        + ''.join(xml_items) +
        '</urlset>'
    )
    return HttpResponse(xml, content_type='application/xml')


def robots_txt(request):
    base_url = request.build_absolute_uri('/').rstrip('/')
    content = (
        'User-agent: *\n'
        'Allow: /\n\n'
        f'Sitemap: {base_url}/sitemap.xml\n'
    )
    return HttpResponse(content, content_type='text/plain')


def site_webmanifest(request):
    """Manifest PWA per installazione su schermata Home Android."""
    icon_192_url = request.build_absolute_uri(static('portal/icons/icon-192.png'))
    icon_512_url = request.build_absolute_uri(static('portal/icons/icon-512.png'))
    manifest = {
        "name": "San Vincenzo S.R.L.",
        "short_name": "San Vincenzo",
        "start_url": "/",
        "scope": "/",
        "display": "standalone",
        "background_color": "#f5f8ff",
        "theme_color": "#0f172a",
        "description": "Servizi professionali e soluzioni operative per aziende.",
        "icons": [
            {
                "src": icon_192_url,
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any"
            },
            {
                "src": icon_512_url,
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any maskable"
            }
        ]
    }
    return HttpResponse(
        json.dumps(manifest),
        content_type='application/manifest+json',
    )


def employee_webmanifest(request):
        """Manifest PWA dedicato al portale dipendenti."""
        icon_192_url = request.build_absolute_uri(static('portal/icons/icon-192.png'))
        icon_512_url = request.build_absolute_uri(static('portal/icons/icon-512.png'))
        dashboard_url = reverse('dashboard')
        timekeeping_url = reverse('timekeeping')
        tutorial_url = reverse('portal_tutorial')
        manifest = {
                "id": dashboard_url,
                "name": "Portale Dipendenti San Vincenzo",
                "short_name": "Cedolini SV",
                "start_url": dashboard_url,
                "scope": "/",
                "display": "standalone",
                "orientation": "portrait",
                "background_color": "#f1f5f9",
                "theme_color": "#0f172a",
                "description": "Cedolini, CUD e marcature per i dipendenti San Vincenzo.",
                "icons": [
                        {
                                "src": icon_192_url,
                                "sizes": "192x192",
                                "type": "image/png",
                                "purpose": "any maskable"
                        },
                        {
                                "src": icon_512_url,
                                "sizes": "512x512",
                                "type": "image/png",
                                "purpose": "any maskable"
                        }
                ],
                "shortcuts": [
                        {
                                "name": "Dashboard",
                                "short_name": "Dashboard",
                                "url": dashboard_url,
                                "icons": [{
                                        "src": icon_192_url,
                                        "sizes": "192x192",
                                        "type": "image/png"
                                }]
                        },
                        {
                                "name": "Marcatura",
                                "short_name": "Marcatura",
                                "url": timekeeping_url,
                                "icons": [{
                                        "src": icon_192_url,
                                        "sizes": "192x192",
                                        "type": "image/png"
                                }]
                        },
                        {
                                "name": "Tutorial",
                                "short_name": "Tutorial",
                                "url": tutorial_url,
                                "icons": [{
                                        "src": icon_192_url,
                                        "sizes": "192x192",
                                        "type": "image/png"
                                }]
                        }
                ]
        }
        return HttpResponse(
                json.dumps(manifest),
                content_type='application/manifest+json',
        )


def service_worker(request):
        """Service worker root-level per installazione PWA del portale."""
        static_assets = [
                static('portal/logo.png'),
                static('portal/icons/icon-192.png'),
                static('portal/icons/icon-512.png'),
                static('portal/icons/apple-touch-icon.png'),
        ]
        script = f"""
const CACHE_NAME = 'cedolini-portal-v1';
const STATIC_ASSETS = {json.dumps(static_assets)};

self.addEventListener('install', (event) => {{
    event.waitUntil(
        caches.open(CACHE_NAME)
            .then((cache) => cache.addAll(STATIC_ASSETS))
            .catch(() => undefined)
            .then(() => self.skipWaiting())
    );
}});

self.addEventListener('activate', (event) => {{
    event.waitUntil(
        caches.keys().then((keys) => Promise.all(
            keys
                .filter((key) => key !== CACHE_NAME)
                .map((key) => caches.delete(key))
        )).then(() => self.clients.claim())
    );
}});

self.addEventListener('fetch', (event) => {{
    const request = event.request;
    if (request.method !== 'GET') {{
        return;
    }}

    const url = new URL(request.url);
    const isSameOrigin = url.origin === self.location.origin;
    const isStaticAsset = isSameOrigin && url.pathname.startsWith('/static/');

    if (request.mode === 'navigate') {{
        event.respondWith(
            fetch(request).catch(async () => {{
                const cachedDashboard = await caches.match('/portal/');
                if (cachedDashboard) {{
                    return cachedDashboard;
                }}
                return new Response(
                    '<!DOCTYPE html><html lang="it"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Offline</title><style>body{{font-family:Segoe UI,sans-serif;background:#f1f5f9;color:#0f172a;padding:24px}}.box{{max-width:420px;margin:10vh auto;background:#fff;border-radius:18px;padding:24px;box-shadow:0 12px 32px rgba(15,23,42,.12)}}h1{{font-size:1.2rem;margin:0 0 12px}}p{{margin:0;line-height:1.5;color:#475569}}</style></head><body><div class="box"><h1>Connessione assente</h1><p>Riapri l\'app quando la rete torna disponibile per consultare il portale dipendenti.</p></div></body></html>',
                    {{ headers: {{ 'Content-Type': 'text/html; charset=utf-8' }} }}
                );
            }})
        );
        return;
    }}

    if (!isStaticAsset) {{
        return;
    }}

    event.respondWith(
        caches.match(request).then((cachedResponse) => {{
            const networkFetch = fetch(request)
                .then((networkResponse) => {{
                    if (networkResponse && networkResponse.ok) {{
                        const responseClone = networkResponse.clone();
                        caches.open(CACHE_NAME).then((cache) => cache.put(request, responseClone));
                    }}
                    return networkResponse;
                }})
                .catch(() => cachedResponse);

            return cachedResponse || networkFetch;
        }})
    );
}});
""".strip()
        response = HttpResponse(script, content_type='application/javascript; charset=utf-8')
        response['Service-Worker-Allowed'] = '/'
        response['Cache-Control'] = 'no-cache'
        return response


def google_site_verification(request):
    """Serve il file di verifica Search Console al percorso richiesto da Google."""
    return HttpResponse(
        'google-site-verification: googlee8ce7f16b7b5fed5.html',
        content_type='text/plain',
    )


def favicon_ico(request):
    """Compatibilita browser: favicon richiesta su /favicon.ico."""
    return HttpResponsePermanentRedirect(static('portal/icons/icon-192.png'))


def favicon_32_png(request):
    """Compatibilita launcher: icona 32x32 su path standard."""
    return HttpResponsePermanentRedirect(static('portal/icons/icon-192.png'))


def favicon_16_png(request):
    """Compatibilita launcher: icona 16x16 su path standard."""
    return HttpResponsePermanentRedirect(static('portal/icons/icon-192.png'))


def apple_touch_icon(request):
    """Compatibilita iOS: apple-touch-icon su root path."""
    return HttpResponsePermanentRedirect(static('portal/icons/apple-touch-icon.png'))


def apple_touch_icon_precomposed(request):
    """Compatibilita iOS legacy: apple-touch-icon-precomposed su root path."""
    return HttpResponsePermanentRedirect(static('portal/icons/apple-touch-icon.png'))


# =========================================================
# REGISTER VIA TOKEN
# =========================================================

def register_with_token(request, token):
    invite = get_object_or_404(InviteToken, token=token)

    if not invite.is_valid():
        return HttpResponse("Token non valido o scaduto.", status=400)

    employee = invite.employee
    user = employee.user

    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        confirm = request.POST.get("confirm_password")
        privacy_accepted = request.POST.get("privacy_accepted") == "on"

        if not first_name or not last_name or not email or not password:
            return render(request, "portal/register.html", {
                "employee": employee,
                "error": "Compila tutti i campi"
            })

        if len(password) < 8:
            return render(request, "portal/register.html", {
                "employee": employee,
                "error": "Password troppo corta (min 8 caratteri)"
            })

        if password != confirm:
            return render(request, "portal/register.html", {
                "employee": employee,
                "error": "Le password non coincidono"
            })

        if not privacy_accepted:
            return render(request, "portal/register.html", {
                "employee": employee,
                "error": "Devi accettare l'informativa privacy per completare la registrazione."
            })

        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.set_password(password)
        user.is_active = True
        user.save()

        employee.first_name = first_name
        employee.last_name = last_name
        employee.email_invio = email
        employee.must_change_password = False
        employee.privacy_accepted = True
        employee.privacy_accepted_at = timezone.now()
        employee.save()

        invite.mark_used()

        return redirect("/login/")

    return render(request, "portal/register.html", {
        "employee": employee
    })


# =========================================================
# DASHBOARD DIPENDENTE
# =========================================================

@login_required
def dashboard(request):
    logger.info("dashboard START user=%s", getattr(request.user, "username", None))

    # Se uno staff arriva direttamente su /portal/, portalo alla dashboard admin.
    if request.user.is_staff:
        return redirect('admin_dashboard')

    employee = Employee.objects.filter(user=request.user).first()
    if not employee:
        logger.warning("dashboard: employee profile missing for user=%s", request.user.id)
        return HttpResponse(
            "Profilo dipendente non trovato. Contatta l'amministratore.",
            status=403,
        )

    logger.info("dashboard employee_id=%s", getattr(employee, "id", None))

    logger.info("constructing payslips queryset")
    payslips = (
        Payslip.objects
        .filter(employee=employee)
        .prefetch_related("payslipview_set")
        .order_by('-year', '-month')
    )

    logger.info("queryset constructed — forcing evaluation with .count()")
    try:
        count = payslips.count()
        logger.info("payslips.count=%d", count)
    except Exception:
        logger.exception("Error while evaluating payslips queryset")

    grouped = {}
    logger.info("entering loop over payslips")
    for idx, p in enumerate(payslips):
        if idx < 5:
            logger.info("processing payslip id=%s year=%s month=%s", getattr(p, "id", None), getattr(p, "year", None), getattr(p, "month", None))

        first_view = None
        if p.payslipview_set.all():
            first_view = p.payslipview_set.order_by('viewed_at').first()

        p.is_viewed = bool(first_view)
        p.viewed_at = first_view.viewed_at if first_view else None

        if p.year not in grouped:
            grouped[p.year] = []

        grouped[p.year].append(p)

    # CUD annuali del dipendente
    cuds = (
        Cud.objects
        .filter(employee=employee)
        .prefetch_related('cudview_set')
        .order_by('-year')
    )

    # Aggiunge info di visualizzazione CUD
    for c in cuds:
        first_view = c.cudview_set.order_by('viewed_at').first() if c.cudview_set.all() else None
        c.is_viewed = bool(first_view)
        c.viewed_at = first_view.viewed_at if first_view else None

        logger.info("about to render template")
    # --- Marcatura ---
    today = timezone.localdate()
    session, _ = WorkSession.objects.get_or_create(employee=employee, work_date=today)
    session.worked_display = session.worked_hours_display()
    active_assignments = _active_assignments_for_employee(employee, today)
    has_active_zone = bool(active_assignments)
    today_requests_qs = WorkMarkRequest.objects.filter(employee=employee, work_date=today).order_by('-created_at')
    today_mark_request_start = (
        today_requests_qs
        .filter(mark_type__in=[WorkMarkRequest.MARK_TYPE_START, WorkMarkRequest.MARK_TYPE_BOTH])
        .first()
    )
    today_mark_request_end = (
        today_requests_qs
        .filter(mark_type__in=[WorkMarkRequest.MARK_TYPE_END, WorkMarkRequest.MARK_TYPE_BOTH])
        .first()
    )
    # --- Fine Marcatura ---
    return render(request, 'portal/dashboard.html', {
        'employee': employee,
        'grouped_payslips': grouped,
        'cuds': cuds,
        'today_session': session,
        'has_active_zone': has_active_zone,
        'today_mark_request_start': today_mark_request_start,
        'today_mark_request_end': today_mark_request_end,
    })


@login_required
def dashboard(request):
    if user_has_full_admin_access(request.user):
        return redirect('admin_dashboard')
    if user_has_today_markings_only_access(request.user):
        return redirect('today_markings_dashboard')

    employee = Employee.objects.filter(user=request.user).first()
    if not employee:
        return HttpResponse("Profilo dipendente non trovato. Contatta l'amministratore.", status=403)

    today = timezone.localdate()
    month_start = today.replace(day=1)
    _sync_approved_requests_for_range(month_start, today, employee=employee)
    _reconcile_overnight_sessions(employee=employee, start_date=month_start, end_date=today)
    _sync_approved_vacations_for_range(month_start, today, employee=employee)

    session, _ = _get_timekeeping_session(employee)
    session.worked_display = session.worked_hours_display()
    active_assignments = _active_assignments_for_employee(employee, today)
    has_active_zone = bool(active_assignments)
    active_zones = _active_zones_for_employee(employee, today)
    is_vacation_today = session.day_type == WorkSession.DAY_TYPE_VACATION

    end_request_work_date = session.work_date if session.effective_started_at() and not session.effective_ended_at() else today
    start_requests_qs = WorkMarkRequest.objects.filter(employee=employee, work_date=today).order_by('-created_at')
    end_requests_qs = WorkMarkRequest.objects.filter(employee=employee, work_date=end_request_work_date).order_by('-created_at')
    today_mark_request_start = (
        start_requests_qs
        .filter(mark_type__in=[WorkMarkRequest.MARK_TYPE_START, WorkMarkRequest.MARK_TYPE_BOTH])
        .first()
    )
    today_mark_request_end = (
        end_requests_qs
        .filter(mark_type__in=[WorkMarkRequest.MARK_TYPE_END, WorkMarkRequest.MARK_TYPE_BOTH])
        .first()
    )
    request_status = request.GET.get('request_status', '')
    vacation_status = request.GET.get('vacation_status', '')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'request_out_of_zone':
            reason = (request.POST.get('reason') or '').strip()
            mark_type = (request.POST.get('mark_type') or '').strip()

            if mark_type not in {WorkMarkRequest.MARK_TYPE_START, WorkMarkRequest.MARK_TYPE_END}:
                return redirect(f"{request.path}?request_status=invalid_type")
            if len(reason) < 8:
                return redirect(f"{request.path}?request_status=invalid")

            target_work_date = end_request_work_date if mark_type == WorkMarkRequest.MARK_TYPE_END else today
            latest_request = (
                WorkMarkRequest.objects
                .filter(employee=employee, work_date=target_work_date, mark_type=mark_type)
                .order_by('-created_at')
                .first()
            )
            if latest_request and latest_request.status == WorkMarkRequest.STATUS_PENDING:
                return redirect(f"{request.path}?request_status=already_pending")

            request_obj = WorkMarkRequest.objects.create(
                employee=employee,
                work_date=target_work_date,
                mark_type=mark_type,
                reason=reason,
            )
            _send_out_of_zone_request_admin_notification(request_obj)

            _create_audit_event(
                request,
                'timekeeping_out_of_zone_requested',
                employee=employee,
                metadata={
                    'work_date': str(target_work_date),
                    'mark_type': mark_type,
                    'reason': reason,
                },
            )
            return redirect(f"{request.path}?request_status=sent_{mark_type}")

        if action == 'request_vacation':
            start_date = _parse_date_or_none(request.POST.get('start_date'))
            end_date = _parse_date_or_none(request.POST.get('end_date'))
            reason = (request.POST.get('vacation_reason') or '').strip()

            if not start_date or not end_date:
                return redirect(f"{request.path}?vacation_status=invalid_dates")
            if end_date < start_date:
                return redirect(f"{request.path}?vacation_status=invalid_range")
            if len(reason) < 8:
                return redirect(f"{request.path}?vacation_status=invalid_reason")

            overlapping_pending = VacationRequest.objects.filter(
                employee=employee,
                status=VacationRequest.STATUS_PENDING,
                start_date__lte=end_date,
                end_date__gte=start_date,
            ).exists()
            if overlapping_pending:
                return redirect(f"{request.path}?vacation_status=already_pending")

            overlapping_approved = VacationRequest.objects.filter(
                employee=employee,
                status=VacationRequest.STATUS_APPROVED,
                start_date__lte=end_date,
                end_date__gte=start_date,
            ).exists()
            if overlapping_approved:
                return redirect(f"{request.path}?vacation_status=already_approved")

            request_obj = VacationRequest.objects.create(
                employee=employee,
                start_date=start_date,
                end_date=end_date,
                reason=reason,
            )
            _send_vacation_request_admin_notification(request_obj)
            _create_audit_event(
                request,
                'vacation_requested',
                employee=employee,
                metadata={
                    'request_id': request_obj.id,
                    'start_date': str(start_date),
                    'end_date': str(end_date),
                    'reason': reason,
                },
            )
            return redirect(f"{request.path}?vacation_status=sent")

    payslips = (
        Payslip.objects
        .filter(employee=employee)
        .prefetch_related('payslipview_set')
        .order_by('-year', '-month')
    )

    payslip_items = []
    for p in payslips:
        first_view = p.payslipview_set.order_by('viewed_at').first() if p.payslipview_set.all() else None
        p.is_viewed = bool(first_view)
        p.viewed_at = first_view.viewed_at if first_view else None
        payslip_items.append(_attach_payslip_display_period(p))

    payslip_items.sort(key=lambda item: (item.display_year, item.display_month, item.id), reverse=True)

    latest_payslip = payslip_items[0] if payslip_items else None
    grouped = {}
    for p in payslip_items:
        grouped.setdefault(p.display_year, []).append(p)

    cuds = (
        Cud.objects
        .filter(employee=employee)
        .prefetch_related('cudview_set')
        .order_by('-year')
    )
    for c in cuds:
        first_view = c.cudview_set.order_by('viewed_at').first() if c.cudview_set.all() else None
        c.is_viewed = bool(first_view)
        c.viewed_at = first_view.viewed_at if first_view else None

    recent_vacation_requests = list(
        VacationRequest.objects
        .filter(employee=employee)
        .order_by('-created_at')[:6]
    )

    today_marked_sessions = []
    if user_has_today_markings_access(request.user):
        today_marked_sessions = _prepare_marked_sessions_for_date(list(_today_marked_sessions_queryset(today)), today)

    published_turni_state = _turni_planner_published_state()
    published_turni_sections = []
    if _user_can_view_published_turni(request.user, employee=employee):
        published_turni_sections = _turni_planner_employee_sections(published_turni_state, include_portineria=True)

    response = render(request, 'portal/dashboard.html', {
        'employee': employee,
        'grouped_payslips': grouped,
        'latest_payslip': latest_payslip,
        'cuds': cuds,
        'today_session': session,
        'has_active_zone': has_active_zone,
        'active_zones': active_zones,
        'today_mark_request_start': today_mark_request_start,
        'today_mark_request_end': today_mark_request_end,
        'request_status': request_status,
        'vacation_status': vacation_status,
        'is_vacation_today': is_vacation_today,
        'published_turni_state': published_turni_state,
        'published_turni_sections': published_turni_sections,
        'recent_vacation_requests': recent_vacation_requests,
        'today': today,
        'today_marked_sessions': today_marked_sessions,
    })
    return _disable_response_cache(response)


@login_required
def personal_asset_dashboard(request):
    denied_response = _patrimonio_allowed_or_403(request)
    if denied_response is not None:
        return denied_response

    if request.method == 'GET' and request.GET.get('report') == 'reimbursement_jpg':
        response, filename = _build_personal_asset_reimbursement_report_response(request.user)
        _create_audit_event(
            request,
            'personal_asset_reimbursement_report_downloaded',
            employee=getattr(request.user, 'employee', None),
            metadata={'filename': filename},
        )
        return response

    status = (request.GET.get('status') or '').strip()
    feedback = None
    feedback_level = 'success'
    if status == 'created':
        feedback = 'Operazione patrimoniale registrata correttamente.'
    elif status == 'deleted':
        feedback = 'Voce eliminata correttamente.'

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        if action == 'delete_entry':
            entry = PersonalAssetEntry.objects.filter(id=request.POST.get('entry_id'), user=request.user).first()
            if entry is not None:
                _create_audit_event(
                    request,
                    'personal_asset_entry_deleted',
                    employee=getattr(request.user, 'employee', None),
                    metadata={
                        'entry_id': entry.id,
                        'operation_type': entry.operation_type,
                        'occurred_on': str(entry.occurred_on),
                        'amount': str(entry.amount),
                    },
                )
                entry.delete()
            return redirect(f'{request.path}?status=deleted')

        form = PersonalAssetEntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.user = request.user
            entry.save()

            _create_audit_event(
                request,
                'personal_asset_entry_created',
                employee=getattr(request.user, 'employee', None),
                metadata={
                    'operation_type': entry.operation_type,
                    'occurred_on': str(entry.occurred_on),
                    'category': entry.category,
                    'amount': str(entry.amount),
                    'reimbursement_amount': str(entry.reimbursement_amount or ''),
                },
            )
            return redirect(f'{request.path}?status=created')

        feedback = 'Correggi i campi evidenziati e riprova.'
        feedback_level = 'danger'
    else:
        form = PersonalAssetEntryForm(initial={'occurred_on': timezone.localdate()})

    entries = list(_personal_asset_history_queryset(request.user)[:100])
    summary = _personal_asset_summary(request.user)
    reimbursement_entries = list(_personal_asset_reimbursement_entries_queryset(request.user)[:200])
    reimbursement_total = sum((entry.reimbursement_amount or entry.amount or Decimal('0.00')) for entry in reimbursement_entries)

    return render(request, 'portal/personal_asset_dashboard.html', {
        'finance_form': form,
        'finance_entries': entries,
        'finance_summary': summary,
        'feedback': feedback,
        'feedback_level': feedback_level,
        'reimbursement_report_entries_count': len(reimbursement_entries),
        'reimbursement_report_total': reimbursement_total,
        'reimbursement_report_recipients': _personal_asset_reimbursement_report_email_recipients(),
    })


def _haversine_meters(lat1, lon1, lat2, lon2):
    """Distanza geodetica approssimata in metri tra due coordinate."""
    earth_radius = 6371000
    d_lat = math.radians(lat2 - lat1)
    d_lon = math.radians(lon2 - lon1)
    a = (
        math.sin(d_lat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(d_lon / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return earth_radius * c


def _parse_coordinate(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _active_assignments_for_employee(employee, on_date):
    assignments = (
        EmployeeWorkZone.objects
        .select_related("zone")
        .filter(
            employee=employee,
            is_active=True,
            zone__is_active=True,
        )
        .filter(Q(valid_to__isnull=True) | Q(valid_to__gte=on_date))
    )
    return list(assignments)


def _active_zones_for_employee(employee, on_date):
    assignments = _active_assignments_for_employee(employee, on_date)
    return [assignment.zone for assignment in assignments if assignment.zone_id]


def _evaluate_location_for_employee_zone(employee, lat, lon, on_date):
    assignments = _active_assignments_for_employee(employee, on_date)
    if lat is None or lon is None:
        return {
            'assignment': assignments[0] if assignments else None,
            'zone': assignments[0].zone if assignments and assignments[0].zone_id else None,
            'distance_meters': None,
            'within': False,
        }
    plat = float(lat)
    plon = float(lon)

    def _distance_to_zone(zone):
        if zone.latitude is None or zone.longitude is None:
            return False, None

        center_dist = _haversine_meters(plat, plon, float(zone.latitude), float(zone.longitude))
        if getattr(zone, 'shape', 'circle') == getattr(WorkZone, 'SHAPE_RECT', 'rect') and \
           zone.rect_north is not None and zone.rect_south is not None and zone.rect_east is not None and zone.rect_west is not None:
            north = float(zone.rect_north)
            south = float(zone.rect_south)
            east = float(zone.rect_east)
            west = float(zone.rect_west)
            within = (south <= plat <= north) and (west <= plon <= east)
            clamped_lat = min(max(plat, south), north)
            clamped_lon = min(max(plon, west), east)
            distance = 0.0 if within else _haversine_meters(plat, plon, clamped_lat, clamped_lon)
            return within, distance

        radius = float(zone.radius_meters or 0)
        within = center_dist <= radius
        distance = 0.0 if within else max(center_dist - radius, 0.0)
        return within, distance

    best_assignment = None
    best_zone = None
    best_distance = None
    best_within = False

    for assignment in assignments:
        zone = assignment.zone
        if zone is None:
            continue
        within, distance = _distance_to_zone(zone)
        if distance is None:
            continue
        if best_distance is None or distance < best_distance:
            best_assignment = assignment
            best_zone = zone
            best_distance = distance
            best_within = within

    return {
        'assignment': best_assignment,
        'zone': best_zone,
        'distance_meters': round(best_distance, 1) if best_distance is not None else None,
        'within': best_within,
    }


def _parse_time_or_none(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        return None


def _session_has_markings(session):
    return bool(
        session.started_at
        or session.ended_at
        or session.corrected_started_at
        or session.corrected_ended_at
    )


def _clear_session_marking(session, delete_target):
    if delete_target == 'start':
        session.started_at = None
        session.corrected_started_at = None
        session.start_latitude = None
        session.start_longitude = None
        session.start_zone = None
        session.start_within_zone = False
        session.save(update_fields=[
            'started_at',
            'corrected_started_at',
            'start_latitude',
            'start_longitude',
            'start_zone',
            'start_within_zone',
            'updated_at',
        ])
        return

    if delete_target == 'end':
        session.ended_at = None
        session.corrected_ended_at = None
        session.end_latitude = None
        session.end_longitude = None
        session.end_zone = None
        session.end_within_zone = False
        session.save(update_fields=[
            'ended_at',
            'corrected_ended_at',
            'end_latitude',
            'end_longitude',
            'end_zone',
            'end_within_zone',
            'updated_at',
        ])
        return

    session.delete()


def _set_session_as_vacation(session):
    session.day_type = WorkSession.DAY_TYPE_VACATION
    session.started_at = None
    session.ended_at = None
    session.start_latitude = None
    session.start_longitude = None
    session.end_latitude = None
    session.end_longitude = None
    session.start_zone = None
    session.end_zone = None
    session.start_within_zone = False
    session.end_within_zone = False
    session.corrected_started_at = None
    session.corrected_ended_at = None
    session.correction_note = None
    session.corrected_by = None
    session.corrected_at = None
    session.save()


def _iter_dates(start_date, end_date):
    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)


def _session_cell_text(session):
    if not session:
        return ''

    if session.day_type == WorkSession.DAY_TYPE_VACATION:
        return 'FERIE'

    def fmt(dt):
        try:
            from django.utils import timezone as djtz
            if djtz.is_aware(dt):
                dt = djtz.localtime(dt)
        except Exception:
            pass
        return dt.strftime('%H:%M')

    start = session.effective_started_at()
    end = session.effective_ended_at()
    if start and end:
        return f"{fmt(start)}-{fmt(end)} ({session.worked_hours_display()})"
    if start:
        return f"IN {fmt(start)}"
    if end:
        return f"OUT {fmt(end)}"
    return ''


def _local_date_from_datetime(value):
    if not value:
        return None
    try:
        if timezone.is_aware(value):
            value = timezone.localtime(value)
    except Exception:
        pass
    return value.date()


def _mark_happened_on_date(value, target_date):
    return _local_date_from_datetime(value) == target_date


def _prepare_marked_sessions_for_date(sessions, target_date):
    prepared = []
    for session in sessions:
        start_dt = session.effective_started_at()
        end_dt = session.effective_ended_at()
        session.display_started_at = start_dt if _mark_happened_on_date(start_dt, target_date) else None
        session.display_ended_at = end_dt if _mark_happened_on_date(end_dt, target_date) else None
        session.display_worked_hours = session.worked_hours_display() if (session.display_started_at or session.display_ended_at) else '00:00'
        prepared.append(session)
    return prepared


def _get_open_shift_session(employee, reference_ts=None):
    reference_ts = timezone.localtime(reference_ts or timezone.now())
    candidate_start_date = reference_ts.date() - timedelta(days=1)
    sessions = (
        WorkSession.objects
        .filter(employee=employee, work_date__gte=candidate_start_date)
        .select_related('start_zone', 'end_zone')
        .order_by('-work_date', '-created_at')
    )

    for session in sessions:
        start_dt = session.effective_started_at()
        end_dt = session.effective_ended_at()
        if not start_dt or end_dt:
            continue

        elapsed_hours = (reference_ts - start_dt).total_seconds() / 3600
        if 0 <= elapsed_hours <= MAX_SHIFT_DURATION_HOURS:
            return session

    return None


def _get_timekeeping_session(employee, reference_ts=None):
    reference_ts = timezone.localtime(reference_ts or timezone.now())
    open_session = _get_open_shift_session(employee, reference_ts)
    if open_session:
        return open_session, False
    return WorkSession.objects.get_or_create(employee=employee, work_date=reference_ts.date())


def _reconcile_overnight_sessions(employee=None, start_date=None, end_date=None):
    sessions_qs = WorkSession.objects.filter(
        Q(started_at__isnull=False)
        | Q(ended_at__isnull=False)
        | Q(corrected_started_at__isnull=False)
        | Q(corrected_ended_at__isnull=False)
    ).select_related('employee', 'end_zone')

    if employee is not None:
        sessions_qs = sessions_qs.filter(employee=employee)
    if start_date and end_date:
        sessions_qs = sessions_qs.filter(work_date__range=(start_date - timedelta(days=1), end_date))

    sessions = list(sessions_qs.order_by('employee_id', 'work_date', 'created_at'))
    previous_by_employee = {}

    for session in sessions:
        previous = previous_by_employee.get(session.employee_id)
        if not previous:
            previous_by_employee[session.employee_id] = session
            continue

        previous_start = previous.effective_started_at()
        previous_end = previous.effective_ended_at()
        current_start = session.effective_started_at()
        current_end = session.effective_ended_at()

        can_merge = (
            previous.work_date + timedelta(days=1) == session.work_date
            and previous_start is not None
            and previous_end is None
            and current_start is None
            and current_end is not None
        )

        if can_merge:
            elapsed_hours = (current_end - previous_start).total_seconds() / 3600
            if 0 <= elapsed_hours <= MAX_SHIFT_DURATION_HOURS:
                previous.ended_at = session.ended_at or previous.ended_at
                previous.corrected_ended_at = session.corrected_ended_at or previous.corrected_ended_at
                previous.end_latitude = session.end_latitude
                previous.end_longitude = session.end_longitude
                previous.end_zone = session.end_zone
                previous.end_within_zone = session.end_within_zone
                previous.correction_note = previous.correction_note or session.correction_note
                previous.corrected_by = previous.corrected_by or session.corrected_by
                previous.corrected_at = previous.corrected_at or session.corrected_at
                previous.save()

                session.delete()
                previous_by_employee[session.employee_id] = previous
                continue

        previous_by_employee[session.employee_id] = session


def _apply_approved_mark_request_to_session(request_obj):
    """Trasforma una richiesta fuori zona approvata in marcatura effettiva.

    Usa il timestamp della richiesta (created_at), non il momento dell'approvazione,
    per mantenere l'orario reale comunicato dal dipendente.
    """
    mark_ts = request_obj.created_at
    if request_obj.mark_type == WorkMarkRequest.MARK_TYPE_END:
        session = _get_open_shift_session(request_obj.employee, mark_ts)
        if session is None:
            session, _ = WorkSession.objects.get_or_create(
                employee=request_obj.employee,
                work_date=request_obj.work_date,
            )
    else:
        session, _ = WorkSession.objects.get_or_create(
            employee=request_obj.employee,
            work_date=request_obj.work_date,
        )

    changed_fields = []

    if request_obj.mark_type in {WorkMarkRequest.MARK_TYPE_START, WorkMarkRequest.MARK_TYPE_BOTH}:
        if not session.started_at:
            session.day_type = WorkSession.DAY_TYPE_WORK
            session.started_at = mark_ts
            session.start_within_zone = False
            changed_fields.extend(['day_type', 'started_at', 'start_within_zone'])

    if request_obj.mark_type in {WorkMarkRequest.MARK_TYPE_END, WorkMarkRequest.MARK_TYPE_BOTH}:
        if not session.ended_at:
            session.day_type = WorkSession.DAY_TYPE_WORK
            session.ended_at = mark_ts
            session.end_within_zone = False
            changed_fields.extend(['day_type', 'ended_at', 'end_within_zone'])

    if changed_fields:
        session.save(update_fields=changed_fields + ['updated_at'])

    return session


def _sync_approved_requests_for_range(start_date, end_date, employee=None):
    """Allinea WorkSession con richieste approvate gia esistenti nel range date."""
    qs = (
        WorkMarkRequest.objects
        .select_related('employee')
        .filter(
            status=WorkMarkRequest.STATUS_APPROVED,
            work_date__range=(start_date, end_date),
        )
        .order_by('work_date', 'created_at')
    )
    if employee is not None:
        qs = qs.filter(employee=employee)

    for req in qs:
        _apply_approved_mark_request_to_session(req)


def _apply_approved_vacation_request_to_sessions(request_obj):
    for current_date in _iter_dates(request_obj.start_date, request_obj.end_date):
        session, _ = WorkSession.objects.get_or_create(
            employee=request_obj.employee,
            work_date=current_date,
        )
        _set_session_as_vacation(session)


def _sync_approved_vacations_for_range(start_date, end_date, employee=None):
    qs = (
        VacationRequest.objects
        .select_related('employee')
        .filter(
            status=VacationRequest.STATUS_APPROVED,
            start_date__lte=end_date,
            end_date__gte=start_date,
        )
        .order_by('start_date', 'created_at')
    )
    if employee is not None:
        qs = qs.filter(employee=employee)

    for req in qs:
        effective_start = max(req.start_date, start_date)
        effective_end = min(req.end_date, end_date)
        for current_date in _iter_dates(effective_start, effective_end):
            session, _ = WorkSession.objects.get_or_create(
                employee=req.employee,
                work_date=current_date,
            )
            _set_session_as_vacation(session)


@login_required
def timekeeping(request):
    """Marcatura dipendente: avvio/fine giornata con supporto geolocalizzazione."""
    if user_has_full_admin_access(request.user):
        return redirect('admin_timekeeping')
    if user_has_today_markings_only_access(request.user):
        return redirect('today_markings_dashboard')

    employee = get_object_or_404(Employee, user=request.user)
    today = timezone.localdate()
    month_start = today.replace(day=1)

    # Rende visibili immediatamente nel riepilogo le richieste gia approvate.
    _sync_approved_requests_for_range(month_start, today, employee=employee)
    _reconcile_overnight_sessions(employee=employee, start_date=month_start, end_date=today)
    _sync_approved_vacations_for_range(month_start, today, employee=employee)

    session, _ = _get_timekeeping_session(employee)
    session.worked_display = session.worked_hours_display()
    active_assignments = _active_assignments_for_employee(employee, today)
    has_active_zone = bool(active_assignments)
    is_vacation_today = session.day_type == WorkSession.DAY_TYPE_VACATION
    end_request_work_date = session.work_date if session.effective_started_at() and not session.effective_ended_at() else today
    start_requests_qs = WorkMarkRequest.objects.filter(employee=employee, work_date=today).order_by('-created_at')
    end_requests_qs = WorkMarkRequest.objects.filter(employee=employee, work_date=end_request_work_date).order_by('-created_at')

    today_mark_request_start = (
        start_requests_qs
        .filter(mark_type__in=[WorkMarkRequest.MARK_TYPE_START, WorkMarkRequest.MARK_TYPE_BOTH])
        .first()
    )
    today_mark_request_end = (
        end_requests_qs
        .filter(mark_type__in=[WorkMarkRequest.MARK_TYPE_END, WorkMarkRequest.MARK_TYPE_BOTH])
        .first()
    )
    request_status = request.GET.get('request_status', '')

    def has_approved_request_for_action(action_name):
        mark_types = [WorkMarkRequest.MARK_TYPE_BOTH]
        target_work_date = today
        if action_name == 'start':
            mark_types.append(WorkMarkRequest.MARK_TYPE_START)
        if action_name == 'end':
            mark_types.append(WorkMarkRequest.MARK_TYPE_END)
            target_work_date = end_request_work_date
        return WorkMarkRequest.objects.filter(
            employee=employee,
            work_date=target_work_date,
            status=WorkMarkRequest.STATUS_APPROVED,
            mark_type__in=mark_types,
        ).exists()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'request_out_of_zone':
            reason = (request.POST.get('reason') or '').strip()
            mark_type = (request.POST.get('mark_type') or '').strip()
            if mark_type not in {WorkMarkRequest.MARK_TYPE_START, WorkMarkRequest.MARK_TYPE_END}:
                return redirect(f"{request.path}?request_status=invalid_type")
            if len(reason) < 8:
                return redirect(f"{request.path}?request_status=invalid")

            target_work_date = end_request_work_date if mark_type == WorkMarkRequest.MARK_TYPE_END else today

            latest_request = (
                WorkMarkRequest.objects
                .filter(employee=employee, work_date=target_work_date, mark_type=mark_type)
                .order_by('-created_at')
                .first()
            )

            if latest_request and latest_request.status == WorkMarkRequest.STATUS_PENDING:
                return redirect(f"{request.path}?request_status=already_pending")

            request_obj = WorkMarkRequest.objects.create(
                employee=employee,
                work_date=target_work_date,
                mark_type=mark_type,
                reason=reason,
            )
            _send_out_of_zone_request_admin_notification(request_obj)

            _create_audit_event(
                request,
                'timekeeping_out_of_zone_requested',
                employee=employee,
                metadata={
                    'work_date': str(target_work_date),
                    'mark_type': mark_type,
                    'reason': reason,
                },
            )
            return redirect(f"{request.path}?request_status=sent_{mark_type}")

        if is_vacation_today:
            return JsonResponse(
                {'ok': False, 'error': 'Marcatura non disponibile: la giornata e segnata come ferie approvate.'},
                status=400,
            )

        latitude = _parse_coordinate(request.POST.get('latitude'))
        longitude = _parse_coordinate(request.POST.get('longitude'))

        if not has_active_zone:
            return JsonResponse(
                {'ok': False, 'error': 'Marcatura non disponibile: nessuna zona attiva assegnata.'},
                status=400,
            )

        if action not in {'start', 'end'}:
            return JsonResponse({'ok': False, 'error': 'Azione non valida.'}, status=400)

        approved_out_of_zone = has_approved_request_for_action(action)

        if action == 'start' and session.effective_started_at() and not session.effective_ended_at():
            return JsonResponse({'ok': False, 'error': 'Hai gia un turno aperto. Completa prima l\'uscita.'}, status=400)

        if action == 'end' and not session.effective_started_at():
            return JsonResponse({'ok': False, 'error': 'Devi marcare prima l\'ingresso.'}, status=400)

        if action == 'end' and session.effective_ended_at():
            return JsonResponse({'ok': False, 'error': 'Uscita gia marcata oggi.'}, status=400)

        now_ts = timezone.now()
        zone_check = _evaluate_location_for_employee_zone(employee, latitude, longitude, today)
        strict_mode = any(a.strict_geofence for a in active_assignments)

        if strict_mode and (latitude is None or longitude is None) and not approved_out_of_zone:
            return JsonResponse(
                {'ok': False, 'error': 'Geolocalizzazione obbligatoria: attiva il GPS per marcare.'},
                status=400,
            )

        if strict_mode and active_assignments and not zone_check['within'] and not approved_out_of_zone:
            return JsonResponse(
                {'ok': False, 'error': 'Marcatura bloccata: sei fuori dalla zona assegnata.'},
                status=400,
            )

        if action == 'start':
            session.started_at = now_ts
            session.start_latitude = latitude
            session.start_longitude = longitude
            session.start_zone = zone_check['zone']
            session.start_within_zone = zone_check['within']
            session.save()

            _create_audit_event(
                request,
                'timekeeping_start',
                employee=employee,
                metadata={
                    'work_date': str(today),
                    'zone': zone_check['zone'].name if zone_check['zone'] else None,
                    'within_zone': zone_check['within'],
                    'distance_meters': zone_check['distance_meters'],
                },
            )

        if action == 'end':
            session.ended_at = now_ts
            session.end_latitude = latitude
            session.end_longitude = longitude
            session.end_zone = zone_check['zone']
            session.end_within_zone = zone_check['within']
            session.save()

            _create_audit_event(
                request,
                'timekeeping_end',
                employee=employee,
                metadata={
                    'work_date': str(today),
                    'zone': zone_check['zone'].name if zone_check['zone'] else None,
                    'within_zone': zone_check['within'],
                    'distance_meters': zone_check['distance_meters'],
                },
            )

        return JsonResponse({
            'ok': True,
            'action': action,
            'started_at': session.started_at.strftime('%H:%M') if session.started_at else None,
            'ended_at': session.ended_at.strftime('%H:%M') if session.ended_at else None,
            'zone': zone_check['zone'].name if zone_check['zone'] else None,
            'within_zone': zone_check['within'],
            'distance_meters': zone_check['distance_meters'],
        })

    response = render(request, 'portal/timekeeping.html', {
        'employee': employee,
        'today_session': session,
        'active_zones': _active_zones_for_employee(employee, today),
        'has_active_zone': has_active_zone,
        'today_mark_request_start': today_mark_request_start,
        'today_mark_request_end': today_mark_request_end,
        'request_status': request_status,
        'is_vacation_today': is_vacation_today,
    })
    return _disable_response_cache(response)


@login_required
def admin_timekeeping(request):
    """Report mensile marcature con dettaglio giornaliero per dipendente."""
    if not request.user.is_staff:
        return redirect('dashboard')

    today = timezone.localdate()
    feedback = None
    feedback_level = 'info'

    if request.method == 'POST':
        action = request.POST.get('action')

        if action in {'approve_mark_request', 'reject_mark_request'}:
            req_id = request.POST.get('request_id')
            request_obj = WorkMarkRequest.objects.filter(id=req_id).select_related('employee').first()
            if request_obj:
                request_obj.status = (
                    WorkMarkRequest.STATUS_APPROVED
                    if action == 'approve_mark_request'
                    else WorkMarkRequest.STATUS_REJECTED
                )
                request_obj.review_note = (request.POST.get('review_note') or '').strip() or None
                request_obj.reviewed_by = request.user
                request_obj.reviewed_at = timezone.now()
                request_obj.save(update_fields=['status', 'review_note', 'reviewed_by', 'reviewed_at', 'updated_at'])

                if request_obj.status == WorkMarkRequest.STATUS_APPROVED:
                    _apply_approved_mark_request_to_session(request_obj)

                _create_audit_event(
                    request,
                    'timekeeping_out_of_zone_reviewed',
                    employee=request_obj.employee,
                    metadata={
                        'request_id': request_obj.id,
                        'work_date': str(request_obj.work_date),
                        'status': request_obj.status,
                        'mark_type': request_obj.mark_type,
                        'marked_at': request_obj.created_at.isoformat() if request_obj.status == WorkMarkRequest.STATUS_APPROVED else None,
                    },
                )

            return redirect(request.get_full_path() or request.path)

        if action in {'approve_vacation_request', 'reject_vacation_request'}:
            req_id = request.POST.get('request_id')
            request_obj = VacationRequest.objects.filter(id=req_id).select_related('employee').first()
            if request_obj:
                request_obj.status = (
                    VacationRequest.STATUS_APPROVED
                    if action == 'approve_vacation_request'
                    else VacationRequest.STATUS_REJECTED
                )
                request_obj.review_note = (request.POST.get('review_note') or '').strip() or None
                request_obj.reviewed_by = request.user
                request_obj.reviewed_at = timezone.now()
                request_obj.save(update_fields=['status', 'review_note', 'reviewed_by', 'reviewed_at', 'updated_at'])

                if request_obj.status == VacationRequest.STATUS_APPROVED:
                    _apply_approved_vacation_request_to_sessions(request_obj)

                _create_audit_event(
                    request,
                    'vacation_reviewed',
                    employee=request_obj.employee,
                    metadata={
                        'request_id': request_obj.id,
                        'start_date': str(request_obj.start_date),
                        'end_date': str(request_obj.end_date),
                        'status': request_obj.status,
                    },
                )

            return redirect(request.get_full_path() or request.path)

        if action == 'correct_day':
            employee_id = request.POST.get('employee_id')
            target_date_raw = request.POST.get('target_date')
            start_time = _parse_time_or_none(request.POST.get('start_time'))
            end_time = _parse_time_or_none(request.POST.get('end_time'))
            note = (request.POST.get('note') or '').strip()

            employee = Employee.objects.filter(id=employee_id).first()
            try:
                target_date = datetime.strptime(target_date_raw or '', '%Y-%m-%d').date()
            except ValueError:
                target_date = None

            if employee and target_date:
                session, _ = WorkSession.objects.get_or_create(employee=employee, work_date=target_date)
                reference_start_dt = session.corrected_started_at or session.started_at

                if start_time or end_time:
                    session.day_type = WorkSession.DAY_TYPE_WORK

                if start_time:
                    corrected_start = datetime.combine(target_date, start_time)
                    session.corrected_started_at = timezone.make_aware(corrected_start, timezone.get_current_timezone())
                    reference_start_dt = session.corrected_started_at
                else:
                    session.corrected_started_at = None
                    reference_start_dt = session.started_at

                if end_time:
                    end_date_for_correction = target_date
                    if reference_start_dt and end_time <= reference_start_dt.timetz().replace(tzinfo=None):
                        end_date_for_correction = target_date + timedelta(days=1)
                    corrected_end = datetime.combine(end_date_for_correction, end_time)
                    session.corrected_ended_at = timezone.make_aware(corrected_end, timezone.get_current_timezone())
                else:
                    session.corrected_ended_at = None

                session.correction_note = note or None
                session.corrected_by = request.user
                session.corrected_at = timezone.now()
                session.save()

                _create_audit_event(
                    request,
                    'timekeeping_corrected',
                    employee=employee,
                    metadata={
                        'work_date': str(target_date),
                        'corrected_started_at': session.corrected_started_at.isoformat() if session.corrected_started_at else None,
                        'corrected_ended_at': session.corrected_ended_at.isoformat() if session.corrected_ended_at else None,
                        'note': session.correction_note,
                    },
                )

                return redirect(
                    f"{request.path}?employee={employee.id}&month={target_date.month}&year={target_date.year}"
                )

        if action == 'delete_marking':
            employee_id = request.POST.get('employee_id')
            target_date_raw = request.POST.get('target_date')
            delete_target = (request.POST.get('delete_target') or '').strip()

            employee = Employee.objects.filter(id=employee_id).first()
            try:
                target_date = datetime.strptime(target_date_raw or '', '%Y-%m-%d').date()
            except ValueError:
                target_date = None

            redirect_employee = request.POST.get('redirect_employee') or 'all'
            redirect_month = request.POST.get('redirect_month') or str(today.month)
            redirect_year = request.POST.get('redirect_year') or str(today.year)

            redirect_url = f"{request.path}?employee={redirect_employee}&month={redirect_month}&year={redirect_year}"

            if not employee or not target_date or delete_target not in {'start', 'end', 'day'}:
                return redirect(f"{redirect_url}&outcome=delete_invalid")

            session = WorkSession.objects.filter(employee=employee, work_date=target_date).first()
            if not session:
                return redirect(f"{redirect_url}&outcome=delete_missing")

            payload = {
                'work_date': str(target_date),
                'delete_target': delete_target,
                'started_at': session.started_at.isoformat() if session.started_at else None,
                'ended_at': session.ended_at.isoformat() if session.ended_at else None,
                'corrected_started_at': session.corrected_started_at.isoformat() if session.corrected_started_at else None,
                'corrected_ended_at': session.corrected_ended_at.isoformat() if session.corrected_ended_at else None,
            }

            request_types = []
            if delete_target in {'start', 'day'}:
                request_types.extend([WorkMarkRequest.MARK_TYPE_START, WorkMarkRequest.MARK_TYPE_BOTH])
            if delete_target in {'end', 'day'}:
                request_types.extend([WorkMarkRequest.MARK_TYPE_END, WorkMarkRequest.MARK_TYPE_BOTH])

            related_requests = list(
                WorkMarkRequest.objects
                .filter(
                    employee=employee,
                    work_date=target_date,
                    mark_type__in=request_types,
                    status__in=[WorkMarkRequest.STATUS_APPROVED, WorkMarkRequest.STATUS_PENDING],
                )
                .order_by('-created_at')
            )

            payload['related_request_ids'] = [req.id for req in related_requests]
            payload['related_request_statuses'] = [req.status for req in related_requests]

            _clear_session_marking(session, delete_target)

            if related_requests:
                WorkMarkRequest.objects.filter(id__in=[req.id for req in related_requests]).delete()

            if delete_target != 'day' and not _session_has_markings(session):
                session.delete()

            _create_audit_event(
                request,
                'timekeeping_marking_deleted',
                employee=employee,
                metadata=payload,
            )

            return redirect(f"{redirect_url}&outcome=deleted_{delete_target}")

    try:
        year = int(request.GET.get('year', today.year))
    except (TypeError, ValueError):
        year = today.year

    try:
        month = int(request.GET.get('month', today.month))
    except (TypeError, ValueError):
        month = today.month

    month = max(1, min(month, 12))

    outcome = (request.GET.get('outcome') or '').strip()
    if outcome == 'deleted_start':
        feedback = 'Marcatura di ingresso eliminata.'
        feedback_level = 'warning'
    elif outcome == 'deleted_end':
        feedback = 'Marcatura di uscita eliminata.'
        feedback_level = 'warning'
    elif outcome == 'deleted_day':
        feedback = 'Giornata di marcatura eliminata.'
        feedback_level = 'warning'
    elif outcome == 'delete_missing':
        feedback = 'Marcatura non trovata o gia eliminata.'
        feedback_level = 'danger'
    elif outcome == 'delete_invalid':
        feedback = 'Richiesta di cancellazione non valida.'
        feedback_level = 'danger'

    employees = list(Employee.objects.select_related('user'))
    _decorate_employee_display_names(employees)
    employees.sort(key=_employee_name_sort_key)

    month_choices = [
        (1, 'Gennaio'),
        (2, 'Febbraio'),
        (3, 'Marzo'),
        (4, 'Aprile'),
        (5, 'Maggio'),
        (6, 'Giugno'),
        (7, 'Luglio'),
        (8, 'Agosto'),
        (9, 'Settembre'),
        (10, 'Ottobre'),
        (11, 'Novembre'),
        (12, 'Dicembre'),
    ]

    employee_filter = (request.GET.get('employee') or 'all').strip()
    all_mode = employee_filter == 'all'

    selected_employee = None
    if not all_mode:
        if employee_filter:
            selected_employee = next((employee for employee in employees if str(employee.id) == employee_filter), None)

    rows = []
    total_minutes = 0
    incomplete_days = 0

    month_last_day = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)
    end_date = date(year, month, month_last_day)

    pending_mark_requests = (
        WorkMarkRequest.objects
        .select_related('employee')
        .filter(status=WorkMarkRequest.STATUS_PENDING)
        .order_by('-created_at')[:20]
    )
    pending_vacation_requests = (
        VacationRequest.objects
        .select_related('employee')
        .filter(status=VacationRequest.STATUS_PENDING)
        .order_by('-created_at')[:20]
    )
    for request_obj in pending_mark_requests:
        request_obj.employee.display_name = _employee_admin_display_name(request_obj.employee)
    for request_obj in pending_vacation_requests:
        request_obj.employee.display_name = _employee_admin_display_name(request_obj.employee)

    # Backfill in lettura: include nel report mensile eventuali approvazioni storiche.
    if all_mode:
        _sync_approved_requests_for_range(start_date, end_date)
        _reconcile_overnight_sessions(start_date=start_date, end_date=end_date)
        _sync_approved_vacations_for_range(start_date, end_date)
    elif selected_employee:
        _sync_approved_requests_for_range(start_date, end_date, employee=selected_employee)
        _reconcile_overnight_sessions(employee=selected_employee, start_date=start_date, end_date=end_date)
        _sync_approved_vacations_for_range(start_date, end_date, employee=selected_employee)

    if all_mode:
        marked_sessions_qs = (
            WorkSession.objects
            .filter(work_date__range=(start_date, end_date))
            .filter(
                Q(day_type=WorkSession.DAY_TYPE_VACATION)
                |
                Q(started_at__isnull=False)
                | Q(ended_at__isnull=False)
                | Q(corrected_started_at__isnull=False)
                | Q(corrected_ended_at__isnull=False)
            )
            .select_related('employee')
            .order_by('employee__last_name', 'employee__first_name', 'work_date')
        )

        sessions_by_employee_day = {}
        employee_ids = []
        seen_ids = set()
        for s in marked_sessions_qs:
            if s.employee_id not in seen_ids:
                seen_ids.add(s.employee_id)
                employee_ids.append(s.employee_id)
            sessions_by_employee_day[(s.employee_id, s.work_date.day)] = s

        if employee_ids:
            matrix_employees = list(
                Employee.objects
                .filter(id__in=employee_ids)
            )
            _decorate_employee_display_names(matrix_employees)
            matrix_employees.sort(key=_employee_name_sort_key)
        else:
            matrix_employees = []

        day_numbers = list(range(1, month_last_day + 1))
        matrix_rows = []
        for emp in matrix_employees:
            cells = []
            employee_total_minutes = 0
            for day_number in day_numbers:
                session = sessions_by_employee_day.get((emp.id, day_number))
                if session:
                    employee_total_minutes += session.worked_minutes()
                cells.append({
                    'day': day_number,
                    'value': _session_cell_text(session),
                })
            matrix_rows.append({
                'employee': emp,
                'cells': cells,
                'month_total': f"{employee_total_minutes // 60:02d}:{employee_total_minutes % 60:02d}",
            })

        export_format = request.GET.get('format')
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="report_marcature_tutti_{year}_{month}.csv"'
            response.write('\ufeff')
            writer = csv.writer(response, delimiter=';')
            day_headers = [date(year, month, d).strftime('%d/%m/%Y') for d in day_numbers]
            writer.writerow(['Dipendente'] + day_headers + ['Totale ore lavorate'])
            for row in matrix_rows:
                writer.writerow(
                    [row['employee'].display_name]
                    + [cell['value'] for cell in row['cells']]
                    + [row['month_total']]
                )
            return response

        return render(request, 'portal/admin_timekeeping.html', {
            'employees': employees,
            'month_choices': month_choices,
            'selected_employee': None,
            'selected_year': year,
            'selected_month': month,
            'rows': [],
            'month_total': '00:00',
            'incomplete_days': 0,
            'all_mode': True,
            'day_numbers': day_numbers,
            'matrix_rows': matrix_rows,
            'employee_filter': 'all',
            'pending_mark_requests': pending_mark_requests,
            'pending_vacation_requests': pending_vacation_requests,
            'feedback': feedback,
            'feedback_level': feedback_level,
        })

    if selected_employee:
        sessions = (
            WorkSession.objects
            .filter(employee=selected_employee, work_date__range=(start_date, end_date))
            .select_related('start_zone', 'end_zone', 'corrected_by')
            .order_by('work_date')
        )
        by_day = {s.work_date: s for s in sessions}

        export_format = request.GET.get('format')
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="marcature_{selected_employee.id}_{year}_{month}.csv"'
            response.write('\ufeff')
            writer = csv.writer(response, delimiter=';')
            writer.writerow([
                'Data',
                'Tipologia giorno',
                'Ingresso',
                'Uscita',
                'Totale',
                'Zona ingresso',
                'Zona uscita',
                'In zona ingresso',
                'In zona uscita',
                'Corretto da admin',
                'Nota correzione',
            ])

            for day in range(1, month_last_day + 1):
                current_date = date(year, month, day)
                session = by_day.get(current_date)
                if not session:
                    writer.writerow([current_date.strftime('%d/%m/%Y'), 'Lavoro', '', '', '00:00', '', '', '', '', '', ''])
                    continue
                writer.writerow([
                    current_date.strftime('%d/%m/%Y'),
                    'Ferie' if session.day_type == WorkSession.DAY_TYPE_VACATION else 'Lavoro',
                    session.effective_started_at().strftime('%H:%M') if session.effective_started_at() else '',
                    session.effective_ended_at().strftime('%H:%M') if session.effective_ended_at() else '',
                    'FERIE' if session.day_type == WorkSession.DAY_TYPE_VACATION else session.worked_hours_display(),
                    session.start_zone.name if session.start_zone else '',
                    session.end_zone.name if session.end_zone else '',
                    'si' if session.start_within_zone else 'no',
                    'si' if session.end_within_zone else 'no',
                    'si' if session.corrected_at else 'no',
                    session.correction_note or '',
                ])
            return response

        if export_format == 'xlsx':
            from openpyxl import Workbook
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = 'Marcature'

            headers = [
                'Data',
                'Tipologia giorno',
                'Ingresso',
                'Uscita',
                'Totale',
                'Zona ingresso',
                'Zona uscita',
                'In zona ingresso',
                'In zona uscita',
                'Corretto da admin',
                'Nota correzione',
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

            total_month_minutes = 0
            for day in range(1, month_last_day + 1):
                current_date = date(year, month, day)
                session = by_day.get(current_date)
                if not session:
                    ws.append([current_date.strftime('%d/%m/%Y'), 'Lavoro', '', '', '00:00', '', '', '', '', '', ''])
                    continue

                minutes = session.worked_minutes()
                total_month_minutes += minutes
                ws.append([
                    current_date.strftime('%d/%m/%Y'),
                    'Ferie' if session.day_type == WorkSession.DAY_TYPE_VACATION else 'Lavoro',
                    session.effective_started_at().strftime('%H:%M') if session.effective_started_at() else '',
                    session.effective_ended_at().strftime('%H:%M') if session.effective_ended_at() else '',
                    'FERIE' if session.day_type == WorkSession.DAY_TYPE_VACATION else session.worked_hours_display(),
                    session.start_zone.name if session.start_zone else '',
                    session.end_zone.name if session.end_zone else '',
                    'si' if session.start_within_zone else 'no',
                    'si' if session.end_within_zone else 'no',
                    'si' if session.corrected_at else 'no',
                    session.correction_note or '',
                ])

            summary_row = month_last_day + 3
            ws.cell(row=summary_row, column=1, value='Totale mese')
            ws.cell(
                row=summary_row,
                column=4,
                value=f"{total_month_minutes // 60:02d}:{total_month_minutes % 60:02d}",
            )
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            ws.cell(row=summary_row, column=4).font = Font(bold=True)

            # Larghezze minime utili per una lettura immediata del file.
            for idx, width in {
                1: 12,
                2: 16,
                3: 10,
                4: 10,
                5: 10,
                6: 22,
                7: 22,
                8: 14,
                9: 14,
                10: 16,
                11: 30,
            }.items():
                ws.column_dimensions[chr(64 + idx)].width = width

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="marcature_{selected_employee.id}_{year}_{month}.xlsx"'
            )
            wb.save(response)
            return response

        for day in range(1, month_last_day + 1):
            current_date = date(year, month, day)
            session = by_day.get(current_date)

            if not session:
                rows.append({
                    'date': current_date,
                    'day_type': WorkSession.DAY_TYPE_WORK,
                    'entry': None,
                    'exit': None,
                    'total': '00:00',
                    'status': 'missing',
                    'start_zone': None,
                    'end_zone': None,
                    'start_latitude': None,
                    'start_longitude': None,
                    'end_latitude': None,
                    'end_longitude': None,
                })
                continue

            worked = session.worked_minutes()
            total_minutes += worked
            effective_start = session.effective_started_at()
            effective_end = session.effective_ended_at()
            if session.day_type != WorkSession.DAY_TYPE_VACATION and effective_start and not effective_end:
                incomplete_days += 1

            row_status = 'vacation' if session.day_type == WorkSession.DAY_TYPE_VACATION else ('ok' if effective_start and effective_end else 'partial')

            rows.append({
                'date': current_date,
                'day_type': session.day_type,
                'entry': effective_start,
                'exit': effective_end,
                'total': 'FERIE' if session.day_type == WorkSession.DAY_TYPE_VACATION else session.worked_hours_display(),
                'status': row_status,
                'start_zone': session.start_zone,
                'end_zone': session.end_zone,
                'start_latitude': session.start_latitude,
                'start_longitude': session.start_longitude,
                'end_latitude': session.end_latitude,
                'end_longitude': session.end_longitude,
                'start_within_zone': session.start_within_zone,
                'end_within_zone': session.end_within_zone,
                'corrected': bool(session.corrected_at),
                'correction_note': session.correction_note,
            })

    return render(request, 'portal/admin_timekeeping.html', {
        'employees': employees,
        'month_choices': month_choices,
        'selected_employee': selected_employee,
        'selected_year': year,
        'selected_month': month,
        'rows': rows,
        'month_total': f"{total_minutes // 60:02d}:{total_minutes % 60:02d}",
        'incomplete_days': incomplete_days,
        'all_mode': False,
        'day_numbers': [],
        'matrix_rows': [],
        'employee_filter': str(selected_employee.id) if selected_employee else '',
        'pending_mark_requests': pending_mark_requests,
        'pending_vacation_requests': pending_vacation_requests,
        'feedback': feedback,
        'feedback_level': feedback_level,
    })


@login_required
def admin_out_of_zone_requests(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    feedback = None
    feedback_level = 'info'

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'delete_out_of_zone_request':
            req_id = request.POST.get('request_id')
            redirect_query = (request.POST.get('redirect_query') or '').strip()
            request_obj = (
                WorkMarkRequest.objects
                .select_related('employee', 'reviewed_by')
                .filter(id=req_id)
                .first()
            )

            if request_obj:
                employee = request_obj.employee
                request_payload = {
                    'request_id': request_obj.id,
                    'work_date': str(request_obj.work_date),
                    'status': request_obj.status,
                    'mark_type': request_obj.mark_type,
                    'reason': request_obj.reason,
                    'review_note': request_obj.review_note,
                    'reviewed_at': request_obj.reviewed_at.isoformat() if request_obj.reviewed_at else None,
                    'reviewed_by': request_obj.reviewed_by.username if request_obj.reviewed_by else None,
                }
                request_obj.delete()

                _create_audit_event(
                    request,
                    'timekeeping_out_of_zone_deleted',
                    employee=employee,
                    metadata=request_payload,
                )

                redirect_url = request.path
                if redirect_query:
                    redirect_url = f'{redirect_url}?{redirect_query}&outcome=deleted'
                else:
                    redirect_url = f'{redirect_url}?outcome=deleted'
                return redirect(redirect_url)

            redirect_url = request.path
            if redirect_query:
                redirect_url = f'{redirect_url}?{redirect_query}&outcome=missing'
            else:
                redirect_url = f'{redirect_url}?outcome=missing'
            return redirect(redirect_url)

    outcome = (request.GET.get('outcome') or '').strip()
    if outcome == 'deleted':
        feedback = 'Richiesta fuori zona eliminata definitivamente.'
        feedback_level = 'warning'
    elif outcome == 'missing':
        feedback = 'Richiesta non trovata o gia eliminata.'
        feedback_level = 'danger'

    employees = Employee.objects.order_by('last_name', 'first_name')
    history_status = (request.GET.get('status') or 'all').strip()
    if history_status not in {
        'all',
        WorkMarkRequest.STATUS_PENDING,
        WorkMarkRequest.STATUS_APPROVED,
        WorkMarkRequest.STATUS_REJECTED,
    }:
        history_status = 'all'

    history_employee_filter = (request.GET.get('employee') or 'all').strip()
    history_requests_qs = (
        WorkMarkRequest.objects
        .select_related('employee', 'reviewed_by')
        .order_by('-created_at')
    )
    if history_status != 'all':
        history_requests_qs = history_requests_qs.filter(status=history_status)
    if history_employee_filter != 'all':
        history_requests_qs = history_requests_qs.filter(employee_id=history_employee_filter)

    history_requests_page = Paginator(history_requests_qs, 25).get_page(request.GET.get('page') or 1)
    pending_requests_count = WorkMarkRequest.objects.filter(status=WorkMarkRequest.STATUS_PENDING).count()

    return render(request, 'portal/admin_out_of_zone_requests.html', {
        'employees': employees,
        'history_requests_page': history_requests_page,
        'history_status': history_status,
        'history_employee_filter': history_employee_filter,
        'pending_requests_count': pending_requests_count,
        'feedback': feedback,
        'feedback_level': feedback_level,
        'current_query_string': request.GET.urlencode,
    })


@login_required
def admin_vacation_requests(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    feedback = None
    feedback_level = 'info'

    if request.method == 'POST':
        action = request.POST.get('action')
        req_id = request.POST.get('request_id')
        redirect_query = (request.POST.get('redirect_query') or '').strip()
        request_obj = (
            VacationRequest.objects
            .select_related('employee', 'reviewed_by')
            .filter(id=req_id)
            .first()
        )

        if action in {'approve_vacation_request', 'reject_vacation_request'} and request_obj:
            request_obj.status = (
                VacationRequest.STATUS_APPROVED
                if action == 'approve_vacation_request'
                else VacationRequest.STATUS_REJECTED
            )
            request_obj.review_note = (request.POST.get('review_note') or '').strip() or None
            request_obj.reviewed_by = request.user
            request_obj.reviewed_at = timezone.now()
            request_obj.save(update_fields=['status', 'review_note', 'reviewed_by', 'reviewed_at', 'updated_at'])

            if request_obj.status == VacationRequest.STATUS_APPROVED:
                _apply_approved_vacation_request_to_sessions(request_obj)

            _create_audit_event(
                request,
                'vacation_reviewed',
                employee=request_obj.employee,
                metadata={
                    'request_id': request_obj.id,
                    'start_date': str(request_obj.start_date),
                    'end_date': str(request_obj.end_date),
                    'status': request_obj.status,
                    'via': 'admin_vacation_requests',
                },
            )

            redirect_url = request.path
            if redirect_query:
                redirect_url = f'{redirect_url}?{redirect_query}&outcome=reviewed'
            else:
                redirect_url = f'{redirect_url}?outcome=reviewed'
            return redirect(redirect_url)

        if action == 'delete_vacation_request':
            if request_obj:
                employee = request_obj.employee
                request_payload = {
                    'request_id': request_obj.id,
                    'start_date': str(request_obj.start_date),
                    'end_date': str(request_obj.end_date),
                    'status': request_obj.status,
                    'reason': request_obj.reason,
                    'review_note': request_obj.review_note,
                }
                request_obj.delete()

                _create_audit_event(
                    request,
                    'vacation_deleted',
                    employee=employee,
                    metadata=request_payload,
                )

                redirect_url = request.path
                if redirect_query:
                    redirect_url = f'{redirect_url}?{redirect_query}&outcome=deleted'
                else:
                    redirect_url = f'{redirect_url}?outcome=deleted'
                return redirect(redirect_url)

            redirect_url = request.path
            if redirect_query:
                redirect_url = f'{redirect_url}?{redirect_query}&outcome=missing'
            else:
                redirect_url = f'{redirect_url}?outcome=missing'
            return redirect(redirect_url)

    outcome = (request.GET.get('outcome') or '').strip()
    if outcome == 'deleted':
        feedback = 'Richiesta ferie eliminata definitivamente.'
        feedback_level = 'warning'
    elif outcome == 'missing':
        feedback = 'Richiesta ferie non trovata o gia eliminata.'
        feedback_level = 'danger'
    elif outcome == 'reviewed':
        feedback = 'Richiesta ferie aggiornata correttamente.'
        feedback_level = 'success'

    employees = Employee.objects.order_by('last_name', 'first_name')
    history_status = (request.GET.get('status') or 'all').strip()
    if history_status not in {
        'all',
        VacationRequest.STATUS_PENDING,
        VacationRequest.STATUS_APPROVED,
        VacationRequest.STATUS_REJECTED,
    }:
        history_status = 'all'

    history_employee_filter = (request.GET.get('employee') or 'all').strip()
    history_requests_qs = (
        VacationRequest.objects
        .select_related('employee', 'reviewed_by')
        .order_by('-created_at')
    )
    if history_status != 'all':
        history_requests_qs = history_requests_qs.filter(status=history_status)
    if history_employee_filter != 'all':
        history_requests_qs = history_requests_qs.filter(employee_id=history_employee_filter)

    history_requests_page = Paginator(history_requests_qs, 25).get_page(request.GET.get('page') or 1)
    pending_requests_count = VacationRequest.objects.filter(status=VacationRequest.STATUS_PENDING).count()

    return render(request, 'portal/admin_vacation_requests.html', {
        'employees': employees,
        'history_requests_page': history_requests_page,
        'history_status': history_status,
        'history_employee_filter': history_employee_filter,
        'pending_requests_count': pending_requests_count,
        'feedback': feedback,
        'feedback_level': feedback_level,
        'current_query_string': request.GET.urlencode,
    })


@login_required
def admin_work_zones(request):
    """Configurazione zone di lavoro e assegnazioni dipendente-zona."""
    if not request.user.is_staff:
        return redirect('dashboard')

    feedback = None
    feedback_level = 'info'

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_zone':
            name = (request.POST.get('name') or '').strip()
            shape = (request.POST.get('shape') or 'circle').strip()

            if shape == 'rect':
                n = _parse_coordinate(request.POST.get('rect_north'))
                s = _parse_coordinate(request.POST.get('rect_south'))
                e = _parse_coordinate(request.POST.get('rect_east'))
                w = _parse_coordinate(request.POST.get('rect_west'))
                if not name or None in (n, s, e, w) or not (n > s and e > w):
                    feedback = 'Rettangolo non valido. Inserisci Nord>Sud ed Est>Ovest.'
                    feedback_level = 'danger'
                else:
                    # Centro per compatibilita con viste esistenti
                    center_lat = (float(n) + float(s)) / 2.0
                    center_lon = (float(e) + float(w)) / 2.0
                    WorkZone.objects.create(
                        name=name,
                        shape=WorkZone.SHAPE_RECT,
                        latitude=center_lat,
                        longitude=center_lon,
                        radius_meters=100,
                        rect_north=n,
                        rect_south=s,
                        rect_east=e,
                        rect_west=w,
                    )
                    feedback = 'Zona rettangolare creata correttamente.'
                    feedback_level = 'success'
            else:
                radius_raw = request.POST.get('radius_meters') or '100'
                latitude = _parse_coordinate(request.POST.get('latitude'))
                longitude = _parse_coordinate(request.POST.get('longitude'))
                try:
                    radius = max(int(radius_raw), 10)
                except (TypeError, ValueError):
                    radius = 100

                if not name or latitude is None or longitude is None:
                    feedback = 'Compila nome e coordinate valide per creare la zona.'
                    feedback_level = 'danger'
                else:
                    WorkZone.objects.create(
                        name=name,
                        shape=WorkZone.SHAPE_CIRCLE,
                        latitude=latitude,
                        longitude=longitude,
                        radius_meters=radius,
                        rect_north=None,
                        rect_south=None,
                        rect_east=None,
                        rect_west=None,
                    )
                    feedback = 'Zona circolare creata correttamente.'
                    feedback_level = 'success'

        if action == 'update_zone':
            zone_id = request.POST.get('zone_id')
            zone = WorkZone.objects.filter(id=zone_id).first()
            if not zone:
                feedback = 'Zona non trovata.'
                feedback_level = 'danger'
            else:
                name = (request.POST.get('name') or '').strip()
                shape = (request.POST.get('shape') or 'circle').strip()
                if not name:
                    feedback = 'Inserisci un nome valido.'
                    feedback_level = 'danger'
                else:
                    zone.name = name
                    if shape == 'rect':
                        n = _parse_coordinate(request.POST.get('rect_north'))
                        s = _parse_coordinate(request.POST.get('rect_south'))
                        e = _parse_coordinate(request.POST.get('rect_east'))
                        w = _parse_coordinate(request.POST.get('rect_west'))
                        if None in (n, s, e, w) or not (n > s and e > w):
                            feedback = 'Rettangolo non valido. Inserisci Nord>Sud ed Est>Ovest.'
                            feedback_level = 'danger'
                        else:
                            center_lat = (float(n) + float(s)) / 2.0
                            center_lon = (float(e) + float(w)) / 2.0
                            zone.shape = WorkZone.SHAPE_RECT
                            zone.latitude = center_lat
                            zone.longitude = center_lon
                            zone.rect_north = n
                            zone.rect_south = s
                            zone.rect_east = e
                            zone.rect_west = w
                            # mantieni radius invariato ma irrilevante
                            zone.save()
                            feedback = 'Zona aggiornata (rettangolo).'
                            feedback_level = 'success'
                    else:
                        latitude = _parse_coordinate(request.POST.get('latitude'))
                        longitude = _parse_coordinate(request.POST.get('longitude'))
                        radius_raw = request.POST.get('radius_meters') or '100'
                        try:
                            radius = max(int(radius_raw), 10)
                        except (TypeError, ValueError):
                            radius = 100
                        if latitude is None or longitude is None:
                            feedback = 'Coordinate non valide.'
                            feedback_level = 'danger'
                        else:
                            zone.shape = WorkZone.SHAPE_CIRCLE
                            zone.latitude = latitude
                            zone.longitude = longitude
                            zone.radius_meters = radius
                            zone.rect_north = None
                            zone.rect_south = None
                            zone.rect_east = None
                            zone.rect_west = None
                            zone.save()
                            feedback = 'Zona aggiornata (cerchio).'
                            feedback_level = 'success'

        if action == 'assign_zone':
            employee_id = request.POST.get('employee_id')
            zone_id = request.POST.get('zone_id')
            valid_from = request.POST.get('valid_from')
            strict_geofence = request.POST.get('strict_geofence') == '1'

            employee = Employee.objects.filter(id=employee_id).first()
            zone = WorkZone.objects.filter(id=zone_id, is_active=True).first()

            if not employee or not zone:
                feedback = 'Seleziona dipendente e zona validi.'
                feedback_level = 'danger'
            else:
                try:
                    EmployeeWorkZone.objects.create(
                        employee=employee,
                        zone=zone,
                        valid_from=valid_from or timezone.localdate(),
                        strict_geofence=strict_geofence,
                    )
                    feedback = 'Zona assegnata al dipendente.'
                    feedback_level = 'success'
                except IntegrityError:
                    feedback = 'Assegnazione gia presente con la stessa data di validita.'
                    feedback_level = 'warning'

        if action == 'deactivate_assignment':
            assignment_id = request.POST.get('assignment_id')
            assignment = EmployeeWorkZone.objects.filter(id=assignment_id).first()

            if not assignment:
                feedback = 'Assegnazione non trovata.'
                feedback_level = 'danger'
            else:
                assignment.is_active = False
                assignment.valid_to = assignment.valid_to or timezone.localdate()
                assignment.save(update_fields=['is_active', 'valid_to'])
                feedback = 'Assegnazione disattivata.'
                feedback_level = 'warning'

        if action == 'delete_zone':
            zone_id = request.POST.get('zone_id')
            zone = WorkZone.objects.filter(id=zone_id).first()

            if not zone:
                feedback = 'Zona non trovata.'
                feedback_level = 'danger'
            else:
                zone_name = zone.name
                zone.delete()
                feedback = f'Zona "{zone_name}" eliminata.'
                feedback_level = 'warning'

        if action == 'delete_assignment':
            assignment_id = request.POST.get('assignment_id')
            assignment = EmployeeWorkZone.objects.select_related('employee', 'zone').filter(id=assignment_id).first()

            if not assignment:
                feedback = 'Assegnazione non trovata.'
                feedback_level = 'danger'
            else:
                employee_name = assignment.employee.full_name
                zone_name = assignment.zone.name
                assignment.delete()
                feedback = f'Assegnazione eliminata: {employee_name} -> {zone_name}.'
                feedback_level = 'warning'

    employees = Employee.objects.order_by('last_name', 'first_name')
    zones = WorkZone.objects.order_by('name')
    assignments = (
        EmployeeWorkZone.objects
        .select_related('employee', 'zone')
        .order_by('employee__last_name', 'employee__first_name', '-is_active', 'zone__name', '-created_at')
    )
    assignment_groups_map = OrderedDict()

    for assignment in assignments:
        employee_id = assignment.employee_id
        group = assignment_groups_map.get(employee_id)
        if group is None:
            group = {
                'employee': assignment.employee,
                'assignments': [],
                'total_count': 0,
                'active_count': 0,
            }
            assignment_groups_map[employee_id] = group

        group['assignments'].append(assignment)
        group['total_count'] += 1
        if assignment.is_active:
            group['active_count'] += 1

    assignment_groups = list(assignment_groups_map.values())

    return render(request, 'portal/admin_work_zones.html', {
        'employees': employees,
        'zones': zones,
        'assignments': assignments,
        'assignment_groups': assignment_groups,
        'feedback': feedback,
        'feedback_level': feedback_level,
    })


# =========================================================
# APERTURA CEDOLINO + EMAIL NOTIFICA LETTURA
# =========================================================

def _append_cache_buster(url, version_token):
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query['v'] = str(version_token)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment))

@login_required
def open_payslip(request, payslip_id):
    payslip = get_object_or_404(Payslip, id=payslip_id)

    if not request.user.is_staff and payslip.employee.user != request.user:
        return HttpResponse("Non autorizzato", status=403)

    if not request.user.is_staff:
        view, created = PayslipView.objects.get_or_create(payslip=payslip)

        # Se è la prima visualizzazione invia email
        if created:
            send_read_notification_email(payslip)

        # Audit: apertura cedolino da parte del dipendente
        _create_audit_event(
            request,
            "payslip_opened",
            employee=payslip.employee,
            payslip=payslip,
            metadata={"first_view": created},
        )

    # Reindirizza sempre all'URL pubblico del PDF (R2 gestisce la visualizzazione/download)
    try:
        version_token = f"{payslip.id}-{int(payslip.uploaded_at.timestamp())}"
        url = _append_cache_buster(payslip.pdf.url, version_token)
        return HttpResponseRedirect(url)
    except Exception:
        logger.exception('open_payslip: failed to build payslip URL id=%s', payslip_id)
        return HttpResponse('Errore nel recupero del file', status=500)


@login_required
def open_cud(request, cud_id):
    """Apertura del PDF CUD annuale."""
    cud = get_object_or_404(Cud, id=cud_id)

    if not request.user.is_staff and cud.employee.user != request.user:
        return HttpResponse("Non autorizzato", status=403)

    # Tracciamento apertura e audit (solo se non staff)
    if not request.user.is_staff:
        view, created = CudView.objects.get_or_create(cud=cud)

        _create_audit_event(
            request,
            "cud_opened",
            employee=cud.employee,
            metadata={
                "year": cud.year,
                "first_view": created,
            },
        )

    try:
        version_token = f"{cud.id}-{int(cud.uploaded_at.timestamp())}"
        url = _append_cache_buster(cud.pdf.url, version_token)
        return HttpResponseRedirect(url)
    except Exception:
        logger.exception('open_cud: failed to build CUD URL id=%s', cud_id)
        return HttpResponse('Errore nel recupero del file CUD', status=500)

# =========================================================
# ADMIN DASHBOARD
# =========================================================

@login_required
def admin_dashboard(request):
    if not user_has_full_admin_access(request.user):
        if user_has_today_markings_only_access(request.user):
            return redirect('today_markings_dashboard')
        return redirect('dashboard')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action in {'approve_mark_request', 'reject_mark_request'}:
            req_id = request.POST.get('request_id')
            request_obj = WorkMarkRequest.objects.filter(id=req_id).select_related('employee').first()
            if request_obj:
                request_obj.status = (
                    WorkMarkRequest.STATUS_APPROVED
                    if action == 'approve_mark_request'
                    else WorkMarkRequest.STATUS_REJECTED
                )
                request_obj.review_note = (request.POST.get('review_note') or '').strip() or None
                request_obj.reviewed_by = request.user
                request_obj.reviewed_at = timezone.now()
                request_obj.save(update_fields=['status', 'review_note', 'reviewed_by', 'reviewed_at', 'updated_at'])

                if request_obj.status == WorkMarkRequest.STATUS_APPROVED:
                    _apply_approved_mark_request_to_session(request_obj)

                _create_audit_event(
                    request,
                    'timekeeping_out_of_zone_reviewed',
                    employee=request_obj.employee,
                    metadata={
                        'request_id': request_obj.id,
                        'work_date': str(request_obj.work_date),
                        'status': request_obj.status,
                        'mark_type': request_obj.mark_type,
                        'marked_at': request_obj.created_at.isoformat() if request_obj.status == WorkMarkRequest.STATUS_APPROVED else None,
                        'via': 'admin_dashboard',
                    },
                )

            return redirect('admin_dashboard')

        if action in {'approve_vacation_request', 'reject_vacation_request'}:
            req_id = request.POST.get('request_id')
            request_obj = VacationRequest.objects.filter(id=req_id).select_related('employee').first()
            if request_obj:
                request_obj.status = (
                    VacationRequest.STATUS_APPROVED
                    if action == 'approve_vacation_request'
                    else VacationRequest.STATUS_REJECTED
                )
                request_obj.review_note = (request.POST.get('review_note') or '').strip() or None
                request_obj.reviewed_by = request.user
                request_obj.reviewed_at = timezone.now()
                request_obj.save(update_fields=['status', 'review_note', 'reviewed_by', 'reviewed_at', 'updated_at'])

                if request_obj.status == VacationRequest.STATUS_APPROVED:
                    _apply_approved_vacation_request_to_sessions(request_obj)

                _create_audit_event(
                    request,
                    'vacation_reviewed',
                    employee=request_obj.employee,
                    metadata={
                        'request_id': request_obj.id,
                        'start_date': str(request_obj.start_date),
                        'end_date': str(request_obj.end_date),
                        'status': request_obj.status,
                        'via': 'admin_dashboard',
                    },
                )

            return redirect('admin_dashboard')

    today = timezone.localdate()
    _sync_approved_requests_for_range(today, today)
    _sync_approved_vacations_for_range(today, today)
    pending_mark_requests = (
        WorkMarkRequest.objects
        .select_related('employee')
        .filter(status=WorkMarkRequest.STATUS_PENDING)
        .order_by('-created_at')[:15]
    )
    pending_vacation_requests = (
        VacationRequest.objects
        .select_related('employee')
        .filter(status=VacationRequest.STATUS_PENDING)
        .order_by('-created_at')[:15]
    )
    pending_mark_requests_count = WorkMarkRequest.objects.filter(status=WorkMarkRequest.STATUS_PENDING).count()
    pending_vacation_requests_count = VacationRequest.objects.filter(status=VacationRequest.STATUS_PENDING).count()

    month_start = today.replace(day=1)
    monthly_requests = WorkMarkRequest.objects.filter(created_at__date__gte=month_start, created_at__date__lte=today)
    monthly_vacation_requests = VacationRequest.objects.filter(created_at__date__gte=month_start, created_at__date__lte=today)

    today_marked_sessions = _prepare_marked_sessions_for_date(list(_today_marked_sessions_queryset(today)), today)

    entered_today_count = sum(1 for session in today_marked_sessions if session.display_started_at)
    completed_today_count = sum(1 for session in today_marked_sessions if session.display_started_at and session.display_ended_at)
    incomplete_today_count = max(entered_today_count - completed_today_count, 0)
    outside_today_count = sum(
        1
        for session in today_marked_sessions
        if (
            session.display_started_at and session.start_zone_id and not session.start_within_zone
        )
        or (
            session.display_ended_at and session.end_zone_id and not session.end_within_zone
        )
    )
    approved_month_count = monthly_requests.filter(status=WorkMarkRequest.STATUS_APPROVED).count()
    rejected_month_count = monthly_requests.filter(status=WorkMarkRequest.STATUS_REJECTED).count()
    approved_vacation_month_count = monthly_vacation_requests.filter(status=VacationRequest.STATUS_APPROVED).count()
    active_zone_count = WorkZone.objects.filter(is_active=True).count()
    active_assignment_count = EmployeeWorkZone.objects.filter(
        is_active=True,
        valid_from__lte=today,
    ).filter(
        Q(valid_to__isnull=True) | Q(valid_to__gte=today)
    ).count()
    employee_count = Employee.objects.count()

    return render(request, "portal/admin_dashboard.html", {
        "entered_today_count": entered_today_count,
        "completed_today_count": completed_today_count,
        "incomplete_today_count": incomplete_today_count,
        "outside_today_count": outside_today_count,
        "approved_month_count": approved_month_count,
        "rejected_month_count": rejected_month_count,
        "pending_mark_requests_count": pending_mark_requests_count,
        "pending_vacation_requests_count": pending_vacation_requests_count,
        "approved_vacation_month_count": approved_vacation_month_count,
        "active_zone_count": active_zone_count,
        "active_assignment_count": active_assignment_count,
        "employee_count": employee_count,
        "today": today,
        "pending_mark_requests": pending_mark_requests,
        "pending_vacation_requests": pending_vacation_requests,
        "today_marked_sessions": today_marked_sessions,
    })


@login_required
def turni_planner_home(request):
    denied = _turni_planner_allowed_or_403(request)
    if denied:
        return denied

    selected_state = None
    selected_week_label = (request.GET.get('week') or request.POST.get('week_label') or '').strip()

    if request.method == 'POST':
        action = (request.POST.get('action') or 'open_week').strip()
        week_label = (request.POST.get('week_label') or '').strip()
        if week_label:
            if action == 'delete_week':
                TurniPlannerWeekState.objects.filter(week_label=week_label).delete()
                return redirect(reverse('turni_planner_home'))
            selected_state, created = TurniPlannerWeekState.objects.get_or_create(
                week_label=week_label,
                defaults={
                    'planner_data': _turni_planner_initial_data_for_new_week(week_label=week_label),
                    'updated_by': request.user,
                },
            )
            if action == 'open_week' and not _turni_planner_data_has_content(selected_state.planner_data or {}):
                planner_data = _turni_planner_initial_data_for_new_week(week_label=week_label)
                if planner_data:
                    selected_state.planner_data = planner_data
                    selected_state.updated_by = request.user
                    selected_state.save(update_fields=['planner_data', 'updated_by', 'updated_at'])
            if action == 'generate_weekend_email':
                try:
                    recipients = _turni_planner_weekend_mail_response(
                        selected_state,
                        recipient_text=(request.POST.get('mail_recipients') or '').strip(),
                        subject_text=(request.POST.get('mail_subject') or '').strip(),
                        body_text=request.POST.get('mail_body') or '',
                        selected_attachment_keys=request.POST.getlist('mail_attachment') or None,
                        selected_file_types=request.POST.getlist('mail_file_type') or None,
                    )
                except ValueError as error:
                    return redirect(
                        f"{reverse('turni_planner_home')}?week={selected_state.week_label}&mail_status=error&mail_message={quote(str(error))}"
                    )
                except Exception:
                    logger.exception('Errore invio mail planner weekend week=%s', selected_state.week_label)
                    return redirect(
                        f"{reverse('turni_planner_home')}?week={selected_state.week_label}&mail_status=error&mail_message={quote('Invio email non riuscito. Controlla configurazione SMTP e destinatari.')}"
                    )
                return redirect(
                    f"{reverse('turni_planner_home')}?week={selected_state.week_label}&mail_status=success&mail_message={quote('Email inviata a: ' + ', '.join(recipients))}"
                )
            if action in ('save_weekly', 'save_planner') or action.startswith('export_'):
                visible_to_employees = request.POST.get('visible_to_employees') == 'on'
                planner_data = dict(selected_state.planner_data or {})
                planner_data[TURNI_PUBLISHED_SECTIONS_KEY] = _normalize_turni_published_sections(
                    request.POST.getlist('published_sections'),
                    include_portineria=True,
                )
                planner_data['weekly_export_week_label'] = (request.POST.get('weekly_export_week_label') or '').strip() or selected_state.week_label
                planner_data['portineria_weekly_export_week_label'] = (request.POST.get('portineria_weekly_export_week_label') or '').strip() or selected_state.week_label
                planner_data['weekly'] = _extract_turni_weekly_data_from_post(request.POST)
                if action == 'save_planner' or action.startswith('export_'):
                    planner_data['saturday'] = _extract_turni_weekend_data_from_post(request.POST, 'saturday')
                    planner_data['sunday'] = _extract_turni_weekend_data_from_post(request.POST, 'sunday')
                    planner_data['jolly_weekend'] = _extract_turni_weekend_data_from_post(request.POST, 'jolly_weekend')
                    planner_data['scorrimento'] = _extract_turni_scorrimento_data_from_post(request.POST)
                    planner_data['portineria_weekly'] = _extract_turni_portineria_weekly_data_from_post(request.POST)
                    planner_data['portineria_weekend'] = _extract_turni_weekend_data_from_post(
                        request.POST,
                        'portineria_weekend',
                        default_row_count=TURNI_PORTINERIA_WEEKEND_DEFAULT_ROW_COUNT,
                        maximum=TURNI_WEEKEND_MAX_SINGLE_PAGE_ROW_COUNT,
                    )
                with transaction.atomic():
                    TurniPlannerWeekState.objects.exclude(pk=selected_state.pk).filter(visible_to_employees=True).update(visible_to_employees=False)
                    selected_state.planner_data = planner_data
                    selected_state.visible_to_employees = visible_to_employees
                    selected_state.updated_by = request.user
                    selected_state.save(update_fields=['planner_data', 'visible_to_employees', 'updated_by', 'updated_at'])
                if action.startswith('export_pdf_'):
                    return _turni_planner_export_response(
                        selected_state,
                        export_format='pdf',
                        export_target=action.removeprefix('export_pdf_'),
                    )
                if action.startswith('export_jpg_'):
                    return _turni_planner_export_response(
                        selected_state,
                        export_format='jpg',
                        export_target=action.removeprefix('export_jpg_'),
                    )
                if action == 'export_all_pdf':
                    return _turni_planner_bulk_export_response(selected_state, export_format='pdf')
                if action == 'export_all_jpg':
                    return _turni_planner_bulk_export_response(selected_state, export_format='jpg')
                if action == 'export_all_pdf_jpg':
                    return _turni_planner_bulk_export_response(selected_state, export_format='all')
                return redirect(f"{reverse('turni_planner_home')}?week={selected_state.week_label}")
            if not created and selected_state.updated_by_id != request.user.id:
                selected_state.updated_by = request.user
                selected_state.save(update_fields=['updated_by', 'updated_at'])
            return redirect(f"{reverse('turni_planner_home')}?week={selected_state.week_label}")

    if selected_week_label:
        selected_state = TurniPlannerWeekState.objects.filter(week_label=selected_week_label).first()

    mail_status = (request.GET.get('mail_status') or '').strip()
    mail_message = (request.GET.get('mail_message') or '').strip()

    recent_weeks = TurniPlannerWeekState.objects.all()[:12]
    planner_data = selected_state.planner_data or {} if selected_state else {}
    weekly_data = _merge_turni_weekly_data(planner_data.get('weekly'))
    saturday_data = _merge_turni_weekend_data(planner_data.get('saturday'))
    sunday_data = _merge_turni_weekend_data(planner_data.get('sunday'))
    jolly_weekend_data = _merge_turni_weekend_data(planner_data.get('jolly_weekend'))
    scorrimento_data = _merge_turni_scorrimento_data(planner_data.get('scorrimento'))
    portineria_weekly_data = _merge_turni_portineria_weekly_data(planner_data.get('portineria_weekly'))
    portineria_weekend_data = _merge_turni_weekend_data(
        planner_data.get('portineria_weekend'),
        default_row_count=TURNI_PORTINERIA_WEEKEND_DEFAULT_ROW_COUNT,
        maximum=TURNI_WEEKEND_MAX_SINGLE_PAGE_ROW_COUNT,
    )

    response = render(request, 'portal/turni_planner.html', {
        'recent_weeks': recent_weeks,
        'selected_state': selected_state,
        'selected_week_label': selected_week_label,
        'visible_to_employees': selected_state.visible_to_employees if selected_state else False,
        'published_turni_section_choices': list(TURNI_MARKINGS_SECTION_META.items()),
        'selected_published_turni_sections': _turni_planner_selected_section_keys(selected_state, include_portineria=True),
        'mail_attachment_options': TURNI_PLANNER_MAIL_ATTACHMENT_OPTIONS,
        'mail_status': mail_status,
        'mail_message': mail_message,
        'weekly_export_week_label': _resolve_turni_export_week_label(planner_data, selected_state.week_label, key='weekly_export_week_label') if selected_state else '',
        'portineria_weekly_export_week_label': _resolve_turni_export_week_label(planner_data, selected_state.week_label, key='portineria_weekly_export_week_label') if selected_state else '',
        'weekly_data': weekly_data,
        'saturday_data': saturday_data,
        'sunday_data': sunday_data,
        'jolly_weekend_data': jolly_weekend_data,
        'scorrimento_data': scorrimento_data,
        'scorrimento_blocks': _turni_scorrimento_template_blocks(scorrimento_data),
        'scorrimento_rows': _turni_scorrimento_template_rows(scorrimento_data),
        'scorrimento_department_blocks': _turni_scorrimento_template_department_blocks(scorrimento_data),
        'scorrimento_stanzette_names': scorrimento_data['department_names'][3] if len(scorrimento_data['department_names']) > 3 else [],
        'portineria_weekly_data': portineria_weekly_data,
        'portineria_weekend_data': portineria_weekend_data,
        'weekend_column_labels': TURNI_WEEKEND_COLUMN_LABELS,
        'weekend_single_page_max_rows': TURNI_WEEKEND_MAX_SINGLE_PAGE_ROW_COUNT,
    })
    return _disable_response_cache(response)


def _today_marked_sessions_queryset(target_date):
    return (
        WorkSession.objects
        .select_related('employee')
        .filter(
            Q(started_at__date=target_date)
            | Q(ended_at__date=target_date)
            | Q(corrected_started_at__date=target_date)
            | Q(corrected_ended_at__date=target_date)
        )
        .order_by('employee__last_name', 'employee__first_name')
    )


@login_required
def riconfezionamento_entry(request):
    denied = _riconfezionamento_allowed_or_403(request)
    if denied is not None:
        return denied
    return redirect('/riconfezionamento/')


@login_required
def today_markings_dashboard(request):
    if not user_has_today_markings_access(request.user) and not user_has_full_admin_access(request.user):
        return redirect('dashboard')

    today = timezone.localdate()
    selected_date_raw = (request.GET.get('date') or '').strip()
    try:
        selected_date = datetime.strptime(selected_date_raw, '%Y-%m-%d').date() if selected_date_raw else today
    except ValueError:
        selected_date = today

    if selected_date > today:
        selected_date = today

    _sync_approved_requests_for_range(today, today)
    today_marked_sessions = _prepare_marked_sessions_for_date(list(_today_marked_sessions_queryset(today)), today)

    selected_marked_sessions = None
    if selected_date != today:
        _sync_approved_requests_for_range(selected_date, selected_date)
        selected_marked_sessions = _prepare_marked_sessions_for_date(list(_today_marked_sessions_queryset(selected_date)), selected_date)

    previous_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1) if selected_date < today else None
    published_turni_state = _turni_planner_published_state()
    published_turni_sections = []
    if _user_can_view_published_turni(request.user):
        published_turni_sections = _turni_planner_employee_sections(published_turni_state, include_portineria=True)

    response = render(request, 'portal/today_markings_dashboard.html', {
        'today': today,
        'selected_date': selected_date,
        'today_marked_sessions': today_marked_sessions,
        'selected_marked_sessions': selected_marked_sessions,
        'show_mark_coordinates': user_has_full_admin_access(request.user),
        'previous_date': previous_date,
        'next_date': next_date,
        'published_turni_state': published_turni_state,
        'published_turni_sections': published_turni_sections,
    })
    return _disable_response_cache(response)


@login_required
def admin_upload_cud(request):
    """Upload CUD multiplo basato sul filename, con fallback al caricamento manuale."""
    if not request.user.is_staff:
        return redirect('dashboard')

    from django import forms

    class CudUploadForm(forms.Form):
        employee = forms.ModelChoiceField(queryset=Employee.objects.order_by('last_name', 'first_name'))
        year = forms.IntegerField(min_value=2000, max_value=2100)
        pdf = forms.FileField()

    if request.method == 'POST':
        if request.POST.get('action') == 'resolve_pending_import':
            session_key = _pending_import_session_key('cud')
            pending_data = request.session.pop(session_key, None)
            if not pending_data:
                return redirect('admin_upload_cud')
            selected_keys = set(request.POST.getlist('create_candidates'))
            return _finalize_pending_import(request, 'cud', pending_data, selected_keys, request.POST)

        if request.FILES and not {'employee', 'year'}.issubset(request.POST.keys()):
            files = request.FILES.getlist('files') or request.FILES.getlist('folder')
            if not files and request.FILES.get('pdf'):
                files = [request.FILES['pdf']]

            _clear_pending_import(request, 'cud')
            employee_lookup = _build_employee_import_lookup()
            records = []

            for f in files:
                year, name_tokens, error_reason = _parse_cud_import_filename(f.name)
                if error_reason:
                    records.append(_build_import_record('cud', f, reason=error_reason))
                    continue

                employee = _find_employee_for_import_tokens(name_tokens, employee_lookup)
                records.append(_build_import_record('cud', f, name_tokens=name_tokens, employee=employee, year=year))

            pending_data = _build_pending_import_data(records, 'cud')
            if pending_data['missing_candidates']:
                request.session[_pending_import_session_key('cud')] = pending_data
                return _render_missing_account_resolution(request, 'cud', pending_data)

            return _finalize_pending_import(request, 'cud', pending_data, set())

        form = CudUploadForm(request.POST, request.FILES)
        success = False
        replaced = False
        cud_obj = None

        if form.is_valid():
            employee = form.cleaned_data['employee']
            year = form.cleaned_data['year']
            pdf_file = form.cleaned_data['pdf']

            with transaction.atomic():
                existing = Cud.objects.filter(employee=employee, year=year).first()
                if existing:
                    try:
                        existing.pdf.delete(save=False)
                    except Exception:
                        logger.exception('Error deleting existing CUD file for %s', existing.id)
                    existing.delete()
                    replaced = True

                cud_obj = Cud(employee=employee, year=year)
                cud_obj.pdf.save(pdf_file.name, pdf_file, save=True)

            _create_audit_event(
                request,
                "cud_uploaded",
                employee=employee,
                metadata={"year": year, "replaced": replaced},
            )

            success = True
    else:
        from django.utils import timezone
        form = CudUploadForm(initial={"year": timezone.now().year})

    return render(request, "portal/admin_upload_cud.html", {
        "form": form,
        "success": locals().get('success', False),
        "replaced": locals().get('replaced', False),
        "cud": locals().get('cud_obj'),
    })


@login_required
def admin_report(request):
    """Report sintetico visualizzazioni cedolini per dipendente."""
    if not request.user.is_staff:
        return redirect('dashboard')

    employees_stats = (
        Employee.objects
        .annotate(
            totale=Count('payslips', distinct=True),
            visualizzati=Count('payslips', filter=Q(payslips__payslipview__isnull=False), distinct=True),
        )
        .order_by('last_name', 'first_name')
    )

    rows = []
    totale_cedolini = 0
    totale_visualizzati = 0

    for emp in employees_stats:
        totale = emp.totale
        visualizzati = emp.visualizzati
        non_visualizzati = max(totale - visualizzati, 0)

        totale_cedolini += totale
        totale_visualizzati += visualizzati

        rows.append({
            'employee': emp,
            'total': totale,
            'visualizzati': visualizzati,
            'non_visualizzati': non_visualizzati,
        })

    totale_non_visualizzati = max(totale_cedolini - totale_visualizzati, 0)

    return render(request, "portal/admin_report.html", {
        "rows": rows,
        "totale_cedolini": totale_cedolini,
        "totale_visualizzati": totale_visualizzati,
        "totale_non_visualizzati": totale_non_visualizzati,
    })


# =========================================================
# ADMIN: CONTROLLO INTEGRITÀ CEDOLINI
# =========================================================

@login_required
def admin_payslip_integrity(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    year = request.GET.get('year')
    employee_id_raw = request.GET.get('employee')
    run_check = request.GET.get('run') == '1'

    try:
        employee_selected = int(employee_id_raw) if employee_id_raw else None
    except (TypeError, ValueError):
        employee_selected = None

    missing = []
    checked = 0
    missing_count = 0
    missing_ratio = 0
    integrity_status = "ok"
    integrity_label = "OK"
    integrity_reason = "Seleziona i filtri e avvia il controllo per generare il report."

    if run_check:
        from django.core.files.storage import default_storage

        payslips_qs = Payslip.objects.all().select_related('employee').order_by('-year', '-month')

        if year:
            payslips_qs = payslips_qs.filter(year=year)
        if employee_selected:
            payslips_qs = payslips_qs.filter(employee_id=employee_selected)

        # Limite di sicurezza per non bombardare lo storage
        payslips_qs = payslips_qs[:2000]

        for payslip in payslips_qs:
            checked += 1
            payslip.month_name = MONTH_LABELS_IT.get(payslip.month, str(payslip.month))
            try:
                exists = default_storage.exists(payslip.pdf.name)
            except Exception:
                logger.exception("Errore nel controllo esistenza file per payslip id=%s", payslip.id)
                exists = False

            if not exists:
                missing.append(payslip)

        missing_count = len(missing)
        missing_ratio = (missing_count / checked) if checked else 0

        if missing_count == 0:
            integrity_status = "ok"
            integrity_label = "OK"
            integrity_reason = "Nessun PDF mancante rilevato."
        elif missing_ratio <= 0.05:
            integrity_status = "warning"
            integrity_label = "Attenzione"
            integrity_reason = "Anomalie limitate: presenza di pochi PDF mancanti."
        else:
            integrity_status = "critical"
            integrity_label = "Critica"
            integrity_reason = "Anomalie elevate: numero significativo di PDF mancanti."

    employees = Employee.objects.order_by('last_name', 'first_name')

    return render(request, "portal/admin_payslip_integrity.html", {
        "employees": employees,
        "year_selected": year,
        "employee_selected": employee_selected,
        "run_check": run_check,
        "checked": checked,
        "missing": missing,
        "missing_count": missing_count,
        "missing_percentage": round(missing_ratio * 100, 1),
        "integrity_status": integrity_status,
        "integrity_label": integrity_label,
        "integrity_reason": integrity_reason,
    })


@login_required
def admin_all_payslips(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    year = (request.GET.get('year') or '').strip()
    month = (request.GET.get('month') or '').strip()
    employee_id = (request.GET.get('employee') or '').strip()

    def _apply_payslip_filters(queryset, selected_year, selected_month, selected_employee_id):
        if selected_year:
            queryset = queryset.filter(year=selected_year)
        if selected_month:
            queryset = queryset.filter(month=selected_month)
        if selected_employee_id:
            queryset = queryset.filter(employee_id=selected_employee_id)
        return queryset

    feedback = ''
    feedback_level = 'info'

    if request.method == 'POST' and request.POST.get('action') == 'bulk_delete_payslips':
        year = (request.POST.get('year') or '').strip()
        month = (request.POST.get('month') or '').strip()
        employee_id = (request.POST.get('employee') or '').strip()

        if not year or not month:
            feedback = 'Per la cancellazione massiva devi indicare almeno anno e mese.'
            feedback_level = 'danger'
        else:
            payslips_to_delete = _apply_payslip_filters(
                Payslip.objects.select_related('employee__user').order_by('-year', '-month'),
                year,
                month,
                employee_id,
            )
            deleted_count = payslips_to_delete.count()

            if deleted_count == 0:
                feedback = 'Nessun cedolino trovato con i filtri selezionati.'
                feedback_level = 'warning'
            else:
                payload = {
                    'year': year,
                    'month': month,
                    'employee_id': employee_id or None,
                    'deleted_count': deleted_count,
                    'payslip_ids': list(payslips_to_delete.values_list('id', flat=True)[:500]),
                }
                employee = Employee.objects.filter(id=employee_id).first() if employee_id else None
                payslips_to_delete.delete()
                _create_audit_event(
                    request,
                    'payslip_bulk_deleted',
                    employee=employee,
                    metadata=payload,
                )
                feedback = f'{deleted_count} cedolini eliminati definitivamente.'
                feedback_level = 'warning'

    payslips = _apply_payslip_filters(
        Payslip.objects.select_related('employee__user').order_by('-year', '-month'),
        year,
        month,
        employee_id,
    )

    # CUD: usiamo gli stessi filtri per anno e dipendente (mese non rilevante)
    cuds = Cud.objects.select_related('employee__user').order_by('-year')
    if year:
        cuds = cuds.filter(year=year)
    if employee_id:
        cuds = cuds.filter(employee_id=employee_id)

    # arricchisci con info di visualizzazione
    cuds_list = []
    for c in cuds:
        view = CudView.objects.filter(cud=c).order_by('viewed_at').first()
        c.is_viewed = bool(view)
        c.viewed_at = view.viewed_at if view else None
        cuds_list.append(c)

    # Export CSV opzionale
    export_format = request.GET.get('format')
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="cedolini.csv"'

        writer = csv.writer(response)
        writer.writerow(['Dipendente', 'Codice esterno', 'Anno', 'Mese', 'Caricato il', 'ID'])
        for p in payslips:
            writer.writerow([
                getattr(p.employee, 'full_name', str(p.employee)),
                getattr(p.employee, 'external_code', ''),
                p.year,
                p.month,
                p.uploaded_at,
                p.id,
            ])
        return response

    employees = Employee.objects.order_by('last_name', 'first_name')

    return render(request, 'portal/admin_all_payslips.html', {
        'payslips': payslips,
        'cuds': cuds_list,
        'employees': employees,
        'feedback': feedback,
        'feedback_level': feedback_level,
        'year_selected': year,
        'month_selected': month,
        'employee_selected': int(employee_id) if employee_id else None,
    })


@login_required
def admin_audit_events(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    qs = AuditEvent.objects.select_related('actor_user', 'employee', 'payslip').order_by('-created_at')

    # simple filtering by action or employee id
    action = request.GET.get('action')
    emp = request.GET.get('employee')
    if action:
        qs = qs.filter(action__icontains=action)
    if emp:
        qs = qs.filter(employee__id=emp)

    paginator = Paginator(qs, 50)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    return render(request, 'portal/admin_audit_events.html', {
        'page_obj': page_obj,
        'action': action or '',
        'employee_filter': emp or '',
    })


@login_required
def admin_import_jobs(request):
    """Storico dei caricamenti cedolini (ImportJob)."""
    if not request.user.is_staff:
        return redirect('dashboard')

    jobs = ImportJob.objects.order_by('-created_at')

    paginator = Paginator(jobs, 50)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    return render(request, 'portal/admin_import_jobs.html', {
        'page_obj': page_obj,
    })


@login_required
def admin_import_job_payslips(request, job_id):
    """Mostra i cedolini caricati nello stesso intervallo temporale di un ImportJob.

    Non avendo un collegamento diretto ImportJob->Payslip, usiamo una finestra
    temporale intorno a created_at, sufficiente per i caricamenti batch normali.
    """
    if not request.user.is_staff:
        return redirect('dashboard')

    job = get_object_or_404(ImportJob, id=job_id)

    # finestra di 10 minuti intorno al job
    window = timedelta(minutes=10)
    start = job.created_at - window
    end = job.created_at + window

    payslips = (
        Payslip.objects
        .select_related('employee__user')
        .filter(uploaded_at__range=(start, end))
        .order_by('-uploaded_at')
    )

    return render(request, 'portal/admin_import_job_payslips.html', {
        'job': job,
        'payslips': payslips,
        'start': start,
        'end': end,
    })


@login_required
def admin_employees(request):
    if not request.user.is_staff:
        return redirect('dashboard')
    employees = list(
        Employee.objects
        .select_related('user')
        .annotate(payslip_count=Count('payslips'))
    )
    riconfezionamento_user_ids = set(
        Group.objects.filter(name=RICONFEZIONAMENTO_GROUP_NAME).values_list('user__id', flat=True)
    )
    patrimonio_user_ids = set(
        Group.objects.filter(name=PATRIMONIO_GROUP_NAME).values_list('user__id', flat=True)
    )
    for employee in employees:
        employee.has_riconfezionamento_access = employee.user_id in riconfezionamento_user_ids
        employee.has_patrimonio_access = employee.user_id in patrimonio_user_ids

    _decorate_employee_display_names(employees)
    employees.sort(key=_employee_name_sort_key)

    registered_employees = [employee for employee in employees if employee.privacy_accepted]
    invited_employees = [employee for employee in employees if not employee.privacy_accepted and employee.invito_inviato]
    not_invited_employees = [employee for employee in employees if not employee.privacy_accepted and not employee.invito_inviato]

    employee_sections = [
        {
            'id': 'registeredEmployees',
            'title': 'Utenti registrati',
            'subtitle': 'Dipendenti che hanno completato la registrazione e accettato la privacy.',
            'employees': registered_employees,
            'count': len(registered_employees),
            'expanded': False,
            'badge_class': 'bg-primary',
            'empty_message': 'Nessun dipendente registrato al momento.',
        },
        {
            'id': 'invitedEmployees',
            'title': 'Invitati',
            'subtitle': 'Dipendenti che hanno ricevuto l\'invito ma non hanno ancora completato la registrazione.',
            'employees': invited_employees,
            'count': len(invited_employees),
            'expanded': False,
            'badge_class': 'bg-success',
            'empty_message': 'Nessun dipendente nello stato invitato.',
        },
        {
            'id': 'notInvitedEmployees',
            'title': 'Non invitati',
            'subtitle': 'Dipendenti da contattare o da invitare per l\'accesso al portale.',
            'employees': not_invited_employees,
            'count': len(not_invited_employees),
            'expanded': False,
            'badge_class': 'bg-warning text-dark',
            'empty_message': 'Nessun dipendente da invitare.',
        },
    ]

    return render(request, 'portal/admin_employees.html', {
        'employees': employees,
        'employee_sections': employee_sections,
        'employee_total_count': len(employees),
    })


@login_required
def admin_employee_detail(request, emp_id):
    if not request.user.is_staff:
        return redirect('dashboard')

    employee = get_object_or_404(Employee, id=emp_id)

    feedback = None
    feedback_level = 'success'
    riconfezionamento_group, _ = Group.objects.get_or_create(name=RICONFEZIONAMENTO_GROUP_NAME)
    patrimonio_group, _ = Group.objects.get_or_create(name=PATRIMONIO_GROUP_NAME)
    has_riconfezionamento_access = employee.user.groups.filter(id=riconfezionamento_group.id).exists()
    has_patrimonio_access = employee.user.groups.filter(id=patrimonio_group.id).exists()

    if request.method == 'POST' and request.POST.get('action') == 'toggle_riconfezionamento_access':
        enable_access = request.POST.get('enable_access') == '1'
        if enable_access:
            employee.user.groups.add(riconfezionamento_group)
            outcome = 'riconfezionamento_enabled'
        else:
            employee.user.groups.remove(riconfezionamento_group)
            outcome = 'riconfezionamento_disabled'

        _create_audit_event(
            request,
            'employee_riconfezionamento_access_updated',
            employee=employee,
            metadata={
                'enabled': enable_access,
                'username': employee.user.username,
            },
        )

        return redirect(f'{request.path}?outcome={outcome}')

    if request.method == 'POST' and request.POST.get('action') == 'toggle_patrimonio_access':
        enable_access = request.POST.get('enable_access') == '1'
        if enable_access:
            employee.user.groups.add(patrimonio_group)
            outcome = 'patrimonio_enabled'
        else:
            employee.user.groups.remove(patrimonio_group)
            outcome = 'patrimonio_disabled'

        _create_audit_event(
            request,
            'employee_patrimonio_access_updated',
            employee=employee,
            metadata={
                'enabled': enable_access,
                'username': employee.user.username,
            },
        )

        return redirect(f'{request.path}?outcome={outcome}')

    if request.method == 'POST' and request.POST.get('action') == 'delete_payslip':
        payslip_id = request.POST.get('payslip_id')
        payslip = Payslip.objects.filter(id=payslip_id, employee=employee).first()

        if not payslip:
            return redirect(f'{request.path}?outcome=missing')

        payload = {
            'payslip_id': payslip.id,
            'year': payslip.year,
            'month': payslip.month,
            'pdf': payslip.pdf.name,
        }

        try:
            if payslip.pdf:
                payslip.pdf.delete(save=False)
        except Exception:
            logger.exception('Errore eliminazione file cedolino id=%s', payslip.id)

        payslip.delete()

        _create_audit_event(
            request,
            'payslip_deleted',
            employee=employee,
            metadata=payload,
        )

        return redirect(f'{request.path}?outcome=deleted')

    outcome = (request.GET.get('outcome') or '').strip()
    if outcome == 'deleted':
        feedback = 'Cedolino eliminato definitivamente.'
        feedback_level = 'warning'
    elif outcome == 'missing':
        feedback = 'Cedolino non trovato o gia eliminato.'
        feedback_level = 'danger'
    elif outcome == 'riconfezionamento_enabled':
        feedback = 'Accesso al riconfezionamento abilitato per questo dipendente.'
        feedback_level = 'success'
    elif outcome == 'riconfezionamento_disabled':
        feedback = 'Accesso al riconfezionamento disattivato per questo dipendente.'
        feedback_level = 'warning'
    elif outcome == 'patrimonio_enabled':
        feedback = 'Accesso alla gestione patrimonio abilitato per questo dipendente.'
        feedback_level = 'success'
    elif outcome == 'patrimonio_disabled':
        feedback = 'Accesso alla gestione patrimonio disattivato per questo dipendente.'
        feedback_level = 'warning'

    payslips = Payslip.objects.filter(employee=employee).order_by('-year', '-month')

    logger.info('admin_employee_detail: employee=%s payslip_count=%d', getattr(employee, 'id', None), payslips.count())
    try:
        ids = list(payslips.values_list('id', flat=True))
        logger.info('admin_employee_detail: payslip_ids=%s', ids)
    except Exception:
        logger.exception('admin_employee_detail: error listing payslip ids for employee=%s', getattr(employee, 'id', None))

    detailed = []
    for p in payslips:
        view = PayslipView.objects.filter(payslip=p).first()
        detailed.append({'payslip': p, 'view': view})

    # CUD del dipendente con info di visualizzazione
    cuds = Cud.objects.filter(employee=employee).order_by('-year')
    cuds_detailed = []
    for c in cuds:
        view = CudView.objects.filter(cud=c).first()
        cuds_detailed.append({'cud': c, 'view': view})

    return render(request, 'portal/admin_employee_detail.html', {
        'employee': employee,
        'detailed': detailed,
        'cuds_detailed': cuds_detailed,
        'feedback': feedback,
        'feedback_level': feedback_level,
        'has_riconfezionamento_access': employee.user.groups.filter(id=riconfezionamento_group.id).exists(),
        'has_patrimonio_access': employee.user.groups.filter(id=patrimonio_group.id).exists(),
    })


@login_required
def admin_reset_payslip_view(request, payslip_id):
    """Consente all'admin di azzerare lo stato di visualizzazione di un cedolino."""
    if not request.user.is_staff:
        return HttpResponse(status=403)

    payslip = get_object_or_404(Payslip, id=payslip_id)

    if request.method == 'POST':
        PayslipView.objects.filter(payslip=payslip).delete()

        _create_audit_event(
            request,
            "payslip_view_reset",
            employee=payslip.employee,
            payslip=payslip,
        )

    return redirect('admin_employee_detail', emp_id=payslip.employee_id)


@login_required
def admin_employee_payslips(request, emp_id):
    if not request.user.is_staff:
        return HttpResponse(status=403)

    employee = get_object_or_404(Employee, id=emp_id)
    payslips = Payslip.objects.filter(employee=employee).order_by('-year', '-month')

    return render(request, 'portal/_employee_payslips.html', {
        'payslips': payslips,
        'employee': employee,
    })


@login_required
def portal_tutorial(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')

    employee = Employee.objects.filter(user=request.user).first()
    if not employee:
        return HttpResponse("Profilo dipendente non trovato. Contatta l'amministratore.", status=403)

    return render(request, 'portal/tutorial.html', {
        'employee': employee,
    })


@login_required
def admin_send_invite(request):
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'forbidden'}, status=403)

    if request.method != 'POST':
            action = request.POST.get('action')

    emp_id = request.POST.get('employee_id')
    email = request.POST.get('email')

    logger.info('admin_send_invite START user=%s emp_id=%s email_present=%s', getattr(request.user, 'username', None), emp_id, bool(email))

    if not emp_id:
        return JsonResponse({'ok': False, 'error': 'missing employee_id'}, status=400)

    try:
        employee = Employee.objects.select_related('user').get(id=emp_id)
    except Employee.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'not_found'}, status=404)

    logger.info('admin_send_invite: target employee=%s username=%s current_email=%s', getattr(employee, 'id', None), getattr(employee.user, 'username', None), employee.email_invio)

    # if email provided, set it
    if email:
        employee.email_invio = email
        employee.save()

        logger.info('admin_send_invite: updated email for employee=%s -> %s', employee.id, email)

    if not employee.email_invio:
        return JsonResponse({'ok': False, 'need_email': True})

    # create token and send email (reuse admin email template)
    from django.utils.crypto import get_random_string
    from django.utils import timezone
    from datetime import timedelta

    try:
        token = get_random_string(64)
        InviteToken.objects.create(employee=employee, token=token, expires_at=timezone.now() + timedelta(days=7))
    except Exception:
        logger.exception('admin_send_invite: error creating InviteToken for employee=%s', employee.id)
        return JsonResponse({'ok': False, 'error': 'token_error'})

    link = f"https://cedolini-web.onrender.com/portal/register/{token}/"
    username = employee.user.username

    subject = "Attivazione account - Portale Cedolini"
    text_content = (
        f"Gentile {employee.first_name or ''} {employee.last_name or ''},\n\n"
        f"è stato creato il tuo accesso al Portale Cedolini.\n\n"
        f"USERNAME: {username}\n\n"
        f"EMAIL: {employee.email_invio}\n\n"
        f"Clicca sul link seguente per attivare il tuo account e creare la password:\n{link}\n\n"
        f"Il link è valido per 7 giorni.\n\n"
        f"Dopo l'attivazione potrai accedere usando il tuo username oppure la tua email.\n\n"
        f"Per accedere al portale utilizza sempre questo indirizzo:\n"
        f"https://cedolini-web.onrender.com/login/\n"
    )

    html_content = f"""
<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background-color:#f3f4f6;font-family:Arial, sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="padding:40px 0;">
<tr>
<td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:8px;padding:40px;">
    <tr>
        <td align="center" style="padding-bottom:20px;">
            <img src="https://cedolini-web.onrender.com/static/portal/logo.png" width="120">
        </td>
    </tr>
    <tr>
        <td style="font-size:24px;font-weight:bold;color:#1f2937;padding-bottom:8px;">Attivazione Portale Cedolini</td>
    </tr>
    <tr>
        <td style="font-size:13px;color:#6b7280;padding-bottom:24px;">
            Accesso sicuro ai tuoi cedolini online
        </td>
    </tr>
    <tr>
        <td style="font-size:14px;color:#374151;padding-bottom:20px;line-height:1.6;">
            Gentile <strong>{employee.first_name or ''} {employee.last_name or ''}</strong>,<br><br>
            è stato creato il tuo accesso al <strong>Portale Cedolini</strong>.<br><br>
            Il tuo <strong style="color:#111827;">USERNAME</strong> per accedere è:<br>
            <span style="display:inline-block;margin-top:10px;padding:10px 14px;background:#eff6ff;border-radius:6px;font-family:monospace;font-size:15px;color:#111827;border:1px solid #bfdbfe;">
                {username}
            </span>
            <br><br>
            La tua <strong style="color:#111827;">EMAIL</strong> associata è:<br>
            <span style="display:inline-block;margin-top:10px;padding:10px 14px;background:#eff6ff;border-radius:6px;font-family:monospace;font-size:15px;color:#111827;border:1px solid #bfdbfe;">
                {employee.email_invio}
            </span>
        </td>
    </tr>
    <tr>
        <td align="center" style="padding:30px 0;">
            <a href="{link}" style="background:#2563eb;color:#ffffff;padding:14px 28px;text-decoration:none;border-radius:6px;font-weight:bold;">Attiva il tuo account</a>
        </td>
    </tr>
    <tr>
        <td style="font-size:12px;color:#6b7280;padding-top:10px;">
            Il link è valido per 7 giorni.
        </td>
    </tr>
    <tr>
        <td style="font-size:13px;color:#374151;padding-top:10px;line-height:1.6;">
            Dopo l'attivazione potrai accedere con <strong>USERNAME</strong> oppure con la tua <strong>EMAIL</strong>.
        </td>
    </tr>
    <tr>
        <td style="font-size:13px;color:#374151;padding-top:10px;">
            Per accedere al portale utilizza sempre questo indirizzo:<br>
            <a href="https://cedolini-web.onrender.com/login/" style="display:inline-block;margin-top:8px;padding:8px 12px;background:#0f172a;color:#ffffff;text-decoration:none;border-radius:4px;font-size:13px;">
                https://cedolini-web.onrender.com/login/
            </a>
        </td>
    </tr>
    <tr>
        <td style="font-size:12px;color:#9ca3af;padding-top:20px;">© San Vincenzo Srl</td>
    </tr>
</table>
</td>
</tr>
</table>
</body>
</html>
"""

    from django.core.mail import EmailMultiAlternatives
    try:
        email_msg = EmailMultiAlternatives(subject, text_content, settings.DEFAULT_FROM_EMAIL, [employee.email_invio], cc=["cedolini@sanvincenzosrl.com"])
        email_msg.attach_alternative(html_content, "text/html")
        email_msg.send()

        employee.invito_inviato = True
        employee.save()

        # Audit: invito inviato a un dipendente
        _create_audit_event(
            request,
            "invite_sent",
            employee=employee,
            metadata={
                "email": employee.email_invio,
            },
        )

        logger.info('admin_send_invite: email sent to %s for employee=%s', employee.email_invio, employee.id)
        return JsonResponse({'ok': True})
    except Exception:
        logger.exception('admin_send_invite: failed sending email to %s for employee=%s', employee.email_invio, employee.id)
        return JsonResponse({'ok': False, 'error': 'send_failed'})


@login_required
def admin_create_invite_link(request):
    """Crea (o riusa) un link di invito senza inviare email.

    Usato dall'area admin per copiare il link e inviarlo manualmente
    via WhatsApp / SMS.
    """
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'forbidden'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'method'}, status=405)

    emp_id = request.POST.get('employee_id')
    if not emp_id:
        return JsonResponse({'ok': False, 'error': 'missing employee_id'}, status=400)

    try:
        employee = Employee.objects.select_related('user').get(id=emp_id)
    except Employee.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'not_found'}, status=404)

    # Riusa un token valido se esiste, altrimenti creane uno nuovo
    invite = (
        InviteToken.objects
        .filter(employee=employee, used=False, expires_at__gt=timezone.now())
        .order_by('-created_at')
        .first()
    )

    if not invite:
        invite = InviteToken.objects.create(employee=employee)

    link = f"https://cedolini-web.onrender.com/portal/register/{invite.token}/"

    # Segna che un invito è stato generato
    employee.invito_inviato = True
    employee.save(update_fields=['invito_inviato'])

    # Testo precompilato in stile email di invito, da incollare in WhatsApp/SMS
    username = employee.user.username
    full_name = employee.full_name
    message = (
        f"Gentile {full_name},\n\n"
        f"è stato creato il tuo accesso al Portale Cedolini.\n\n"
        f"USERNAME: {username}\n\n"
        f"Per attivare il tuo account e creare la password utilizza questo link:\n"
        f"{link}\n\n"
        f"Il link è valido per 7 giorni.\n\n"
        f"Per i prossimi accessi al portale utilizza sempre questo indirizzo:\n"
        f"https://cedolini-web.onrender.com/login/\n\n"
        f"Cordiali saluti\n"
        f"San Vincenzo Srl"
    )

    # Audit: link invito creato per invio manuale
    _create_audit_event(
        request,
        "invite_link_created",
        employee=employee,
        metadata={
            "invite_token_id": invite.id,
            "via": "manual",
        },
    )

    return JsonResponse({
        'ok': True,
        'link': link,
        'username': employee.user.username,
        'full_name': employee.full_name,
        'message': message,
    })

def send_read_notification_email(payslip):

    employee = payslip.employee
    user = employee.user

    subject = "Cedolino visualizzato - Portale Cedolini"

    text_content = f"""
Il cedolino di {employee.first_name} {employee.last_name}
({payslip.month}/{payslip.year})
è stato visualizzato in data {timezone.now().strftime("%d/%m/%Y %H:%M")}.
"""

    html_content = f"""
<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background-color:#f3f4f6;font-family:Arial, sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="padding:40px 0;">
<tr>
<td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:8px;padding:40px;">
    
    <tr>
        <td align="center" style="padding-bottom:20px;">
            <img src="https://cedolini-web.onrender.com/static/portal/logo.png" width="120">
        </td>
    </tr>

    <tr>
        <td style="font-size:20px;font-weight:bold;color:#1f2937;padding-bottom:20px;">
            Cedolino Visualizzato
        </td>
    </tr>

    <tr>
        <td style="font-size:14px;color:#374151;padding-bottom:20px;">
            Il dipendente <strong>{employee.first_name} {employee.last_name}</strong><br>
            ha visualizzato il cedolino di <strong>{payslip.month}/{payslip.year}</strong>.
        </td>
    </tr>

    <tr>
        <td style="font-size:13px;color:#6b7280;">
            Data e ora: {timezone.now().strftime("%d/%m/%Y %H:%M")}
        </td>
    </tr>

</table>
</td>
</tr>
</table>
</body>
</html>
"""

    email = EmailMultiAlternatives(
        subject,
        text_content,
        settings.DEFAULT_FROM_EMAIL,
        ["cedolini@sanvincenzosrl.com"],
    )

    email.attach_alternative(html_content, "text/html")
    email.send()


# =========================================================
# ADMIN: Upload folder / multiple payslips import
# =========================================================

@login_required
def admin_upload_period_folder(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    months_map = {
        'gennaio': 1, 'febbraio': 2, 'marzo': 3, 'aprile': 4,
        'maggio': 5, 'giugno': 6, 'luglio': 7, 'agosto': 8,
        'settembre': 9, 'ottobre': 10, 'novembre': 11, 'dicembre': 12,
    }
    if request.method == 'POST' and request.POST.get('action') == 'resolve_pending_import':
        session_key = _pending_import_session_key('payslip')
        pending_data = request.session.pop(session_key, None)
        if not pending_data:
            return redirect('admin_upload_period_folder')
        selected_keys = set(request.POST.getlist('create_candidates'))
        return _finalize_pending_import(request, 'payslip', pending_data, selected_keys, request.POST)

    if request.method == 'POST' and request.FILES:
        files = request.FILES.getlist('files') or request.FILES.getlist('folder')
        employee_lookup = _build_employee_import_lookup()
        records = []

        _clear_pending_import(request, 'payslip')

        for f in files:
            name = os.path.splitext(f.name)[0].strip()
            parts = name.split()
            if len(parts) < 3:
                records.append(_build_import_record('payslip', f, reason='filename too short'))
                continue

            year_token = parts[-1]
            try:
                year = int(year_token)
            except Exception:
                records.append(_build_import_record('payslip', f, reason='invalid year'))
                continue

            month_token = parts[-2].lower()
            month = months_map.get(month_token)
            if not month:
                records.append(_build_import_record('payslip', f, reason=f'unknown month "{parts[-2]}"'))
                continue

            name_tokens = parts[:-2]
            employee = _find_employee_for_import_tokens(name_tokens, employee_lookup)
            records.append(_build_import_record('payslip', f, name_tokens=name_tokens, employee=employee, year=year, month=month))

        pending_data = _build_pending_import_data(records, 'payslip')
        if pending_data['missing_candidates']:
            request.session[_pending_import_session_key('payslip')] = pending_data
            return _render_missing_account_resolution(request, 'payslip', pending_data)

        return _finalize_pending_import(request, 'payslip', pending_data, set())

    recent_jobs = ImportJob.objects.order_by('-created_at')[:20]

    return render(request, 'portal/admin_upload_period_folder.html', {
        'import_jobs': recent_jobs,
    })


@login_required
def admin_cancel_import(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('admin_upload_period_folder')

    created_usernames = request.session.pop('last_import_created_users', []) or []
    created_payslip_ids = request.session.pop('last_import_created_payslips', []) or []

    # delete payslips first
    try:
        Payslip.objects.filter(id__in=created_payslip_ids).delete()
    except Exception:
        logger.exception('Error deleting created payslips during cancel-import')

    # delete users (which will cascade to Employee)
    try:
        User.objects.filter(username__in=created_usernames).delete()
    except Exception:
        logger.exception('Error deleting created users during cancel-import')

    # Audit: annullamento ultimo import cedolini
    _create_audit_event(
        request,
        "payslip_import_cancelled",
        metadata={
            "deleted_users": len(created_usernames),
            "deleted_payslips": len(created_payslip_ids),
        },
    )

    return redirect('admin_upload_period_folder')