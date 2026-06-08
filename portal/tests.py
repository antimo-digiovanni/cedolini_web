import importlib
from io import BytesIO
import os
from pathlib import Path
import sqlite3
from tempfile import NamedTemporaryFile
from django.conf import settings
from django.core import mail
from django.contrib.auth import BACKEND_SESSION_KEY, HASH_SESSION_KEY, SESSION_KEY, authenticate, get_user_model
from django.contrib.auth.models import Group
from django.contrib.sessions.backends.db import SessionStore
from django.core.files.uploadedfile import SimpleUploadedFile
import tempfile
from django.test import Client
from django.test import TestCase
from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from datetime import datetime
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from starlette.testclient import TestClient as AsgiTestClient

from .access import TODAY_MARKINGS_GROUP_NAME, RICONFEZIONAMENTO_GROUP_NAME, TURNI_PLANNER_GROUP_NAME
from .models import Cud, Employee, ImportJob, Payslip, PortalUserSetting, TurniPlannerWeekState, VacationRequest, WorkSession


class EmailOrUsernameBackendTests(TestCase):
	def setUp(self):
		self.user = get_user_model().objects.create_user(
			username="mario.rossi",
			email="mario.rossi@example.com",
			password="Password123!",
		)

	def test_login_with_username(self):
		logged_user = authenticate(username="mario.rossi", password="Password123!")
		self.assertIsNotNone(logged_user)
		self.assertEqual(logged_user.pk, self.user.pk)

	def test_login_with_email(self):
		logged_user = authenticate(username="mario.rossi@example.com", password="Password123!")
		self.assertIsNotNone(logged_user)
		self.assertEqual(logged_user.pk, self.user.pk)

	def test_login_with_username_case_insensitive(self):
		logged_user = authenticate(username="Mario.Rossi", password="Password123!")
		self.assertIsNotNone(logged_user)
		self.assertEqual(logged_user.pk, self.user.pk)


class TodayMarkingsAccessTests(TestCase):
	def setUp(self):
		self.client = Client()
		self.group = Group.objects.create(name=TODAY_MARKINGS_GROUP_NAME)
		self.owner_user = get_user_model().objects.create_user(
			username="titolare",
			password="Password123!",
			first_name="Mario",
			last_name="Bianchi",
		)
		self.owner_user.groups.add(self.group)

		self.employee_user = get_user_model().objects.create_user(
			username="dipendente",
			password="Password123!",
		)
		self.employee = Employee.objects.create(
			user=self.employee_user,
			first_name="Luca",
			last_name="Verdi",
		)
		WorkSession.objects.create(
			employee=self.employee,
			work_date=timezone.localdate(),
			started_at=timezone.now(),
		)
		self.turni_state = TurniPlannerWeekState.objects.create(
			week_label="WEEK OWNER",
			visible_to_employees=True,
			planner_data={
				"weekly": {
					"headers": ["Reparto A", "Reparto B", "Reparto C", "", "", "", "", "", "", ""],
					"central_departments": [""] * 10,
					"sections": [
						{"label": "1 turno", "time_values": ["06:00", "06:00", "06:00", "", "", "", "", "", "", ""], "rows": [["Mario", "Luca", "Anna", "", "", "", "", "", "", ""], [""] * 10, [""] * 10]},
						{"label": "2 turno", "time_values": [""] * 10, "rows": [[""] * 10, [""] * 10, [""] * 10]},
						{"label": "3 turno", "time_values": [""] * 10, "rows": [[""] * 10, [""] * 10, [""] * 10]},
						{"label": "turno centrale", "time_values": [""] * 10, "rows": [[""] * 10, [""] * 10, [""] * 10]},
					],
				},
				"saturday": {"base_date": "24/05/2026", "rows": [["24/05/2026", "Mattina", "Mario", "Capo A", "Presidio", "Reparto A"]]},
				"sunday": {"base_date": "25/05/2026", "rows": [["25/05/2026", "Sera", "Luca", "Capo B", "Supporto", "Reparto B"]]},
				"scorrimento": {"title": "Scorrimento demo", "base_date": "08/05/2026", "rows": [["08/05/2026", "Mattina", "Mario Rossi", "Capo A", "Scorrimento", "Reparto A"]]},
				"portineria_weekly": {
					"headers": ["PORTINERIA CENTRALE", "CENTRALINISTA", "PORTINERIA CELLA"],
					"sections": [
						{"label": "1 turno", "time_values": ["06:14", "08:17", "06:14"], "rows": [["A", "B", "C"], ["", "", ""], ["", "", ""]]},
						{"label": "2 turno", "time_values": ["14:22", "", "14:22"], "rows": [["", "", ""], ["", "", ""], ["", "", ""]]},
						{"label": "3 turno", "time_values": ["22:06", "", "22:06"], "rows": [["", "", ""], ["", "", ""], ["", "", ""]]},
					],
				},
				"portineria_weekend": {"base_date": "24/05/2026", "rows": [["24/05/2026", "Mattina", "Port A", "Resp A", "Controllo", "Portineria"]]},
			},
		)

	def test_home_redirects_limited_user_to_today_markings(self):
		self.client.force_login(self.owner_user)
		response = self.client.get(reverse("home"))
		self.assertRedirects(response, reverse("today_markings_dashboard"))

	def test_dashboard_redirects_limited_user_to_today_markings(self):
		self.client.force_login(self.owner_user)
		response = self.client.get(reverse("dashboard"))
		self.assertRedirects(response, reverse("today_markings_dashboard"))

	def test_admin_dashboard_redirects_limited_user_to_today_markings(self):
		self.client.force_login(self.owner_user)
		response = self.client.get(reverse("admin_dashboard"))
		self.assertRedirects(response, reverse("today_markings_dashboard"))

	def test_employee_with_today_markings_group_keeps_employee_home(self):
		self.employee_user.groups.add(self.group)
		self.client.force_login(self.employee_user)
		response = self.client.get(reverse("home"))
		self.assertRedirects(response, reverse("dashboard"))

	def test_employee_with_today_markings_group_keeps_timekeeping_page(self):
		self.employee_user.groups.add(self.group)
		self.client.force_login(self.employee_user)
		response = self.client.get(reverse("timekeeping"))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Marcatura")

	def test_employee_with_today_markings_group_can_still_open_today_markings_page(self):
		self.employee_user.groups.add(self.group)
		self.client.force_login(self.employee_user)
		response = self.client.get(reverse("today_markings_dashboard"))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Chi ha marcato oggi")

	def test_employee_with_today_markings_group_sees_markings_open_in_dashboard(self):
		self.employee_user.groups.add(self.group)
		self.client.force_login(self.employee_user)
		response = self.client.get(reverse("dashboard"))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Chi ha marcato oggi")
		self.assertContains(response, "Ingresso")
		self.assertContains(response, "Luca Verdi")

	def test_limited_user_can_view_today_markings_page(self):
		self.client.force_login(self.owner_user)
		response = self.client.get(reverse("today_markings_dashboard"))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Chi ha marcato oggi")
		self.assertContains(response, "Luca Verdi")

	def test_limited_user_does_not_see_mark_coordinates(self):
		session = WorkSession.objects.filter(employee=self.employee, work_date=timezone.localdate()).first()
		session.start_latitude = "40.123456"
		session.start_longitude = "14.654321"
		session.save(update_fields=["start_latitude", "start_longitude"])

		self.client.force_login(self.owner_user)
		response = self.client.get(reverse("today_markings_dashboard"))

		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, "Coordinate ingresso")
		self.assertNotContains(response, "40.123456, 14.654321")

	def test_limited_user_sees_published_turni_on_today_markings_by_default(self):
		self.client.force_login(self.owner_user)
		response = self.client.get(reverse("today_markings_dashboard"))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Turni della settimana")
		self.assertContains(response, self.turni_state.week_label)
		self.assertContains(response, reverse("employee_turni_published_image", args=["weekly"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["portineria_weekly"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["scorrimento"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["portineria_weekend"]))
		self.assertContains(response, "Scorrimento")
		self.assertContains(response, "Portineria settimana")
		self.assertContains(response, "Portineria weekend")

		image_response = self.client.get(reverse("employee_turni_published_image", args=["weekly"]))
		self.assertEqual(image_response.status_code, 200)
		portineria_response = self.client.get(reverse("employee_turni_published_image", args=["portineria_weekly"]))
		self.assertEqual(portineria_response.status_code, 200)

	def test_limited_user_sees_only_selected_published_turni_sections(self):
		self.turni_state.planner_data["published_sections"] = ["weekly", "saturday", "sunday"]
		self.turni_state.save(update_fields=["planner_data"])
		self.client.force_login(self.owner_user)

		response = self.client.get(reverse("today_markings_dashboard"))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, reverse("employee_turni_published_image", args=["weekly"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["saturday"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["sunday"]))
		self.assertNotContains(response, reverse("employee_turni_published_image", args=["portineria_weekly"]))
		self.assertNotContains(response, reverse("employee_turni_published_image", args=["portineria_weekend"]))

		portineria_response = self.client.get(reverse("employee_turni_published_image", args=["portineria_weekly"]))
		self.assertEqual(portineria_response.status_code, 404)

	def test_limited_user_can_be_disabled_from_published_turni(self):
		PortalUserSetting.objects.update_or_create(
			user=self.owner_user,
			defaults={"show_published_turni": False},
		)
		self.client.force_login(self.owner_user)

		response = self.client.get(reverse("today_markings_dashboard"))
		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, "Turni della settimana")

		image_response = self.client.get(reverse("employee_turni_published_image", args=["weekly"]))
		self.assertEqual(image_response.status_code, 404)

	def test_limited_user_can_view_previous_day_markings(self):
		WorkSession.objects.create(
			employee=self.employee,
			work_date=timezone.localdate() - timezone.timedelta(days=1),
			started_at=timezone.now() - timezone.timedelta(days=1),
		)
		self.client.force_login(self.owner_user)
		response = self.client.get(reverse("today_markings_dashboard"), {"date": (timezone.localdate() - timezone.timedelta(days=1)).isoformat()})
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Elenco marcature del")
		self.assertContains(response, "Luca Verdi")

	def test_today_markings_includes_overnight_exit_on_next_day(self):
		WorkSession.objects.all().delete()
		yesterday = timezone.localdate() - timezone.timedelta(days=1)
		today = timezone.localdate()
		tz = timezone.get_current_timezone()
		started_at = timezone.make_aware(datetime.combine(yesterday, datetime.strptime("17:00", "%H:%M").time()), tz)
		ended_at = timezone.make_aware(datetime.combine(today, datetime.strptime("01:00", "%H:%M").time()), tz)
		WorkSession.objects.create(
			employee=self.employee,
			work_date=yesterday,
			started_at=started_at,
			ended_at=ended_at,
		)

		self.client.force_login(self.owner_user)
		response_today = self.client.get(reverse("today_markings_dashboard"))
		self.assertEqual(response_today.status_code, 200)
		self.assertContains(response_today, "Luca Verdi")
		self.assertContains(response_today, "01:00")
		self.assertContains(response_today, "--:--")

		response_yesterday = self.client.get(reverse("today_markings_dashboard"), {"date": yesterday.isoformat()})
		self.assertEqual(response_yesterday.status_code, 200)
		self.assertContains(response_yesterday, "Luca Verdi")
		self.assertContains(response_yesterday, "17:00")

class PublicMachineryPageTests(TestCase):
	def test_public_machinery_page_shows_real_vehicle_cards(self):
		response = self.client.get(reverse("public_machinery"))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Bobcat")
		self.assertContains(response, "Gruppo elettrogeno industriale")
		self.assertContains(response, "Autospurgo canal-jet su Iveco Stralis a 4 assi")
		self.assertContains(response, "Piattaforma aerea semovente a braccio articolato JLG E300")
		self.assertContains(response, "Spazzatrice stradale aspirante Dulevo D6")
		self.assertContains(response, "Trattore stradale con semirimorchio")
		self.assertContains(response, "Magazzino operativo con carrelli")
		self.assertContains(response, "Autocarro con gru retrocabina")
		self.assertContains(response, 'data-image-count="3"')
		self.assertContains(response, 'data-image-count="4"')
		self.assertGreaterEqual(response.content.decode().count('data-image-count="2"'), 5)
		self.assertContains(response, 'alt="Bobcat"')


class VacationRequestFlowTests(TestCase):
	def setUp(self):
		self.client = Client()
		self.employee_user = get_user_model().objects.create_user(
			username="operaio",
			password="Password123!",
		)
		self.employee = Employee.objects.create(
			user=self.employee_user,
			first_name="Giovanni",
			last_name="Neri",
		)
		self.admin_user = get_user_model().objects.create_user(
			username="admin",
			password="Password123!",
			is_staff=True,
		)

	def test_employee_can_submit_vacation_request_from_dashboard(self):
		self.client.force_login(self.employee_user)
		start_date = timezone.localdate() + timezone.timedelta(days=2)
		end_date = start_date + timezone.timedelta(days=1)

		response = self.client.post(
			reverse("dashboard"),
			{
				"action": "request_vacation",
				"start_date": start_date.isoformat(),
				"end_date": end_date.isoformat(),
				"vacation_reason": "Ferie programmate con la famiglia.",
			},
		)

		self.assertRedirects(response, f"{reverse('dashboard')}?vacation_status=sent")
		request_obj = VacationRequest.objects.get(employee=self.employee)
		self.assertEqual(request_obj.start_date, start_date)
		self.assertEqual(request_obj.end_date, end_date)
		self.assertEqual(request_obj.status, VacationRequest.STATUS_PENDING)

	def test_admin_dashboard_approval_marks_days_as_vacation_in_report(self):
		start_date = timezone.localdate() + timezone.timedelta(days=1)
		end_date = start_date + timezone.timedelta(days=1)
		request_obj = VacationRequest.objects.create(
			employee=self.employee,
			start_date=start_date,
			end_date=end_date,
			reason="Ferie gia concordate.",
		)

		self.client.force_login(self.admin_user)
		response = self.client.post(
			reverse("admin_dashboard"),
			{
				"action": "approve_vacation_request",
				"request_id": str(request_obj.id),
				"review_note": "Approvato",
			},
		)

		self.assertRedirects(response, reverse("admin_dashboard"))
		request_obj.refresh_from_db()
		self.assertEqual(request_obj.status, VacationRequest.STATUS_APPROVED)

		sessions = WorkSession.objects.filter(employee=self.employee, work_date__range=(start_date, end_date)).order_by("work_date")
		self.assertEqual(sessions.count(), 2)
		self.assertTrue(all(session.day_type == WorkSession.DAY_TYPE_VACATION for session in sessions))
		self.assertTrue(all(session.started_at is None and session.ended_at is None for session in sessions))

		report_response = self.client.get(
			reverse("admin_timekeeping"),
			{
				"employee": str(self.employee.id),
				"month": start_date.month,
				"year": start_date.year,
			},
		)
		self.assertEqual(report_response.status_code, 200)
		self.assertContains(report_response, "Ferie")
		self.assertContains(report_response, "FERIE")

	def test_admin_can_see_mark_coordinates_in_reports(self):
		work_date = timezone.localdate()
		tz = timezone.get_current_timezone()
		WorkSession.objects.create(
			employee=self.employee,
			work_date=work_date,
			started_at=timezone.make_aware(datetime.combine(work_date, datetime.strptime("08:00", "%H:%M").time()), tz),
			ended_at=timezone.make_aware(datetime.combine(work_date, datetime.strptime("17:00", "%H:%M").time()), tz),
			start_latitude="40.123456",
			start_longitude="14.654321",
			end_latitude="40.123400",
			end_longitude="14.654300",
		)

		self.client.force_login(self.admin_user)

		report_response = self.client.get(
			reverse("admin_timekeeping"),
			{
				"employee": str(self.employee.id),
				"month": work_date.month,
				"year": work_date.year,
			},
		)
		self.assertEqual(report_response.status_code, 200)
		self.assertContains(report_response, "Coordinate ingresso")
		self.assertContains(report_response, "40.123456, 14.654321")
		self.assertContains(report_response, "40.123400, 14.654300")

		today_response = self.client.get(reverse("today_markings_dashboard"))
		self.assertEqual(today_response.status_code, 200)
		self.assertContains(today_response, "Coordinate ingresso")
		self.assertContains(today_response, "40.123456, 14.654321")
		self.assertContains(today_response, "40.123400, 14.654300")

	def test_admin_views_order_employees_by_first_name_without_changing_data(self):
		first_user = get_user_model().objects.create_user(
			username="zeno.alfa",
			password="Password123!",
		)
		second_user = get_user_model().objects.create_user(
			username="anna.zulu",
			password="Password123!",
		)
		first_employee = Employee.objects.create(
			user=first_user,
			first_name="Zeno",
			last_name="Alfa",
		)
		second_employee = Employee.objects.create(
			user=second_user,
			first_name="Anna",
			last_name="Zulu",
		)
		WorkSession.objects.create(
			employee=first_employee,
			work_date=timezone.localdate(),
			started_at=timezone.now(),
		)
		WorkSession.objects.create(
			employee=second_employee,
			work_date=timezone.localdate(),
			started_at=timezone.now(),
		)

		self.client.force_login(self.admin_user)

		employees_response = self.client.get(reverse("admin_employees"))
		self.assertEqual(employees_response.status_code, 200)
		employees_html = employees_response.content.decode()
		self.assertLess(employees_html.index("Anna Zulu"), employees_html.index("Zeno Alfa"))

		timekeeping_response = self.client.get(
			reverse("admin_timekeeping"),
			{
				"employee": "all",
				"month": timezone.localdate().month,
				"year": timezone.localdate().year,
			},
		)
		self.assertEqual(timekeeping_response.status_code, 200)
		timekeeping_html = timekeeping_response.content.decode()
		self.assertLess(timekeeping_html.index("Anna Zulu"), timekeeping_html.index("Zeno Alfa"))
		self.assertEqual(Employee.objects.get(id=first_employee.id).last_name, "Alfa")
		self.assertEqual(Employee.objects.get(id=second_employee.id).last_name, "Zulu")


class AdminRiconfezionamentoAccessManagementTests(TestCase):
	def setUp(self):
		self.client = Client()
		self.admin_user = get_user_model().objects.create_user(
			username="staff.riconf.manage",
			password="Password123!",
			is_staff=True,
		)
		self.client.force_login(self.admin_user)
		self.employee_user = get_user_model().objects.create_user(
			username="operatore.riconf",
			password="Password123!",
		)
		self.employee = Employee.objects.create(
			user=self.employee_user,
			first_name="Operatore",
			last_name="Test",
		)
		self.group, _ = Group.objects.get_or_create(name=RICONFEZIONAMENTO_GROUP_NAME)

	def test_admin_employee_detail_shows_enable_button_when_access_is_disabled(self):
		response = self.client.get(reverse("admin_employee_detail", args=[self.employee.id]))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Accesso riconfezionamento")
		self.assertContains(response, "Abilita accesso")

	def test_admin_can_enable_riconfezionamento_access(self):
		response = self.client.post(
			reverse("admin_employee_detail", args=[self.employee.id]),
			{
				"action": "toggle_riconfezionamento_access",
				"enable_access": "1",
			},
		)
		self.assertRedirects(response, f"{reverse('admin_employee_detail', args=[self.employee.id])}?outcome=riconfezionamento_enabled")
		self.assertTrue(self.employee_user.groups.filter(name=RICONFEZIONAMENTO_GROUP_NAME).exists())

	def test_admin_can_disable_riconfezionamento_access(self):
		self.employee_user.groups.add(self.group)
		response = self.client.post(
			reverse("admin_employee_detail", args=[self.employee.id]),
			{
				"action": "toggle_riconfezionamento_access",
				"enable_access": "0",
			},
		)
		self.assertRedirects(response, f"{reverse('admin_employee_detail', args=[self.employee.id])}?outcome=riconfezionamento_disabled")
		self.assertFalse(self.employee_user.groups.filter(name=RICONFEZIONAMENTO_GROUP_NAME).exists())

	def test_admin_employees_lists_riconfezionamento_badge(self):
		self.employee_user.groups.add(self.group)
		response = self.client.get(reverse("admin_employees"))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Riconfezionamento")
		self.assertContains(response, "Abilitato")


class EmployeePublishedTurniDashboardTests(TestCase):
	def setUp(self):
		self.client = Client()
		self.antimo_user = get_user_model().objects.create_user(
			username="antimo",
			password="Password123!",
			is_staff=True,
		)
		self.other_user = get_user_model().objects.create_user(
			username="mario",
			password="Password123!",
		)
		self.user = get_user_model().objects.create_user(
			username="employee.turni",
			password="Password123!",
		)
		self.employee = Employee.objects.create(
			user=self.user,
			first_name="Mario",
			last_name="Rossi",
		)
		self.state = TurniPlannerWeekState.objects.create(
			week_label="WEEK 21",
			visible_to_employees=True,
			planner_data={
				"weekly": {
					"headers": ["Reparto A", "Reparto B", "Reparto C", "", "", "", "", "", "", ""],
					"central_departments": [""] * 10,
					"sections": [
						{"label": "1 turno", "time_values": ["06:00", "06:00", "06:00", "", "", "", "", "", "", ""], "rows": [["Mario", "Luca", "Anna", "", "", "", "", "", "", ""], [""] * 10, [""] * 10]},
						{"label": "2 turno", "time_values": [""] * 10, "rows": [[""] * 10, [""] * 10, [""] * 10]},
						{"label": "3 turno", "time_values": [""] * 10, "rows": [[""] * 10, [""] * 10, [""] * 10]},
						{"label": "turno centrale", "time_values": [""] * 10, "rows": [[""] * 10, [""] * 10, [""] * 10]},
					],
				},
				"saturday": {
					"base_date": "24/05/2026",
					"rows": [["24/05/2026", "Mattina", "Mario", "Capo A", "Presidio", "Reparto A"]],
				},
				"sunday": {
					"base_date": "25/05/2026",
					"rows": [["25/05/2026", "Sera", "Luca", "Capo B", "Supporto", "Reparto B"]],
				},
				"jolly_weekend": {
					"title": "Comandata jolly demo",
					"base_date": "26/05/2026",
					"rows": [["26/05/2026", "Mattina", "Jolly A", "Capo J", "Presidio", "Reparto J"]],
				},
				"scorrimento": {
					"title": "Scorrimento demo",
					"base_date": "08/05/2026",
					"rows": [["08/05/2026", "Mattina", "Mario Rossi", "Capo A", "Scorrimento", "Reparto A"]],
				},
				"portineria_weekly": {
					"headers": ["PORTINERIA CENTRALE", "CENTRALINISTA", "PORTINERIA CELLA"],
					"sections": [
						{"label": "1 turno", "time_values": ["06:14", "08:17", "06:14"], "rows": [["A", "B", "C"], ["", "", ""], ["", "", ""]]},
						{"label": "2 turno", "time_values": ["14:22", "", "14:22"], "rows": [["", "", ""], ["", "", ""], ["", "", ""]]},
						{"label": "3 turno", "time_values": ["22:06", "", "22:06"], "rows": [["", "", ""], ["", "", ""], ["", "", ""]]},
					],
				},
				"portineria_weekend": {
					"base_date": "24/05/2026",
					"rows": [["24/05/2026", "Mattina", "Port A", "Resp A", "Controllo", "Portineria"]],
				},
			},
		)

	def test_employee_dashboard_shows_only_published_turni_images(self):
		self.client.force_login(self.user)
		response = self.client.get(reverse("dashboard"))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Turni della settimana")
		self.assertContains(response, self.state.week_label)
		self.assertContains(response, reverse("employee_turni_published_image", args=["weekly"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["portineria_weekly"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["saturday"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["sunday"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["jolly_weekend"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["scorrimento"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["portineria_weekend"]))
		self.assertContains(response, "Scorrimento")
		self.assertContains(response, "Portineria settimana")
		self.assertContains(response, "Portineria weekend")

		image_response = self.client.get(reverse("employee_turni_published_image", args=["weekly"]))
		self.assertEqual(image_response.status_code, 200)
		self.assertEqual(image_response["Content-Type"], "image/jpeg")
		self.assertTrue(image_response.content.startswith(b"\xff\xd8\xff"))

		portineria_response = self.client.get(reverse("employee_turni_published_image", args=["portineria_weekly"]))
		self.assertEqual(portineria_response.status_code, 200)

	def test_employee_dashboard_shows_only_selected_published_turni_images(self):
		self.state.planner_data["published_sections"] = ["weekly", "saturday", "sunday"]
		self.state.save(update_fields=["planner_data"])
		self.client.force_login(self.user)

		response = self.client.get(reverse("dashboard"))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, reverse("employee_turni_published_image", args=["weekly"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["saturday"]))
		self.assertContains(response, reverse("employee_turni_published_image", args=["sunday"]))
		self.assertNotContains(response, reverse("employee_turni_published_image", args=["portineria_weekly"]))
		self.assertNotContains(response, reverse("employee_turni_published_image", args=["portineria_weekend"]))

		portineria_response = self.client.get(reverse("employee_turni_published_image", args=["portineria_weekly"]))
		self.assertEqual(portineria_response.status_code, 404)

	def test_employee_dashboard_hides_turni_when_nothing_is_published(self):
		self.state.visible_to_employees = False
		self.state.save(update_fields=["visible_to_employees"])
		self.client.force_login(self.user)

		response = self.client.get(reverse("dashboard"))
		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, "Turni della settimana")

		image_response = self.client.get(reverse("employee_turni_published_image", args=["weekly"]))
		self.assertEqual(image_response.status_code, 404)

	def test_employee_dashboard_hides_turni_for_employee_disabled_in_admin(self):
		self.employee.show_published_turni = False
		self.employee.save(update_fields=["show_published_turni"])
		self.client.force_login(self.user)

		response = self.client.get(reverse("dashboard"))
		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, "Turni della settimana")

		image_response = self.client.get(reverse("employee_turni_published_image", args=["weekly"]))
		self.assertEqual(image_response.status_code, 404)

	def test_staff_can_still_open_published_turni_images(self):
		self.client.force_login(self.antimo_user)

		image_response = self.client.get(reverse("employee_turni_published_image", args=["weekly"]))
		self.assertEqual(image_response.status_code, 200)
		self.assertEqual(image_response["Content-Type"], "image/jpeg")
class TurniPlannerAccessTests(TestCase):
	def setUp(self):
		self.client = Client()
		self.group, _ = Group.objects.get_or_create(name=TURNI_PLANNER_GROUP_NAME)
		self.allowed_user = get_user_model().objects.create_user(
			username="planner.user",
			password="Password123!",
		)
		self.allowed_user.groups.add(self.group)
		self.denied_user = get_user_model().objects.create_user(
			username="basic.user",
			password="Password123!",
		)

	def test_turni_planner_allows_large_post_payloads(self):
		self.assertGreaterEqual(settings.DATA_UPLOAD_MAX_NUMBER_FIELDS, 20000)

	def test_home_redirects_turni_planner_user_to_planner(self):
		self.client.force_login(self.allowed_user)
		response = self.client.get(reverse("home"))
		self.assertRedirects(response, reverse("turni_planner_home"))

	def test_turni_planner_denies_non_authorized_user(self):
		self.client.force_login(self.denied_user)
		response = self.client.get(reverse("turni_planner_home"))
		self.assertEqual(response.status_code, 403)

	def test_turni_planner_disables_cache_headers(self):
		self.client.force_login(self.allowed_user)
		response = self.client.get(reverse("turni_planner_home"))
		self.assertEqual(response.status_code, 200)
		self.assertEqual(response["Cache-Control"], "no-store, no-cache, must-revalidate, max-age=0")
		self.assertEqual(response["Pragma"], "no-cache")
		self.assertEqual(response["Expires"], "0")

	def test_turni_planner_allows_group_user_and_creates_shared_week(self):
		self.client.force_login(self.allowed_user)
		response = self.client.post(
			reverse("turni_planner_home"),
			{"action": "open_week", "week_label": "Week 28 da Lunedi 06/07/2026 a Sabato 11/07/2026"},
		)

		state = TurniPlannerWeekState.objects.get()
		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={state.week_label}")
		self.assertEqual(state.updated_by, self.allowed_user)

	def test_turni_planner_deletes_requested_week(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 28 da Lunedi 06/07/2026 a Sabato 11/07/2026",
			planner_data={"weekly": {"headers": ["A"]}},
		)
		TurniPlannerWeekState.objects.create(
			week_label="Week 29 da Lunedi 13/07/2026 a Sabato 18/07/2026",
			planner_data={"weekly": {"headers": ["B"]}},
		)
		self.client.force_login(self.allowed_user)
		response = self.client.post(
			reverse("turni_planner_home"),
			{"action": "delete_week", "week_label": state.week_label},
		)

		self.assertRedirects(response, reverse("turni_planner_home"))
		self.assertFalse(TurniPlannerWeekState.objects.filter(week_label=state.week_label).exists())
		self.assertTrue(TurniPlannerWeekState.objects.filter(week_label="Week 29 da Lunedi 13/07/2026 a Sabato 18/07/2026").exists())


class RiconfezionamentoAccessTests(TestCase):
	@classmethod
	def setUpClass(cls):
		super().setUpClass()
		cls._original_data_dir = os.environ.get('APP_RICONFEZIONAMENTO_DATA_DIR')
		cls._original_products_xlsx = os.environ.get('APP_RICONFEZIONAMENTO_PRODUCTS_XLSX')
		cls._temp_data_dir = tempfile.TemporaryDirectory()
		cls._products_catalog_path = os.path.join(cls._temp_data_dir.name, 'Prodotti.xlsx')
		cls._write_products_catalog([
			('ART-001', 'Prodotto corretto'),
		])
		os.environ['APP_RICONFEZIONAMENTO_DATA_DIR'] = cls._temp_data_dir.name
		os.environ['APP_RICONFEZIONAMENTO_PRODUCTS_XLSX'] = cls._products_catalog_path
		cls.riconf_main = importlib.reload(importlib.import_module('riconfezionamento_app.main'))
		asgi_module = importlib.import_module('config.asgi')
		cls.asgi_application = importlib.reload(asgi_module).application

	@classmethod
	def tearDownClass(cls):
		if cls._original_data_dir is None:
			os.environ.pop('APP_RICONFEZIONAMENTO_DATA_DIR', None)
		else:
			os.environ['APP_RICONFEZIONAMENTO_DATA_DIR'] = cls._original_data_dir
		if cls._original_products_xlsx is None:
			os.environ.pop('APP_RICONFEZIONAMENTO_PRODUCTS_XLSX', None)
		else:
			os.environ['APP_RICONFEZIONAMENTO_PRODUCTS_XLSX'] = cls._original_products_xlsx
		cls._temp_data_dir.cleanup()
		super().tearDownClass()

	@classmethod
	def _write_products_catalog(cls, rows):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = 'Prodotti'
		worksheet.append(['Codice prodotto', 'Prodotto'])
		for product_code, product_name in rows:
			worksheet.append([product_code, product_name])
		workbook.save(cls._products_catalog_path)

	def _build_lot_excel(self, rows):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = 'Lotto'
		worksheet.append(['Fiche', 'Prodotto', 'Codice prodotto', 'Motivo riconfezionamento', 'Lotto di produzione', 'ZUN'])
		for row in rows:
			worksheet.append(row)
		buffer = BytesIO()
		workbook.save(buffer)
		return buffer.getvalue()

	def _build_lot_excel_with_merged_reason(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = 'Lotto'
		worksheet.append(['Fiche', 'Prodotto', 'Codice prodotto', 'Motivo riconfezionamento', 'Lotto di produzione', 'ZUN'])
		worksheet.append(['FICHE-001', 'Prodotto corretto', 'ART-001', 'Etichetta errata', 'LOT-001', 4])
		worksheet.append(['FICHE-002', 'Prodotto corretto', 'ART-001', None, 'LOT-001', 5])
		worksheet.append(['FICHE-003', 'Prodotto corretto', 'ART-001', None, 'LOT-001', 6])
		worksheet.merge_cells('D2:D4')
		buffer = BytesIO()
		workbook.save(buffer)
		return buffer.getvalue()

	def setUp(self):
		self.client = Client()
		self.group, _ = Group.objects.get_or_create(name=RICONFEZIONAMENTO_GROUP_NAME)
		self.allowed_user = get_user_model().objects.create_user(
			username='riconf.user',
			password='Password123!',
		)
		self.allowed_user.groups.add(self.group)
		self.denied_user = get_user_model().objects.create_user(
			username='no.riconf',
			password='Password123!',
		)
		Employee.objects.create(
			user=self.allowed_user,
			first_name='Operatore',
			last_name='Riconfezionamento',
		)

	def _build_asgi_client(self, user=None):
		client = AsgiTestClient(self.asgi_application)
		if user is None:
			return client

		session = SessionStore()
		session[SESSION_KEY] = str(user.pk)
		session[BACKEND_SESSION_KEY] = settings.AUTHENTICATION_BACKENDS[0]
		session[HASH_SESSION_KEY] = user.get_session_auth_hash()
		session.save()
		client.cookies.set(settings.SESSION_COOKIE_NAME, session.session_key)
		return client

	def test_portal_entry_denies_non_authorized_user(self):
		self.client.force_login(self.denied_user)
		response = self.client.get(reverse('riconfezionamento_entry'))
		self.assertEqual(response.status_code, 403)

	def test_portal_entry_redirects_authorized_user_to_mounted_app(self):
		self.client.force_login(self.allowed_user)
		response = self.client.get(reverse('riconfezionamento_entry'))
		self.assertRedirects(response, '/riconfezionamento/')

	def test_dashboard_shows_riconfezionamento_link_for_authorized_user(self):
		self.client.force_login(self.allowed_user)
		response = self.client.get(reverse('dashboard'))
		self.assertContains(response, reverse('riconfezionamento_entry'))

	def test_mounted_app_redirects_unauthenticated_user_to_login(self):
		with self._build_asgi_client() as client:
			response = client.get('/riconfezionamento/', follow_redirects=False)
		self.assertEqual(response.status_code, 302)
		self.assertEqual(response.headers['location'], f"{reverse('login')}?next=%2Friconfezionamento%2F")

	def test_mounted_app_denies_logged_user_without_group(self):
		with self._build_asgi_client(self.denied_user) as client:
			response = client.get('/riconfezionamento/')
		self.assertEqual(response.status_code, 403)

	def test_mounted_app_allows_group_user(self):
		with self._build_asgi_client(self.allowed_user) as client:
			response = client.get('/riconfezionamento/')
		self.assertEqual(response.status_code, 200)

	def test_mounted_app_downloads_product_catalog_workbook(self):
		with self._build_asgi_client(self.allowed_user) as client:
			response = client.get('/riconfezionamento/api/product-catalog/download')
		self.assertEqual(response.status_code, 200)
		self.assertIn('attachment;', response.headers.get('content-disposition', '').lower())
		self.assertIn('prodotti.xlsx', response.headers.get('content-disposition', '').lower())

	def test_import_check_blocks_code_product_mismatch_and_syncs_catalog(self):
		lot_bytes = self._build_lot_excel([
			['FICHE-001', 'Prodotto errato', 'ART-001', 'Etichetta rovinata', 'LOT-001', 4],
		])
		_, _, headers, rows = self.riconf_main.read_excel(lot_bytes, 1, None)
		self.riconf_main.sync_product_catalog()
		reason_column = self.riconf_main.validate_import_columns(
			headers,
			'Fiche',
			'',
			'',
			'Prodotto',
			'Codice prodotto',
			'Motivo riconfezionamento',
			'Lotto di produzione',
			'ZUN',
		)
		product_catalog, product_catalog_by_name = self.riconf_main.validate_product_catalog_for_rows(rows, 'Codice prodotto', 'Prodotto')

		with self.assertRaises(HTTPException) as exc:
			self.riconf_main.build_import_rows(
				rows,
				'',
				'Fiche',
				'',
				'Prodotto',
				'Codice prodotto',
				reason_column,
				'Lotto di produzione',
				'ZUN',
				strict_empty=False,
				product_catalog_by_code=product_catalog,
				product_catalog_by_name=product_catalog_by_name,
			)

		self.assertEqual(exc.exception.status_code, 400)
		self.assertEqual(exc.exception.detail['error_code'], 'product_catalog_mismatch')
		self.assertEqual(exc.exception.detail['mismatch_rows'][0]['product_code'], 'ART-001')
		self.assertEqual(exc.exception.detail['mismatch_rows'][0]['expected_product_name'], 'Prodotto corretto')

		db_path = os.path.join(self._temp_data_dir.name, 'repackaging.db')
		with sqlite3.connect(db_path) as connection:
			row = connection.execute(
				'SELECT product_name FROM product_catalog WHERE product_code = ?',
				('ART-001',),
			).fetchone()
		self.assertEqual(row[0], 'Prodotto corretto')

	def test_sync_product_catalog_for_import_allows_empty_catalog(self):
		self._write_products_catalog([])
		self.assertEqual(self.riconf_main.sync_product_catalog_for_import(), 0)

	def test_import_rows_auto_enrich_missing_catalog_entry(self):
		self._write_products_catalog([
			('ART-001', 'Prodotto corretto'),
		])
		lot_bytes = self._build_lot_excel([
			['FICHE-001', 'Prodotto nuovo', 'ART-999', 'Etichetta rovinata', 'LOT-002', 4],
		])
		_, _, headers, rows = self.riconf_main.read_excel(lot_bytes, 1, None)
		self.riconf_main.sync_product_catalog()
		reason_column = self.riconf_main.validate_import_columns(
			headers,
			'Fiche',
			'',
			'',
			'Prodotto',
			'Codice prodotto',
			'Motivo riconfezionamento',
			'Lotto di produzione',
			'ZUN',
		)
		product_catalog, product_catalog_by_name = self.riconf_main.validate_product_catalog_for_rows(rows, 'Codice prodotto', 'Prodotto')
		imported_rows, skipped_rows, catalog_rows_to_add = self.riconf_main.build_import_rows(
			rows,
			'',
			'Fiche',
			'',
			'Prodotto',
			'Codice prodotto',
			reason_column,
			'Lotto di produzione',
			'ZUN',
			strict_empty=False,
			product_catalog_by_code=product_catalog,
			product_catalog_by_name=product_catalog_by_name,
		)
		self.assertEqual(len(skipped_rows), 0)
		self.assertEqual(len(imported_rows), 1)
		self.assertEqual(catalog_rows_to_add, [{'product_code': 'ART-999', 'product_name': 'Prodotto nuovo'}])

		self.riconf_main._append_product_catalog_entries(catalog_rows_to_add)
		self.riconf_main.sync_product_catalog()

		catalog_workbook = load_workbook(self._products_catalog_path, read_only=True)
		try:
			rows_values = list(catalog_workbook.active.iter_rows(values_only=True))
		finally:
			catalog_workbook.close()
		self.assertIn(('ART-999', 'Prodotto nuovo'), rows_values)

	def test_read_excel_repeats_merged_reason_cells(self):
		lot_bytes = self._build_lot_excel_with_merged_reason()
		_, _, headers, rows = self.riconf_main.read_excel(lot_bytes, 1, None)
		self.riconf_main.sync_product_catalog()
		reason_column = self.riconf_main.validate_import_columns(
			headers,
			'Fiche',
			'',
			'',
			'Prodotto',
			'Codice prodotto',
			'Motivo riconfezionamento',
			'Lotto di produzione',
			'ZUN',
		)
		product_catalog, product_catalog_by_name = self.riconf_main.validate_product_catalog_for_rows(rows, 'Codice prodotto', 'Prodotto')

		imported_rows, skipped_rows, catalog_rows_to_add = self.riconf_main.build_import_rows(
			rows,
			'',
			'Fiche',
			'',
			'Prodotto',
			'Codice prodotto',
			reason_column,
			'Lotto di produzione',
			'ZUN',
			strict_empty=False,
			product_catalog_by_code=product_catalog,
			product_catalog_by_name=product_catalog_by_name,
		)

		self.assertEqual(len(skipped_rows), 0)
		self.assertEqual(len(imported_rows), 3)
		self.assertEqual(len(catalog_rows_to_add), 0)
		self.assertTrue(all(row['repackaging_reason'] == 'Etichetta errata' for row in imported_rows))

	def test_import_product_catalog_rows_adds_only_new_codes(self):
		result = self.riconf_main.import_product_catalog_rows([
			{'product_code': 'ART-001', 'product_name': 'Prodotto corretto'},
			{'product_code': 'ART-002', 'product_name': 'Secondo prodotto'},
		])
		self.assertEqual(result.added, 1)
		self.assertEqual(result.existing, 1)
		self.assertEqual(len(result.conflicts), 0)

	def test_import_product_catalog_rows_bulk_adds_multiple_codes(self):
		result = self.riconf_main.import_product_catalog_rows([
			{'product_code': 'ART-010', 'product_name': 'Prodotto dieci'},
			{'product_code': 'ART-011', 'product_name': 'Prodotto undici'},
		])
		self.assertEqual(result.added, 2)
		self.assertEqual(result.existing, 0)
		self.assertEqual(len(result.conflicts), 0)

		catalog_workbook = load_workbook(self._products_catalog_path, read_only=True)
		try:
			rows_values = list(catalog_workbook.active.iter_rows(values_only=True))
		finally:
			catalog_workbook.close()
		self.assertIn(('ART-010', 'Prodotto dieci'), rows_values)
		self.assertIn(('ART-011', 'Prodotto undici'), rows_values)

	def test_load_catalog_import_rows_accepts_mrdr_headers(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.append(['MRDR', 'MRDR Description'])
		worksheet.append(['ART-777', 'Prodotto De Rosa'])
		buffer = BytesIO()
		workbook.save(buffer)

		rows = self.riconf_main._load_catalog_import_rows(buffer.getvalue(), 1, None)

		self.assertEqual(rows, [{'product_code': 'ART-777', 'product_name': 'Prodotto De Rosa'}])

	def test_load_catalog_import_rows_reports_empty_file(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.append(['Codice', 'Nome'])
		buffer = BytesIO()
		workbook.save(buffer)

		with self.assertRaises(HTTPException) as context:
			self.riconf_main._load_catalog_import_rows(buffer.getvalue(), 1, None)

		self.assertEqual(context.exception.status_code, 400)
		self.assertEqual(context.exception.detail, "Il file anagrafica e' vuoto.")

	def test_sync_product_catalog_creates_missing_catalog_file(self):
		if os.path.exists(self._products_catalog_path):
			os.remove(self._products_catalog_path)

		with self.assertRaises(HTTPException) as context:
			self.riconf_main.sync_product_catalog()

		self.assertEqual(context.exception.status_code, 500)
		self.assertEqual(context.exception.detail, 'Anagrafica prodotti senza righe valide.')
		self.assertTrue(os.path.exists(self._products_catalog_path))

		catalog_workbook = load_workbook(self._products_catalog_path, read_only=True)
		try:
			rows_values = list(catalog_workbook.active.iter_rows(values_only=True, max_row=2))
		finally:
			catalog_workbook.close()
		self.assertEqual(rows_values[0], ('Codice prodotto', 'Prodotto'))

	def test_sync_product_catalog_accepts_mrdr_headers(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.append(['MRDR', 'MRDR Description'])
		worksheet.append(['ART-888', 'Catalogo De Rosa'])
		workbook.save(self._products_catalog_path)

		self.riconf_main.sync_product_catalog()

		db_path = os.path.join(self._temp_data_dir.name, 'repackaging.db')
		with sqlite3.connect(db_path) as connection:
			row = connection.execute(
				'SELECT product_name FROM product_catalog WHERE product_code = ?',
				('ART-888',),
			).fetchone()
		self.assertEqual(row[0], 'Catalogo De Rosa')

	def test_sync_product_catalog_ignores_conflicting_duplicate_codes(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.append(['Codice', 'Nome'])
		worksheet.append(['ART-999', 'Prodotto Uno'])
		worksheet.append(['ART-999', 'Prodotto Due'])
		workbook.save(self._products_catalog_path)

		self.riconf_main.sync_product_catalog()

		db_path = os.path.join(self._temp_data_dir.name, 'repackaging.db')
		with sqlite3.connect(db_path) as connection:
			rows = connection.execute(
				'SELECT product_code, product_name FROM product_catalog WHERE product_code = ?',
				('ART-999',),
			).fetchall()
		self.assertEqual(rows, [('ART-999', 'Prodotto Uno')])

	def test_sync_product_catalog_deduplicates_identical_codes(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.append(['Codice', 'Nome'])
		worksheet.append(['ART-111', 'Prodotto Uno'])
		worksheet.append(['ART-111', 'Prodotto Uno'])
		workbook.save(self._products_catalog_path)

		self.riconf_main.sync_product_catalog()

		db_path = os.path.join(self._temp_data_dir.name, 'repackaging.db')
		with sqlite3.connect(db_path) as connection:
			count = connection.execute(
				'SELECT COUNT(*) FROM product_catalog WHERE product_code = ?',
				('ART-111',),
			).fetchone()[0]
		self.assertEqual(count, 1)

	def test_clear_product_catalog_empties_workbook_and_database(self):
		self._write_products_catalog([
			('ART-001', 'Prodotto corretto'),
			('ART-002', 'Prodotto persistente'),
		])
		self.riconf_main.sync_product_catalog()

		cleared = self.riconf_main.clear_product_catalog()

		self.assertEqual(cleared, 0)
		self.assertEqual(self.riconf_main.list_product_catalog(limit=10), [])

		catalog_workbook = load_workbook(self._products_catalog_path, read_only=True)
		try:
			rows_values = list(catalog_workbook.active.iter_rows(values_only=True))
		finally:
			catalog_workbook.close()
		self.assertEqual(rows_values, [('Codice prodotto', 'Prodotto')])

	def test_resolve_product_catalog_entry_without_force_keeps_conflict_blocking(self):
		with self.assertRaises(HTTPException) as exc:
			self.riconf_main.resolve_product_catalog_entry(
				'ART-001',
				'Prodotto corretto',
				'ART-777',
				'Prodotto corretto',
				force=False,
			)

		self.assertEqual(exc.exception.status_code, 400)
		self.assertIn("gia' presente", exc.exception.detail)

	def test_resolve_product_catalog_entry_with_force_replaces_existing_mapping(self):
		result = self.riconf_main.resolve_product_catalog_entry(
			'ART-001',
			'Prodotto corretto',
			'ART-777',
			'Prodotto corretto',
			force=True,
		)

		self.assertTrue(result['forced'])
		catalog_workbook = load_workbook(self._products_catalog_path, read_only=True)
		try:
			rows_values = list(catalog_workbook.active.iter_rows(values_only=True))
		finally:
			catalog_workbook.close()
		self.assertIn(('ART-777', 'Prodotto corretto'), rows_values)
		self.assertNotIn(('ART-001', 'Prodotto corretto'), rows_values)

	def test_scan_incoming_recognizes_pallet_in_non_current_open_batch(self):
		batch_eight = self.riconf_main.import_items('Lotto n° 8.xlsx', [{
			'pallet_code': 'PALLET-8',
			'incoming_fiche': 'PALLET-8',
			'outgoing_fiche': 'OUT-8',
			'product_name': 'Prodotto corretto',
			'product_code': 'ART-001',
			'production_lot': 'LOT-008',
			'repackaging_reason': 'Controllo lotto 8',
			'zun_quantity': 5,
		}])
		self.riconf_main.import_items('Lotto n° 9.xlsx', [{
			'pallet_code': 'PALLET-9',
			'incoming_fiche': 'PALLET-9',
			'outgoing_fiche': 'OUT-9',
			'product_name': 'Prodotto corretto',
			'product_code': 'ART-001',
			'production_lot': 'LOT-009',
			'repackaging_reason': 'Controllo lotto 9',
			'zun_quantity': 6,
		}])

		success, message, item = self.riconf_main.register_incoming('PALLET-8', 'Operatore Test')

		self.assertTrue(success)
		self.assertEqual(message, "OK entrata: pallet registrato in lavorazione.")
		self.assertEqual(item['batch_id'], batch_eight['batch_id'])
		self.assertEqual(item['state'], 'in_progress')

		self.riconf_main.reset_pallet('PALLET-8', batch_id=batch_eight['batch_id'])
		success_with_selected_batch, _, item_with_selected_batch = self.riconf_main.register_incoming(
			'PALLET-8',
			'Operatore Test',
			batch_id=batch_eight['batch_id'] + 1,
		)
		self.assertTrue(success_with_selected_batch)
		self.assertEqual(item_with_selected_batch['batch_id'], batch_eight['batch_id'])

	def test_scan_incoming_uses_selected_open_batch_without_crashing(self):
		batch_eight = self.riconf_main.import_items('Lotto n° 8.xlsx', [{
			'pallet_code': 'PALLET-8',
			'incoming_fiche': 'PALLET-8',
			'outgoing_fiche': '',
			'product_name': 'Prodotto corretto',
			'product_code': 'ART-001',
			'production_lot': 'LOT-008',
			'repackaging_reason': 'Controllo lotto 8',
			'zun_quantity': 5,
		}])

		success, message, item = self.riconf_main.register_incoming(
			'PALLET-8',
			'Operatore Test',
			batch_id=batch_eight['batch_id'],
		)

		self.assertTrue(success)
		self.assertEqual(message, 'OK entrata: pallet registrato in lavorazione.')
		self.assertEqual(item['batch_id'], batch_eight['batch_id'])
		self.assertEqual(item['state'], 'in_progress')

	def test_scan_outgoing_uses_selected_batch_when_latest_batch_is_different(self):
		batch_eight = self.riconf_main.import_items('Lotto n° 8.xlsx', [{
			'pallet_code': 'PALLET-8',
			'incoming_fiche': 'PALLET-8',
			'outgoing_fiche': '',
			'product_name': 'Prodotto corretto',
			'product_code': 'ART-001',
			'production_lot': 'LOT-008',
			'repackaging_reason': 'Controllo lotto 8',
			'zun_quantity': 5,
		}])
		self.riconf_main.import_items('Lotto n° 9.xlsx', [{
			'pallet_code': 'PALLET-9',
			'incoming_fiche': 'PALLET-9',
			'outgoing_fiche': 'OUT-9',
			'product_name': 'Prodotto corretto',
			'product_code': 'ART-001',
			'production_lot': 'LOT-009',
			'repackaging_reason': 'Controllo lotto 9',
			'zun_quantity': 6,
		}])

		incoming_success, _, incoming_item = self.riconf_main.register_incoming('PALLET-8', 'Operatore Test')
		self.assertTrue(incoming_success)

		outgoing_success, _, outgoing_item, error_code = self.riconf_main.register_outgoing(
			'PALLET-8',
			'NUOVA-OUT-8',
			4,
			'ART-001',
			'Operatore Test',
			batch_id=batch_eight['batch_id'],
		)

		self.assertTrue(outgoing_success)
		self.assertIsNone(error_code)
		self.assertEqual(incoming_item['batch_id'], batch_eight['batch_id'])
		self.assertEqual(outgoing_item['batch_id'], batch_eight['batch_id'])
		self.assertEqual(outgoing_item['state'], 'completed')

	def test_scan_incoming_reports_similar_codes_when_exact_code_is_missing(self):
		self.riconf_main.import_items('Lotto n° 8.xlsx', [{
			'pallet_code': '180011905525294048',
			'incoming_fiche': '180011905525294048',
			'outgoing_fiche': '',
			'product_name': 'Prodotto corretto',
			'product_code': 'ART-001',
			'production_lot': 'LOT-008',
			'repackaging_reason': 'Controllo lotto 8',
			'zun_quantity': 5,
		}, {
			'pallet_code': '180011905525294123',
			'incoming_fiche': '180011905525294123',
			'outgoing_fiche': '',
			'product_name': 'Prodotto corretto',
			'product_code': 'ART-001',
			'production_lot': 'LOT-008',
			'repackaging_reason': 'Controllo lotto 8',
			'zun_quantity': 5,
		}])

		success, message, item = self.riconf_main.register_incoming('180011905525294000', 'Operatore Test')

		self.assertFalse(success)
		self.assertIsNone(item)
		self.assertIn("Codice non presente in nessun lotto aperto.", message)
		self.assertIn("Lotto n° 8.xlsx", message)
		self.assertIn("180011905525294048", message)
		self.assertIn("180011905525294123", message)

	def test_wipe_all_data_keeps_product_catalog(self):
		self._write_products_catalog([
			('ART-001', 'Prodotto corretto'),
			('ART-002', 'Prodotto persistente'),
		])
		self.riconf_main.sync_product_catalog()
		self.riconf_main.import_items('Lotto n° 8.xlsx', [{
			'pallet_code': 'PALLET-8',
			'incoming_fiche': 'PALLET-8',
			'outgoing_fiche': 'OUT-8',
			'product_name': 'Prodotto corretto',
			'product_code': 'ART-001',
			'production_lot': 'LOT-008',
			'repackaging_reason': 'Controllo lotto 8',
			'zun_quantity': 5,
		}])

		deleted = self.riconf_main.wipe_all_data()

		self.assertEqual(self.riconf_main.list_items(), [])
		catalog_rows = self.riconf_main.list_product_catalog(limit=10)
		self.assertEqual(len(catalog_rows), 2)
		self.assertEqual({row['product_code'] for row in catalog_rows}, {'ART-001', 'ART-002'})
		self.assertEqual(deleted['backups_deleted'], 0)

		catalog_workbook = load_workbook(self._products_catalog_path, read_only=True)
		try:
			rows_values = list(catalog_workbook.active.iter_rows(values_only=True))
		finally:
			catalog_workbook.close()
		self.assertIn(('ART-001', 'Prodotto corretto'), rows_values)
		self.assertIn(('ART-002', 'Prodotto persistente'), rows_values)

	def test_multiple_catalog_conflicts_can_be_forced_in_sequence(self):
		self._write_products_catalog([
			('65676623', 'HB Remix 65ml Sandwich 4MPCL1x4x155EB'),
			('65283476', 'CORN 125ml HQFM CLASC Promo CL1x24X162EB'),
		])
		self.riconf_main.sync_product_catalog()

		result = self.riconf_main.import_product_catalog_rows([
			{'product_code': '99900001', 'product_name': 'HB Remix 65ml Sandwich 4MPCL1x4x155EB'},
			{'product_code': '99900002', 'product_name': 'CORN 125ml HQFM CLASC Promo CL1x24X162EB'},
		])

		self.assertEqual(result.added, 0)
		self.assertEqual(result.existing, 0)
		self.assertEqual(len(result.conflicts), 2)

		for conflict in result.conflicts:
			self.riconf_main.resolve_product_catalog_entry(
				conflict['current_product_code'],
				conflict['current_product_name'],
				conflict['product_code'],
				conflict['product_name'],
				force=True,
			)

		catalog_rows = self.riconf_main.list_product_catalog(limit=10)
		catalog_codes = {row['product_code'] for row in catalog_rows}
		self.assertIn('99900001', catalog_codes)
		self.assertIn('99900002', catalog_codes)
		self.assertNotIn('65676623', catalog_codes)
		self.assertNotIn('65283476', catalog_codes)

	def test_batch_report_highlights_manual_changes_and_forced_changes_with_distinct_row_colors(self):
		reports_dir = Path(self._temp_data_dir.name) / 'reports'
		report_path = self.riconf_main.generate_batch_report(
			reports_dir,
			{
				'filename': 'Lotto n° 9.xlsx',
				'imported_at': timezone.now().isoformat(),
				'completed_at': timezone.now().isoformat(),
			},
			[
				{
					'pallet_code': 'PALLET-MOD',
					'product_name': 'Prodotto corretto',
					'original_product_code': 'ART-001',
					'product_code': 'ART-001',
					'zun_quantity': 5,
					'repackaging_reason': 'Motivo modificato',
					'incoming_fiche': 'PALLET-MOD',
					'incoming_operator': 'Operatore Test',
					'scanned_incoming_at': timezone.now().isoformat(),
					'state': 'completed',
					'manual_reason_override': 1,
					'product_code_changed': 0,
					'product_code_change_operator': '',
					'waiting_operator': '',
					'outgoing_fiche': 'OUT-MOD',
					'outgoing_operator': 'Operatore Test',
					'scanned_outgoing_at': timezone.now().isoformat(),
				},
				{
					'pallet_code': 'PALLET-FORCE',
					'product_name': 'Prodotto corretto',
					'original_product_code': 'ART-001',
					'product_code': 'ART-999',
					'zun_quantity': 6,
					'repackaging_reason': 'Motivo invariato',
					'incoming_fiche': 'PALLET-FORCE',
					'incoming_operator': 'Operatore Test',
					'scanned_incoming_at': timezone.now().isoformat(),
					'state': 'completed',
					'manual_reason_override': 0,
					'product_code_changed': 1,
					'product_code_change_operator': 'Operatore Test',
					'waiting_operator': '',
					'outgoing_fiche': 'OUT-FORCE',
					'outgoing_operator': 'Operatore Test',
					'scanned_outgoing_at': timezone.now().isoformat(),
				},
			],
			{
				'total_items': 2,
				'completed': 2,
				'waiting_fiche': 0,
				'registered': 0,
			},
		)

		workbook = load_workbook(report_path, read_only=False)
		try:
			detail_sheet = workbook['Dettaglio pedane']
			manual_fill = detail_sheet['A6'].fill.fgColor.rgb
			forced_fill = detail_sheet['A7'].fill.fgColor.rgb
			note_text = detail_sheet.parent['Riepilogo lotto']['B20'].value
		finally:
			workbook.close()

		self.assertTrue(str(manual_fill or '').upper().endswith('DDEBFF'))
		self.assertTrue(str(forced_fill or '').upper().endswith('FFF59D'))
		self.assertIn('modifica manuale', note_text)
		self.assertIn('forzatura', note_text)

	def test_turni_planner_new_week_clones_latest_planner_data(self):
		previous_state = TurniPlannerWeekState.objects.create(
			week_label="Week 17: da Lunedi 20/04/2026 a Venerdi 24/04/2026",
			planner_data={
				"weekly_export_week_label": "Titolo export settimana 17",
				"portineria_weekly_export_week_label": "Titolo export portineria 17",
				"weekly": {
					"headers": [f"Reparto {index}" for index in range(1, 11)],
					"central_departments": [""] * 10,
					"sections": [
						{
							"label": "1 turno",
							"time_values": ["06:00"] * 10,
							"rows": [["Mario"] * 10, ["Luigi"] * 10, ["Anna"] * 10],
						},
						{
							"label": "2 turno",
							"time_values": ["14:00"] * 10,
							"rows": [["Paolo"] * 10, ["Gina"] * 10, ["Luca"] * 10],
						},
						{
							"label": "3 turno",
							"time_values": ["22:00"] * 10,
							"rows": [["Sara"] * 10, ["Piero"] * 10, ["Marta"] * 10],
						},
						{
							"label": "4 turno",
							"time_values": ["00:00"] * 10,
							"rows": [["Notte A"] * 10, ["Notte B"] * 10, ["Notte C"] * 10],
						},
					],
				},
				"saturday": {
					"base_date": "25/04/2026",
					"rows": [["25/04/2026", "Mattina", "Mario", "Capo A", "Lavaggio", "Reparto A"]],
				},
				"sunday": {
					"base_date": "26/04/2026",
					"rows": [["26/04/2026", "Sera", "Luigi", "Capo B", "Controllo", "Reparto B"]],
				},
				"scorrimento": {
					"title": "Scorrimento 24/04/2026",
					"base_date": "24/04/2026",
					"rows": [["24/04/2026", "Mattina", "Mario Rossi", "Capo A", "Scorrimento", "Reparto A"]],
				},
				"portineria_weekly": {
					"headers": ["Portineria Centrale", "Centralinista", "Portineria Cella"],
					"sections": [
						{"label": "1 turno", "time_values": ["06:14", "08:17", "06:14"], "rows": [["A", "B", "C"], ["D", "E", "F"], ["G", "H", "I"]]},
						{"label": "2 turno", "time_values": ["14:22", "", "14:22"], "rows": [["L", "M", "N"], ["O", "P", "Q"], ["R", "S", "T"]]},
						{"label": "3 turno", "time_values": ["22:06", "", "22:06"], "rows": [["U", "V", "Z"], ["AA", "AB", "AC"], ["AD", "AE", "AF"]]},
					],
				},
				"portineria_weekend": {
					"base_date": "25/04/2026",
					"rows": [["25/04/2026", "Mattina", "Port A", "Resp A", "Controllo", "Portineria"]],
				},
			},
			updated_by=self.allowed_user,
		)

		self.client.force_login(self.allowed_user)
		new_week_label = "Week 19: da Lunedi 04/05/2026 a Venerdi 08/05/2026"
		response = self.client.post(
			reverse("turni_planner_home"),
			{"action": "open_week", "week_label": new_week_label},
		)

		new_state = TurniPlannerWeekState.objects.get(week_label=new_week_label)
		previous_state.refresh_from_db()
		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={new_state.week_label}")
		self.assertEqual(new_state.updated_by, self.allowed_user)
		self.assertEqual(new_state.planner_data["weekly"], previous_state.planner_data["weekly"])
		self.assertEqual(new_state.planner_data["saturday"], previous_state.planner_data["saturday"])
		self.assertEqual(new_state.planner_data["scorrimento"], previous_state.planner_data["scorrimento"])
		self.assertEqual(new_state.planner_data["portineria_weekly"], previous_state.planner_data["portineria_weekly"])
		self.assertEqual(new_state.planner_data["portineria_weekend"], previous_state.planner_data["portineria_weekend"])
		self.assertEqual(new_state.planner_data["weekly_export_week_label"], new_week_label)
		self.assertEqual(new_state.planner_data["portineria_weekly_export_week_label"], new_week_label)
		self.assertIsNot(new_state.planner_data, previous_state.planner_data)

	def test_turni_planner_open_week_backfills_existing_empty_week_from_latest_non_empty_state(self):
		previous_state = TurniPlannerWeekState.objects.create(
			week_label="Week 17: da Lunedi 20/04/2026 a Venerdi 24/04/2026",
			planner_data={
				"weekly_export_week_label": "Titolo export settimana 17",
				"portineria_weekly_export_week_label": "Titolo export portineria 17",
				"weekly": {
					"headers": [f"Reparto {index}" for index in range(1, 11)],
					"central_departments": [""] * 10,
					"sections": [
						{"label": "1 turno", "time_values": ["06:00"] * 10, "rows": [["Mario"] * 10, ["Luigi"] * 10, ["Anna"] * 10]},
						{"label": "2 turno", "time_values": ["14:00"] * 10, "rows": [["Paolo"] * 10, ["Gina"] * 10, ["Luca"] * 10]},
						{"label": "3 turno", "time_values": ["22:00"] * 10, "rows": [["Sara"] * 10, ["Piero"] * 10, ["Marta"] * 10]},
						{"label": "4 turno", "time_values": ["00:00"] * 10, "rows": [["Notte A"] * 10, ["Notte B"] * 10, ["Notte C"] * 10]},
					],
				},
				"saturday": {"base_date": "25/04/2026", "rows": [["25/04/2026", "Mattina", "Mario", "Capo A", "Lavaggio", "Reparto A"]]},
			},
			updated_by=self.allowed_user,
		)
		TurniPlannerWeekState.objects.create(
			week_label="WEEK 18",
			planner_data={},
			updated_by=self.allowed_user,
		)
		empty_state = TurniPlannerWeekState.objects.create(
			week_label="Week 19: da Lunedì 04/05/2026 a Venerdì 08/05/2026",
			planner_data={},
		)

		self.client.force_login(self.allowed_user)
		response = self.client.post(
			reverse("turni_planner_home"),
			{"action": "open_week", "week_label": empty_state.week_label},
		)

		empty_state.refresh_from_db()
		previous_state.refresh_from_db()
		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={empty_state.week_label}")
		self.assertEqual(empty_state.planner_data["weekly"], previous_state.planner_data["weekly"])
		self.assertEqual(empty_state.planner_data["saturday"], previous_state.planner_data["saturday"])
		self.assertEqual(empty_state.planner_data["weekly_export_week_label"], empty_state.week_label)
		self.assertEqual(empty_state.planner_data["portineria_weekly_export_week_label"], empty_state.week_label)

	def test_turni_planner_saves_shared_planner_data(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 29 da Lunedi 13/07/2026 a Sabato 18/07/2026",
			planner_data={},
		)
		self.client.force_login(self.allowed_user)
		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "save_planner",
				"week_label": state.week_label,
				"weekly_export_week_label": "Week 29 sabato settimana",
				"portineria_weekly_export_week_label": "Week 29 venerdi portineria",
				"weekly_headers": [f"Reparto {index}" for index in range(1, 11)],
				"weekly_time_0": ["06:00"] * 10,
				"weekly_time_1": ["14:00"] * 10,
				"weekly_time_2": ["22:00"] * 10,
				"weekly_time_3": ["00:00"] * 10,
				"weekly_row_0_0": ["Mario"] * 10,
				"weekly_row_0_1": ["Luigi"] * 10,
				"weekly_row_0_2": ["Anna"] * 10,
				"weekly_row_1_0": ["Paolo"] * 10,
				"weekly_row_1_1": ["Gina"] * 10,
				"weekly_row_1_2": ["Luca"] * 10,
				"weekly_row_2_0": ["Sara"] * 10,
				"weekly_row_2_1": ["Piero"] * 10,
				"weekly_row_2_2": ["Marta"] * 10,
				"weekly_row_3_0": ["Notte A"] * 10,
				"weekly_row_3_1": ["Notte B"] * 10,
				"weekly_row_3_2": ["Notte C"] * 10,
				"saturday_base_date": "18/07/2026",
				"saturday_row_0": ["18/07/2026", "Mattina", "Mario", "Capo A", "Lavaggio", "Reparto A"],
				"sunday_base_date": "19/07/2026",
				"sunday_row_0": ["19/07/2026", "Notte", "Luigi", "Capo B", "Sanificazione", "Reparto B"],
				"jolly_weekend_title": "Comandata primo maggio",
				"jolly_weekend_base_date": "01/05/2026",
				"jolly_weekend_row_0": ["01/05/2026", "Mattina", "Jolly A", "Capo J", "Presidio", "Reparto J"],
				"scorrimento_title": "Scorrimento 08/05/2026",
				"scorrimento_base_date": "08/05/2026",
				"scorrimento_row_0": ["08/05/2026", "Mattina", "Mario Rossi", "Capo A", "Scorrimento", "Reparto A"],
				"portineria_weekly_headers": ["Portineria Centrale", "Centralinista", "Portineria Cella"],
				"portineria_weekly_time_0": ["06:14", "08:17", "06:14"],
				"portineria_weekly_time_1": ["14:22", "", "14:22"],
				"portineria_weekly_time_2": ["22:06", "", "22:06"],
				"portineria_weekly_row_0_0": ["Persona A", "Persona B", "Persona C"],
				"portineria_weekly_row_0_1": ["Persona D", "Persona E", "Persona F"],
				"portineria_weekly_row_0_2": ["Persona G", "Persona H", "Persona I"],
				"portineria_weekly_row_1_0": ["Persona L", "Persona M", "Persona N"],
				"portineria_weekly_row_1_1": ["Persona O", "Persona P", "Persona Q"],
				"portineria_weekly_row_1_2": ["Persona R", "Persona S", "Persona T"],
				"portineria_weekly_row_2_0": ["Persona U", "Persona V", "Persona Z"],
				"portineria_weekly_row_2_1": ["Persona AA", "Persona AB", "Persona AC"],
				"portineria_weekly_row_2_2": ["Persona AD", "Persona AE", "Persona AF"],
				"portineria_weekend_base_date": "18/07/2026",
				"portineria_weekend_row_0": ["18/07/2026", "Mattina", "Port A", "Resp A", "Controllo", "Portineria"],
			},
		)

		state.refresh_from_db()
		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={state.week_label}")
		self.assertEqual(state.updated_by, self.allowed_user)
		self.assertEqual(state.planner_data["weekly_export_week_label"], "Week 29 sabato settimana")
		self.assertEqual(state.planner_data["portineria_weekly_export_week_label"], "Week 29 venerdi portineria")
		self.assertEqual(state.planner_data["weekly"]["headers"][0], "Reparto 1")
		self.assertEqual(state.planner_data["weekly"]["sections"][0]["rows"][0][0], "Mario")
		self.assertEqual(state.planner_data["weekly"]["sections"][3]["rows"][2][9], "Notte C")
		self.assertEqual(state.planner_data["saturday"]["base_date"], "18/07/2026")
		self.assertEqual(state.planner_data["saturday"]["rows"][0][2], "Mario")
		self.assertEqual(state.planner_data["sunday"]["rows"][0][4], "Sanificazione")
		self.assertEqual(state.planner_data["jolly_weekend"]["title"], "Comandata primo maggio")
		self.assertEqual(state.planner_data["jolly_weekend"]["rows"][0][2], "Jolly A")
		self.assertEqual(state.planner_data["scorrimento"]["title"], "Scorrimento 08/05/2026")
		self.assertEqual(state.planner_data["scorrimento"]["rows"][0][2], "Mario Rossi")
		self.assertEqual(state.planner_data["portineria_weekly"]["sections"][0]["rows"][0][1], "Persona B")
		self.assertEqual(state.planner_data["portineria_weekend"]["rows"][0][5], "Portineria")

	def test_turni_planner_saves_dynamic_weekend_row_counts(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 29 bis da Lunedi 13/07/2026 a Sabato 18/07/2026",
			planner_data={},
		)
		self.client.force_login(self.allowed_user)
		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "save_planner",
				"week_label": state.week_label,
				"weekly_headers": [f"Reparto {index}" for index in range(1, 11)],
				"weekly_time_0": [""] * 10,
				"weekly_time_1": [""] * 10,
				"weekly_time_2": [""] * 10,
				"weekly_time_3": [""] * 10,
				"weekly_row_0_0": [""] * 10,
				"weekly_row_0_1": [""] * 10,
				"weekly_row_0_2": [""] * 10,
				"weekly_row_1_0": [""] * 10,
				"weekly_row_1_1": [""] * 10,
				"weekly_row_1_2": [""] * 10,
				"weekly_row_2_0": [""] * 10,
				"weekly_row_2_1": [""] * 10,
				"weekly_row_2_2": [""] * 10,
				"weekly_row_3_0": [""] * 10,
				"weekly_row_3_1": [""] * 10,
				"weekly_row_3_2": [""] * 10,
				"saturday_base_date": "18/07/2026",
				"saturday_row_count": "22",
				"saturday_row_0": ["18/07/2026", "Mattina", "Mario", "Capo A", "Lavaggio", "Reparto A"],
				"saturday_row_21": ["18/07/2026", "Sera", "Ultimo Sabato", "Capo Z", "Chiusura", "Reparto Z"],
				"sunday_base_date": "19/07/2026",
				"sunday_row_count": "21",
				"sunday_row_0": ["19/07/2026", "Notte", "Luigi", "Capo B", "Sanificazione", "Reparto B"],
				"sunday_row_20": ["19/07/2026", "Tardo", "Ultima Domenica", "Capo Y", "Controllo", "Reparto Y"],
				"jolly_weekend_title": "Comandata ferragosto",
				"jolly_weekend_base_date": "15/08/2026",
				"jolly_weekend_row_count": "24",
				"jolly_weekend_row_0": ["15/08/2026", "Mattina", "Jolly Inizio", "Capo J", "Presidio", "Reparto J"],
				"jolly_weekend_row_23": ["15/08/2026", "Sera", "Jolly Fine", "Capo K", "Supporto", "Reparto K"],
				"scorrimento_title": "Scorrimento ferragosto",
				"scorrimento_base_date": "15/08/2026",
				"scorrimento_row_count": "25",
				"scorrimento_row_0": ["15/08/2026", "Mattina", "Scorrimento Inizio", "Capo S", "Supporto", "Reparto S"],
				"scorrimento_row_24": ["15/08/2026", "Sera", "Scorrimento Fine", "Capo T", "Chiusura", "Reparto T"],
				"portineria_weekly_headers": ["Portineria Centrale", "Centralinista", "Portineria Cella"],
				"portineria_weekly_time_0": ["06:14", "08:17", "06:14"],
				"portineria_weekly_time_1": ["14:22", "", "14:22"],
				"portineria_weekly_time_2": ["22:06", "", "22:06"],
				"portineria_weekly_row_0_0": ["", "", ""],
				"portineria_weekly_row_0_1": ["", "", ""],
				"portineria_weekly_row_0_2": ["", "", ""],
				"portineria_weekly_row_1_0": ["", "", ""],
				"portineria_weekly_row_1_1": ["", "", ""],
				"portineria_weekly_row_1_2": ["", "", ""],
				"portineria_weekly_row_2_0": ["", "", ""],
				"portineria_weekly_row_2_1": ["", "", ""],
				"portineria_weekly_row_2_2": ["", "", ""],
				"portineria_weekend_base_date": "18/07/2026",
				"portineria_weekend_row_count": "23",
				"portineria_weekend_row_0": ["18/07/2026", "Mattina", "Port A", "Resp A", "Controllo", "Portineria"],
				"portineria_weekend_row_22": ["18/07/2026", "Notte", "Port Ultima", "Resp Ultimo", "Sorveglianza", "Portineria"],
			},
		)

		state.refresh_from_db()
		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={state.week_label}")
		self.assertEqual(len(state.planner_data["saturday"]["rows"]), 22)
		self.assertEqual(len(state.planner_data["sunday"]["rows"]), 21)
		self.assertEqual(len(state.planner_data["jolly_weekend"]["rows"]), 24)
		self.assertEqual(len(state.planner_data["scorrimento"]["rows"]), 25)
		self.assertEqual(len(state.planner_data["portineria_weekend"]["rows"]), 23)
		self.assertEqual(state.planner_data["saturday"]["rows"][21][2], "Ultimo Sabato")
		self.assertEqual(state.planner_data["sunday"]["rows"][20][2], "Ultima Domenica")
		self.assertEqual(state.planner_data["jolly_weekend"]["title"], "Comandata ferragosto")
		self.assertEqual(state.planner_data["jolly_weekend"]["rows"][23][2], "Jolly Fine")
		self.assertEqual(state.planner_data["scorrimento"]["title"], "Scorrimento ferragosto")
		self.assertEqual(state.planner_data["scorrimento"]["rows"][24][2], "Scorrimento Fine")
		self.assertEqual(state.planner_data["portineria_weekend"]["rows"][22][2], "Port Ultima")

	def test_turni_planner_save_sets_employee_visibility_exclusively(self):
		old_state = TurniPlannerWeekState.objects.create(
			week_label="WEEK 21",
			planner_data={},
			visible_to_employees=True,
		)
		new_state = TurniPlannerWeekState.objects.create(
			week_label="WEEK 22",
			planner_data={},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "save_planner",
				"week_label": new_state.week_label,
				"visible_to_employees": "on",
				"weekly_headers": [""] * 10,
				"weekly_time_0": [""] * 10,
				"weekly_time_1": [""] * 10,
				"weekly_time_2": [""] * 10,
				"weekly_time_3": [""] * 10,
				"weekly_row_0_0": [""] * 10,
				"weekly_row_0_1": [""] * 10,
				"weekly_row_0_2": [""] * 10,
				"weekly_row_1_0": [""] * 10,
				"weekly_row_1_1": [""] * 10,
				"weekly_row_1_2": [""] * 10,
				"weekly_row_2_0": [""] * 10,
				"weekly_row_2_1": [""] * 10,
				"weekly_row_2_2": [""] * 10,
				"weekly_row_3_0": [""] * 10,
				"weekly_row_3_1": [""] * 10,
				"weekly_row_3_2": [""] * 10,
				"saturday_base_date": "",
				"sunday_base_date": "",
				"portineria_weekly_headers": [""] * 3,
				"portineria_weekly_time_0": [""] * 3,
				"portineria_weekly_time_1": [""] * 3,
				"portineria_weekly_time_2": [""] * 3,
				"portineria_weekend_base_date": "",
			},
		)

		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={new_state.week_label}")
		old_state.refresh_from_db()
		new_state.refresh_from_db()
		self.assertFalse(old_state.visible_to_employees)
		self.assertTrue(new_state.visible_to_employees)
		self.assertEqual(new_state.planner_data["published_sections"], [])

	def test_turni_planner_save_persists_selected_published_sections(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="WEEK 23",
			planner_data={},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "save_planner",
				"week_label": state.week_label,
				"visible_to_employees": "on",
				"published_sections": ["weekly", "saturday", "sunday"],
				"weekly_headers": [""] * 10,
				"weekly_time_0": [""] * 10,
				"weekly_time_1": [""] * 10,
				"weekly_time_2": [""] * 10,
				"weekly_time_3": [""] * 10,
				"weekly_row_0_0": [""] * 10,
				"weekly_row_0_1": [""] * 10,
				"weekly_row_0_2": [""] * 10,
				"weekly_row_1_0": [""] * 10,
				"weekly_row_1_1": [""] * 10,
				"weekly_row_1_2": [""] * 10,
				"weekly_row_2_0": [""] * 10,
				"weekly_row_2_1": [""] * 10,
				"weekly_row_2_2": [""] * 10,
				"weekly_row_3_0": [""] * 10,
				"weekly_row_3_1": [""] * 10,
				"weekly_row_3_2": [""] * 10,
				"saturday_base_date": "",
				"sunday_base_date": "",
				"portineria_weekly_headers": [""] * 3,
				"portineria_weekly_time_0": [""] * 3,
				"portineria_weekly_time_1": [""] * 3,
				"portineria_weekly_time_2": [""] * 3,
				"portineria_weekend_base_date": "",
			},
		)

		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={state.week_label}")
		state.refresh_from_db()
		self.assertEqual(state.planner_data["published_sections"], ["weekly", "saturday", "sunday"])

	def test_turni_planner_save_without_checkbox_hides_turni_from_employees(self):
		old_state = TurniPlannerWeekState.objects.create(
			week_label="WEEK 21",
			planner_data={},
			visible_to_employees=True,
		)
		new_state = TurniPlannerWeekState.objects.create(
			week_label="WEEK 22",
			planner_data={},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "save_planner",
				"week_label": new_state.week_label,
				"weekly_headers": [""] * 10,
				"weekly_time_0": [""] * 10,
				"weekly_time_1": [""] * 10,
				"weekly_time_2": [""] * 10,
				"weekly_time_3": [""] * 10,
				"weekly_row_0_0": [""] * 10,
				"weekly_row_0_1": [""] * 10,
				"weekly_row_0_2": [""] * 10,
				"weekly_row_1_0": [""] * 10,
				"weekly_row_1_1": [""] * 10,
				"weekly_row_1_2": [""] * 10,
				"weekly_row_2_0": [""] * 10,
				"weekly_row_2_1": [""] * 10,
				"weekly_row_2_2": [""] * 10,
				"weekly_row_3_0": [""] * 10,
				"weekly_row_3_1": [""] * 10,
				"weekly_row_3_2": [""] * 10,
				"saturday_base_date": "",
				"sunday_base_date": "",
				"portineria_weekly_headers": [""] * 3,
				"portineria_weekly_time_0": [""] * 3,
				"portineria_weekly_time_1": [""] * 3,
				"portineria_weekly_time_2": [""] * 3,
				"portineria_weekend_base_date": "",
			},
		)

		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={new_state.week_label}")
		old_state.refresh_from_db()
		new_state.refresh_from_db()
		self.assertFalse(old_state.visible_to_employees)
		self.assertFalse(new_state.visible_to_employees)

	def test_turni_planner_exports_jolly_weekend_pdf_download(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 32 da Lunedi 03/08/2026 a Sabato 08/08/2026",
			planner_data={},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "export_pdf_jolly_weekend",
				"week_label": state.week_label,
				"weekly_export_week_label": state.week_label,
				"portineria_weekly_export_week_label": state.week_label,
				"weekly_headers": [""] * 10,
				"weekly_time_0": [""] * 10,
				"weekly_time_1": [""] * 10,
				"weekly_time_2": [""] * 10,
				"weekly_time_3": [""] * 10,
				"weekly_row_0_0": [""] * 10,
				"weekly_row_0_1": [""] * 10,
				"weekly_row_0_2": [""] * 10,
				"weekly_row_1_0": [""] * 10,
				"weekly_row_1_1": [""] * 10,
				"weekly_row_1_2": [""] * 10,
				"weekly_row_2_0": [""] * 10,
				"weekly_row_2_1": [""] * 10,
				"weekly_row_2_2": [""] * 10,
				"weekly_row_3_0": [""] * 10,
				"weekly_row_3_1": [""] * 10,
				"weekly_row_3_2": [""] * 10,
				"saturday_base_date": "08/08/2026",
				"sunday_base_date": "09/08/2026",
				"jolly_weekend_title": "Comandata primo maggio",
				"jolly_weekend_base_date": "01/05/2026",
				"jolly_weekend_row_0": ["01/05/2026", "Mattina", "Jolly A", "Capo J", "Presidio", "Reparto J"],
				"portineria_weekly_headers": ["Portineria Centrale", "Centralinista", "Portineria Cella"],
				"portineria_weekly_time_0": ["06:14", "08:17", "06:14"],
				"portineria_weekly_time_1": ["14:22", "", "14:22"],
				"portineria_weekly_time_2": ["22:06", "", "22:06"],
				"portineria_weekend_base_date": "08/08/2026",
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response["Content-Type"], "application/pdf")
		self.assertIn("Comandata jolly.pdf", response["Content-Disposition"])
		self.assertTrue(response.content.startswith(b"%PDF"))

	def test_turni_planner_exports_scorrimento_pdf_download(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 32 bis da Lunedi 03/08/2026 a Sabato 08/08/2026",
			planner_data={},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "export_pdf_scorrimento",
				"week_label": state.week_label,
				"weekly_export_week_label": state.week_label,
				"portineria_weekly_export_week_label": state.week_label,
				"weekly_headers": [""] * 10,
				"weekly_time_0": [""] * 10,
				"weekly_time_1": [""] * 10,
				"weekly_time_2": [""] * 10,
				"weekly_time_3": [""] * 10,
				"weekly_row_0_0": [""] * 10,
				"weekly_row_0_1": [""] * 10,
				"weekly_row_0_2": [""] * 10,
				"weekly_row_1_0": [""] * 10,
				"weekly_row_1_1": [""] * 10,
				"weekly_row_1_2": [""] * 10,
				"weekly_row_2_0": [""] * 10,
				"weekly_row_2_1": [""] * 10,
				"weekly_row_2_2": [""] * 10,
				"weekly_row_3_0": [""] * 10,
				"weekly_row_3_1": [""] * 10,
				"weekly_row_3_2": [""] * 10,
				"saturday_base_date": "08/08/2026",
				"sunday_base_date": "09/08/2026",
				"scorrimento_title": "Scorrimento 08/05/2026",
				"scorrimento_base_date": "08/05/2026",
				"scorrimento_row_0": ["08/05/2026", "Mattina", "Mario Rossi", "Capo A", "Scorrimento", "Reparto A"],
				"portineria_weekly_headers": ["Portineria Centrale", "Centralinista", "Portineria Cella"],
				"portineria_weekly_time_0": ["06:14", "08:17", "06:14"],
				"portineria_weekly_time_1": ["14:22", "", "14:22"],
				"portineria_weekly_time_2": ["22:06", "", "22:06"],
				"portineria_weekend_base_date": "08/08/2026",
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response["Content-Type"], "application/pdf")
		self.assertIn("Scorrimento.pdf", response["Content-Disposition"])
		self.assertTrue(response.content.startswith(b"%PDF"))

	def test_turni_planner_generates_mail_with_all_pdf_and_jpg_attachments(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 33 da Lunedi 10/08/2026 a Sabato 15/08/2026",
			planner_data={
				"weekly": {
					"headers": [f"Reparto {index}" for index in range(1, 11)],
					"central_departments": [""] * 10,
					"sections": [
						{"label": "1 turno", "time_values": ["06:00"] * 10, "rows": [["Mario"] * 10, ["Luigi"] * 10, ["Anna"] * 10]},
						{"label": "2 turno", "time_values": ["14:00"] * 10, "rows": [["Paolo"] * 10, ["Gina"] * 10, ["Luca"] * 10]},
						{"label": "3 turno", "time_values": ["22:00"] * 10, "rows": [["Sara"] * 10, ["Piero"] * 10, ["Marta"] * 10]},
						{"label": "turno centrale", "time_values": [""] * 10, "rows": [[""] * 10, [""] * 10, [""] * 10]},
					],
				},
				"portineria_weekly": {
					"headers": ["PORTINERIA CENTRALE", "CENTRALINISTA", "PORTINERIA CELLA"],
					"sections": [
						{"label": "1 turno", "time_values": ["06:14", "08:17", "06:14"], "rows": [["A", "B", "C"], ["", "", ""], ["", "", ""]]},
						{"label": "2 turno", "time_values": ["14:22", "", "14:22"], "rows": [["", "", ""], ["", "", ""], ["", "", ""]]},
						{"label": "3 turno", "time_values": ["22:06", "", "22:06"], "rows": [["", "", ""], ["", "", ""], ["", "", ""]]},
					],
				},
				"saturday": {
					"base_date": "15/08/2026",
					"rows": [["15/08/2026", "Mattina", "Sabato A", "Capo A", "Presidio", "Reparto A"]],
				},
				"sunday": {
					"base_date": "16/08/2026",
					"rows": [["16/08/2026", "Notte", "Domenica A", "Capo B", "Supporto", "Reparto B"]],
				},
				"jolly_weekend": {
					"title": "Comandata ferragosto",
					"base_date": "15/08/2026",
					"rows": [["15/08/2026", "Sera", "Jolly A", "Capo J", "Controllo", "Reparto J"]],
				},
				"scorrimento": {
					"title": "Scorrimento ferragosto",
					"base_date": "15/08/2026",
					"rows": [["15/08/2026", "Mattina", "Mario Rossi", "Capo S", "Scorrimento", "Reparto S"]],
				},
				"portineria_weekend": {
					"base_date": "15/08/2026",
					"rows": [["15/08/2026", "Mattina", "Port A", "Resp A", "Vigilanza", "Portineria"]],
				},
			},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "generate_weekend_email",
				"week_label": state.week_label,
				"mail_recipients": "turni@example.com;caposervizio@example.com",
				"mail_subject": "Turni weekend Ferragosto",
				"mail_body": "Buongiorno team,\ninvio i PDF weekend aggiornati.",
			},
		)

		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={state.week_label}&mail_status=success&mail_message=Email%20inviata%20a%3A%20turni%40example.com%2C%20caposervizio%40example.com")
		self.assertEqual(len(mail.outbox), 1)
		email_message = mail.outbox[0]
		attachment_names = sorted(attachment[0] for attachment in email_message.attachments)
		self.assertEqual(
			attachment_names,
			sorted([
				"Turno settimanale.pdf",
				"Turno settimanale.jpg",
				"Turno settimanale portineria.pdf",
				"Turno settimanale portineria.jpg",
				"Comandata sabato.pdf",
				"Comandata sabato.jpg",
				"Comandata domenica.pdf",
				"Comandata domenica.jpg",
				"Comandata jolly.pdf",
				"Comandata jolly.jpg",
				"Scorrimento.pdf",
				"Scorrimento.jpg",
				"Comandata Sabato - Domenica e festivi Portineria.pdf",
				"Comandata Sabato - Domenica e festivi Portineria.jpg",
			]),
		)
		self.assertEqual(email_message.to, ["turni@example.com", "caposervizio@example.com"])
		self.assertEqual(email_message.subject, "Turni weekend Ferragosto")
		self.assertIn("Buongiorno team,", email_message.body)
		self.assertIn("invio i PDF weekend aggiornati.", email_message.body)
		self.assertIn("Turno settimanale", email_message.body)
		self.assertIn("Turno settimanale portineria", email_message.body)
		self.assertIn("Comandata ferragosto", email_message.body)
		self.assertIn("Scorrimento ferragosto", email_message.body)

	def test_turni_planner_mail_can_send_selected_attachments_only(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 34 da Lunedi 17/08/2026 a Sabato 22/08/2026",
			planner_data={
				"weekly": {
					"headers": [f"Reparto {index}" for index in range(1, 11)],
					"central_departments": [""] * 10,
					"sections": [
						{"label": "1 turno", "time_values": ["06:00"] * 10, "rows": [["Mario"] * 10, ["Luigi"] * 10, ["Anna"] * 10]},
						{"label": "2 turno", "time_values": ["14:00"] * 10, "rows": [["Paolo"] * 10, ["Gina"] * 10, ["Luca"] * 10]},
						{"label": "3 turno", "time_values": ["22:00"] * 10, "rows": [["Sara"] * 10, ["Piero"] * 10, ["Marta"] * 10]},
						{"label": "turno centrale", "time_values": [""] * 10, "rows": [[""] * 10, [""] * 10, [""] * 10]},
					],
				},
				"portineria_weekly": {
					"headers": ["PORTINERIA CENTRALE", "CENTRALINISTA", "PORTINERIA CELLA"],
					"sections": [
						{"label": "1 turno", "time_values": ["06:14", "08:17", "06:14"], "rows": [["A", "B", "C"], ["", "", ""], ["", "", ""]]},
						{"label": "2 turno", "time_values": ["14:22", "", "14:22"], "rows": [["", "", ""], ["", "", ""], ["", "", ""]]},
						{"label": "3 turno", "time_values": ["22:06", "", "22:06"], "rows": [["", "", ""], ["", "", ""], ["", "", ""]]},
					],
				},
				"saturday": {
					"base_date": "22/08/2026",
					"rows": [["22/08/2026", "Mattina", "Sabato A", "Capo A", "Presidio", "Reparto A"]],
				},
				"sunday": {
					"base_date": "23/08/2026",
					"rows": [["23/08/2026", "Notte", "Domenica A", "Capo B", "Supporto", "Reparto B"]],
				},
				"jolly_weekend": {
					"title": "Comandata estate",
					"base_date": "22/08/2026",
					"rows": [["22/08/2026", "Sera", "Jolly A", "Capo J", "Controllo", "Reparto J"]],
				},
				"scorrimento": {
					"title": "Scorrimento estate",
					"base_date": "22/08/2026",
					"rows": [["22/08/2026", "Mattina", "Mario Rossi", "Capo S", "Scorrimento", "Reparto S"]],
				},
				"portineria_weekend": {
					"base_date": "22/08/2026",
					"rows": [["22/08/2026", "Mattina", "Port A", "Resp A", "Vigilanza", "Portineria"]],
				},
			},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "generate_weekend_email",
				"week_label": state.week_label,
				"mail_recipients": "turni@example.com",
				"mail_subject": "Turni selezionati",
				"mail_body": "Invio solo gli allegati richiesti.",
				"mail_attachment": ["scorrimento", "portineria_weekend"],
				"mail_file_type": ["pdf", "jpg"],
			},
		)

		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={state.week_label}&mail_status=success&mail_message=Email%20inviata%20a%3A%20turni%40example.com")
		self.assertEqual(len(mail.outbox), 1)
		email_message = mail.outbox[0]
		attachment_names = sorted(attachment[0] for attachment in email_message.attachments)
		self.assertEqual(
			attachment_names,
			sorted([
				"Scorrimento.pdf",
				"Scorrimento.jpg",
				"Comandata Sabato - Domenica e festivi Portineria.pdf",
				"Comandata Sabato - Domenica e festivi Portineria.jpg",
			]),
		)
		self.assertIn("- Scorrimento estate", email_message.body)
		self.assertIn("- Sabato - Domenica e festivi Portineria", email_message.body)
		self.assertNotIn("Turno settimanale", email_message.body)

	def test_turni_planner_mail_can_send_pdf_only_for_selected_attachments(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 35 da Lunedi 24/08/2026 a Sabato 29/08/2026",
			planner_data={
				"weekly": {
					"headers": [f"Reparto {index}" for index in range(1, 11)],
					"central_departments": [""] * 10,
					"sections": [
						{"label": "1 turno", "time_values": ["06:00"] * 10, "rows": [["Mario"] * 10, ["Luigi"] * 10, ["Anna"] * 10]},
						{"label": "2 turno", "time_values": ["14:00"] * 10, "rows": [["Paolo"] * 10, ["Gina"] * 10, ["Luca"] * 10]},
						{"label": "3 turno", "time_values": ["22:00"] * 10, "rows": [["Sara"] * 10, ["Piero"] * 10, ["Marta"] * 10]},
						{"label": "turno centrale", "time_values": [""] * 10, "rows": [[""] * 10, [""] * 10, [""] * 10]},
					],
				},
				"scorrimento": {
					"title": "Scorrimento fine agosto",
					"base_date": "29/08/2026",
					"rows": [["29/08/2026", "Mattina", "Mario Rossi", "Capo S", "Scorrimento", "Reparto S"]],
				},
			},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "generate_weekend_email",
				"week_label": state.week_label,
				"mail_recipients": "turni@example.com",
				"mail_subject": "Solo PDF",
				"mail_body": "Invio il solo PDF richiesto.",
				"mail_attachment": ["scorrimento"],
				"mail_file_type": ["pdf"],
			},
		)

		self.assertRedirects(response, f"{reverse('turni_planner_home')}?week={state.week_label}&mail_status=success&mail_message=Email%20inviata%20a%3A%20turni%40example.com")
		self.assertEqual(len(mail.outbox), 1)
		email_message = mail.outbox[0]
		attachment_names = sorted(attachment[0] for attachment in email_message.attachments)
		self.assertEqual(attachment_names, ["Scorrimento.pdf"])
		self.assertIn("- Scorrimento fine agosto", email_message.body)

	def test_turni_planner_exports_weekly_pdf_download(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 30 da Lunedi 20/07/2026 a Sabato 25/07/2026",
			planner_data={},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "export_pdf_weekly",
				"week_label": state.week_label,
				"weekly_export_week_label": "Week 30 titolo PDF personalizzato",
				"portineria_weekly_export_week_label": "Week 30 portineria venerdi",
				"weekly_headers": [f"Reparto {index}" for index in range(1, 11)],
				"weekly_time_0": ["06:00"] * 10,
				"weekly_time_1": ["14:00"] * 10,
				"weekly_time_2": ["22:00"] * 10,
				"weekly_time_3": ["00:00"] * 10,
				"weekly_row_0_0": ["Mario"] * 10,
				"weekly_row_0_1": ["Luigi"] * 10,
				"weekly_row_0_2": ["Anna"] * 10,
				"weekly_row_1_0": ["Paolo"] * 10,
				"weekly_row_1_1": ["Gina"] * 10,
				"weekly_row_1_2": ["Luca"] * 10,
				"weekly_row_2_0": ["Sara"] * 10,
				"weekly_row_2_1": ["Piero"] * 10,
				"weekly_row_2_2": ["Marta"] * 10,
				"weekly_row_3_0": ["Notte A"] * 10,
				"weekly_row_3_1": ["Notte B"] * 10,
				"weekly_row_3_2": ["Notte C"] * 10,
				"saturday_base_date": "25/07/2026",
				"sunday_base_date": "26/07/2026",
				"portineria_weekly_headers": ["Portineria Centrale", "Centralinista", "Portineria Cella"],
				"portineria_weekly_time_0": ["06:14", "08:17", "06:14"],
				"portineria_weekly_time_1": ["14:22", "", "14:22"],
				"portineria_weekly_time_2": ["22:06", "", "22:06"],
				"portineria_weekend_base_date": "25/07/2026",
			},
		)

		self.assertEqual(response.status_code, 200)
		state.refresh_from_db()
		self.assertEqual(state.planner_data["weekly_export_week_label"], "Week 30 titolo PDF personalizzato")
		self.assertEqual(state.planner_data["portineria_weekly_export_week_label"], "Week 30 portineria venerdi")
		self.assertEqual(response["Content-Type"], "application/pdf")
		self.assertIn("Turno settimanale.pdf", response["Content-Disposition"])
		self.assertTrue(response.content.startswith(b"%PDF"))

	def test_turni_planner_exports_portineria_weekend_jpg_download(self):
		state = TurniPlannerWeekState.objects.create(
			week_label="Week 31 da Lunedi 27/07/2026 a Sabato 01/08/2026",
			planner_data={},
		)
		self.client.force_login(self.allowed_user)

		response = self.client.post(
			reverse("turni_planner_home"),
			{
				"action": "export_jpg_portineria_weekend",
				"week_label": state.week_label,
				"weekly_headers": [""] * 10,
				"weekly_time_0": [""] * 10,
				"weekly_time_1": [""] * 10,
				"weekly_time_2": [""] * 10,
				"weekly_time_3": [""] * 10,
				"weekly_row_0_0": [""] * 10,
				"weekly_row_0_1": [""] * 10,
				"weekly_row_0_2": [""] * 10,
				"weekly_row_1_0": [""] * 10,
				"weekly_row_1_1": [""] * 10,
				"weekly_row_1_2": [""] * 10,
				"weekly_row_2_0": [""] * 10,
				"weekly_row_2_1": [""] * 10,
				"weekly_row_2_2": [""] * 10,
				"weekly_row_3_0": [""] * 10,
				"weekly_row_3_1": [""] * 10,
				"weekly_row_3_2": [""] * 10,
				"saturday_base_date": "01/08/2026",
				"sunday_base_date": "02/08/2026",
				"portineria_weekly_headers": ["Portineria Centrale", "Centralinista", "Portineria Cella"],
				"portineria_weekly_time_0": ["06:14", "08:17", "06:14"],
				"portineria_weekly_time_1": ["14:22", "", "14:22"],
				"portineria_weekly_time_2": ["22:06", "", "22:06"],
				"portineria_weekend_base_date": "01/08/2026",
				"portineria_weekend_row_0": ["01/08/2026", "Mattina", "Port A", "Resp A", "Controllo", "Portineria"],
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response["Content-Type"], "image/jpeg")
		self.assertIn("Comandata Sabato - Domenica e festivi Portineria.jpg", response["Content-Disposition"])
		self.assertTrue(response.content.startswith(b"\xff\xd8\xff"))


class PayslipUploadImportTests(TestCase):
	@classmethod
	def setUpClass(cls):
		super().setUpClass()
		cls._temp_media = tempfile.TemporaryDirectory()
		cls._override = override_settings(
			MEDIA_ROOT=cls._temp_media.name,
			STORAGES={
				"default": {
					"BACKEND": "django.core.files.storage.FileSystemStorage",
				},
				"staticfiles": {
					"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
				},
			},
		)
		cls._override.enable()

	@classmethod
	def tearDownClass(cls):
		cls._override.disable()
		cls._temp_media.cleanup()
		super().tearDownClass()

	def setUp(self):
		self.client = Client()
		self.admin_user = get_user_model().objects.create_user(
			username="staff.upload",
			password="Password123!",
			is_staff=True,
		)
		self.client.force_login(self.admin_user)

		self.active_user = get_user_model().objects.create_user(
			username="mario.rossi",
			password="Password123!",
			is_active=True,
		)
		self.active_employee = Employee.objects.create(
			user=self.active_user,
			first_name="Mario",
			last_name="Rossi",
		)

		self.inactive_user = get_user_model().objects.create_user(
			username="anna.bianchi",
			password="Password123!",
			is_active=False,
		)
		self.inactive_employee = Employee.objects.create(
			user=self.inactive_user,
			first_name="Anna",
			last_name="Bianchi",
		)

	def _pdf_file(self, name):
		return SimpleUploadedFile(name, b"%PDF-1.4\n%test pdf\n", content_type="application/pdf")

	def test_upload_imports_only_employees_with_active_account(self):
		response = self.client.post(
			reverse("admin_upload_period_folder"),
			{
				"folder": [
					self._pdf_file("Rossi Mario Gennaio 2026.pdf"),
					self._pdf_file("Bianchi Anna Gennaio 2026.pdf"),
					self._pdf_file("Verdi Luca Gennaio 2026.pdf"),
				]
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Conferma Account Mancanti")

		confirm_response = self.client.post(
			reverse("admin_upload_period_folder"),
			{
				"action": "resolve_pending_import",
			},
		)

		self.assertEqual(confirm_response.status_code, 200)
		self.assertEqual(Payslip.objects.filter(employee=self.active_employee, year=2026, month=1).count(), 1)
		self.assertEqual(Payslip.objects.filter(employee=self.inactive_employee, year=2026, month=1).count(), 0)
		self.assertEqual(Payslip.objects.count(), 1)
		self.assertEqual(Employee.objects.count(), 2)

		job = ImportJob.objects.latest("created_at")
		self.assertEqual(job.created_users, 0)
		self.assertEqual(job.created_payslips, 1)
		self.assertEqual(job.skipped, 2)
		self.assertEqual(job.status, "completed")

		self.assertContains(confirm_response, "account non attivo")
		self.assertContains(confirm_response, "account non creato")

	def test_upload_can_create_missing_employee_and_save_payslip(self):
		response = self.client.post(
			reverse("admin_upload_period_folder"),
			{
				"folder": [self._pdf_file("Verdi Luca Gennaio 2026.pdf")]
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Conferma Account Mancanti")

		confirm_response = self.client.post(
			reverse("admin_upload_period_folder"),
			{
				"action": "resolve_pending_import",
				"create_candidates": ["verdi-luca"],
				"first_name_verdi-luca": "Luca",
				"last_name_verdi-luca": "Verdi",
			},
		)

		self.assertEqual(confirm_response.status_code, 200)
		created_employee = Employee.objects.get(last_name="Verdi", first_name="Luca")
		self.assertFalse(created_employee.user.is_active)
		self.assertEqual(Payslip.objects.filter(employee=created_employee, year=2026, month=1).count(), 1)
		self.assertContains(confirm_response, "Account creati: 1")

	def test_upload_prefers_existing_active_registered_employee_when_duplicate_name_exists(self):
		duplicate_user = get_user_model().objects.create_user(
			username="mario.rossi.duplicate",
			password="Password123!",
			is_active=False,
		)
		duplicate_employee = Employee.objects.create(
			user=duplicate_user,
			first_name="Mario",
			last_name="Rossi",
		)

		self.active_employee.privacy_accepted = True
		self.active_employee.save(update_fields=["privacy_accepted"])

		response = self.client.post(
			reverse("admin_upload_period_folder"),
			{
				"folder": [self._pdf_file("Mario Rossi Gennaio 2026.pdf")]
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(Payslip.objects.filter(employee=self.active_employee, year=2026, month=1).count(), 1)
		self.assertEqual(Payslip.objects.filter(employee=duplicate_employee, year=2026, month=1).count(), 0)

	def test_upload_matches_employee_from_username_when_names_are_missing(self):
		username_user = get_user_model().objects.create_user(
			username="daponte-giuseppe",
			password="Password123!",
			is_active=True,
		)
		username_employee = Employee.objects.create(
			user=username_user,
			first_name="",
			last_name="",
		)

		response = self.client.post(
			reverse("admin_upload_period_folder"),
			{
				"folder": [self._pdf_file("D'Aponte Giuseppe Gennaio 2026.pdf")]
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(Payslip.objects.filter(employee=username_employee, year=2026, month=1).count(), 1)


class CudUploadImportTests(TestCase):
	@classmethod
	def setUpClass(cls):
		super().setUpClass()
		cls._temp_media = tempfile.TemporaryDirectory()
		cls._override = override_settings(
			MEDIA_ROOT=cls._temp_media.name,
			STORAGES={
				"default": {
					"BACKEND": "django.core.files.storage.FileSystemStorage",
				},
				"staticfiles": {
					"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
				},
			},
		)
		cls._override.enable()

	@classmethod
	def tearDownClass(cls):
		cls._override.disable()
		cls._temp_media.cleanup()
		super().tearDownClass()

	def setUp(self):
		self.client = Client()
		self.admin_user = get_user_model().objects.create_user(
			username="staff.cud",
			password="Password123!",
			is_staff=True,
		)
		self.client.force_login(self.admin_user)

		self.active_user = get_user_model().objects.create_user(
			username="mario.rossi",
			password="Password123!",
			is_active=True,
		)
		self.active_employee = Employee.objects.create(
			user=self.active_user,
			first_name="Mario",
			last_name="Rossi",
		)

		self.inactive_user = get_user_model().objects.create_user(
			username="anna.bianchi",
			password="Password123!",
			is_active=False,
		)
		self.inactive_employee = Employee.objects.create(
			user=self.inactive_user,
			first_name="Anna",
			last_name="Bianchi",
		)

	def _pdf_file(self, name, content=b"%PDF-1.4\n%test pdf\n"):
		return SimpleUploadedFile(name, content, content_type="application/pdf")

	def test_upload_imports_only_cuds_for_active_accounts(self):
		response = self.client.post(
			reverse("admin_upload_cud"),
			{
				"files": [
					self._pdf_file("CU2026_ROSSI_MARIO.pdf"),
					self._pdf_file("CU2026_BIANCHI_ANNA.pdf"),
					self._pdf_file("CU2026_VERDI_LUCA.pdf"),
				]
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Conferma Account Mancanti")

		confirm_response = self.client.post(
			reverse("admin_upload_cud"),
			{
				"action": "resolve_pending_import",
			},
		)

		self.assertEqual(confirm_response.status_code, 200)
		self.assertEqual(Cud.objects.filter(employee=self.active_employee, year=2026).count(), 1)
		self.assertEqual(Cud.objects.filter(employee=self.inactive_employee, year=2026).count(), 0)
		self.assertEqual(Cud.objects.count(), 1)
		self.assertContains(confirm_response, "account non attivo")
		self.assertContains(confirm_response, "account non creato")

	def test_upload_can_create_missing_employee_and_save_cud(self):
		response = self.client.post(
			reverse("admin_upload_cud"),
			{
				"files": [self._pdf_file("CU2026_VERDI_LUCA.pdf")]
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Conferma Account Mancanti")

		confirm_response = self.client.post(
			reverse("admin_upload_cud"),
			{
				"action": "resolve_pending_import",
				"create_candidates": ["verdi-luca"],
				"first_name_verdi-luca": "Luca",
				"last_name_verdi-luca": "Verdi",
			},
		)

		self.assertEqual(confirm_response.status_code, 200)
		created_employee = Employee.objects.get(last_name="Verdi", first_name="Luca")
		self.assertFalse(created_employee.user.is_active)
		self.assertEqual(Cud.objects.filter(employee=created_employee, year=2026).count(), 1)
		self.assertContains(confirm_response, "Account creati: 1")

	def test_upload_matches_employee_from_username_when_names_are_missing(self):
		username_user = get_user_model().objects.create_user(
			username="daponte-giuseppe",
			password="Password123!",
			is_active=True,
		)
		username_employee = Employee.objects.create(
			user=username_user,
			first_name="",
			last_name="",
		)

		response = self.client.post(
			reverse("admin_upload_cud"),
			{
				"files": [self._pdf_file("CU2026_D'APONTE_GIUSEPPE.pdf")]
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(Cud.objects.filter(employee=username_employee, year=2026).count(), 1)

	def test_upload_replaces_existing_cud_for_same_employee_and_year(self):
		existing = Cud.objects.create(
			employee=self.active_employee,
			year=2026,
			pdf=self._pdf_file("old.pdf", content=b"%PDF-1.4\n%old\n"),
		)

		response = self.client.post(
			reverse("admin_upload_cud"),
			{
				"files": [self._pdf_file("CU2026_ROSSI_MARIO.pdf", content=b"%PDF-1.4\n%new\n")]
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(Cud.objects.filter(employee=self.active_employee, year=2026).count(), 1)
		self.assertFalse(Cud.objects.filter(id=existing.id).exists())
		self.assertContains(response, "Sostituiti: 1")
