from __future__ import annotations

from io import BytesIO
import json
import mimetypes
from pathlib import Path
import re

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from .reporting import generate_batch_report
from .runtime_paths import PRODUCTS_CATALOG_PATH, STATIC_DIR, TEMPLATES_DIR
from .store import (
    REPORTS_DIR,
    active_pallets,
    current_batch,
    delete_batch,
    find_item_by_scan,
    finish_batch,
    get_item_by_pallet,
    get_product_catalog_by_codes,
    get_product_catalog_by_names,
    import_items,
    init_db,
    list_batches,
    list_product_catalog,
    list_items,
    list_items_for_batch,
    mark_waiting_fiche,
    normalize_product_name,
    reset_pallet,
    replace_product_catalog,
    register_incoming,
    register_outgoing,
    summary,
    update_pallet,
    wipe_all_data,
)


app = FastAPI(title="App riconfezionamento")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
init_db()


class IncomingScan(BaseModel):
    code: str
    operator_name: str
    batch_id: int | None = None


class OutgoingScan(BaseModel):
    pallet_code: str
    outgoing_code: str
    outgoing_zun: int
    outgoing_product_code: str
    operator_name: str
    allow_product_code_change: bool = False
    batch_id: int | None = None


class WaitingFicheAction(BaseModel):
    pallet_code: str
    operator_name: str
    batch_id: int | None = None


class CompletedPalletEdit(BaseModel):
    pallet_code: str
    incoming_code: str
    outgoing_code: str
    outgoing_zun: int
    product_code: str = ""
    operator_name: str
    batch_id: int | None = None


class PalletResetRequest(BaseModel):
    batch_id: int | None = None


class AdminWipeRequest(BaseModel):
    password: str


class ProductCatalogEntryRequest(BaseModel):
    product_code: str
    product_name: str


class ProductCatalogResolveRequest(BaseModel):
    current_product_code: str = ""
    current_product_name: str = ""
    product_code: str
    product_name: str
    force: bool = False


class ProductCatalogImportRequest(BaseModel):
    added: int
    existing: int
    conflicts: list[dict[str, str]]


ADMIN_WIPE_PASSWORD = "Unilever_1992"


@app.on_event("startup")
def startup() -> None:
    init_db()


def _normalize_headers(values: tuple[object, ...] | list[object]) -> list[str]:
    return [str(value).strip() if value is not None else "" for value in values]


def _score_header_row(headers: list[str]) -> int:
    normalized = [header.lower() for header in headers if header]
    if not normalized:
        return 0

    score = len(normalized)
    score += sum(
        3
        for header in normalized
        if any(term in header for term in ["fiche", "prodotto", "codice", "motivo", "riconfezionamento", "q.ta", "zun"])
    )
    return score


def _worksheet_values_with_merged_cells(worksheet) -> list[tuple[object, ...]]:
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    if not rows:
        return []

    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = rows[min_row - 1][min_col - 1] if min_row - 1 < len(rows) and min_col - 1 < len(rows[min_row - 1]) else None
        if top_left_value is None:
            continue
        for row_index in range(min_row - 1, min(max_row, len(rows))):
            row_values = rows[row_index]
            for col_index in range(min_col - 1, min(max_col, len(row_values))):
                if row_values[col_index] is None:
                    row_values[col_index] = top_left_value

    return [tuple(row) for row in rows]


def _resolve_sheet_and_header(
    workbook,
    sheet_name: str | None,
    header_row: int,
):
    if sheet_name and sheet_name not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail="Foglio Excel non trovato.")

    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = _worksheet_values_with_merged_cells(worksheet)
    header_index = header_row - 1

    if header_index < 0:
        raise HTTPException(status_code=400, detail="Riga intestazioni non valida.")

    if header_index < len(rows):
        headers = _normalize_headers(rows[header_index])
        if any(headers):
            return worksheet, rows, header_index, headers

    if sheet_name:
        raise HTTPException(status_code=400, detail="Intestazioni non trovate nella riga selezionata.")

    best_match: tuple[int, object, list[tuple[object, ...]], int, list[str]] | None = None
    for candidate_sheet in workbook.worksheets:
        candidate_rows = _worksheet_values_with_merged_cells(candidate_sheet)
        for candidate_index, candidate_row in enumerate(candidate_rows[:25]):
            candidate_headers = _normalize_headers(candidate_row)
            score = _score_header_row(candidate_headers)
            if score <= 0:
                continue
            if best_match is None or score > best_match[0]:
                best_match = (score, candidate_sheet, candidate_rows, candidate_index, candidate_headers)

    if best_match is None:
        raise HTTPException(status_code=400, detail="Intestazioni non trovate nella riga selezionata.")

    _, worksheet, rows, header_index, headers = best_match
    return worksheet, rows, header_index, headers


def read_excel(file_bytes: bytes, header_row: int, sheet_name: str | None) -> tuple[str, int, list[str], list[dict[str, str]]]:
    normalized_sheet_name = (sheet_name or "").strip() or None
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet, rows, header_index, headers = _resolve_sheet_and_header(workbook, normalized_sheet_name, header_row)

    data_rows: list[dict[str, str]] = []
    for row_offset, values in enumerate(rows[header_index + 1 :], start=1):
        record = {
            headers[index]: "" if value is None else str(value).strip()
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        if any(record.values()):
            record["__row_number"] = str(header_index + 1 + row_offset)
            data_rows.append(record)

    return worksheet.title, header_index + 1, headers, data_rows


def parse_zun(value: str) -> int:
    text = value.strip()
    if not text:
        return 0
    try:
        return int(float(text.replace(",", ".")))
    except ValueError:
        return 0


def _normalize_header_label(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _choose_catalog_column(headers: list[str], keywords: tuple[str, ...]) -> str:
    normalized = {header: _normalize_header_label(header) for header in headers if header}
    for header, lower in normalized.items():
        if any(keyword == lower for keyword in keywords):
            return header
    for header, lower in normalized.items():
        if any(keyword in lower for keyword in keywords):
            return header
    return ""


def _ensure_product_catalog_workbook() -> None:
    catalog_path = PRODUCTS_CATALOG_PATH
    catalog_path.parent.mkdir(parents=True, exist_ok=True)
    if catalog_path.exists():
        return

    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = "Prodotti"
        worksheet.append(["Codice prodotto", "Prodotto"])
        workbook.save(catalog_path)
    finally:
        workbook.close()


def _load_product_catalog_rows() -> list[dict[str, str]]:
    _ensure_product_catalog_workbook()
    catalog_path = PRODUCTS_CATALOG_PATH

    workbook = load_workbook(catalog_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=500, detail="Anagrafica prodotti vuota.")

        headers = _normalize_headers(rows[0])
        product_code_column = _choose_catalog_column(headers, ("codice", "codice prodotto", "sku", "articolo", "code", "mrdr"))
        product_name_column = _choose_catalog_column(
            headers,
            ("prodotto", "nome", "nome prodotto", "descrizione", "descrizione prodotto", "mrdr description", "description"),
        )
        if not product_code_column or not product_name_column:
            raise HTTPException(
                status_code=500,
                detail="Impossibile leggere l'anagrafica prodotti: servono colonne codice e prodotto.",
            )

        catalog_rows: list[dict[str, str]] = []
        known_products: dict[str, str] = {}
        for values in rows[1:]:
            record = {
                headers[index]: "" if value is None else str(value).strip()
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
            product_code = record.get(product_code_column, "").strip()
            product_name = record.get(product_name_column, "").strip()
            if not product_code or not product_name:
                continue

            normalized_name = normalize_product_name(product_name)
            previous_name = known_products.get(product_code)
            if previous_name and previous_name != normalized_name:
                continue
            if previous_name == normalized_name:
                continue

            known_products[product_code] = normalized_name
            catalog_rows.append({"product_code": product_code, "product_name": product_name})
        if not catalog_rows:
            raise HTTPException(status_code=500, detail="Anagrafica prodotti senza righe valide.")
        return catalog_rows
    finally:
        workbook.close()


def sync_product_catalog() -> int:
    return replace_product_catalog(_load_product_catalog_rows())


def clear_product_catalog() -> int:
    catalog_path = PRODUCTS_CATALOG_PATH
    catalog_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = "Prodotti"
        worksheet.append(["Codice prodotto", "Prodotto"])
        workbook.save(catalog_path)
    finally:
        workbook.close()

    return replace_product_catalog([])


def sync_product_catalog_for_import() -> int:
    try:
        return sync_product_catalog()
    except HTTPException as exc:
        if exc.detail == "Anagrafica prodotti senza righe valide.":
            return 0
        raise


def add_product_to_catalog(product_code: str, product_name: str) -> dict[str, object]:
    normalized_product_code = str(product_code or "").strip()
    normalized_product_name = str(product_name or "").strip()
    if not normalized_product_code:
        raise HTTPException(status_code=400, detail="Codice prodotto mancante.")
    if not normalized_product_name or normalize_product_name(normalized_product_name) == "prodotto non indicato":
        raise HTTPException(status_code=400, detail="Nome prodotto non valido per l'anagrafica.")

    _ensure_product_catalog_workbook()
    catalog_path = PRODUCTS_CATALOG_PATH

    workbook = load_workbook(catalog_path)
    result: dict[str, object] | None = None
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        headers = _normalize_headers(header_values)
        if not any(headers):
            worksheet.cell(row=1, column=1, value="Codice prodotto")
            worksheet.cell(row=1, column=2, value="Prodotto")
            headers = ["Codice prodotto", "Prodotto"]

        product_code_column = _choose_catalog_column(headers, ("codice", "codice prodotto", "sku", "articolo", "code", "mrdr")) or "Codice prodotto"
        product_name_column = _choose_catalog_column(
            headers,
            ("prodotto", "nome", "nome prodotto", "descrizione", "descrizione prodotto", "mrdr description", "description"),
        ) or "Prodotto"

        header_indexes = {header: index + 1 for index, header in enumerate(headers) if header}
        if product_code_column not in header_indexes:
            product_code_index = max(header_indexes.values(), default=0) + 1
            worksheet.cell(row=1, column=product_code_index, value=product_code_column)
            header_indexes[product_code_column] = product_code_index
        if product_name_column not in header_indexes:
            product_name_index = max(header_indexes.values(), default=0) + 1
            worksheet.cell(row=1, column=product_name_index, value=product_name_column)
            header_indexes[product_name_column] = product_name_index

        product_code_index = header_indexes[product_code_column]
        product_name_index = header_indexes[product_name_column]
        normalized_new_name = normalize_product_name(normalized_product_name)

        for row_index in range(2, worksheet.max_row + 1):
            existing_code = str(worksheet.cell(row=row_index, column=product_code_index).value or "").strip()
            existing_name = str(worksheet.cell(row=row_index, column=product_name_index).value or "").strip()
            if existing_code != normalized_product_code:
                if existing_name and normalize_product_name(existing_name) == normalized_new_name:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"Il prodotto {normalized_product_name} e' gia' presente in anagrafica con codice '{existing_code}'. "
                            "Modifica il codice oppure usa la forzatura."
                        ),
                    )
                continue

            if not existing_name:
                worksheet.cell(row=row_index, column=product_name_index, value=normalized_product_name)
                workbook.save(catalog_path)
                result = {
                    "product_code": normalized_product_code,
                    "product_name": normalized_product_name,
                    "created": False,
                }
                break
            if normalize_product_name(existing_name) == normalized_new_name:
                result = {
                    "product_code": normalized_product_code,
                    "product_name": existing_name,
                    "created": False,
                }
                break
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Il codice {normalized_product_code} e' gia' presente in anagrafica come '{existing_name}'. "
                    "Non posso sovrascriverlo automaticamente."
                ),
            )

        if result is None:
            next_row = worksheet.max_row + 1 if worksheet.max_row >= 1 else 2
            worksheet.cell(row=next_row, column=product_code_index, value=normalized_product_code)
            worksheet.cell(row=next_row, column=product_name_index, value=normalized_product_name)
            workbook.save(catalog_path)
            result = {
                "product_code": normalized_product_code,
                "product_name": normalized_product_name,
                "created": True,
            }
    finally:
        workbook.close()

    sync_product_catalog()
    return result or {
        "product_code": normalized_product_code,
        "product_name": normalized_product_name,
        "created": False,
    }


def resolve_product_catalog_entry(
    current_product_code: str,
    current_product_name: str,
    product_code: str,
    product_name: str,
    *,
    force: bool = False,
) -> dict[str, object]:
    normalized_product_code = str(product_code or "").strip()
    normalized_product_name = str(product_name or "").strip()
    normalized_current_code = str(current_product_code or "").strip()
    normalized_current_name = normalize_product_name(current_product_name or "")
    if not normalized_product_code:
        raise HTTPException(status_code=400, detail="Codice prodotto mancante.")
    if not normalized_product_name or normalize_product_name(normalized_product_name) == "prodotto non indicato":
        raise HTTPException(status_code=400, detail="Nome prodotto non valido per l'anagrafica.")

    if not force:
        return add_product_to_catalog(normalized_product_code, normalized_product_name)

    desired_name = normalize_product_name(normalized_product_name)
    _ensure_product_catalog_workbook()
    catalog_path = PRODUCTS_CATALOG_PATH

    workbook = load_workbook(catalog_path)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        headers = _normalize_headers(header_values)
        if not any(headers):
            worksheet.cell(row=1, column=1, value="Codice prodotto")
            worksheet.cell(row=1, column=2, value="Prodotto")
            headers = ["Codice prodotto", "Prodotto"]

        product_code_column = _choose_catalog_column(headers, ("codice", "codice prodotto", "sku", "articolo", "code", "mrdr")) or "Codice prodotto"
        product_name_column = _choose_catalog_column(
            headers,
            ("prodotto", "nome", "nome prodotto", "descrizione", "descrizione prodotto", "mrdr description", "description"),
        ) or "Prodotto"

        header_indexes = {header: index + 1 for index, header in enumerate(headers) if header}
        if product_code_column not in header_indexes:
            header_indexes[product_code_column] = max(header_indexes.values(), default=0) + 1
            worksheet.cell(row=1, column=header_indexes[product_code_column], value=product_code_column)
        if product_name_column not in header_indexes:
            header_indexes[product_name_column] = max(header_indexes.values(), default=0) + 1
            worksheet.cell(row=1, column=header_indexes[product_name_column], value=product_name_column)

        product_code_index = header_indexes[product_code_column]
        product_name_index = header_indexes[product_name_column]
        keep_row_index = None
        rows_to_delete: list[int] = []

        for row_index in range(2, worksheet.max_row + 1):
            existing_code = str(worksheet.cell(row=row_index, column=product_code_index).value or "").strip()
            existing_name = str(worksheet.cell(row=row_index, column=product_name_index).value or "").strip()
            normalized_existing_name = normalize_product_name(existing_name)
            matches_current = (
                (normalized_current_code and existing_code == normalized_current_code)
                or (normalized_current_name and normalized_existing_name == normalized_current_name)
            )
            matches_desired = existing_code == normalized_product_code or normalized_existing_name == desired_name
            if not matches_current and not matches_desired:
                continue
            if keep_row_index is None:
                keep_row_index = row_index
            else:
                rows_to_delete.append(row_index)

        if keep_row_index is None:
            keep_row_index = worksheet.max_row + 1 if worksheet.max_row >= 1 else 2

        worksheet.cell(row=keep_row_index, column=product_code_index, value=normalized_product_code)
        worksheet.cell(row=keep_row_index, column=product_name_index, value=normalized_product_name)
        for row_index in sorted(rows_to_delete, reverse=True):
            worksheet.delete_rows(row_index, 1)

        workbook.save(catalog_path)
    finally:
        workbook.close()

    sync_product_catalog()
    return {
        "product_code": normalized_product_code,
        "product_name": normalized_product_name,
        "created": True,
        "forced": True,
    }


def _append_product_catalog_entries(rows: list[dict[str, str]]) -> int:
    if not rows:
        return 0

    _ensure_product_catalog_workbook()
    catalog_path = PRODUCTS_CATALOG_PATH

    workbook = load_workbook(catalog_path)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        headers = _normalize_headers(header_values)
        if not any(headers):
            worksheet.cell(row=1, column=1, value="Codice prodotto")
            worksheet.cell(row=1, column=2, value="Prodotto")
            headers = ["Codice prodotto", "Prodotto"]

        product_code_column = _choose_catalog_column(headers, ("codice", "codice prodotto", "sku", "articolo", "code", "mrdr")) or "Codice prodotto"
        product_name_column = _choose_catalog_column(
            headers,
            ("prodotto", "nome", "nome prodotto", "descrizione", "descrizione prodotto", "mrdr description", "description"),
        ) or "Prodotto"

        header_indexes = {header: index + 1 for index, header in enumerate(headers) if header}
        if product_code_column not in header_indexes:
            product_code_index = max(header_indexes.values(), default=0) + 1
            worksheet.cell(row=1, column=product_code_index, value=product_code_column)
            header_indexes[product_code_column] = product_code_index
        if product_name_column not in header_indexes:
            product_name_index = max(header_indexes.values(), default=0) + 1
            worksheet.cell(row=1, column=product_name_index, value=product_name_column)
            header_indexes[product_name_column] = product_name_index

        product_code_index = header_indexes[product_code_column]
        product_name_index = header_indexes[product_name_column]

        next_row = worksheet.max_row + 1 if worksheet.max_row >= 1 else 2
        added = 0
        for row in rows:
            worksheet.cell(row=next_row, column=product_code_index, value=str(row["product_code"]).strip())
            worksheet.cell(row=next_row, column=product_name_index, value=str(row["product_name"]).strip())
            next_row += 1
            added += 1

        workbook.save(catalog_path)
        return added
    finally:
        workbook.close()


def validate_product_catalog_for_rows(
    rows: list[dict[str, str]],
    product_code_column: str,
    product_column: str,
) -> tuple[dict[str, dict[str, str]], dict[str, list[dict[str, str]]]]:
    if not product_column.strip():
        raise HTTPException(
            status_code=400,
            detail="Seleziona anche la colonna del prodotto per verificare la corrispondenza codice-prodotto.",
        )

    product_codes = [row.get(product_code_column, "").strip() for row in rows if product_code_column]
    product_names = [row.get(product_column, "").strip() for row in rows if product_column]
    return get_product_catalog_by_codes(product_codes), get_product_catalog_by_names(product_names)


def _load_catalog_import_rows(file_bytes: bytes, header_row: int, sheet_name: str | None) -> list[dict[str, str]]:
    _, _, headers, rows = read_excel(file_bytes, header_row, sheet_name)
    product_code_column = _choose_catalog_column(headers, ("codice", "codice prodotto", "sku", "articolo", "code", "mrdr"))
    product_name_column = _choose_catalog_column(
        headers,
        ("prodotto", "nome", "nome prodotto", "descrizione", "descrizione prodotto", "mrdr description", "description"),
    )
    if not product_code_column or not product_name_column:
        raise HTTPException(status_code=400, detail="Nel file anagrafica servono le colonne codice e prodotto.")

    catalog_rows: list[dict[str, str]] = []
    for row in rows:
        product_code = row.get(product_code_column, "").strip()
        product_name = row.get(product_name_column, "").strip()
        if not product_code or not product_name:
            continue
        catalog_rows.append({"product_code": product_code, "product_name": product_name})
    if not catalog_rows:
        if not rows:
            raise HTTPException(status_code=400, detail="Il file anagrafica e' vuoto.")
        raise HTTPException(status_code=400, detail="Nel file anagrafica non ci sono righe valide.")
    return catalog_rows


def import_product_catalog_rows(rows: list[dict[str, str]]) -> ProductCatalogImportRequest:
    try:
        sync_product_catalog()
    except HTTPException as exc:
        if exc.detail != "Anagrafica prodotti senza righe valide.":
            raise
    existing_by_code = get_product_catalog_by_codes([row["product_code"] for row in rows])
    existing_by_name = get_product_catalog_by_names([row["product_name"] for row in rows])

    added = 0
    existing = 0
    conflicts: list[dict[str, str]] = []
    rows_to_add: list[dict[str, str]] = []
    for row in rows:
        product_code = row["product_code"]
        product_name = row["product_name"]
        normalized_name = normalize_product_name(product_name)
        existing_code_entry = existing_by_code.get(product_code)
        if existing_code_entry:
            if existing_code_entry["normalized_product_name"] == normalized_name:
                existing += 1
                continue
            conflicts.append({
                "product_code": product_code,
                "product_name": product_name,
                "current_product_code": existing_code_entry["product_code"],
                "current_product_name": existing_code_entry["product_name"],
                "conflict_type": "code_already_present",
                "message": f"Codice gia' presente con prodotto '{existing_code_entry['product_name']}'.",
            })
            continue
        same_name_entries = existing_by_name.get(normalized_name, [])
        if same_name_entries:
            conflicts.append({
                "product_code": product_code,
                "product_name": product_name,
                "current_product_code": same_name_entries[0]["product_code"],
                "current_product_name": same_name_entries[0]["product_name"],
                "conflict_type": "product_already_present",
                "message": f"Prodotto gia' presente con codice '{same_name_entries[0]['product_code']}'.",
            })
            continue
        rows_to_add.append({"product_code": product_code, "product_name": product_name})
        existing_by_code[product_code] = {
            "product_code": product_code,
            "product_name": product_name,
            "normalized_product_name": normalized_name,
            "synced_at": "",
        }
        existing_by_name.setdefault(normalized_name, []).append(existing_by_code[product_code])
        added += 1

    if rows_to_add:
        _append_product_catalog_entries(rows_to_add)
        sync_product_catalog()

    return ProductCatalogImportRequest(added=added, existing=existing, conflicts=conflicts)


def resolve_reason_column(headers: list[str], requested_column: str) -> str:
    requested = requested_column.strip()
    normalized = {header: header.lower() for header in headers if header}
    preferred = [
        header
        for header, lower in normalized.items()
        if lower.startswith("motivo") or "motivo " in lower or "motivo_" in lower
    ]
    if not preferred:
        preferred = [
            header
            for header, lower in normalized.items()
            if "riconfezionamento" in lower and "costo" not in lower
        ]

    if preferred:
        if requested in preferred:
            return requested
        if not requested or "nota" in requested.lower() or "costo" in requested.lower():
            return preferred[0]

    return requested


def parse_row_actions(raw_actions: str) -> dict[int, dict[str, object]]:
    if not raw_actions.strip():
        return {}

    try:
        payload = json.loads(raw_actions)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Formato correzioni righe non valido.") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Formato correzioni righe non valido.")

    normalized: dict[int, dict[str, object]] = {}
    for raw_key, raw_value in payload.items():
        try:
            row_number = int(str(raw_key).strip())
        except ValueError:
            continue
        if not isinstance(raw_value, dict):
            continue

        normalized[row_number] = {
            "reason": str(raw_value.get("reason", "")).strip(),
            "discard": bool(raw_value.get("discard", False)),
        }

    return normalized


def validate_import_columns(
    headers: list[str],
    incoming_column: str,
    pallet_column: str,
    outgoing_column: str,
    product_column: str,
    product_code_column: str,
    reason_column: str,
    production_lot_column: str,
    zun_column: str,
) -> str:
    if not reason_column.strip():
        raise HTTPException(status_code=400, detail="Seleziona la colonna del motivo riconfezionamento.")
    if not product_column.strip():
        raise HTTPException(status_code=400, detail="Seleziona la colonna del prodotto.")
    if not product_code_column.strip():
        raise HTTPException(status_code=400, detail="Seleziona la colonna del codice prodotto.")
    if not production_lot_column.strip():
        raise HTTPException(status_code=400, detail="Seleziona la colonna del lotto di produzione.")

    resolved_reason_column = resolve_reason_column(headers, reason_column)
    missing_columns = [
        column
        for column in [incoming_column, pallet_column, outgoing_column, product_column, product_code_column, resolved_reason_column, production_lot_column, zun_column]
        if column and column not in headers
    ]
    if missing_columns:
        raise HTTPException(status_code=400, detail=f"Colonne non trovate: {', '.join(missing_columns)}")

    return resolved_reason_column


def build_import_rows(
    rows: list[dict[str, str]],
    pallet_column: str,
    incoming_column: str,
    outgoing_column: str,
    product_column: str,
    product_code_column: str,
    reason_column: str,
    production_lot_column: str,
    zun_column: str,
    row_actions: dict[int, dict[str, object]] | None = None,
    product_catalog_by_code: dict[str, dict[str, str]] | None = None,
    product_catalog_by_name: dict[str, list[dict[str, str]]] | None = None,
    strict_empty: bool = True,
) -> tuple[list[dict[str, str | int]], list[dict[str, str]], list[dict[str, str]]]:
    imported_rows: list[dict[str, str | int]] = []
    last_product_name = ""
    last_product_code = ""
    last_repackaging_reason = ""
    missing_reason_rows: list[dict[str, str]] = []
    missing_product_code_rows: list[dict[str, str]] = []
    missing_production_lot_rows: list[dict[str, str]] = []
    mismatched_product_rows: list[dict[str, str]] = []
    catalog_rows_to_add: list[dict[str, str]] = []
    actions = row_actions or {}
    catalog_by_code = dict(product_catalog_by_code or {})
    catalog_by_name = {key: list(value) for key, value in (product_catalog_by_name or {}).items()}
    for row in rows:
        try:
            row_number = int(row.get("__row_number", "0") or 0)
        except ValueError:
            row_number = 0
        row_action = actions.get(row_number, {})
        selected_pallet_value = row.get(pallet_column, "").strip() if pallet_column else ""
        incoming_fiche = row.get(incoming_column, "").strip()
        outgoing_fiche = row.get(outgoing_column, "").strip() if outgoing_column else ""
        raw_product_name = row.get(product_column, "").strip() if product_column else ""
        raw_product_code = row.get(product_code_column, "").strip() if product_code_column else ""
        production_lot = row.get(production_lot_column, "").strip() if production_lot_column else ""
        raw_repackaging_reason = str(row_action.get("reason", "")).strip() or (row.get(reason_column, "").strip() if reason_column else "")
        product_name = raw_product_name or last_product_name
        product_code = raw_product_code or last_product_code
        if raw_repackaging_reason:
            repackaging_reason = raw_repackaging_reason
        elif product_name and product_name == last_product_name:
            repackaging_reason = last_repackaging_reason
        else:
            repackaging_reason = ""
        zun_quantity = parse_zun(row.get(zun_column, "").strip()) if zun_column else 0
        if not incoming_fiche or re.search(r"\d", incoming_fiche) is None:
            continue
        # The operator flow needs a unique row key. The incoming fiche is the only
        # stable per-row identifier, while product codes can repeat across many rows.
        pallet_code = incoming_fiche
        if not pallet_code:
            continue
        if bool(row_action.get("discard", False)):
            continue
        if not repackaging_reason:
            missing_reason_rows.append(
                {
                    "row_number": str(row_number),
                    "fiche": incoming_fiche or pallet_code,
                    "pallet": selected_pallet_value or pallet_code,
                    "product_name": product_name or "prodotto non indicato",
                    "zun_quantity": str(zun_quantity),
                    "reason": "",
                }
            )
            continue
        if not product_code:
            missing_product_code_rows.append(
                {
                    "row_number": str(row_number),
                    "fiche": incoming_fiche or pallet_code,
                    "pallet": selected_pallet_value or pallet_code,
                    "product_name": product_name or "prodotto non indicato",
                    "zun_quantity": str(zun_quantity),
                    "reason": repackaging_reason or "-",
                }
            )
            continue
        if not production_lot:
            missing_production_lot_rows.append(
                {
                    "row_number": str(row_number),
                    "fiche": incoming_fiche or pallet_code,
                    "pallet": selected_pallet_value or pallet_code,
                    "product_name": product_name or "prodotto non indicato",
                    "zun_quantity": str(zun_quantity),
                    "reason": repackaging_reason or "-",
                }
            )
            continue

        catalog_entry = catalog_by_code.get(product_code)
        normalized_row_product_name = normalize_product_name(product_name)
        if catalog_entry is None:
            if not normalized_row_product_name or normalized_row_product_name == "prodotto non indicato":
                mismatched_product_rows.append(
                    {
                        "row_number": str(row_number),
                        "fiche": incoming_fiche or pallet_code,
                        "pallet": selected_pallet_value or pallet_code,
                        "product_name": product_name or "prodotto non indicato",
                        "zun_quantity": str(zun_quantity),
                        "reason": repackaging_reason or "-",
                        "expected_product_name": "prodotto non presente in anagrafica",
                        "product_code": product_code,
                        "catalog_missing": True,
                    }
                )
                continue
            same_name_entries = catalog_by_name.get(normalized_row_product_name, [])
            expected_name = "codice non presente in anagrafica"
            catalog_missing = True
            if same_name_entries:
                expected_name = f"prodotto gia' presente con codice {same_name_entries[0]['product_code']}"
                catalog_missing = False
            if same_name_entries:
                mismatched_product_rows.append(
                    {
                        "row_number": str(row_number),
                        "fiche": incoming_fiche or pallet_code,
                        "pallet": selected_pallet_value or pallet_code,
                        "product_name": product_name or "prodotto non indicato",
                        "zun_quantity": str(zun_quantity),
                        "reason": repackaging_reason or "-",
                        "expected_product_name": expected_name,
                        "product_code": product_code,
                        "catalog_missing": catalog_missing,
                    }
                )
                continue

            new_catalog_entry = {
                "product_code": product_code,
                "product_name": product_name,
                "normalized_product_name": normalized_row_product_name,
                "synced_at": "",
            }
            catalog_by_code[product_code] = new_catalog_entry
            catalog_by_name.setdefault(normalized_row_product_name, []).append(new_catalog_entry)
            catalog_rows_to_add.append({"product_code": product_code, "product_name": product_name})
            catalog_entry = new_catalog_entry
        if not normalized_row_product_name or normalized_row_product_name != catalog_entry["normalized_product_name"]:
            mismatched_product_rows.append(
                {
                    "row_number": str(row_number),
                    "fiche": incoming_fiche or pallet_code,
                    "pallet": selected_pallet_value or pallet_code,
                    "product_name": product_name or "prodotto non indicato",
                    "zun_quantity": str(zun_quantity),
                    "reason": repackaging_reason or "-",
                    "expected_product_name": catalog_entry["product_name"],
                    "product_code": product_code,
                    "catalog_missing": False,
                }
            )
            continue

        if product_name:
            last_product_name = product_name
        if product_code:
            last_product_code = product_code
        if repackaging_reason:
            last_repackaging_reason = repackaging_reason

        imported_rows.append(
            {
                "pallet_code": pallet_code,
                "incoming_fiche": incoming_fiche,
                "outgoing_fiche": outgoing_fiche,
                "product_name": product_name,
                "product_code": product_code,
                "production_lot": production_lot,
                "repackaging_reason": repackaging_reason,
                "zun_quantity": zun_quantity,
                "manual_reason_override": 1 if str(row_action.get("reason", "")).strip() else 0,
            }
        )

    if not imported_rows and strict_empty:
        if missing_product_code_rows:
            preview = ", ".join(row["fiche"] for row in missing_product_code_rows[:4])
            product_preview = missing_product_code_rows[0]["product_name"]
            raise HTTPException(
                status_code=400,
                detail={
                    "message": (
                        f"Codice prodotto mancante per {len(missing_product_code_rows)} righe del lotto. "
                        f"Prodotto coinvolto: {product_preview}. Fiches: {preview}. Correggi il file Excel e reimporta."
                    ),
                    "skipped_rows": missing_product_code_rows,
                },
            )
        if missing_reason_rows:
            preview = ", ".join(row["fiche"] for row in missing_reason_rows[:4])
            product_preview = missing_reason_rows[0]["product_name"]
            raise HTTPException(
                status_code=400,
                detail={
                    "message": (
                        f"Motivo riconfezionamento mancante per {len(missing_reason_rows)} righe del lotto. "
                        f"Prodotto coinvolto: {product_preview}. Fiches: {preview}."
                    ),
                    "skipped_rows": missing_reason_rows,
                },
            )
        if missing_production_lot_rows:
            preview = ", ".join(row["fiche"] for row in missing_production_lot_rows[:4])
            product_preview = missing_production_lot_rows[0]["product_name"]
            raise HTTPException(
                status_code=400,
                detail={
                    "message": (
                        f"Lotto di produzione mancante per {len(missing_production_lot_rows)} righe del lotto. "
                        f"Prodotto coinvolto: {product_preview}. Fiches: {preview}."
                    ),
                    "skipped_rows": missing_production_lot_rows,
                },
            )
        raise HTTPException(status_code=400, detail="Nessun pallet valido trovato nel file Excel.")
    if missing_product_code_rows:
        raise HTTPException(
            status_code=400,
            detail={
                "message": (
                    f"Ci sono {len(missing_product_code_rows)} righe senza codice prodotto. "
                    "Correggi il file Excel e reimporta il lotto."
                ),
                "skipped_rows": missing_product_code_rows,
            },
        )
    if missing_production_lot_rows:
        raise HTTPException(
            status_code=400,
            detail={
                "message": (
                    f"Ci sono {len(missing_production_lot_rows)} righe senza lotto di produzione. "
                    "Correggi il file Excel e reimporta il lotto."
                ),
                "skipped_rows": missing_production_lot_rows,
            },
        )
    if mismatched_product_rows:
        preview = ", ".join(row["fiche"] for row in mismatched_product_rows[:4])
        first_row = mismatched_product_rows[0]
        raise HTTPException(
            status_code=400,
            detail={
                "error_code": "product_catalog_mismatch",
                "message": (
                    f"Ci sono {len(mismatched_product_rows)} righe con codice prodotto non coerente con l'anagrafica aggiornata. "
                    f"Codice {first_row['product_code']}: Excel '{first_row['product_name']}', anagrafica '{first_row['expected_product_name']}'. "
                    f"Fiches: {preview}."
                ),
                "mismatch_rows": mismatched_product_rows,
            },
        )
    return imported_rows, missing_reason_rows, catalog_rows_to_add


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    base_path = str(request.scope.get("root_path") or "").rstrip("/")
    html = (TEMPLATES_DIR / "index.html").read_text(encoding="utf-8")
    return HTMLResponse(html.replace("__APP_BASE_PATH__", base_path))


@app.post("/api/preview")
async def preview_excel(
    file: UploadFile = File(...),
    header_row: int = Form(1),
    sheet_name: str = Form(""),
) -> dict[str, object]:
    content = await file.read()
    resolved_sheet_name, resolved_header_row, headers, rows = read_excel(content, header_row, sheet_name or None)
    return {
        "filename": file.filename,
        "sheet_names": load_workbook(BytesIO(content), read_only=True).sheetnames,
        "resolved_sheet_name": resolved_sheet_name,
        "resolved_header_row": resolved_header_row,
        "headers": headers,
        "preview": rows[:5],
        "row_count": len(rows),
    }


@app.post("/api/import/check")
async def check_import_excel(
    file: UploadFile = File(...),
    pallet_column: str = Form(""),
    incoming_column: str = Form(...),
    outgoing_column: str = Form(""),
    product_column: str = Form(""),
    product_code_column: str = Form(""),
    reason_column: str = Form(...),
    production_lot_column: str = Form(...),
    zun_column: str = Form(""),
    header_row: int = Form(1),
    sheet_name: str = Form(""),
) -> dict[str, object]:
    content = await file.read()
    _, _, headers, rows = read_excel(content, header_row, sheet_name or None)
    sync_product_catalog_for_import()
    reason_column = validate_import_columns(
        headers,
        incoming_column,
        pallet_column,
        outgoing_column,
        product_column,
        product_code_column,
        reason_column,
        production_lot_column,
        zun_column,
    )
    product_catalog_by_code, product_catalog_by_name = validate_product_catalog_for_rows(rows, product_code_column, product_column)

    imported_rows, skipped_rows, catalog_rows_to_add = build_import_rows(
        rows,
        pallet_column,
        incoming_column,
        outgoing_column,
        product_column,
        product_code_column,
        reason_column,
        production_lot_column,
        zun_column,
        strict_empty=False,
        product_catalog_by_code=product_catalog_by_code,
        product_catalog_by_name=product_catalog_by_name,
    )
    if not imported_rows and not skipped_rows:
        raise HTTPException(status_code=400, detail="Nessun pallet valido trovato nel file Excel.")

    if skipped_rows:
        return {
            "message": f"Controllo completato: {len(skipped_rows)} righe da correggere o scartare prima dell'import.",
            "issues": skipped_rows,
            "valid_rows": len(imported_rows),
        }

    return {
        "message": (
            f"Controllo completato: {len(imported_rows)} righe pronte per l'import. "
            f"{len(catalog_rows_to_add)} nuovi codici saranno aggiunti in anagrafica."
            if catalog_rows_to_add
            else f"Controllo completato: {len(imported_rows)} righe pronte per l'import."
        ),
        "issues": [],
        "valid_rows": len(imported_rows),
    }


@app.post("/api/import")
async def import_excel(
    file: UploadFile = File(...),
    pallet_column: str = Form(""),
    incoming_column: str = Form(...),
    outgoing_column: str = Form(""),
    product_column: str = Form(""),
    product_code_column: str = Form(""),
    reason_column: str = Form(...),
    production_lot_column: str = Form(...),
    zun_column: str = Form(""),
    header_row: int = Form(1),
    sheet_name: str = Form(""),
    row_actions: str = Form(""),
) -> dict[str, object]:
    content = await file.read()
    _, _, headers, rows = read_excel(content, header_row, sheet_name or None)
    sync_product_catalog_for_import()

    reason_column = validate_import_columns(
        headers,
        incoming_column,
        pallet_column,
        outgoing_column,
        product_column,
        product_code_column,
        reason_column,
        production_lot_column,
        zun_column,
    )
    parsed_row_actions = parse_row_actions(row_actions)
    product_catalog_by_code, product_catalog_by_name = validate_product_catalog_for_rows(rows, product_code_column, product_column)

    imported_rows, skipped_rows, catalog_rows_to_add = build_import_rows(
        rows,
        pallet_column,
        incoming_column,
        outgoing_column,
        product_column,
        product_code_column,
        reason_column,
        production_lot_column,
        zun_column,
        row_actions=parsed_row_actions,
        product_catalog_by_code=product_catalog_by_code,
        product_catalog_by_name=product_catalog_by_name,
    )
    if skipped_rows:
        raise HTTPException(
            status_code=400,
            detail={
                "message": f"Ci sono ancora {len(skipped_rows)} righe da correggere o scartare prima dell'import.",
                "skipped_rows": skipped_rows,
            },
        )

    if catalog_rows_to_add:
        _append_product_catalog_entries(catalog_rows_to_add)
        sync_product_catalog()

    result = import_items(file.filename or "lotto.xlsx", imported_rows)
    discarded_count = sum(1 for action in parsed_row_actions.values() if bool(action.get("discard", False)))
    message = f"Lotto importato: {result['total_items']} pedane registrate."
    if discarded_count:
        message += f" {discarded_count} righe scartate su richiesta."
    if catalog_rows_to_add:
        message += f" Anagrafica arricchita con {len(catalog_rows_to_add)} nuovi codici-prodotto."
    return {
        "message": message,
        "summary": summary(result["batch_id"]),
        "skipped_rows": [],
        "partial_import": False,
    }


@app.post("/api/product-catalog")
def create_product_catalog_entry(payload: ProductCatalogEntryRequest) -> dict[str, object]:
    result = add_product_to_catalog(payload.product_code, payload.product_name)
    if bool(result.get("created")):
        message = f"Codice {result['product_code']} aggiunto in anagrafica prodotti."
    else:
        message = f"Codice {result['product_code']} gia' presente in anagrafica prodotti."
    return {
        "message": message,
        "entry": result,
    }


@app.post("/api/product-catalog/resolve")
def resolve_product_catalog_conflict(payload: ProductCatalogResolveRequest) -> dict[str, object]:
    result = resolve_product_catalog_entry(
        payload.current_product_code,
        payload.current_product_name,
        payload.product_code,
        payload.product_name,
        force=payload.force,
    )
    action = "forzata" if payload.force else "aggiornata"
    return {
        "message": f"Anagrafica {action}: {result['product_code']} - {result['product_name']}",
        "entry": result,
    }


@app.get("/api/product-catalog")
def get_product_catalog(limit: int = 500) -> dict[str, object]:
    rows = list_product_catalog(limit=limit)
    return {
        "rows": rows,
        "count": len(rows),
    }


@app.get("/api/product-catalog/download")
def download_product_catalog() -> FileResponse:
    _ensure_product_catalog_workbook()
    catalog_path = PRODUCTS_CATALOG_PATH.resolve()
    media_type, _ = mimetypes.guess_type(catalog_path.name)
    return FileResponse(
        path=catalog_path,
        media_type=media_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=catalog_path.name,
    )


@app.post("/api/product-catalog/clear")
def clear_product_catalog_endpoint() -> dict[str, object]:
    clear_product_catalog()
    return {
        "message": "Anagrafica prodotti cancellata.",
        "rows": [],
        "count": 0,
    }


@app.post("/api/product-catalog/import")
async def import_product_catalog_excel(
    file: UploadFile = File(...),
    header_row: int = Form(1),
    sheet_name: str = Form(""),
) -> dict[str, object]:
    content = await file.read()
    result = import_product_catalog_rows(_load_catalog_import_rows(content, header_row, sheet_name or None))
    message = f"Anagrafica aggiornata. Nuovi codici: {result.added}. Gia' presenti: {result.existing}."
    if result.conflicts:
        message += f" Conflitti: {len(result.conflicts)}."
    return {
        "message": message,
        "added": result.added,
        "existing": result.existing,
        "conflicts": result.conflicts,
    }


@app.delete("/api/batches/current")
def remove_current_batch() -> dict[str, object]:
    batch = current_batch()
    if batch is None:
        raise HTTPException(status_code=404, detail="Nessun lotto da cancellare.")

    deleted = delete_batch(int(batch["id"]))
    if deleted is None:
        raise HTTPException(status_code=404, detail="Nessun lotto da cancellare.")

    return {
        "message": f"Lotto eliminato: {deleted['filename']}",
        "summary": summary(),
        "items": list_items(),
        "active_pallets": active_pallets(),
    }


@app.post("/api/admin/wipe-all")
def admin_wipe_all(payload: AdminWipeRequest) -> dict[str, object]:
    if payload.password != ADMIN_WIPE_PASSWORD:
        raise HTTPException(status_code=403, detail="Password amministratore non valida.")

    deleted = wipe_all_data()
    return {
        "message": (
            "Tutti i lotti importati e i backup disponibili sono stati cancellati. "
            "L'anagrafica prodotti e' stata mantenuta. "
            f"Report eliminati: {deleted['reports_deleted']}. Backup eliminati: {deleted['backups_deleted']}."
        ),
        "summary": summary(),
        "items": list_items(),
        "active_pallets": active_pallets(),
        "batches": list_batches(),
    }


@app.get("/api/dashboard")
def dashboard(batch_id: int | None = None) -> dict[str, object]:
    active_batch = current_batch()
    selected_batch_id = batch_id if batch_id is not None else (int(active_batch["id"]) if active_batch else None)
    return {
        "current_batch": active_batch,
        "selected_batch_id": selected_batch_id,
        "summary": summary(selected_batch_id),
        "items": list_items(batch_id=selected_batch_id),
        "active_pallets": active_pallets(batch_id=selected_batch_id),
        "batches": list_batches(),
        "can_operate": bool(
            active_batch
            and selected_batch_id == int(active_batch["id"])
            and active_batch.get("completed_at") is None
        ),
    }


@app.get("/api/items/{pallet_code}")
def item_detail(pallet_code: str, batch_id: int | None = None) -> dict[str, object]:
    item = get_item_by_pallet(pallet_code.strip(), batch_id=batch_id)
    if item is None and batch_id is None:
        item = find_item_by_scan(pallet_code.strip())
    if item is None:
        raise HTTPException(status_code=404, detail="Pallet non presente.")
    return {"item": item}


@app.post("/api/scan/incoming")
def scan_incoming(payload: IncomingScan) -> dict[str, object]:
    success, message, item = register_incoming(
        payload.code.strip(),
        payload.operator_name.strip(),
        batch_id=payload.batch_id,
    )
    if not success:
        raise HTTPException(status_code=400, detail={"message": message, "item": item})
    return {"message": message, "item": item, "summary": summary(), "active_pallets": active_pallets()}


@app.post("/api/scan/waiting-fiche")
def set_waiting_fiche(payload: WaitingFicheAction) -> dict[str, object]:
    success, message, item = mark_waiting_fiche(
        payload.pallet_code.strip(),
        payload.operator_name.strip(),
        batch_id=payload.batch_id,
    )
    if not success:
        raise HTTPException(status_code=400, detail={"message": message, "item": item})
    return {"message": message, "item": item, "summary": summary(), "active_pallets": active_pallets()}


@app.post("/api/scan/outgoing")
def scan_outgoing(payload: OutgoingScan) -> dict[str, object]:
    success, message, item, error_code = register_outgoing(
        payload.pallet_code.strip(),
        payload.outgoing_code.strip(),
        payload.outgoing_zun,
        payload.outgoing_product_code.strip(),
        payload.operator_name.strip(),
        allow_product_code_change=payload.allow_product_code_change,
        batch_id=payload.batch_id,
    )
    if not success:
        raise HTTPException(status_code=400, detail={"message": message, "item": item, "error_code": error_code})
    return {"message": message, "item": item, "summary": summary(), "active_pallets": active_pallets()}


@app.put("/api/items/{pallet_code}/completed")
def edit_completed_pallet(pallet_code: str, payload: CompletedPalletEdit) -> dict[str, object]:
    success, message, item = update_pallet(
        pallet_code.strip(),
        payload.pallet_code.strip(),
        payload.incoming_code.strip(),
        payload.outgoing_code.strip(),
        payload.outgoing_zun,
        payload.product_code.strip(),
        payload.operator_name.strip(),
        batch_id=payload.batch_id,
    )
    if not success:
        raise HTTPException(status_code=400, detail={"message": message, "item": item})
    return {"message": message, "item": item, "summary": summary(payload.batch_id), "active_pallets": active_pallets(payload.batch_id)}


@app.post("/api/items/{pallet_code}/reset")
def reset_item(pallet_code: str, payload: PalletResetRequest) -> dict[str, object]:
    success, message, item = reset_pallet(pallet_code.strip(), batch_id=payload.batch_id)
    if not success:
        raise HTTPException(status_code=400, detail={"message": message, "item": item})
    return {"message": message, "item": item, "summary": summary(payload.batch_id), "active_pallets": active_pallets(payload.batch_id)}


@app.post("/api/batches/current/finish")
def close_current_batch() -> dict[str, object]:
    batch = current_batch()
    if batch is None:
        raise HTTPException(status_code=400, detail="Nessun lotto disponibile.")

    batch_id = int(batch["id"])
    batch_summary = summary(batch_id)
    batch_items = list_items_for_batch(batch_id)
    report_path = generate_batch_report(REPORTS_DIR, batch, batch_items, batch_summary)
    finish_batch(str(report_path), batch_id=batch_id)
    updated_summary = summary(batch_id)
    return {
        "message": "Report Excel lotto generato con successo.",
        "summary": updated_summary,
        "report_url": f"/api/reports/{report_path.name}",
    }


@app.get("/api/reports/{report_name}")
def download_report(report_name: str) -> FileResponse:
    report_path = (REPORTS_DIR / report_name).resolve()
    if not report_path.exists() or report_path.parent != REPORTS_DIR.resolve():
        raise HTTPException(status_code=404, detail="Report non trovato.")
    media_type, _ = mimetypes.guess_type(report_path.name)
    return FileResponse(
        path=report_path,
        media_type=media_type or "application/octet-stream",
        filename=report_path.name,
    )