from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .runtime_paths import STATIC_DIR


ROME_TZ = ZoneInfo("Europe/Rome")
ASSETS_DIR = STATIC_DIR / "assets"


def _safe_filename(value: str) -> str:
    sanitized = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_")
    return sanitized or "lotto"


def _format_timestamp(value: object) -> str:
    if not value:
        return "-"
    text = str(value)
    try:
        parsed = datetime.fromisoformat(text)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=ROME_TZ)
        else:
            parsed = parsed.astimezone(ROME_TZ)
        return parsed.strftime("%d/%m/%Y %H:%M:%S")
    except ValueError:
        return text


def _add_logo(sheet, asset_name: str, anchor: str, width: int) -> None:
    asset_path = ASSETS_DIR / asset_name
    if not asset_path.exists():
        return

    image = OpenpyxlImage(str(asset_path))
    ratio = image.height / image.width if image.width else 1
    image.width = width
    image.height = int(width * ratio)
    sheet.add_image(image, anchor)


def generate_batch_report(
    reports_dir: Path,
    batch: dict[str, object],
    items: list[dict[str, object]],
    summary: dict[str, object],
) -> Path:
    reports_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(ROME_TZ).strftime("%Y%m%d_%H%M%S")
    filename = f"{_safe_filename(str(batch['filename']))}_{timestamp}.xlsx"
    report_path = reports_dir / filename

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Riepilogo lotto"
    detail_sheet = workbook.create_sheet("Dettaglio pedane")
    completed_sheet = workbook.create_sheet("Pedane completate")
    open_sheet = workbook.create_sheet("Pedane aperte")

    header_fill = PatternFill(fill_type="solid", fgColor="1666C1")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill(fill_type="solid", fgColor="E8F2FF")
    title_fill = PatternFill(fill_type="solid", fgColor="F5F9FF")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    status_fills = {
        "registered": PatternFill(fill_type="solid", fgColor="EEF4FC"),
        "in_progress": PatternFill(fill_type="solid", fgColor="FCE4D6"),
        "waiting_fiche": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        "completed": PatternFill(fill_type="solid", fgColor="DDEBFF"),
    }
    manual_override_fill = PatternFill(fill_type="solid", fgColor="FFF59D")

    summary_sheet.row_dimensions[1].height = 52
    summary_sheet.row_dimensions[4].height = 26
    summary_sheet.row_dimensions[5].height = 22
    summary_sheet.row_dimensions[6].height = 22
    _add_logo(summary_sheet, "logo-san-vincenzo.png", "A1", 92)
    _add_logo(summary_sheet, "logo-magnum.jpg", "H1", 126)
    summary_sheet.merge_cells("C1:G1")
    summary_sheet["C1"] = "San Vincenzo S.R.L. | The Magnum Ice Cream Company"
    summary_sheet["C1"].font = Font(bold=True, size=12)
    summary_sheet["C1"].alignment = Alignment(horizontal="center", vertical="center")
    for cell_range in ["A4:I4", "A5:I5", "A6:I6"]:
        summary_sheet.merge_cells(cell_range)
    summary_sheet["A4"] = "Report lotto riconfezionamento"
    summary_sheet["A4"].font = Font(bold=True, size=16)
    summary_sheet["A4"].fill = title_fill
    summary_sheet["A5"] = f"File: {str(batch.get('filename') or '-')}"
    summary_sheet["A5"].font = Font(bold=True, size=11)
    summary_sheet["A6"] = f"Generato il: {_format_timestamp(datetime.now(ROME_TZ).isoformat(timespec='seconds'))} - Fuso orario: Europe/Rome"
    summary_sheet["A6"].font = Font(italic=True, size=10)
    total_zun = sum(int(item.get("zun_quantity") or 0) for item in items)
    completed_items = [item for item in items if str(item.get("state") or "") == "completed"]
    completed_zun = sum(int(item.get("zun_quantity") or 0) for item in completed_items)
    open_items = [item for item in items if str(item.get("state") or "") != "completed"]
    open_zun = sum(int(item.get("zun_quantity") or 0) for item in open_items)

    summary_rows = [
        ("Importato il", _format_timestamp(batch.get("imported_at") or "-")),
        ("Chiuso il", _format_timestamp(batch.get("completed_at") or "-")),
        ("Pedane totali", summary.get("total_items") or 0),
        ("ZUN totali", total_zun),
        ("Completate", summary.get("completed") or 0),
        ("ZUN completati", completed_zun),
        ("Pedane aperte", len(open_items)),
        ("ZUN aperti", open_zun),
        ("In attesa fiches", summary.get("waiting_fiche") or 0),
        ("Non lavorate", summary.get("registered") or 0),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=9):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=value)
        summary_sheet.cell(row=row_index, column=1).fill = section_fill
        summary_sheet.cell(row=row_index, column=2).alignment = wrap_alignment

    summary_sheet.column_dimensions["A"].width = 22
    summary_sheet.column_dimensions["B"].width = 42

    headers = [
        "Pedana",
        "Prodotto",
        "Codice ingresso",
        "Codice finale",
        "ZUN",
        "Motivo",
        "Fiche entrata",
        "Operatore entrata",
        "Entrata",
        "Stato",
        "Controllo import",
        "Cambio codice",
        "Operatore cambio codice",
        "Postilla",
        "Operatore attesa",
        "Nuova fiche",
        "Operatore uscita",
        "Uscita",
    ]

    def setup_table(sheet, title: str) -> None:
        sheet.row_dimensions[1].height = 48
        sheet.row_dimensions[3].height = 24
        _add_logo(sheet, "logo-san-vincenzo.png", "A1", 82)
        _add_logo(sheet, "logo-magnum.jpg", "K1", 110)
        sheet.merge_cells("C1:J1")
        sheet["C1"] = title
        sheet["C1"].font = Font(bold=True, size=13)
        sheet["C1"].fill = title_fill
        sheet["C1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells("C2:J2")
        sheet["C2"] = f"Generato il: {_format_timestamp(datetime.now(ROME_TZ).isoformat(timespec='seconds'))}"
        sheet["C2"].font = Font(italic=True, size=10)
        sheet["C2"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.append([])
        sheet.append([])
        sheet.append(headers)
        for col_index, _ in enumerate(headers, start=1):
            cell = sheet.cell(row=5, column=col_index)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap_alignment

    setup_table(detail_sheet, "Dettaglio completo pedane")
    setup_table(completed_sheet, "Pedane completate")
    setup_table(open_sheet, "Pedane aperte")

    def append_item(sheet, item: dict[str, object]) -> None:
        sheet.append(
            [
                str(item.get("pallet_code") or "-"),
                str(item.get("product_name") or "-"),
                str(item.get("original_product_code") or item.get("product_code") or "-"),
                str(item.get("product_code") or "-"),
                item.get("zun_quantity") or 0,
                str(item.get("repackaging_reason") or "-"),
                str(item.get("incoming_fiche") or "-"),
                str(item.get("incoming_operator") or "-"),
                _format_timestamp(item.get("scanned_incoming_at") or "-"),
                str(item.get("state") or "-"),
                "Corretto manualmente" if int(item.get("manual_reason_override") or 0) else "-",
                "Cambio codice prodotto" if int(item.get("product_code_changed") or 0) else "-",
                str(item.get("product_code_change_operator") or "-"),
                (
                    f"Cambio codice prodotto autorizzato da {item.get('product_code_change_operator')}"
                    if int(item.get("product_code_changed") or 0) and item.get("product_code_change_operator")
                    else "Cambio codice prodotto autorizzato"
                    if int(item.get("product_code_changed") or 0)
                    else "-"
                ),
                str(item.get("waiting_operator") or "-"),
                str(item.get("outgoing_fiche") or "-"),
                str(item.get("outgoing_operator") or "-"),
                _format_timestamp(item.get("scanned_outgoing_at") or "-"),
            ]
        )
        row_index = sheet.max_row
        fill = manual_override_fill if int(item.get("manual_reason_override") or 0) else status_fills.get(str(item.get("state") or ""))
        for cell in sheet[row_index]:
            cell.alignment = wrap_alignment
            if fill is not None:
                cell.fill = fill

    for item in items:
        append_item(detail_sheet, item)
        if str(item.get("state") or "") == "completed":
            append_item(completed_sheet, item)
        else:
            append_item(open_sheet, item)

    widths = [18, 38, 18, 18, 10, 42, 22, 18, 22, 18, 24, 20, 22, 32, 22, 18, 22]
    for sheet in [detail_sheet, completed_sheet, open_sheet]:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    summary_sheet["A20"] = "Note"
    summary_sheet["A20"].fill = section_fill
    summary_sheet["B20"] = "Colori stato: azzurro registrato, arancio in lavorazione, giallo attesa fiches, blu completato. Le righe corrette manualmente in import sono evidenziate in giallo chiaro."
    summary_sheet["B20"].alignment = wrap_alignment

    for sheet in [detail_sheet, completed_sheet, open_sheet]:
        sheet.freeze_panes = "A6"
        sheet.auto_filter.ref = f"A5:Q{sheet.max_row}"
    summary_sheet.freeze_panes = "A9"

    workbook.save(report_path)
    return report_path