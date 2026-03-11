from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


EXCEL_DIR = Path(__file__).resolve().parent.parent / "generated_excels"


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


def create_quote_excel(quote_row, item_rows) -> Path:
    EXCEL_DIR.mkdir(exist_ok=True)
    excel_path = EXCEL_DIR / f"{quote_row['quote_code']}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Preventivo"

    sheet["A1"] = "Preventivo"
    sheet["A1"].font = Font(size=18, bold=True)

    fields = [
        ("Numero", quote_row["quote_code"]),
        ("Cliente", quote_row["client_name"]),
        ("Referente", quote_row["client_contact_person"]),
        ("Email", quote_row["client_email"]),
        ("Telefono", quote_row["client_phone"]),
        ("Indirizzo", quote_row["client_address"]),
        ("Oggetto", quote_row["title"]),
        ("Descrizione", quote_row["description"]),
        ("Pagamento", quote_row["payment_status"]),
        ("Stato", quote_row["quote_status"]),
        ("Creato il", quote_row["created_at"]),
    ]

    start_row = 3
    for index, (label, value) in enumerate(fields, start=start_row):
        sheet[f"A{index}"] = label
        sheet[f"A{index}"].font = Font(bold=True)
        sheet[f"B{index}"] = value

    items_start_row = start_row + len(fields) + 2
    headers = ["Riga", "Descrizione", "Quantita", "Prezzo unitario", "Totale"]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=items_start_row, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    current_row = items_start_row + 1
    for item_row in item_rows:
        sheet.cell(row=current_row, column=1).value = item_row["line_number"]
        sheet.cell(row=current_row, column=2).value = item_row["description"]
        sheet.cell(row=current_row, column=3).value = item_row["quantity"]
        sheet.cell(row=current_row, column=4).value = item_row["unit_price"]
        sheet.cell(row=current_row, column=5).value = item_row["total_amount"]
        current_row += 1

    if not item_rows:
        sheet.cell(row=current_row, column=2).value = "Nessuna riga dettaglio inserita"
        current_row += 1

    sheet.cell(row=current_row + 1, column=4).value = "Totale"
    sheet.cell(row=current_row + 1, column=4).font = Font(bold=True)
    sheet.cell(row=current_row + 1, column=5).value = quote_row["amount"]
    sheet.cell(row=current_row + 1, column=5).font = Font(bold=True)

    if quote_row["notes"]:
        note_row = current_row + 3
        sheet.cell(row=note_row, column=1).value = "Note"
        sheet.cell(row=note_row, column=1).font = Font(bold=True)
        sheet.cell(row=note_row, column=2).value = quote_row["notes"]

    _apply_sheet_layout(sheet, [16, 44, 12, 16, 16])
    workbook.save(excel_path)
    return excel_path


def create_quotes_registry_excel(quote_rows: Iterable) -> Path:
    EXCEL_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = EXCEL_DIR / f"registro_preventivi_{timestamp}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registro"

    headers = [
        "Numero",
        "Cliente",
        "Referente",
        "Email",
        "Telefono",
        "Oggetto",
        "Importo",
        "Pagamento",
        "Stato",
        "PDF",
        "Excel",
        "Creato il",
    ]

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    total_amount = 0.0
    for row_index, quote_row in enumerate(quote_rows, start=2):
        values = [
            quote_row["quote_code"],
            quote_row["client_name"],
            quote_row["client_contact_person"],
            quote_row["client_email"],
            quote_row["client_phone"],
            quote_row["title"],
            float(quote_row["amount"]),
            quote_row["payment_status"],
            quote_row["quote_status"],
            "Si" if quote_row["pdf_path"] else "No",
            "Si" if quote_row["excel_path"] else "No",
            quote_row["created_at"],
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index).value = value
        total_amount += float(quote_row["amount"])

    total_row = sheet.max_row + 2
    sheet.cell(row=total_row, column=6).value = "Totale importi"
    sheet.cell(row=total_row, column=6).font = Font(bold=True)
    sheet.cell(row=total_row, column=7).value = total_amount
    sheet.cell(row=total_row, column=7).font = Font(bold=True)

    _apply_sheet_layout(sheet, [16, 24, 20, 28, 18, 28, 14, 14, 14, 10, 10, 22])
    workbook.save(excel_path)
    return excel_path


def _apply_sheet_layout(sheet, widths) -> None:
    for column_letter, width in zip("ABCDEFGHIJKLMNOPQRSTUVWXYZ", widths):
        sheet.column_dimensions[column_letter].width = width
