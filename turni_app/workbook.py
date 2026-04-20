from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
import re
import shutil
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


WEEKLY_SHEET_NAME = " Turni settimanali"
WEEKEND_SHEET_NAMES = (
    "Comandata pulizie Sabato",
    "Comandata pulizie Domenica",
)
WEEKLY_COLUMN_RANGE = range(3, 13)
WEEKLY_HEADER_ROW = 5
WEEKLY_LABEL_CELL = "B3"
WEEKLY_SIGNATURE_CELL = "B18"
WEEKEND_BASE_DATE_CELL = "C4"
WEEKEND_EDITABLE_COLUMNS = (2, 3, 4, 5, 6, 7)
WEEKEND_COLUMN_LABELS = ("Data", "Turno", "Nominativo", "Preposto", "Attivita", "Reparto")
WEEKEND_FOOTER_MARKER = "Firma per Autorizzazione servizi generali"
KEYWORD_EXCLUSIONS = {
    "SCORRIMENTO",
    "CARRELLISTI",
    "PULIZIA FABBRICA",
    "SPOGLIATOIO",
    "PALAZZINA",
    "LAVORI VARI",
    "PRODUZIONE",
    "DE NIGRIS",
    "NAVETTA",
    "AIMPIANTO",
    "TUNNEL",
    "BULK",
    "FREEZER",
    "BALLATOIO",
    "B.S.C.M.",
    "STANZETTE",
    "LARA",
    "GIARDINAGGIO",
    "M.M.D. BUILER",
    "M.M.D BUILER",
}


@dataclass(frozen=True)
class ShiftSection:
    label: str
    time_row: int
    assignment_rows: tuple[int, int, int]


@dataclass(frozen=True)
class WeeklySectionData:
    label: str
    time_label: str
    time_values: list[str]
    rows: list[list[str]]


@dataclass(frozen=True)
class CellSnapshot:
    raw: Any
    display: str


@dataclass(frozen=True)
class WeekendRowData:
    row_number: int
    cells: list[CellSnapshot]


@dataclass(frozen=True)
class WeekendSheetData:
    name: str
    base_date: CellSnapshot
    rows: list[WeekendRowData]


SHIFT_SECTIONS = (
    ShiftSection(label="1 turno", time_row=6, assignment_rows=(7, 8, 9)),
    ShiftSection(label="2 turno", time_row=10, assignment_rows=(11, 12, 13)),
    ShiftSection(label="3 turno", time_row=14, assignment_rows=(15, 16, 17)),
    ShiftSection(label="4 turno", time_row=23, assignment_rows=(24, 25, 26)),
)


def format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).replace("\n", " ").strip()


def looks_like_person_name(value: str) -> bool:
    text = value.strip().upper()
    if not text or text in KEYWORD_EXCLUSIONS:
        return False
    if any(char.isdigit() for char in text):
        return False
    if text.startswith("FIRMA") or text.startswith("TELEFONO"):
        return False
    if "SERVIZI GENERALI" in text or "ORGANIZZAZIONE TURNI" in text:
        return False
    if len(text) > 28:
        return False
    return bool(re.fullmatch(r"[A-Z' .]+", text))


class TurniWorkbook:
    def __init__(self, path: Path | str):
        self.path = Path(path)
        self.workbook: Workbook | None = None
        self.value_workbook: Workbook | None = None
        self.load()

    def load(self) -> None:
        if self.workbook is not None:
            try:
                self.workbook.close()
            except Exception:
                pass
        if self.value_workbook is not None:
            try:
                self.value_workbook.close()
            except Exception:
                pass
        self.workbook = load_workbook(self.path, data_only=False)
        self.value_workbook = load_workbook(self.path, data_only=True)
        self._validate()

    def _validate(self) -> None:
        available = set(self.workbook.sheetnames)
        required = {WEEKLY_SHEET_NAME, *WEEKEND_SHEET_NAMES}
        missing = sorted(required - available)
        if missing:
            raise ValueError(f"Workbook non valido. Mancano i fogli: {', '.join(missing)}")

    @property
    def weekly_sheet(self) -> Worksheet:
        return self.workbook[WEEKLY_SHEET_NAME]

    @property
    def weekly_value_sheet(self) -> Worksheet:
        return self.value_workbook[WEEKLY_SHEET_NAME]

    def weekly_title(self) -> str:
        return format_cell_value(self.weekly_value_sheet[WEEKLY_LABEL_CELL].value)

    def signature(self) -> str:
        return format_cell_value(self.weekly_value_sheet[WEEKLY_SIGNATURE_CELL].value)

    def department_headers(self) -> list[str]:
        return [
            format_cell_value(self.weekly_value_sheet.cell(WEEKLY_HEADER_ROW, column).value)
            for column in WEEKLY_COLUMN_RANGE
        ]

    def weekly_sections(self) -> list[WeeklySectionData]:
        sections: list[WeeklySectionData] = []
        value_sheet = self.weekly_value_sheet
        for section in SHIFT_SECTIONS:
            time_values = [
                format_cell_value(value_sheet.cell(section.time_row, column).value)
                for column in WEEKLY_COLUMN_RANGE
            ]
            rows: list[list[str]] = []
            for row_number in section.assignment_rows:
                rows.append(
                    [
                        format_cell_value(value_sheet.cell(row_number, column).value)
                        for column in WEEKLY_COLUMN_RANGE
                    ]
                )
            time_label = format_cell_value(value_sheet.cell(section.time_row, 3).value)
            sections.append(
                WeeklySectionData(
                    label=section.label,
                    time_label=time_label,
                    time_values=time_values,
                    rows=rows,
                )
            )
        return sections

    def weekend_sheets(self) -> list[WeekendSheetData]:
        result: list[WeekendSheetData] = []
        for sheet_name in WEEKEND_SHEET_NAMES:
            raw_sheet = self.workbook[sheet_name]
            value_sheet = self.value_workbook[sheet_name]
            footer_row = self._find_footer_row(value_sheet)
            base_date = CellSnapshot(
                raw=raw_sheet[WEEKEND_BASE_DATE_CELL].value,
                display=format_cell_value(value_sheet[WEEKEND_BASE_DATE_CELL].value),
            )
            rows: list[WeekendRowData] = []
            for row_number in range(6, footer_row):
                cells: list[CellSnapshot] = []
                for column in WEEKEND_EDITABLE_COLUMNS:
                    raw_value = raw_sheet.cell(row_number, column).value
                    display_value = value_sheet.cell(row_number, column).value
                    cells.append(CellSnapshot(raw=raw_value, display=format_cell_value(display_value)))
                rows.append(WeekendRowData(row_number=row_number, cells=cells))
            result.append(WeekendSheetData(name=sheet_name, base_date=base_date, rows=rows))
        return result

    def people_palette(self) -> list[str]:
        values: set[str] = set()
        for section in self.weekly_sections():
            for row in section.rows:
                for cell_value in row:
                    if looks_like_person_name(cell_value):
                        values.add(cell_value.strip())
        for sheet in self.weekend_sheets():
            for row in sheet.rows:
                nominativo = row.cells[2].display
                preposto = row.cells[3].display
                if looks_like_person_name(nominativo):
                    values.add(nominativo.strip())
                if looks_like_person_name(preposto):
                    values.add(preposto.strip())
        return sorted(values)

    def summary(self) -> dict[str, Any]:
        departments = [header for header in self.department_headers() if header]
        weekend_rows = sum(
            1
            for sheet in self.weekend_sheets()
            for row in sheet.rows
            if any(cell.display for cell in row.cells)
        )
        return {
            "week": self.weekly_title(),
            "signature": self.signature(),
            "departments": len(departments),
            "people": len(self.people_palette()),
            "weekend_rows": weekend_rows,
        }

    def save(
        self,
        *,
        week_label: str,
        signature: str,
        department_headers: list[str],
        weekly_time_values: list[list[str]],
        weekly_sections: list[list[list[str]]],
        weekend_values: dict[str, list[list[str]]],
        weekend_base_dates: dict[str, str],
        backup: bool = False,
    ) -> Path | None:
        backup_path: Path | None = None
        if backup:
            backup_path = self._create_backup()

        self.weekly_sheet[WEEKLY_LABEL_CELL] = week_label.strip() or None
        self.weekly_sheet[WEEKLY_SIGNATURE_CELL] = signature.strip() or None
        for local_column_index, column_number in enumerate(WEEKLY_COLUMN_RANGE):
            header_value = department_headers[local_column_index].strip()
            self.weekly_sheet.cell(WEEKLY_HEADER_ROW, column_number).value = header_value or None

        for section_index, section in enumerate(SHIFT_SECTIONS):
            for local_column_index, column_number in enumerate(WEEKLY_COLUMN_RANGE):
                time_value = weekly_time_values[section_index][local_column_index].strip()
                original_value = self.weekly_sheet.cell(section.time_row, column_number).value
                original_display = format_cell_value(self.weekly_value_sheet.cell(section.time_row, column_number).value)
                self.weekly_sheet.cell(section.time_row, column_number).value = self._parse_text_value(
                    time_value,
                    original_value,
                    original_display,
                )
            for local_row_index, row_number in enumerate(section.assignment_rows):
                for local_column_index, column_number in enumerate(WEEKLY_COLUMN_RANGE):
                    value = weekly_sections[section_index][local_row_index][local_column_index].strip()
                    self.weekly_sheet.cell(row_number, column_number).value = value or None

        weekend_data = {sheet.name: sheet for sheet in self.weekend_sheets()}
        for sheet_name, rows in weekend_values.items():
            raw_sheet = self.workbook[sheet_name]
            original_rows = weekend_data[sheet_name].rows
            base_date_text = weekend_base_dates.get(sheet_name, "")
            raw_sheet[WEEKEND_BASE_DATE_CELL] = self._parse_text_value(
                base_date_text,
                weekend_data[sheet_name].base_date.raw,
                weekend_data[sheet_name].base_date.display,
            )
            for row_index, row_values in enumerate(rows):
                original_row = original_rows[row_index]
                for column_index, column_number in enumerate(WEEKEND_EDITABLE_COLUMNS):
                    original_cell = original_row.cells[column_index]
                    display_value = row_values[column_index].strip()
                    raw_sheet.cell(original_row.row_number, column_number).value = self._parse_text_value(
                        display_value,
                        original_cell.raw,
                        original_cell.display,
                    )

        if self.value_workbook is not None:
            try:
                self.value_workbook.close()
            except Exception:
                pass
            self.value_workbook = None
        self.workbook.save(self.path)
        self.load()
        return backup_path

    def _find_footer_row(self, sheet: Worksheet) -> int:
        for row_number in range(sheet.max_row, 0, -1):
            text = format_cell_value(sheet.cell(row_number, 2).value)
            if WEEKEND_FOOTER_MARKER in text:
                return row_number
        return sheet.max_row + 1

    def _create_backup(self) -> Path:
        backup_dir = self.path.parent / "_backup_turni"
        backup_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{self.path.stem}_{timestamp}{self.path.suffix}"
        shutil.copy2(self.path, backup_path)
        return backup_path

    def _parse_text_value(self, text: str, original_value: Any, original_display: str) -> Any:
        clean_text = text.strip()
        if clean_text == "":
            return None
        if clean_text == original_display.strip():
            return original_value
        if clean_text.startswith("="):
            return clean_text
        if isinstance(original_value, datetime):
            parsed = self._parse_datetime(clean_text)
            return parsed if parsed is not None else clean_text
        if isinstance(original_value, time):
            parsed = self._parse_time(clean_text)
            return parsed if parsed is not None else clean_text
        if re.fullmatch(r"\d{2}/\d{2}/\d{4}", clean_text) or re.fullmatch(r"\d{4}-\d{2}-\d{2}", clean_text):
            parsed = self._parse_datetime(clean_text)
            if parsed is not None:
                return parsed
        if re.fullmatch(r"\d{1,2}:\d{2}", clean_text):
            parsed = self._parse_time(clean_text)
            if parsed is not None:
                return parsed
        return clean_text

    @staticmethod
    def _parse_datetime(text: str) -> datetime | None:
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_time(text: str) -> time | None:
        for fmt in ("%H:%M", "%H.%M"):
            try:
                return datetime.strptime(text, fmt).time()
            except ValueError:
                continue
        return None